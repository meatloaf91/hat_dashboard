import os
import re
from datetime import datetime

import pandas as pd
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import stylesheet as ss


class PackshotNamingGenerator:
    def __init__(self, ui):
        self.ui = ui

        # PNM main panel controls
        self.radio_from_clipboard = getattr(ui, "radioButton_pnm_from_clipboard", None)
        self.radio_from_tracker = getattr(ui, "radioButton_pnm_generate_from_tracker", None)
        self.btn_paste_on_table = getattr(ui, "btn_pnm_paste_on_table", None)
        self.btn_excel_tracker = getattr(ui, "btn_pg6_excel_tracker", None)
        self.input_excel_tracker = getattr(ui, "input_pg6_excel_tracker", None)
        self.btn_output_location = getattr(ui, "btn_pg6_output", None)
        self.input_output_location = getattr(ui, "input_pg6_output", None)
        self.btn_run_process = getattr(ui, "btn_run_process_generate_packshot_naming", None)

        # Clipboard table dialog controls
        self.btn_generate_packshot_naming = getattr(ui, "btn_generate_packshot_naming", None)
        self.btn_import_from_tracker = getattr(ui, "btn_import_from_tracker", None)
        self.btn_export_table = getattr(ui, "btn_export_table", None)

        self.excel_values = GetExcelValues()
        self.error_checker = CheckErrors()

        self.row_threshold = 25
        self.name_variations_idh = ["idh number", "idh", "idh num", "idh no.", "idh no"]
        self.name_variations_product_name = [
            "variant name / base product name",
            "base product name",
            "base product",
            "variant name",
            "product name",
            "product",
            "variant",
        ]
        self.name_variations_packaging_type = ["packaging type", "packaging"]
        self.name_variations_packaging_size = ["capacity", "packaging size", "size"]
        self.name_variations_view = ["view", "views","packshot view", "angle"]

        self.column_name_variations = {
            "idh": self.name_variations_idh,
            "product_name": self.name_variations_product_name,
            "packaging_type": self.name_variations_packaging_type,
            "packaging_size": self.name_variations_packaging_size,
            "view": self.name_variations_view,
        }

    def run_process(self) -> None:
        if self.btn_excel_tracker is not None:
            self.btn_excel_tracker.clicked.connect(self.name_gen_get_tracker_file)
        if self.btn_output_location is not None:
            self.btn_output_location.clicked.connect(self.name_gen_get_output_location)
        if self.btn_run_process is not None:
            self.btn_run_process.clicked.connect(self.name_gen_run_process)

    def attach_table_dialog(self, dialog) -> None:
        """Attach clipboard table dialog actions to module logic."""
        self.btn_generate_packshot_naming = getattr(dialog, "btn_generate_packshot_naming", None)
        self.btn_import_from_tracker = getattr(dialog, "btn_import_from_tracker", None)
        self.btn_export_table = getattr(dialog, "btn_export_table", None)
        if self.btn_generate_packshot_naming is not None:
            self.btn_generate_packshot_naming.clicked.connect(self.generate_from_clipboard_table)
        if self.btn_import_from_tracker is not None:
            _start = getattr(dialog, "_browse_start_dir", "")
            self.btn_import_from_tracker.clicked.connect(lambda: self.import_from_tracker_to_table(_start))
        if self.btn_export_table is not None:
            self.btn_export_table.clicked.connect(self.export_table_to_excel)

    def _show_alert(self, messages, title="Alert") -> None:
        if isinstance(messages, list):
            text = "\n".join([f"- {msg}" for msg in messages if msg])
        else:
            text = str(messages)

        msg = QMessageBox()
        msg.setWindowTitle(title)
        msg.setText(text)
        msg.setIcon(QMessageBox.Icon.Warning)
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def name_gen_get_tracker_file(self, start_dir: str = ""):
        self.excel_values.fn_get_excel_file(start_dir=start_dir)
        if self.excel_values.check_tracker_loaded and self.input_excel_tracker is not None:
            self.input_excel_tracker.setText(self.excel_values.out_excel_file_name)

    def name_gen_get_output_location(self, start_dir: str = ""):
        path = QFileDialog.getExistingDirectory(
            self.btn_run_process,
            "Select Output Location",
            start_dir,
            QFileDialog.Option.ShowDirsOnly,
        )
        if path and self.input_output_location is not None:
            self.input_output_location.setText(path)

    def name_gen_run_process(self):
        # Tracker mode only (Run Process button is disabled in clipboard mode by UI logic).
        errors = []
        if not self.excel_values.check_tracker_loaded or not self.excel_values.out_excel_file:
            errors.append("Tracker excel file missing")

        output_dir = self.input_output_location.text().strip() if self.input_output_location is not None else ""
        if not output_dir:
            errors.append("Output location missing")

        if errors:
            self._show_alert(errors)
            return

        records, missing = self._extract_records_from_tracker(self.excel_values.out_excel_file, starting_row=1)
        if missing:
            self._show_alert(missing, title="Error")
            return

        records = [record for record in records if not self._should_skip_table_import_row(record)]
        if not records:
            self._show_alert("No usable rows found in tracker.", title="Information")
            return

        headers = [
            "Product Name",
            "IDH",
            "Packaging Type",
            "Packaging Size",
            "View",
            "Packshot Naming",
        ]

        rows_to_export = []
        for rec in records:
            generated_name = self._compose_packshot_name(rec)
            rows_to_export.append(
                [
                    str(rec.get("product_name_raw", "") or ""),
                    str(rec.get("idh_raw", "") or ""),
                    str(rec.get("pack_type_raw", "") or ""),
                    str(rec.get("pack_size_raw", "") or ""),
                    str(rec.get("view_raw", "") or ""),
                    generated_name,
                ]
            )

        output_path = os.path.join(output_dir, self._build_packshot_naming_filename())
        self._write_styled_table_excel(output_path=output_path, headers=headers, rows=rows_to_export, sheet_title="Packshot Table")
        self._show_styled_info("table successfully exported")

    def generate_from_clipboard_table(self):
        dialog = getattr(self.ui, "_packshot_clipboard_dialog", None)
        if dialog is None or not hasattr(dialog, "table"):
            self._show_alert("Open 'Paste on Table' first.", title="Information")
            return

        table = dialog.table
        for row in range(table.rowCount()):
            product_name = self._get_table_text(table, row, 0)
            idh = self._get_table_text(table, row, 1)
            pack_type = self._get_table_text(table, row, 2)
            pack_size = self._get_table_text(table, row, 3)
            view = self._get_table_text(table, row, 4)

            if not any([product_name, idh, pack_type, pack_size, view]):
                continue

            record = {
                "product_name_raw": product_name,
                "idh_raw": idh,
                "pack_type_raw": pack_type,
                "pack_size_raw": pack_size,
                "view_raw": view,
            }
            generated_name = self._compose_packshot_name(record)
            item_col5 = table.item(row, 5)
            if item_col5 is None:
                from PySide6.QtWidgets import QTableWidgetItem as _TWI
                item_col5 = _TWI("")
                table.setItem(row, 5, item_col5)
            item_col5.setText(generated_name)

    def _get_table_text(self, table, row: int, col: int) -> str:
        item = table.item(row, col)
        return item.text().strip() if item is not None else ""

    def _show_info(self, text: str, title: str = "Done") -> None:
        msg = QMessageBox()
        msg.setWindowTitle(title)
        msg.setText(text)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def _show_styled_info(self, message: str, parent=None) -> None:
        msg = QMessageBox(parent or self.btn_export_table)
        msg.setWindowTitle("Information")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setTextFormat(Qt.TextFormat.RichText)
        msg.setStyleSheet(ss.msg_stylesheet)
        msg.setText(
            "<p style='color:#FFFFFF; font-family:\"Inter\"; font-weight:bold; "
            "font-size:12px; margin-bottom:6px;'>"
            f"{message}</p>"
        )
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def export_table_to_excel(self) -> None:
        dialog = getattr(self.ui, "_packshot_clipboard_dialog", None)
        if dialog is None or not hasattr(dialog, "table"):
            self._show_alert("Open 'Paste on Table' first.", title="Information")
            return

        table = dialog.table
        headers = [
            "Product Name",
            "IDH",
            "Packaging Type",
            "Packaging Size",
            "View",
            "Packshot Naming",
        ]

        rows_to_export = []
        for row in range(table.rowCount()):
            values = [self._get_table_text(table, row, col) for col in range(table.columnCount())]
            if any(values):
                rows_to_export.append(values)

        if not rows_to_export:
            self._show_alert("No table data to export.", title="Information")
            return

        output_path, _ = QFileDialog.getSaveFileName(
            self.btn_export_table,
            "Save Exported Table",
            self._build_packshot_naming_filename(),
            "Excel Files (*.xlsx)",
        )
        if not output_path:
            return

        if not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"

        self._write_styled_table_excel(output_path=output_path, headers=headers, rows=rows_to_export, sheet_title="Packshot Table")

        self._show_styled_info("table successfully exported")

    def _build_packshot_naming_filename(self) -> str:
        timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
        return f"packshot_naming_{timestamp}.xlsx"

    def _write_styled_table_excel(self, output_path: str, headers: list[str], rows: list[list[str]], sheet_title: str = "Packshot Table") -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_title

        worksheet.append(headers)
        for row_values in rows:
            worksheet.append(row_values)

        header_fill = PatternFill(fill_type="solid", fgColor="111F35")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        body_font = Font(name="Calibri", size=11, color="1F2937")
        border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        band_fill = PatternFill(fill_type="solid", fgColor="F7F9FC")

        for col in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in range(2, worksheet.max_row + 1):
            is_banded = row % 2 == 0
            for col in range(1, len(headers) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = body_font
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = border
                if is_banded:
                    cell.fill = band_fill

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=len(headers)).coordinate}"

        for col_cells in worksheet.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 48)

        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        workbook.save(output_path)

    def import_from_tracker_to_table(self, start_dir: str = ""):
        dialog = getattr(self.ui, "_packshot_clipboard_dialog", None)
        if dialog is None or not hasattr(dialog, "table"):
            self._show_alert("Open 'Paste on Table' first.", title="Information")
            return

        starting_row = 1
        if hasattr(dialog, "starting_row_input"):
            raw_starting_row = dialog.starting_row_input.text().strip()
            try:
                starting_row = max(1, int(raw_starting_row))
            except ValueError:
                starting_row = 1
                dialog.starting_row_input.setText("1")

        pick = QFileDialog()
        pick.setWindowTitle("Select Excel File")
        pick.setNameFilter("Excel Files (*.xlsx *.xlsm *.xltx *.xltm)")
        pick.setFileMode(QFileDialog.FileMode.ExistingFile)
        if start_dir:
            pick.setDirectory(start_dir)

        if not pick.exec():
            return

        selected = pick.selectedFiles()
        if not selected:
            return

        tracker_path = selected[0]
        records, missing_report = self._extract_records_from_tracker(tracker_path, starting_row=starting_row)
        if missing_report:
            self._show_alert(missing_report, title="Error")
            return

        records = [record for record in records if not self._should_skip_table_import_row(record)]

        if not records:
            self._show_alert("No usable rows found in tracker.", title="Information")
            return

        table = dialog.table
        row_count = len(records)
        table.clearSelection()
        table.clearContents()
        table.setRowCount(0)  # hard reset old inputs/rows before importing
        table.setRowCount(row_count)

        for row_idx, rec in enumerate(records):
            values = [
                str(rec.get("product_name_raw", "") or ""),
                str(rec.get("idh_raw", "") or ""),
                str(rec.get("pack_type_raw", "") or ""),
                str(rec.get("pack_size_raw", "") or ""),
                str(rec.get("view_raw", "") or ""),
                "",
            ]
            for col_idx, value in enumerate(values):
                cell = table.item(row_idx, col_idx)
                if cell is None:
                    from PySide6.QtWidgets import QTableWidgetItem
                    cell = QTableWidgetItem("")
                    table.setItem(row_idx, col_idx, cell)
                cell.setText(value)

        if hasattr(dialog, "row_count_input"):
            dialog.row_count_input.setText(str(row_count))

        # Keep selected tracker in module state and panel field for continuity.
        self.excel_values.out_excel_file = tracker_path
        self.excel_values.out_excel_file_name = os.path.basename(tracker_path)
        self.excel_values.check_tracker_loaded = True
        if self.input_excel_tracker is not None:
            self.input_excel_tracker.setText(self.excel_values.out_excel_file_name)

        self._show_info(f"Imported {row_count} row(s) from tracker.")

    def _should_skip_table_import_row(self, record: dict) -> bool:
        product_name = str(record.get("product_name_raw", "") or "").strip()
        idh = str(record.get("idh_raw", "") or "").strip()

        if idh == "":
            return True

        if product_name.lower() == "admin" or idh.lower() == "admin":
            return True

        return False

    def _normalize_general(self, value) -> str:
        text = str(value or "").strip().lower()
        text = re.sub(r"[^a-z0-9]+", "_", text)
        return re.sub(r"_+", "_", text).strip("_")

    def _normalize_pack_type(self, value) -> str:
        """Apply special packaging-type rules before general normalisation."""
        raw = str(value or "").strip()
        lower = raw.lower()
        # Rule: ignore un-identifiable types
        if "other pack" in lower or "not spec" in lower:
            return ""
        # Rule: 'Set (1 cartridge)' -> 'cartridge'
        if re.search(r"set\s*\(.*cartridge.*\)", lower):
            return "cartridge"
        return self._normalize_general(raw)

    def _normalize_pack_size(self, value) -> str:
        text = str(value or "").strip().lower()
        # Packaging size rule: no spaces; keep punctuation as-is (e.g., 2.80g, 2x12ml).
        return re.sub(r"\s+", "", text)

    def _compose_packshot_name(self, record: dict) -> str:
        product_name = self._normalize_general(record.get("product_name_raw", ""))
        idh = self._normalize_general(record.get("idh_raw", ""))
        pack_type = self._normalize_pack_type(record.get("pack_type_raw", ""))
        pack_size_raw = self._normalize_pack_size(record.get("pack_size_raw", ""))
        # Rule: ignore packaging size if value is '0'
        pack_size = "" if pack_size_raw == "0" else pack_size_raw
        view = self._normalize_general(record.get("view_raw", ""))

        combined = "_".join(part for part in [product_name, idh, pack_type, pack_size, view] if part)
        combined = re.sub(r"_+", "_", combined).strip("_")
        return combined

    def _normalize_header(self, value) -> str:
        return str(value or "").strip().lower()

    def _find_col(self, columns, variations) -> str | None:
        normalized = {self._normalize_header(c): c for c in columns}
        for var in variations:
            key = self._normalize_header(var)
            if key in normalized:
                return normalized[key]
        return None

    def _detect_header_row(self, tracker_path: str, sheet_name: str) -> int | None:
        preview = pd.read_excel(
            tracker_path,
            sheet_name=sheet_name,
            nrows=self.row_threshold,
            header=None,
            dtype=str,
            na_filter=False,
            engine="openpyxl",
        )
        all_variations = (
            self.name_variations_product_name
            + self.name_variations_idh
            + self.name_variations_packaging_type
            + self.name_variations_packaging_size
            + self.name_variations_view
        )
        expected = {self._normalize_header(v) for v in all_variations}

        for row_idx in range(len(preview)):
            row_vals = {self._normalize_header(v) for v in preview.iloc[row_idx].tolist() if str(v).strip()}
            matches = row_vals.intersection(expected)
            if len(matches) >= 2:
                return row_idx
        return None

    def _extract_records_from_tracker(self, tracker_path: str, starting_row: int = 1) -> tuple[list[dict], list[str]]:
        records = []
        if not tracker_path:
            return records, ["No tracker file selected."]

        starting_row = max(1, int(starting_row))

        sheet_names = pd.ExcelFile(tracker_path, engine="openpyxl").sheet_names
        target_sheets = [s for s in sheet_names if "tracker" in s.lower()] or [sheet_names[0]]
        required_labels = {
            "product_name": "Product Name",
            "idh": "IDH",
            "packaging_type": "Packaging Type",
            "packaging_size": "Packaging Size",
        }

        missing_report = []

        for sheet in target_sheets:
            header_row = self._detect_header_row(tracker_path, sheet)
            if header_row is None:
                missing_report.append(f"{sheet}: Unable to detect headers")
                continue

            df = pd.read_excel(
                tracker_path,
                sheet_name=sheet,
                header=header_row,
                dtype=str,
                na_filter=False,
                engine="openpyxl",
            )
            df.columns = [self._normalize_header(c) for c in df.columns]

            col_product = self._find_col(df.columns, self.column_name_variations["product_name"])
            col_idh = self._find_col(df.columns, self.column_name_variations["idh"])
            col_pack_type = self._find_col(df.columns, self.column_name_variations["packaging_type"])
            col_pack_size = self._find_col(df.columns, self.column_name_variations["packaging_size"])
            col_view = self._find_col(df.columns, self.column_name_variations["view"])

            missing_cols = []
            if col_product is None:
                missing_cols.append(required_labels["product_name"])
            if col_idh is None:
                missing_cols.append(required_labels["idh"])
            if col_pack_type is None:
                missing_cols.append(required_labels["packaging_type"])
            if col_pack_size is None:
                missing_cols.append(required_labels["packaging_size"])
            if missing_cols:
                missing_report.append(f"{sheet}: missing {', '.join(missing_cols)}")
                continue

            for excel_row_num, (_, row) in enumerate(df.iterrows(), start=header_row + 2):
                if excel_row_num < starting_row:
                    continue

                product_name = row.get(col_product, "")
                idh = row.get(col_idh, "")
                pack_type = row.get(col_pack_type, "")
                pack_size = row.get(col_pack_size, "")
                view = row.get(col_view, "") if col_view is not None else "front"

                if all(str(v).strip() == "" for v in [product_name, idh, pack_type, pack_size, view]):
                    continue

                records.append(
                    {
                        "product_name_raw": product_name,
                        "idh_raw": idh,
                        "pack_type_raw": pack_type,
                        "pack_size_raw": pack_size,
                        "view_raw": view,
                    }
                )

        if missing_report:
            formatted = ["Tracker sheets w/ missing required column(s):", ""]
            file_name = os.path.basename(tracker_path)
            formatted.append(f"{file_name}:")
            for line in missing_report:
                formatted.append(f"\t{line}")
            return [], formatted

        return records, []

    def _get_records_from_tracker(self) -> list[dict]:
        tracker_path = self.excel_values.out_excel_file
        records, missing = self._extract_records_from_tracker(tracker_path)
        if missing:
            self._show_alert(missing, title="Error")
            return []
        return records

    def _write_packshot_excel(self, idhs, packnames_final, output_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Packshot Names"
        ws.append(["IDH", "Packshot Naming"])

        for idh, name in zip(idhs, packnames_final):
            ws.append([idh, name])

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)

    # Old tab flow kept as comments/reference in git history; replaced by normalized pipeline above.


class GetExcelValues:
    def __init__(self):
        self.out_column_values = []  # output, all values in multiple columns
        self.out_multiple_column_numbers = []
        self.out_excel_file = None
        self.out_excel_file_name = ""
        self.out_error_message = ""
        self.check_tracker_loaded = False

        self.name_gen_capacity_col = []

    def fn_get_excel_file(self, start_dir: str = ""):
        dialog = QFileDialog()
        dialog.setWindowTitle("Select Excel File")
        dialog.setNameFilter("Excel Files (*.xlsx *.xlsm *.xltx *.xltm)")
        dialog.setFileMode(QFileDialog.FileMode.ExistingFile)
        if start_dir:
            dialog.setDirectory(start_dir)

        if dialog.exec():
            file_paths = dialog.selectedFiles()
            self.out_excel_file = file_paths[0]
            self.out_excel_file_name = os.path.basename(self.out_excel_file)
            self.check_tracker_loaded = True

            return self.out_excel_file
        self.out_error_message = "No file selected."
        self.fn_show_alert()
        return []

    def fn_get_values_of_multiple_columns(self, list_of_columns, minrow, maxrow):
        if not self.out_excel_file:
            self.out_error_message = "No excel file to evaluate."
            self.fn_show_alert()
            return []

        self.out_column_values = []
        wb = load_workbook(self.out_excel_file, data_only=True)
        ws = wb.active

        for col in list_of_columns:
            data_collected_from_column = []

            for row in ws.iter_rows(min_col=col, max_col=col, min_row=minrow, max_row=maxrow):
                cell = row[0]
                if cell.value is None or str(cell.value).strip() == "":
                    data_collected_from_column.append("undefined")
                else:
                    val = str(cell.value).replace(" ", "").lower()
                    val = val.replace(",", ".")
                    data_collected_from_column.append(val)

            self.out_column_values.append(data_collected_from_column)

        return self.out_column_values

    def fn_multiple_letters_to_numbers(self, list_of_letters):
        self.out_multiple_column_numbers = []
        for letter in list_of_letters:
            letter = letter.upper()
            result = 0
            for char in letter:
                if "A" <= char <= "Z":
                    result = result * 26 + (ord(char) - ord("A") + 1)
                else:
                    raise ValueError(f"Invalid column letter: {letter}")

            self.out_multiple_column_numbers.append(result)

        return self.out_multiple_column_numbers

    def fn_show_alert(self, title="Alert"):
        msg_box = QMessageBox()
        msg_box.setWindowTitle(title)
        msg_box.setText(self.out_error_message)
        msg_box.setIcon(QMessageBox.Icon.Warning)
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.exec()

    # Fix and choose which one of these 2 functions you will use
    def fn_name_gen_get_values_of_mult_columns(self, list_of_columns, minrow, maxrow):
        if not self.out_excel_file:
            self.out_error_message = "No excel file to evaluate."
            self.fn_show_alert()
            return []

        self.out_column_values = []
        wb = load_workbook(self.out_excel_file, data_only=True)
        ws = wb.active

        for col in list_of_columns:
            data_collected_from_column = []

            for row in ws.iter_rows(min_col=col, max_col=col, min_row=minrow, max_row=maxrow):
                cell = row[0]
                raw_val = str(cell.value).strip() if cell.value is not None else ""
                lower_raw_val = raw_val.lower()

                if lower_raw_val == "":
                    data_collected_from_column.append("_")
                else:
                    val = str(cell.value)

                    r_val = re.sub(r"[®™@*]", "_", val)  # reformatted value
                    r_val = r_val.replace(",", ".")
                    r_val = r_val.replace("/", "_")

                    data_collected_from_column.append(r_val)

            self.out_column_values.append(data_collected_from_column)

        return self.out_column_values

    def fn_name_gen_fix_values(self, list_of_columns, minrow, maxrow):
        self.out_column_values = []

        if not self.out_excel_file:
            self.out_error_message = "No excel file to evaluate."
            self.fn_show_alert()
            return []

        wb = load_workbook(self.out_excel_file, data_only=True)
        ws = wb.active
        excel_errors = {"#N/A", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NULL!", "#NUM!"}

        for col in list_of_columns:
            data_collected_from_column = []

            for row in ws.iter_rows(min_col=col, max_col=col, min_row=minrow, max_row=maxrow):
                cell = row[0]
                cell_val = str(cell.value).strip().upper() if cell.value is not None else ""

                if cell_val in excel_errors or cell_val == "":
                    data_collected_from_column.append("")
                else:
                    val = str(cell.value).lower()

                    if val.strip() == "otherpack.types,notspec.":  # Other pack. types, not spec.
                        val = ""

                    r_val = re.sub(r"[®™@*]", "", val)
                    r_val = r_val.replace(",", ".")
                    r_val = r_val.replace("/", "_")
                    r_val = r_val.replace("-", "_")
                    r_val = r_val.replace("#n/a", "_")
                    r_val = r_val.replace("(", "").replace(")", "")
                    r_val = r_val.replace(" ", "_")
                    data_collected_from_column.append(r_val)

            self.out_column_values.append(data_collected_from_column)

        return self.out_column_values


def fn_fix_capacity(value_to_evaluate):
    value_to_evaluate = value_to_evaluate.replace("_", "")
    return value_to_evaluate


class CheckErrors:
    def __init__(self):
        self.input_error_count = 0
        self.input_error_descriptions = []
        self.can_proceed = True

    def _get_text(self, value):
        if hasattr(value, "text") and callable(value.text):
            return value.text().strip()
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def check_if_value_is_a_letter(self, value_to_evaluate, error_message):
        """
        Returns the value if it's a letter and "" if it's not.
        """
        text = self._get_text(value_to_evaluate)

        if text.isalpha():
            return text
        self.input_error_descriptions.append(error_message)
        self.input_error_count += 1
        return "_"

    def check_if_value_is_nonzero_digit(self, value_to_evaluate, error_message):
        """
        Returns the value if it's a number and "" if it's not.
        """
        val = self._get_text(value_to_evaluate)
        if not val.isdigit() or val == str(0):  # 0 is not a valid input
            self.input_error_descriptions.append(error_message)
            self.input_error_count += 1
            return ""
        return val

    # Note: .text() is UI related; using this on non-UI values will cause errors.
    def check_if_first_value_is_lower_than_second(self, value1, value2, error_message):
        # this function is connected to check_if_value_is_a_number
        # this will only run if the 2 values are confirmed non-zero digits
        val1 = self._get_text(value1)
        val2 = self._get_text(value2)
        if val1.isdigit() and val2.isdigit():
            if int(val1) >= int(val2):
                self.input_error_descriptions.append(error_message)
                self.input_error_count += 1
                return False
            return True
        return None

    # Used for Tab1
    # Exception: values do not use .text() because these are plain strings.
    def check_if_two_values_are_same(self, value1, value2, error_message):
        val1 = value1.strip().lower()
        val2 = value2.strip().lower()

        if val1.isalpha() and val2.isalpha():
            if val1 == val2:
                self.input_error_descriptions.append(error_message)
                self.input_error_count += 1
                return False
            return True
        return None
