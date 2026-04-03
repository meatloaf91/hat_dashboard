"""
reference_collector.py
======================
Generates per-IDH PDF reference sheets from an SDC output Excel file.

For every row whose Hit Type is "Total Hit" or "Partial Hit" the PDF contains:

  Section A – target IDH summary (IDH, Basic Num, Basic Desc, Hit Type, Basic Missed)
  Section B – thumbnail grid of Master IDH images with numbered basic-number details

Public API
----------
    run_reference_collector(params: RefCollectorParams) -> RefCollectorResult
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from pathlib import Path
import os
import re

import openpyxl
from PIL import Image as PILImage
from reportlab.lib.pagesizes import letter as _LETTER_PAGE
from reportlab.lib.colors import HexColor
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen.canvas import Canvas
from body_mapper import _read_tsc_idh_build as _bm_read_tsc_idh_build
import pandas as _pd
import re as _re_bm


# ── public dataclasses ────────────────────────────────────────────────────────

@dataclass
class RefCollectorParams:
    sdc_output_paths:   list[str]   # SDC xlsx files produced by run_comparison()
    image_library_path: str         # root folder that contains packshot images (searched recursively)
    output_dir:         str         # base folder; a timestamped subfolder is created inside
    tsc_data_path:      str = ""    # path to TSC Data xlsx — used to resolve Build Type per Master IDH
    max_images:         int = 5     # maximum Master IDH images to include per SDC row


@dataclass
class RefCollectorResult:
    output_folder:  str = ""                           # last folder created (compat)
    output_folders: list[str] = field(default_factory=list)  # one per SDC file processed
    pdf_paths:      list[str] = field(default_factory=list)
    warnings:       list[str] = field(default_factory=list)


# ── internal row model ────────────────────────────────────────────────────────

@dataclass
class _SdcRow:
    head_bom_mat:       str
    basic_number:       str
    hit_type:           str
    master_idh:         str
    basic_matched:      str
    basic_matched_desc: str
    basic_missed:       str
    basic_missed_desc:  str


# ── image search ──────────────────────────────────────────────────────────────

_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"}


def _find_image_for_idh(idh: str, library_root: str) -> str | None:
    """
    Recursively walk *library_root* and return the path of the most-recently
    modified image file whose name contains *idh* as a standalone digit token
    (i.e. not immediately preceded or followed by another digit).

    Returns None if no match is found.
    """
    pattern = re.compile(r"(?<!\d)" + re.escape(idh.strip()) + r"(?!\d)")
    best_path:  str | None = None
    best_mtime: float      = -1.0

    for dirpath, _, filenames in os.walk(library_root):
        for fname in filenames:
            if Path(fname).suffix.lower() not in _IMAGE_EXTS:
                continue
            if pattern.search(fname):
                fpath = os.path.join(dirpath, fname)
                try:
                    mtime = os.path.getmtime(fpath)
                except OSError:
                    mtime = 0.0
                if mtime > best_mtime:
                    best_mtime = mtime
                    best_path  = fpath

    return best_path


def _safe_mtime(path: str | None) -> float:
    """Return the file's modification time (epoch seconds), or 0.0 on error."""
    if not path:
        return 0.0
    try:
        return os.path.getmtime(path)
    except OSError:
        return 0.0


# ── SDC output reader ─────────────────────────────────────────────────────────

def _cell_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v != v:              # NaN
            return ""
        return str(int(v)) if v.is_integer() else str(v)
    return str(v).strip()


def _read_sdc_rows(sdc_path: str) -> list[_SdcRow]:
    """Read the *body_map_IDH* sheet and return only Total/Partial Hit rows."""
    wb = openpyxl.load_workbook(sdc_path, read_only=True, data_only=True)
    sheet_name = next(
        (s for s in wb.sheetnames if s.lower() == "body_map_idh"),
        wb.sheetnames[0] if wb.sheetnames else None,
    )
    if sheet_name is None:
        wb.close()
        return []

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        hdr = [_cell_str(v).lower() for v in next(rows_iter)]
    except StopIteration:
        wb.close()
        return []

    def _ci(name: str) -> int | None:
        try:    return hdr.index(name)
        except ValueError: return None

    ci_hbm   = _ci("head bom mat")
    ci_bn    = _ci("basic number")
    ci_ht    = _ci("hit type")
    ci_midh  = _ci("master idh")
    ci_bm    = _ci("basic matched")
    ci_bmd   = _ci("basic matched desc")
    ci_bmiss = _ci("basic missed")
    ci_bmisd = _ci("basic missed desc")

    def _get(row: tuple, idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        return _cell_str(row[idx])

    rows: list[_SdcRow] = []
    for row in rows_iter:
        ht = _get(row, ci_ht)
        if ht not in ("Total Hit", "Partial Hit"):
            continue
        rows.append(_SdcRow(
            head_bom_mat       = _get(row, ci_hbm),
            basic_number       = _get(row, ci_bn),
            hit_type           = ht,
            master_idh         = _get(row, ci_midh),
            basic_matched      = _get(row, ci_bm),
            basic_matched_desc = _get(row, ci_bmd),
            basic_missed       = _get(row, ci_bmiss),
            basic_missed_desc  = _get(row, ci_bmisd),
        ))

    wb.close()
    return rows


# ── PDF layout constants ──────────────────────────────────────────────────────

_PW, _PH   = _LETTER_PAGE           # 612 × 792 pt
_MARGIN    = 36.0
_AVAIL_W   = _PW - 2 * _MARGIN      # 540 pt

_FONT_B    = "Helvetica-Bold"
_FONT_R    = "Helvetica"
_FS_DETAIL = 6.5                    # body text inside cells / section A
_FS_LABEL  = 8.5                    # section headers ("Master IDHs:")

_C_BOX_BG      = HexColor("#EFF3F8")
_C_BOX_BORDER  = HexColor("#8FAABF")
_C_LBL         = HexColor("#444444")
_C_VAL         = HexColor("#111111")
_C_RED_BOLD    = HexColor("#CC0000")   # Hit Type / Build Type NA
_C_GREEN_BOLD  = HexColor("#1A7A1A")   # Build Type 3D
_C_ORANGE_BOLD = HexColor("#CC6600")   # Build Type 2D
_C_DARK        = HexColor("#111111")
_C_PLACEHOLDER = HexColor("#C8C8C8")
_C_PLHLD_TXT   = HexColor("#666666")


# ── TSC build-type classifier ─────────────────────────────────────────────────

_3D_TYPES = {"master", "clone", "master/clone", "3d"}
_2D_TYPES = {"resizing", "upload"}


def _classify_build(raw: str) -> str:
    """Map a raw TSC build-type string to '3D', '2D', or 'NA'."""
    v = raw.strip().lower()
    if v in _3D_TYPES:
        return "3D"
    if v in _2D_TYPES:
        return "2D"
    return "NA"


def _load_tsc_build_map(tsc_path: str) -> dict[str, str]:
    """
    Returns a mapping  IDH → classified build type ('3D' / '2D' / 'NA').
    Falls back to empty dict if the file cannot be read.
    """
    if not tsc_path:
        return {}
    try:
        raw_map = _bm_read_tsc_idh_build(tsc_path)
        return {idh: _classify_build(raw) for idh, raw in raw_map.items()}
    except Exception:
        return {}


def _load_tsc_extra_map(tsc_path: str) -> dict[str, dict[str, str]]:
    """
    Returns a mapping  IDH → {'project_name': ..., 'label_size': ...}
    Reads 'Tracker Data' sheet (or first sheet) of the TSC Data file.
    Values are empty string when column is absent or cell is blank.
    """
    if not tsc_path:
        return {}
    try:
        xl = _pd.ExcelFile(tsc_path, engine="openpyxl")
    except Exception:
        return {}

    try:
        df = _pd.read_excel(xl, sheet_name="Tracker Data", dtype=object)
    except Exception:
        try:
            df = _pd.read_excel(xl, sheet_name=xl.sheet_names[0], dtype=object)
        except Exception:
            return {}

    # Strip TSC export artefacts from column names
    df.columns = [_re_bm.sub(r"\s*▾\*?\s*$", "", str(c)).strip() for c in df.columns]

    def _ci(col_name: str) -> str | None:
        for c in df.columns:
            if c.strip().lower() == col_name.lower():
                return c
        return None

    def _cell(val: object) -> str:
        if val is None:
            return ""
        if isinstance(val, float):
            if val != val:
                return ""
            return str(int(val)) if val.is_integer() else str(val)
        return str(val).strip()

    idh_col  = _ci("IDH Number")
    proj_col = _ci("Project Name")
    lbl_col  = _ci("Label Size")

    if idh_col is None:
        return {}

    mapping: dict[str, dict[str, str]] = {}
    for _, row_s in df.iterrows():
        idh = _cell(row_s.get(idh_col, ""))
        if not idh:
            continue
        mapping[idh] = {
            "project_name": _cell(row_s.get(proj_col, "")) if proj_col else "",
            "label_size":   _cell(row_s.get(lbl_col, ""))  if lbl_col  else "",
        }
    return mapping


# ── text word-wrap helper ─────────────────────────────────────────────────────

def _wrap(text: str, max_chars: int) -> list[str]:
    """Split *text* into lines of at most *max_chars* characters."""
    if not text:
        return [""]
    words = text.split()
    lines: list[str] = []
    cur = ""
    for w in words:
        candidate = (cur + " " + w).lstrip()
        if len(candidate) <= max_chars:
            cur = candidate
        else:
            if cur:
                lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines or [""]


# ── section A ─────────────────────────────────────────────────────────────────

_A_PAD    = 8.0
_A_LBL_W  = 76.0
_A_LINE_H = 11.5
_A_FLD_GAP = 2.5


def _section_a_content(row: _SdcRow, avail_w: float):
    """Return list of (label_str, wrapped_value_lines) for Section A."""
    val_chars = max(20, int((avail_w - _A_PAD * 2 - _A_LBL_W) / 3.55))

    desc_parts = []
    if row.basic_matched_desc and row.basic_matched_desc.upper() != "NA":
        desc_parts.append(row.basic_matched_desc)
    if row.basic_missed_desc and row.basic_missed_desc.upper() != "NA":
        desc_parts.append(row.basic_missed_desc)
    basic_desc = " | ".join(desc_parts)

    return [
        ("IDH:",          _wrap(row.head_bom_mat, val_chars)),
        ("Basic Num:",    _wrap(row.basic_number, val_chars)),
        ("Basic Desc:",   _wrap(basic_desc, val_chars)),
        ("Hit Type:",     _wrap(row.hit_type, val_chars)),
        ("Basic Missed:", _wrap(row.basic_missed, val_chars)),
    ]


def _section_a_height(row: _SdcRow, avail_w: float) -> float:
    content = _section_a_content(row, avail_w)
    h = _A_PAD
    for _, lines in content:
        h += len(lines) * _A_LINE_H + _A_FLD_GAP
    h += _A_PAD
    return h


def _draw_section_a(c: Canvas, y_top: float, row: _SdcRow) -> float:
    """
    Draw Section A box flush against the left margin, full available width.
    Returns the y coordinate directly below the box (with an 8 pt gap).
    """
    x       = _MARGIN
    avail_w = _AVAIL_W
    content = _section_a_content(row, avail_w)
    box_h   = _section_a_height(row, avail_w)
    y_bot   = y_top - box_h

    # background + border
    c.setFillColor(_C_BOX_BG)
    c.setStrokeColor(_C_BOX_BORDER)
    c.setLineWidth(0.5)
    c.rect(x, y_bot, avail_w, box_h, fill=1, stroke=1)

    # text rows (baselines descend from y_top - padding)
    base_y = y_top - _A_PAD - _FS_DETAIL
    val_x  = x + _A_PAD + _A_LBL_W

    for lbl, lines in content:
        c.setFont(_FONT_B, _FS_DETAIL)
        c.setFillColor(_C_LBL)
        c.drawString(x + _A_PAD, base_y, lbl)

        # Hit Type value → bold red; all other values → normal dark
        is_hit_type = lbl.startswith("Hit Type")
        if is_hit_type:
            c.setFont(_FONT_B, _FS_DETAIL)
            c.setFillColor(_C_RED_BOLD)
        else:
            c.setFont(_FONT_R, _FS_DETAIL)
            c.setFillColor(_C_VAL)
        for i, line in enumerate(lines):
            c.drawString(val_x, base_y - i * _A_LINE_H, line)

        base_y -= len(lines) * _A_LINE_H + _A_FLD_GAP

    return y_bot - 8.0          # 8 pt gap below box


# ── image cell ────────────────────────────────────────────────────────────────

def _load_reader(img_path: str) -> ImageReader | None:
    """Open image via Pillow, convert to JPEG in-memory, return ImageReader."""
    try:
        pil = PILImage.open(img_path)
        if pil.mode == "RGBA":
            bg = PILImage.new("RGB", pil.size, (255, 255, 255))
            bg.paste(pil, mask=pil.split()[3])
            pil = bg
        elif pil.mode not in ("RGB", "L"):
            pil = pil.convert("RGB")
        buf = BytesIO()
        pil.save(buf, format="JPEG", quality=55, optimize=True)
        buf.seek(0)
        return ImageReader(buf)
    except Exception:
        return None


def _draw_placeholder(c: Canvas, x: float, y: float, w: float, h: float) -> None:
    c.setFillColor(_C_PLACEHOLDER)
    c.setStrokeColor(HexColor("#999999"))
    c.setLineWidth(0.5)
    c.rect(x, y, w, h, fill=1, stroke=1)
    c.setFillColor(_C_PLHLD_TXT)
    c.setFont(_FONT_B, 7.5)
    c.drawCentredString(x + w / 2, y + h / 2 - 4.0, "No Image Found")


def _draw_image_cell(
    c:            Canvas,
    x:            float,
    y_top:        float,
    cell_w:       float,
    cell_h:       float,
    img_path:     str | None,
    idh_label:    str,
    basics:       list[str],
    build_type:   str = "NA",
    project_name: str = "NA",
    label_size:   str = "NA",
    packshot_name: str = "",
) -> None:
    """Draw one thumbnail cell (image area + text block) at (x, y_top)."""
    PAD = 5.0

    # Text block height estimation (now includes Build Type, Packshot Name, Project Name, Label Size)
    text_h = (
        _FS_DETAIL + 3.0        # "IDH: …" line
        + _FS_DETAIL + 2.0      # "Build Type: …" line
        + _FS_DETAIL + 2.0      # "Packshot Name: …" line
        + _FS_DETAIL + 2.0      # "Project Name: …" line
        + _FS_DETAIL + 3.0      # "Label Size: …" line
        + _FS_DETAIL + 3.0      # "Basic Matched:" label
        + max(0, len(basics)) * (_FS_DETAIL + 2.0)
        + 4.0                   # bottom padding
    )
    img_h = max(40.0, cell_h - text_h - PAD * 2)
    img_w = cell_w - PAD * 2

    img_y = y_top - PAD - img_h     # bottom of image area

    # Draw image or placeholder
    if img_path:
        reader = _load_reader(img_path)
        if reader:
            ow, oh = reader.getSize()
            ratio  = min(img_w / ow, img_h / oh)
            dw, dh = ow * ratio, oh * ratio
            c.drawImage(
                reader,
                x + PAD + (img_w - dw) / 2,
                img_y  + (img_h - dh) / 2,
                dw, dh, mask="auto",
            )
        else:
            _draw_placeholder(c, x + PAD, img_y, img_w, img_h)
    else:
        _draw_placeholder(c, x + PAD, img_y, img_w, img_h)

    # Text below image
    ty = img_y - 3.0 - _FS_DETAIL

    # IDH line
    c.setFont(_FONT_B, _FS_DETAIL)
    c.setFillColor(_C_DARK)
    c.drawString(x + PAD, ty, f"IDH: {idh_label}")
    ty -= _FS_DETAIL + 2.0

    # Build Type line (label in dark, value colour depends on type)
    c.setFont(_FONT_B, _FS_DETAIL)
    c.setFillColor(_C_DARK)
    c.drawString(x + PAD, ty, "Build Type: ")
    lbl_w = c.stringWidth("Build Type: ", _FONT_B, _FS_DETAIL)
    bt_color = (
        _C_GREEN_BOLD  if build_type == "3D"
        else _C_ORANGE_BOLD if build_type == "2D"
        else _C_RED_BOLD
    )
    c.setFillColor(bt_color)
    c.drawString(x + PAD + lbl_w, ty, build_type)
    ty -= _FS_DETAIL + 2.0

    # Packshot Name line
    c.setFont(_FONT_B, _FS_DETAIL)
    c.setFillColor(_C_DARK)
    c.drawString(x + PAD, ty, "Packshot Name: ")
    psn_w = c.stringWidth("Packshot Name: ", _FONT_B, _FS_DETAIL)
    c.setFont(_FONT_R, _FS_DETAIL)
    c.drawString(x + PAD + psn_w, ty, packshot_name or "NA")
    ty -= _FS_DETAIL + 2.0

    # Project Name line
    c.setFont(_FONT_B, _FS_DETAIL)
    c.setFillColor(_C_DARK)
    c.drawString(x + PAD, ty, "Project Name: ")
    pn_w = c.stringWidth("Project Name: ", _FONT_B, _FS_DETAIL)
    c.setFont(_FONT_R, _FS_DETAIL)
    c.drawString(x + PAD + pn_w, ty, project_name or "NA")
    ty -= _FS_DETAIL + 2.0

    # Label Size line
    c.setFont(_FONT_B, _FS_DETAIL)
    c.setFillColor(_C_DARK)
    c.drawString(x + PAD, ty, "Label Size: ")
    ls_w = c.stringWidth("Label Size: ", _FONT_B, _FS_DETAIL)
    c.setFont(_FONT_R, _FS_DETAIL)
    c.drawString(x + PAD + ls_w, ty, label_size or "NA")
    ty -= _FS_DETAIL + 3.0

    # Basic label + numbered lines
    c.setFont(_FONT_B, _FS_DETAIL)
    c.setFillColor(_C_DARK)
    c.drawString(x + PAD, ty, "Basic Matched:")
    ty -= _FS_DETAIL + 2.0

    c.setFont(_FONT_R, _FS_DETAIL)
    for bline in basics:
        c.drawString(x + PAD + 3.0, ty, bline)
        ty -= _FS_DETAIL + 2.0


# ── basic list builder ────────────────────────────────────────────────────────

def _build_basics(basic_matched: str, basic_matched_desc: str) -> list[str]:
    """
    Build numbered 'BN  –  Description' lines from the SDC matched columns.
    Example output: ['1. 821633  BOTT_PE_91,5x45,3x168,0MM_36g__DG', ...]
    """
    bns: list[str] = (
        [v.strip() for v in basic_matched.split(",") if v.strip()]
        if basic_matched.upper() != "NA" else []
    )
    descs_raw = basic_matched_desc if basic_matched_desc.upper() != "NA" else ""
    descs: list[str] = (
        [v.strip() for v in descs_raw.split("|")] if descs_raw else []
    )
    lines: list[str] = []
    for i, bn in enumerate(bns):
        desc = descs[i] if i < len(descs) else ""
        lines.append(f"{i + 1}. {bn}  {desc}")
    return lines


# ── single-row PDF generator ──────────────────────────────────────────────────

_NCOLS   = 2
_COL_GAP = 8.0
_ROW_GAP = 8.0
_CELL_H  = 185.0                         # fixed cell height (image + text)
_CELL_W  = (_AVAIL_W - _COL_GAP) / _NCOLS   # ≈ 266 pt


def _generate_pdf(
    row:            _SdcRow,
    image_library:  str,
    max_images:     int,
    out_path:       str,
    tsc_build_map:  dict[str, str] | None = None,
    tsc_extra_map:  dict[str, dict[str, str]] | None = None,
) -> None:
    """Build and save a one-PDF-per-row reference sheet."""

    # Resolve image paths ---------------------------------------------------
    raw_idhs = (
        [v.strip() for v in row.master_idh.split(",") if v.strip()]
        if row.master_idh else []
    )
    with_img:    list[tuple[str, str | None]] = []
    without_img: list[tuple[str, str | None]] = []
    for idh in raw_idhs:
        img = _find_image_for_idh(idh, image_library) if image_library else None
        (with_img if img else without_img).append((idh, img))

    # Prioritise IDHs that have images.
    # When more IDHs have images than max_images allows, keep the ones with
    # the most-recently *modified* image (Date modified, not Date created).
    with_img.sort(key=lambda t: _safe_mtime(t[1]), reverse=True)
    cells = list(with_img[:max_images])
    needed = max_images - len(cells)
    cells += [(idh, None) for idh, _ in without_img[:needed]]
    cells  = cells[:max_images]

    basics = _build_basics(row.basic_matched, row.basic_matched_desc)

    # Page geometry ---------------------------------------------------------
    a_h     = _section_a_height(row, _AVAIL_W)
    lbl_blk = _FS_LABEL + 8.0          # "Master IDHs:" + gap to first image row

    # Available height on page 1 (after Section A + label block)
    avail_p1   = (_PH - _MARGIN) - a_h - 8.0 - lbl_blk - _MARGIN
    avail_rest = _PH - 2 * _MARGIN

    rows_p1   = max(1, int((avail_p1   + _ROW_GAP) / (_CELL_H + _ROW_GAP)))
    rows_rest = max(1, int((avail_rest + _ROW_GAP) / (_CELL_H + _ROW_GAP)))
    cpc_p1    = rows_p1   * _NCOLS
    cpc_rest  = rows_rest * _NCOLS

    # Draw ------------------------------------------------------------------
    c = Canvas(out_path, pagesize=_LETTER_PAGE)
    idx      = 0
    page_num = 0

    while idx < len(cells):
        if page_num > 0:
            c.showPage()

        if page_num == 0:
            # Section A
            grid_y = _PH - _MARGIN
            grid_y = _draw_section_a(c, grid_y, row)
            # "Master IDHs:" label
            c.setFont(_FONT_B, _FS_LABEL)
            c.setFillColor(_C_DARK)
            c.drawString(_MARGIN, grid_y - _FS_LABEL, "Master IDHs:")
            grid_y -= lbl_blk
            batch  = cpc_p1
        else:
            grid_y = _PH - _MARGIN
            batch  = cpc_rest

        for i, (idh, img_path) in enumerate(cells[idx: idx + batch]):
            col = i % _NCOLS
            row_i = i // _NCOLS
            cx = _MARGIN + col * (_CELL_W + _COL_GAP)
            cy = grid_y  - row_i * (_CELL_H + _ROW_GAP)
            bt = (tsc_build_map or {}).get(idh, "NA")
            extra = (tsc_extra_map or {}).get(idh, {})
            pn = extra.get("project_name", "") or "NA"
            ls = extra.get("label_size",   "") or "NA"
            psn = Path(img_path).stem if img_path else ""
            _draw_image_cell(c, cx, cy, _CELL_W, _CELL_H, img_path, idh, basics, bt, pn, ls, psn)

        idx      += batch
        page_num += 1

    c.save()


# ── folder name helper ───────────────────────────────────────────────────────

_DATE_SUFFIX_RE = re.compile(r"_\d{4}_\d{2}_\d{2}_\d{2}_\d{2}$")


def _rc_folder_name(sdc_path: str, ts: str) -> str:
    """
    Derive the RC output folder name from an SDC file path.
    Example:
      sdc_path stem : sdc_bma_sap_data_raw_cu1_2026_03_16_01_13
      result        : rc_bma_sap_data_raw_cu1_2026_03_16_01_14   (ts = RC run time)
    Steps:
      1. Take the stem of the filename.
      2. Strip a leading 'sdc_' prefix (case-insensitive).
      3. Strip a trailing date pattern _YYYY_MM_DD_HH_MM.
      4. Prepend 'rc_' and append '_<ts>'.
    """
    stem = Path(sdc_path).stem
    if stem.lower().startswith("sdc_"):
        stem = stem[4:]
    stem = _DATE_SUFFIX_RE.sub("", stem)
    return f"rc_{stem}_{ts}"


# ── public entry point ────────────────────────────────────────────────────────

def run_reference_collector(params: RefCollectorParams) -> RefCollectorResult:
    """
    For every Total/Partial Hit row across all *sdc_output_paths*, generate a
    PDF named ``rc_<IDH>.pdf`` and save it into a timestamped subfolder of
    *output_dir*.
    """
    result = RefCollectorResult()

    if not params.sdc_output_paths:
        result.warnings.append("No SDC output files to process.")
        return result
    if not params.output_dir:
        result.warnings.append("No output directory specified.")
        return result

    ts = datetime.now().strftime("%Y_%m_%d_%H_%M")

    # Load TSC build-type map once (shared across all SDC output files)
    tsc_build_map = _load_tsc_build_map(params.tsc_data_path)
    tsc_extra_map = _load_tsc_extra_map(params.tsc_data_path)

    for sdc_path in params.sdc_output_paths:
        # Each SDC file gets its own output folder
        folder_name = _rc_folder_name(sdc_path, ts)
        folder = Path(params.output_dir) / folder_name
        try:
            folder.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            result.warnings.append(f"Cannot create folder '{folder_name}': {exc}")
            continue
        result.output_folders.append(str(folder))
        result.output_folder = str(folder)   # keep compat (last folder)

        try:
            rows = _read_sdc_rows(sdc_path)
        except Exception as exc:
            result.warnings.append(f"Cannot read {Path(sdc_path).name}: {exc}")
            continue

        for row in rows:
            if not row.head_bom_mat:
                continue
            safe_name  = re.sub(r'[<>:"/\\|?*\s]+', "_", row.head_bom_mat)
            ht_suffix  = "_t" if row.hit_type == "Total Hit" else "_p"
            pdf_path   = folder / f"rc_{safe_name}{ht_suffix}.pdf"
            try:
                _generate_pdf(
                    row,
                    params.image_library_path,
                    params.max_images,
                    str(pdf_path),
                    tsc_build_map=tsc_build_map,
                    tsc_extra_map=tsc_extra_map,
                )
                result.pdf_paths.append(str(pdf_path))
            except Exception as exc:
                result.warnings.append(
                    f"PDF failed for IDH '{row.head_bom_mat}': {exc}"
                )

    return result
