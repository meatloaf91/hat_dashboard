import html
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QDialog,
    QStyle,
)
from PySide6.QtCore import Qt
import os
from datetime import datetime
import re
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from collections import defaultdict
from datetime import datetime


import stylesheet as ss

#IN PROGRESS: SBU Column
class StatusCollector:

    # Column name variations used when importing from the Tracker Window dialog.
    # Order must match _TscOption1TableDialog.TSC_HEADERS exactly.
    TRACKER_WINDOW_COLUMN_VARIATIONS: dict[str, list[str]] = {
        "SBU":            ["sbu", "idh sbu", "idh accounting sbu"],
        "Product Name":   ["variant name / base product name", "base product name", "base product",
                           "variant name", "product name", "product", "variant"],
        "IDH Number":     ["idh number", "idh", "idh num", "idh no.", "idh no"],
        "Build Type":     ["build type", "master / clone", "master/clone"],
        "Status":         ["status"],
        "Packaging Type": ["packaging type", "packaging", "pack type", "packtype"],
        "Packaging Size": ["capacity", "packaging size", "size"],
        "Project Name":   ["project name", "project", "wrike name"],
        "Basic Number":   ["basic number", "basic num", "basic no.", "basic no", "basic", "basic #"],
        "Label Size":     ["label size", "label size mm(h x w)", "label size mm (h x w)",
                           "label size mm (hxw)", "label size mm(hxw)", "label"],
        "Is Deployment":  ["is deployment", "deployment", "is_deployment", "isdeployment"],
    }

    def __init__(self, ui):
        self.excel_trackers = [] #SAP data for masterfile, -all existing packshots
        self.excel_tracker_names = []
        self.output_location = ""
        self.statuses = []
        self.row_threshold = 15  # max rows to search for header
        self.timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        self.output_name = "CPD" #collected packshot data

        #has all the needed info
        #format: dictionary of dictionaries
        """
        self.target_final_info = [
            { file: tracker_2025_CHINA, sheet: tracker_China, index: 11 },
            { file: tracker_2025_EE, sheet: tracker_PL, index: 11 },
        ]
        """
        self.target_final_info = []

        #------------------------------------------------------------------------------------------------

        self.name_variations_idh = ["idh number", "idh", "idh num", "idh no.", "idh no"]
        self.name_variations_product_name = ["variant name / base product name",  
                                             "base product name", "base product", "variant name",
                                             "product name", "product", "variant"]
        self.name_variations_build_type = ["build type", "master / clone", "master/clone"]
        self.name_variations_status = ["status"]
        self.name_variations_project_name = ["project name", "project", "wrike name"]
        self.name_variations_sbu = ["sbu", "idh sbu", "idh accounting sbu"]

        #-----new addition:
        self.name_variations_packaging_type = ["packaging type", "packaging"]
        self.name_variations_packaging_size = ["capacity", "packaging size", "size"]
        self.name_variations_label_size = ["label size", "label size mm(h x w)", "label size mm (h x w)", "label size mm (hxw)", "label size mm(hxw)"]
        self.name_variations_is_deployment = ["is deployment", "deployment", "is_deployment", "isdeployment"]
        self.name_variations_basic_number = ["basic number", "basic num", "basic no.", "basic no", "basic"]


        #column naming variation can be edited above
        self.column_name_variations = {
            "idh":self.name_variations_idh,
            "product_name":self.name_variations_product_name,
            "build_type":self.name_variations_build_type,
            "status":self.name_variations_status,
            "project_name":self.name_variations_project_name,
            "sbu": self.name_variations_sbu,
            #---new addition:
            "packaging_type": self.name_variations_packaging_type,
            "packaging_size":self.name_variations_packaging_size,
            "label_size":self.name_variations_label_size,
            "is_deployment": self.name_variations_is_deployment,
            "basic_number": self.name_variations_basic_number
        }
        #---------------------------------------------------------------------------------------------------

        self.trackers_warnings_count = 0
        self.status_error_count = 0
        self.status_missing_on_some_sheets = "none" #can be none, partial or total. This is for what type of alert message to show. Ex, if partial, then generate only a warning
        self.successful_run = False
        self.can_proceed = False
        self.general_report = ""
        # self.trackers_missing_columns_warning = ""
        # self.final_status_description = ""
        # self.fields_errors_description = ""

        self.btn_select_trackers = ui.btn_sc_select_trackers
        self.btn_output_location = ui.btn_sc_output_location
        self.editText_selected_trackers = ui.textEdit_sc_selected_trackers #textEdit_pg10_select_trackers
        self.editText_output_location = ui.textEdit_sc_output_location #textEdit_pg10_output

        self.checkboxes = [
            ui.checkbox_status_collector_cancelled,
            ui.checkbox_status_collector_in_progress,
            ui.checkbox_status_collector_completed,
            ui.checkbox_status_collector_on_hold,
            ui.checkbox_status_collector_to_do,
        ]
        self.radiobtn01 = ui.radioButton_status_select
        self.radiobtn02 = ui.radioButton_status_input
        self.textEdit01 = ui.textEdit_status_collector_status_input
        self.btn_run_process = ui.btn_run_process_all_trackers_status_collector
        self.checkbox_apply_cleanup = getattr(ui, "checkbox_sc_apply_cleanup", None)
        self.btn_menu_clear_all_fields = ui.btn_menu_clear_all_fields
        self.progress_bar = getattr(ui, "collector_progress_bar", None)

    def _set_processing_state(self, running: bool) -> None:
        self.btn_run_process.setEnabled(not running)
        if self.progress_bar is not None:
            self.progress_bar.setVisible(running)
            if running:
                self.progress_bar.setRange(0, 0)  # indeterminate busy animation
            else:
                self.progress_bar.setRange(0, 100)
                self.progress_bar.setValue(0)

        app = QApplication.instance()
        if app is not None:
            app.processEvents()

    def show_alert_old(self, message: str, alert_type: str = "info", parent: QWidget | None = None):
        parent = parent or self.btn_run_process  # keep it on top of your window
        msg = QMessageBox(parent)
        msg.setText(message)
        msg.setStyleSheet("QLabel{ color: white; }")

        if alert_type == "info":
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Information")
            msg.setStandardButtons(QMessageBox.Ok)
            return msg.exec()

        if alert_type == "warning":
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Warning")
            msg.setStandardButtons(QMessageBox.Ok)
            return msg.exec()

        if alert_type == "error":
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            return msg.exec()

        if alert_type == "question":
            msg.setIcon(QMessageBox.Question)
            msg.setWindowTitle("Question")
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            return msg.exec()

    def status_behavior(self) -> None:
        if self.radiobtn01.isChecked():
            for cb in self.checkboxes:
                cb.setEnabled(True)
            self.textEdit01.setEnabled(False)
            self.textEdit01.clear()
        else:
            for cb in self.checkboxes:
                cb.setEnabled(False)
                cb.setChecked(False)
            self.textEdit01.setEnabled(True)
    
    def statuses_selected(self) -> None:
        self.statuses = [cb.text().strip().lower() for cb in self.checkboxes if cb.isChecked()]

        text_edit_val = self.textEdit01.toPlainText().strip()
        if text_edit_val:
            parts = [p.strip().lower() for p in text_edit_val.split(",")]
            self.statuses.extend([p for p in parts if p])

        # Keep order while removing duplicates.
        self.statuses = list(dict.fromkeys(self.statuses))

    def bind_input_type_to_radioboxes(self) -> None:
        # start in 'checkbox select' mode disabled for text
        self.textEdit01.setEnabled(False)

        # React to either radio toggling
        self.radiobtn01.toggled.connect(lambda _checked: self.status_behavior())
        self.radiobtn02.toggled.connect(lambda _checked: self.status_behavior())

        # Apply initial state once
        self.status_behavior()

    def get_output_location(self, start_dir: str = ""):
        path = QFileDialog.getExistingDirectory(parent=None, caption="", dir=start_dir, options=QFileDialog.Options())
        if path:
            self.output_location = path
            self.editText_output_location.setPlainText(f"{path}")
        else:
            self.folder_location = None

    def get_tracker_files(self, start_dir: str = ""):
        dialog = QFileDialog()
        dialog.setWindowTitle("Select Excel File(s)")
        dialog.setNameFilter("Excel Files (*.xlsx *.xlsm *.xltx *.xltm *.xls)")
        dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)
        if start_dir:
            dialog.setDirectory(start_dir)

        if dialog.exec():
            file_paths = dialog.selectedFiles()
            self.excel_trackers = file_paths or []
            self.excel_tracker_names = [os.path.basename(excel_file) for excel_file in self.excel_trackers] or []
            # self.excel_tracker_names = os.path.basename(self.out_excel_file)
        
        if self.excel_tracker_names:
            self.editText_selected_trackers.setPlainText(",\n".join(self.excel_tracker_names))

    def check_if_columns_needed_exist(self):
        #normalizing column names, remove space and convert to lowercase
        idh_name_variations = [s.lower().strip() for s in self.name_variations_idh]
        product_name_variations = [s.lower().strip() for s in self.name_variations_product_name]
        build_type_variations = [s.lower().strip() for s in self.name_variations_build_type]
        status_variations = [s.lower().strip() for s in self.name_variations_status]
        proj_name_variations = [s.lower().strip() for s in self.name_variations_project_name]
        sbu_name_variations = [s.lower().strip() for s in self.name_variations_sbu]
        #------new addition:
        packtype_variations = [s.lower().strip() for s in self.name_variations_packaging_type]
        packsize_variations = [s.lower().strip() for s in self.name_variations_packaging_size]
        labelsize_variations = [s.lower().strip() for s in self.name_variations_label_size]
        basicnum_variations = [s.lower().strip() for s in self.name_variations_basic_number]

        #add all column names in one list
        all_columns = (
            idh_name_variations +
            product_name_variations +
            build_type_variations +
            status_variations +
            proj_name_variations +
            sbu_name_variations +
            #-------new addition
            packtype_variations +
            packsize_variations +
            labelsize_variations +
            basicnum_variations
        )

        formatted_message = "Tracker sheets w/ missing required column(s):\n"
        warning_count = 0

        for tracker in self.excel_trackers:
            tracker_warnings = {}
            sheet_warnings = {} 
            sheet_warning_count = 0
            has_sheet_with_tracker_naming = False

            current_tracker = pd.ExcelFile(tracker)#getting 1 tracker at a time
            current_tracker_name = os.path.basename(tracker) #getting tracker name only

            for sheet in current_tracker.sheet_names:
                #if a sheet dont have "tracker" in the naming, then ignore it
                if "tracker" not in sheet.lower():
                    continue
                
                #panda reading excel, nrows is the threshold of until which row will the system look for columns
                df = pd.read_excel(
                    current_tracker,
                    sheet_name=sheet,
                    nrows=self.row_threshold,
                    header=None,
                    dtype=str,
                    na_filter=False,     # skip parsing for NaN-like strings
                    engine="openpyxl",
                )
                has_sheet_with_tracker_naming = True

                # Normalize cell value: lowercase, strip, collapse internal whitespace
                def _norm_cell(val: object) -> str:
                    return re.sub(r'\s+', ' ', str(val).strip().lower())

                #variable for target index – reset for every sheet
                index = None
                header_row_index = -1
                for i in range(len(df)):
                    row_vals = [_norm_cell(v) for v in df.iloc[i].values if pd.notna(v)]

                    #checking a potential header. Require at least 2 matching column names
                    #so a stray keyword in a title row is not mistaken for the header.
                    match_count = sum(1 for col in all_columns if col in row_vals)
                    if match_count >= 2:
                        header_row_index = i
                        index = header_row_index
                        break
                  
                #if no potential header is found, then generate error description
                if index is None:
                    sheet_warning_count += 1
                    sheet_warnings[sheet] = ["IDH Number", "Product Name", "Build Type", "Status", "Project Name", "SBU", "Packaging Type", "Packaging Size", "Label Size", "Is Deployment", "Basic Number"]
                else:
                    #if a sheet have very less number of columns, then generate same error
                    if len(df.columns) < 6:
                        sheet_warnings[sheet] = ["IDH Number", "Product Name", "Build Type", "Status", "Project Name", "SBU", "Packaging Type", "Packaging Size", "Label Size", "Is Deployment", "Basic Number"]
                        continue
                    
                    missing_cols = [] #list of missing columns for current sheet

                    # reuse the first read (df: header=None, limited rows)
                    col_names = [_norm_cell(v) for v in df.iloc[index].tolist()]

                    #checking if column names does not exist in each sheet
                    if not any(var in col_names for var in idh_name_variations):
                        missing_cols.append("IDH Number")
                    if not any(var in col_names for var in product_name_variations):
                        missing_cols.append("Product Name")
                    if not any(var in col_names for var in build_type_variations):
                        missing_cols.append("Build Type")
                    if not any(var in col_names for var in status_variations):
                        missing_cols.append("Status")
                    if not any(var in col_names for var in proj_name_variations):
                        missing_cols.append("Project Name")
                    if not any(var in col_names for var in sbu_name_variations):
                        missing_cols.append("SBU")
                    #-------new addition
                    if not any(var in col_names for var in packtype_variations):
                        missing_cols.append("Packaging Type")
                    if not any(var in col_names for var in packsize_variations):
                        missing_cols.append("Packaging Size")
                    if not any(var in col_names for var in labelsize_variations):
                        missing_cols.append("Label Size")
                    if not any(var in col_names for var in self.name_variations_is_deployment):
                        missing_cols.append("Is Deployment")
                    if not any(var in col_names for var in basicnum_variations):
                        missing_cols.append("Basic Number")

                    #if there is missing columns, add to sheet_warnings dict
                    if missing_cols:
                        sheet_warnings[sheet] = missing_cols
                        sheet_warning_count += 1

                    # Always register the sheet so "Ignore errors" can still read it.
                    # Columns that are missing will resolve to None; get_all_inputs handles that.
                    def pick(actual_vals, variations):
                        want = {s.strip().lower() for s in variations}
                        for v in actual_vals:
                            if v.lower() in want:
                                return v  # return the original header text (not lowercased)
                        return None

                    actual_idh_col          = pick(col_names, self.name_variations_idh)
                    actual_product_name_col = pick(col_names, self.name_variations_product_name)
                    actual_build_type_col   = pick(col_names, self.name_variations_build_type)
                    actual_status_col       = pick(col_names, self.name_variations_status)
                    actual_project_name_col = pick(col_names, self.name_variations_project_name)
                    actual_sbu_name_col     = pick(col_names, self.name_variations_sbu)
                    actual_packtype         = pick(col_names, self.name_variations_packaging_type)
                    actual_packsize         = pick(col_names, self.name_variations_packaging_size)
                    actual_labelsize        = pick(col_names, self.name_variations_label_size)
                    actual_is_deployment    = pick(col_names, self.name_variations_is_deployment)
                    actual_basicnum         = pick(col_names, self.name_variations_basic_number)

                    # Skip sheets where Status column could not be found – nothing to filter on
                    if actual_status_col is not None:
                        self.target_final_info.append({
                            "file":  current_tracker,
                            "sheet": sheet,
                            "index": index,
                            "cols": {
                                "idh":          actual_idh_col,
                                "product_name": actual_product_name_col,
                                "build_type":   actual_build_type_col,
                                "status":       actual_status_col,
                                "project_name": actual_project_name_col,
                                "sbu_name":     actual_sbu_name_col,
                                "packaging_type":   actual_packtype,
                                "packaging_size":   actual_packsize,
                                "label_size":       actual_labelsize,
                                "is_deployment":    actual_is_deployment,
                                "basic_number":     actual_basicnum,
                            }
                        })

                #if there are any sheets with missing columns, add to main warning dict
                if sheet_warning_count > 0:
                    tracker_warnings[current_tracker_name] = sheet_warnings
            
            #checking if after evaluating all sheets on a tracker, if no sheet named "tracker" is found,
            #add a specific error description on formatted message
            if not has_sheet_with_tracker_naming:
                warning_count += 1
                formatted_message += f"\n{current_tracker_name}:\nNo sheet named 'tracker' found. Unable to check columns.\n"
                
            #else if a tracker sheet is found but has missing columns, add the missing columns on formatted message
            if tracker_warnings:
                warning_count += 1
                for tracker_name, sheet_names in tracker_warnings.items():
                    formatted_message += f"\n{tracker_name}:\n"
                    for sheet_name, cols in sheet_names.items():
                        formatted_message += f"\t{sheet_name}: {', '.join(cols)}\n"
        
        self.trackers_warnings_count = warning_count  

        if self.trackers_warnings_count > 0:              
            self.general_report  = f"{formatted_message}\nKindly correct all errors inorder to proceed.\nIf a sheet is not relevant, remove 'tracker' from the name so that it will not be evaluated."

    def find_col(self, df: pd.DataFrame, candidates: list[str]) -> str | None:
        """Return the actual df column whose normalized name matches any candidate (case-insensitive)."""
        # Normalize actual DataFrame column labels to strings
        normalized_column_names = {str(c).strip().lower(): c for c in df.columns}

        # Flatten if a nested list accidentally slips in
        flat_candidates: list[str] = []
        for cand in candidates:
            if isinstance(cand, (list, tuple)):
                flat_candidates.extend(cand)
            else:
                flat_candidates.append(cand)

        # Match
        for cand in flat_candidates:
            if not isinstance(cand, str):
                continue
            key = cand.strip().lower()
            if key in normalized_column_names:
                return normalized_column_names[key]
        return None

    def get_all_inputs_according_to_status_chatgpt_version_2(self, statuses):
        all_data: list[pd.DataFrame] = []
        info_message = ""
        status_errors_dict = {}

        # pre-normalize statuses once
        status_set = {s.strip().lower() for s in statuses}
        found_statuses = set()       # track which were actually found

        for item in self.target_final_info:
            tracker_file = item["file"]
            tracker_sheet = item["sheet"]
            sheet_header = item["index"]

            # Faster read
            df = pd.read_excel(
                tracker_file,
                sheet_name=tracker_sheet,
                header=sheet_header,
                dtype=str,
                na_filter=False,
                engine="openpyxl",
            )

            # Resolve columns once
            idh          = self.find_col(df, self.column_name_variations["idh"])
            product_name = self.find_col(df, self.column_name_variations["product_name"])
            build_type   = self.find_col(df, self.column_name_variations["build_type"])
            status_col   = self.find_col(df, self.column_name_variations["status"])
            project_name = self.find_col(df, self.column_name_variations["project_name"])
            sbu = self.find_col(df, self.column_name_variations["sbu"])
            #--------new addition:
            packaging_type = self.find_col(df, self.column_name_variations["packaging_type"])
            packaging_size = self.find_col(df, self.column_name_variations["packaging_size"])
            label_size = self.find_col(df, self.column_name_variations["label_size"])
            is_deployment = self.find_col(df, self.column_name_variations.get("is_deployment", []))
            basic_number = self.find_col(df, self.column_name_variations["basic_number"])


            # Keep only relevant cols (order matches TRACKER_WINDOW_COLUMN_VARIATIONS)
            # Any column that wasn't found (None) is replaced with a blank-string series
            display_names = list(self.TRACKER_WINDOW_COLUMN_VARIATIONS.keys())
            col_sources = [sbu, product_name, idh, build_type, status_col,
                           packaging_type, packaging_size, project_name,
                           basic_number, label_size, is_deployment]
            df_selected = pd.DataFrame(index=df.index)
            for dname, src in zip(display_names, col_sources):
                df_selected[dname] = df[src].astype(str) if src is not None else ""
            df = df_selected

            # If Apply Cleanup is checked, run cleanup on ALL rows NOW – before status
            # filtering – so that remapped statuses (e.g. "uploaded"→"completed",
            # "descoped"→"cancelled") are visible to the filter below.
            apply_cleanup = (
                self.checkbox_apply_cleanup is not None
                and self.checkbox_apply_cleanup.isChecked()
            )
            if apply_cleanup:
                col_names = list(df.columns)
                cleaned_rows = [
                    self._apply_row_cleanup(list(row), col_names)
                    for _, row in df.iterrows()
                ]
                df = pd.DataFrame(cleaned_rows, columns=col_names)

            # Normalize statuses once (against already-cleaned values when applicable)
            norm_status = df["Status"].map(lambda v: v.strip().lower() if isinstance(v, str) else "")
            unique_in_sheet = set(norm_status.unique()) - {""}

            # Track which of the requested statuses exist in this sheet
            found_here = unique_in_sheet.intersection(status_set)
            found_statuses.update(found_here)

            # Which requested statuses are missing from this sheet?
            missing_here = status_set - unique_in_sheet
            this_tracker_name = os.path.basename(getattr(tracker_file, "io", tracker_file))

            # Only record a partial warning if none of the rows here will contribute data
            if missing_here and not found_here:
                status_errors_dict[this_tracker_name] = sorted(missing_here)
                self.status_missing_on_some_sheets = "partial"

            # Filter to only the requested ones
            mask = norm_status.isin(status_set)
            subset = df.loc[mask].copy()

            if subset.empty:
                # still proceed to next sheet; we've already recorded missing_here above
                self.status_error_count += 1
                continue

            # Exclude rows where IDH is blank, or IDH/Build Type is "admin" (case-insensitive)
            idh_norm = subset["IDH Number"].map(lambda v: v.strip().lower() if isinstance(v, str) else "")
            build_type_norm = subset["Build Type"].map(lambda v: v.strip().lower() if isinstance(v, str) else "")
            valid_rows = (idh_norm != "") & (idh_norm != "admin") & (build_type_norm != "admin")
            subset = subset.loc[valid_rows].copy()

            # When Apply Cleanup is active: also exclude rows whose IDH Number is not purely numeric
            if apply_cleanup:
                idh_digits_mask = subset["IDH Number"].map(
                    lambda v: bool(re.match(r'^\d+$', v.strip())) if isinstance(v, str) and v.strip() else False
                )
                subset = subset.loc[idh_digits_mask].copy()

            if subset.empty:
                continue

            # Cleanup already applied above; columns are already display names
            all_data.append(subset)

        try:
            #if all is well
            if all_data:
                ts = datetime.now().strftime("%Y_%m_%d_%H_%M")
                output_name = f"{self.output_location}/tsc_data_{ts}.xlsx"
                out = pd.concat(all_data, ignore_index=True)
                with pd.ExcelWriter(output_name, engine="openpyxl") as writer:
                    out.to_excel(writer, index=False, sheet_name="Tracker Data")
                    self._style_cpd_sheet(writer.book["Tracker Data"], out)

                self.successful_run = True
                info_message = "✅ Excel file saved successfully."

                if status_errors_dict:
                    info_message = "✅ CPD report saved successfully but with below warnings:\n\n"
                    lines = ["⚠️ Below trackers contain sheet(s) with missing status(es):"]
                    for file, missing_stat in status_errors_dict.items():
                        lines.append(f"\n{file}:")
                        for stat in missing_stat:
                            lines.append(f"{stat}")
                    info_message += "\n".join(lines) 

                self.general_report  = info_message
                return out
            else:
                info_message = "⚠️ No matching status(es) found on tracker(s)."
                self.status_missing_on_some_sheets = "total"
                self.status_error_count += 1

                self.general_report  = info_message
                return None
        except ValueError as e:
            #if file is too large, example scenario: if user highlighted a column and automaticall filled all
            if "sheet is too large" in str(e).lower():
                self.show_alert(
                    "The generated report is too large for Excel (over 1,048,576 rows).\n"
                    "Please refine your filters.",
                    alert_type="error",
                    first_line_col="#FF5555"
                )
            else:
                # Handle other ValueError types gracefully
                self.show_alert(
                    (
                        "An unknown error occurred while exporting the Excel file.\n"
                        "Please contact the tool creator to check the issue."
                    ),
                    alert_type="error",
                    first_line_col="#FF5555"
                )
                raise  # re-raise so you can debug if needed

    def _style_cpd_sheet(self, ws, out_df: pd.DataFrame) -> None:
        """Apply a modern, readable style to the CPD worksheet."""
        total_cols = len(out_df.columns)
        total_rows = len(out_df) + 1  # include header

        header_fill = PatternFill(fill_type="solid", fgColor="111F35")
        header_font = Font(color="FFFFFF", bold=True, size=11, name="Segoe UI")
        body_font = Font(color="1B1F24", size=10, name="Segoe UI")
        alt_fill = PatternFill(fill_type="solid", fgColor="F3F7FC")
        thin_border = Border(
            left=Side(style="thin", color="D9E1EA"),
            right=Side(style="thin", color="D9E1EA"),
            top=Side(style="thin", color="D9E1EA"),
            bottom=Side(style="thin", color="D9E1EA"),
        )

        # Header styling
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Body styling (zebra rows + borders)
        for row_idx in range(2, total_rows + 1):
            use_alt = (row_idx % 2 == 0)
            for col_idx in range(1, total_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border
                if use_alt:
                    cell.fill = alt_fill

        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=total_cols).coordinate}"
        ws.row_dimensions[1].height = 24

        # Auto-fit columns with practical min/max width
        for idx, col_name in enumerate(out_df.columns, start=1):
            max_len = len(str(col_name))
            for value in out_df[col_name].head(5000):
                value_len = len(str(value)) if value is not None else 0
                if value_len > max_len:
                    max_len = value_len
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max(14, max_len + 2), 44)

    def check_fields(self):
        empty_fields = []
        if not self.excel_trackers:
            empty_fields.append("Select Trackers")
        if not self.output_location:
            empty_fields.append("Output Report Location")
        if not self.statuses:
            empty_fields.append(r"Select Status / Input Status")
        
        info_message = ""

        if empty_fields:
            lines = ["Unable to proceed. Kindly fill below missing field(s)"]
            for field in empty_fields:
                lines.append(f"*{field}:")
            info_message += "\n".join(lines)

            self.general_report = info_message
        else:
            self.can_proceed = True

    def export_txt_report(self, message: str):
        """Always ask the user for a folder and export a .txt report there."""
        parent_widget = getattr(self, "btn_run_process", None)

        try:
            # Always start from current directory (or change to desktop if preferred)
            start_dir = os.getcwd()  
            # Optional alternative: start_dir = os.path.expanduser("~/Desktop")

            # Ask user for a folder
            folder = QFileDialog.getExistingDirectory(
                parent_widget,
                "Select Folder to Save Report",
                start_dir
            )

            # If user cancels
            if not folder:
                self.show_alert(message="✅ Report export was cancelled", alert_type="info")
                return None

            os.makedirs(folder, exist_ok=True)

            # Build file name
            # timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"tracker_errors_{self.timestamp}.txt"
            filepath = os.path.join(folder, filename)

            # Write the file
            with open(filepath, "w", encoding="utf-8") as f:
                f.write((message or "").strip())

            # Success confirmation
            self.show_alert(message="✅ Text report saved successfully", alert_type="info")

            return filepath

        except Exception as e:
            self.show_alert(message="❌ Failed to export text report", alert_type="info")
            return None

    def show_alert(self, message: str, alert_type: str = "info", first_line_col = "#E1AA36", parent: QWidget | None = None):
        parent = parent or self.btn_run_process
        if alert_type in ("error", "warning"):
            first_line_col = "#F63049"
            return self._show_scrollable_alert(
                message=message,
                alert_type=alert_type,
                first_line_col=first_line_col,
                parent=parent,
            )
        elif alert_type == "info" and first_line_col == "#E1AA36":
            first_line_col = "#FFFFFF"

        msg = QMessageBox(parent)

        # --- Separate first line and the rest ---
        lines = message.split("\n", 1)
        first_line = lines[0]
        rest_text = lines[1] if len(lines) > 1 else ""

        # ✅ Convert any "\n" in the rest_text into HTML line breaks
        rest_html = rest_text.replace("\n", "<br>")

        # --- HTML formatting (bold yellow header + gray body) ---
        styled_message = (
            f"<p style='color:{first_line_col}; font-family:\"Inter\"; font-weight:bold; font-size:12px; margin-bottom:6px;'>"
            f"{first_line}</p>"
            f"<p style='color:#B2B2B2; font-family:\"Inter\"; font-size:12px; line-height:1.4;'>"
            f"{rest_html}</p>"
        )
        
        msg.setText(styled_message)
        msg.setTextFormat(Qt.TextFormat.RichText)  # enable HTML rendering
        msg.setStyleSheet(ss.msg_stylesheet)

        # --- Icon & titles ---
        if alert_type == "error" or alert_type == "warning":
            if alert_type == "error":
                msg.setIcon(QMessageBox.Icon.Critical)
                msg.setWindowTitle("Error")
            elif alert_type == "warning":
                msg.setIcon(QMessageBox.Icon.Warning)
                msg.setWindowTitle("Warning")

            # Remove default buttons
            msg.setStandardButtons(QMessageBox.StandardButton.NoButton)

            # Add custom buttons
            export_btn = msg.addButton("Export Error Report", QMessageBox.ButtonRole.AcceptRole)
            close_btn = msg.addButton("Close", QMessageBox.ButtonRole.RejectRole)

            msg.exec()

            # Handle which button was clicked
            if msg.clickedButton() == export_btn:
                # Directly call your export function — no need to check output_location
                self.export_txt_report(message)

            return
        
        if alert_type == "info":
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setWindowTitle("Information")
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        elif alert_type == "question":
            msg.setIcon(QMessageBox.Icon.Question)
            msg.setWindowTitle("Question")
            msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        return msg.exec()

    def _show_scrollable_alert(
        self,
        message: str,
        alert_type: str,
        first_line_col: str,
        parent: QWidget | None = None,
    ) -> int:
        dialog = QDialog(parent)
        dialog.setModal(True)
        dialog.setWindowTitle("Error" if alert_type == "error" else "Warning")
        dialog.setStyleSheet(
            """
            QDialog {
                background-color: #111F35;
                border: 1px solid #2D3E58;
            }
            QLabel {
                color: #F3F6F8;
                font-family: "Inter";
                font-size: 12px;
            }
            QTextEdit {
                background-color: #111F35;
                color: #B2B2B2;
                border: none;
                font-family: "Inter";
                font-size: 12px;
                line-height: 1.4;
                padding: 0px;
            }
            QScrollBar:vertical {
                border: none;
                background: #0D1728;
                width: 8px;
                margin: 0px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #42546F;
                min-height: 24px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical:hover {
                background: #657A98;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
            QPushButton {
                background-color: #7A808A;
                border: 1px solid #7A808A;
                border-radius: 10px;
                color: #F3F6F8;
                font-weight: 700;
                min-width: 170px;
                max-width: 170px;
                min-height: 30px;
                max-height: 30px;
            }
            QPushButton:hover {
                background-color: #8B929D;
                border: 1px solid #8B929D;
            }
            QPushButton:pressed {
                background-color: #666C75;
                border: 1px solid #666C75;
            }
            """
        )

        lines = (message or "").split("\n", 1)
        first_line = lines[0] if lines else ""
        rest_text = lines[1] if len(lines) > 1 else ""
        escaped_header = html.escape(first_line)
        escaped_body = html.escape(rest_text).replace("\n", "<br>")

        screen = None
        if parent is not None and parent.windowHandle() is not None:
            screen = parent.windowHandle().screen()
        if screen is None:
            app = QApplication.instance()
            if app is not None:
                screen = app.primaryScreen()
        if screen is not None:
            available = screen.availableGeometry()
            dialog.setMaximumSize(int(available.width() * 0.95), int(available.height() * 0.92))
            dialog.resize(min(920, int(available.width() * 0.84)), min(740, int(available.height() * 0.82)))
        dialog.setMinimumSize(560, 420)

        root = QVBoxLayout(dialog)
        root.setContentsMargins(16, 12, 16, 12)
        root.setSpacing(10)

        top_row = QHBoxLayout()
        top_row.setSpacing(12)

        icon_label = QLabel()
        if alert_type == "error":
            icon = dialog.style().standardIcon(QStyle.StandardPixmap.SP_MessageBoxCritical)
        else:
            icon = dialog.style().standardIcon(QStyle.StandardPixmap.SP_MessageBoxWarning)
        icon_label.setPixmap(icon.pixmap(34, 34))
        icon_label.setAlignment(Qt.AlignTop)
        top_row.addWidget(icon_label, 0, Qt.AlignTop)

        text_view = QTextEdit()
        text_view.setReadOnly(True)
        text_view.setHtml(
            f"<p style='color:{first_line_col}; font-weight:bold; margin-bottom:6px;'>{escaped_header}</p>"
            f"<p style='color:#B2B2B2; margin-top:0px;'>{escaped_body}</p>"
        )
        top_row.addWidget(text_view, 1)
        root.addLayout(top_row, 1)

        button_row = QHBoxLayout()
        button_row.addStretch(1)
        export_btn = QPushButton("Export Error Report")
        close_btn = QPushButton("Close")
        button_row.addWidget(export_btn)
        button_row.addSpacing(10)
        button_row.addWidget(close_btn)
        button_row.addStretch(1)
        root.addLayout(button_row)

        export_btn.clicked.connect(dialog.accept)
        close_btn.clicked.connect(dialog.reject)

        result = dialog.exec()
        if result == QDialog.DialogCode.Accepted:
            self.export_txt_report(message)
        return result

    def clear_all_fields(self):
        """Clears all input fields, checkboxes, and text values in the Status Collector section."""
        if hasattr(self, "editText_selected_trackers"):
            self.editText_selected_trackers.clear()
        if hasattr(self, "editText_output_location"):
            self.editText_output_location.clear()
        if hasattr(self, "textEdit01"):
            self.textEdit01.clear()
        
        self.radiobtn01.setChecked(True)   # Status Select
        self.radiobtn02.setChecked(False)  # Input Status
        self.status_behavior()

        for checkbox in self.checkboxes:
            checkbox.setChecked(False)

        self.excel_trackers = [] #SAP data for masterfile, -all existing packshots
        self.excel_tracker_names = []
        self.output_location = ""
        self.statuses = []
        self.row_threshold = 15  # max rows to search for header

        self.trackers_warnings_count = 0
        self.status_error_count = 0
        self.status_missing_on_some_sheets = "none" #can be none, partial or total. This is for what type of alert message to show. Ex, if partial, then generate only a warning
        self.successful_run = False
        self.can_proceed = False
        self.general_report = ""





    def process_steps(self) -> None:
        alert_args = None
        should_clear_fields = False

        self._set_processing_state(True)
        try:
            # Reset run-specific state to avoid carrying data across repeated runs.
            self.can_proceed = False
            self.successful_run = False
            self.general_report = ""
            self.statuses = []
            self.target_final_info = []
            self.trackers_warnings_count = 0
            self.status_error_count = 0
            self.status_missing_on_some_sheets = "none"

            self.statuses_selected()
            self.check_fields()

            if not self.can_proceed:
                alert_args = {"message": self.general_report, "alert_type": "error"}
            else:
                #if fields are all filled, proceed to check for tracker errors:
                #all the 5 necessary column names (idh number, product name, status, build type or if master/clone/resizing, project name)
                #check if excel file have "tracker" named sheet(s), if none, will also generate an error
                self.check_if_columns_needed_exist()

                if self.trackers_warnings_count == 0:
                    #collects data according to status specified by user
                    #check if status specified exists in all or some of the trackers
                    self.get_all_inputs_according_to_status_chatgpt_version_2(self.statuses)

                    if self.status_missing_on_some_sheets == "none":
                        alert_args = {
                            "message": self.general_report,
                            "alert_type": "info",
                            "first_line_col": "#1bab02",
                        }
                    else:
                        alert_args = {
                            "message": self.general_report,
                            "alert_type": "warning",
                            "first_line_col": "#1bab02",
                        }

                    should_clear_fields = True
                else:
                    # Show the 3-button error dialog (same as Tracker Window Import Trackers)
                    dlg_result = self._show_tracker_import_error_dialog(self.general_report)
                    if dlg_result == 1:   # Export error details
                        self.export_txt_report(self.general_report)
                    if dlg_result in (0, 1):  # Close or Export → abort run
                        should_clear_fields = True
                    else:  # result == 2 → Ignore errors, run with available data
                        self.get_all_inputs_according_to_status_chatgpt_version_2(self.statuses)
                        if self.status_missing_on_some_sheets == "none":
                            alert_args = {
                                "message": self.general_report,
                                "alert_type": "info",
                                "first_line_col": "#1bab02",
                            }
                        else:
                            alert_args = {
                                "message": self.general_report,
                                "alert_type": "warning",
                                "first_line_col": "#1bab02",
                            }
                        should_clear_fields = True
        finally:
            self._set_processing_state(False)

        if alert_args:
            self.show_alert(**alert_args)
        if should_clear_fields:
            #clear fields after first run. In case user runs the app continuously for another batch
            self.clear_all_fields()
                

    # -----------------------------------------------------------------------
    # Tracker Window (Option 1) – Import Trackers
    # -----------------------------------------------------------------------

    @staticmethod
    def _apply_row_cleanup(row_data: list, col_display_names: list) -> list:
        """Apply data-cleanup transformations to a single row.

        Transformations:
        - Lowercase: Status, Build Type, Packaging Type, Packaging Size, Label Size
        - Build Type "resizing" misspelling fix (resize / resized / reszing / resizng …)
        - Build Type "Master Variant" → "master"
        - Build Type containing "clone" → "clone"
        - Build Type not in valid set → "na"
        - Status "not needed" / "existing" / "duplicate" → "cancelled"
        - Status "cannot proceed" → "on hold"
        """
        _LOWERCASE_COLS = {"Status", "Build Type", "Packaging Type", "Packaging Size", "Label Size"}
        _RESIZING_RE = re.compile(r'^res[ie]?z[iyenigds]{0,6}$', re.IGNORECASE)
        _VALID_BUILD_TYPES = {"master", "clone", "resizing", "upload", "admin"}

        result = list(row_data)
        for i, name in enumerate(col_display_names):
            val = str(result[i]) if not isinstance(result[i], str) else result[i]

            if name in _LOWERCASE_COLS:
                val = val.lower()

            if name == "Build Type":
                val_l = val.lower()
                if "clone" in val_l:
                    val = "clone"
                elif val_l == "master variant":
                    val = "master"
                elif _RESIZING_RE.match(val_l):
                    val = "resizing"
                # After normalisation, reject anything not in the valid set
                if val not in _VALID_BUILD_TYPES:
                    val = "na"

            elif name == "Status":
                val_l = val.lower()
                # Substring-based normalisations (case-insensitive)
                if "descoped" in val_l:
                    val = "cancelled"
                # Phrases meaning no action → cancelled
                elif any(kw in val_l for kw in ("no action needed", "no action", "no need", "no needed")):
                    val = "cancelled"
                # 'moved' → cancelled
                elif "moved" in val_l:
                    val = "cancelled"
                # Phrases meaning on hold
                elif any(kw in val_l for kw in ("can not start", "cannot start", "cant start", "can't start")):
                    val = "on hold"
                elif "waiting" in val_l:
                    val = "on hold"
                elif "tbc" in val_l or "to confirm" in val_l:
                    val = "on hold"
                # Upload-related → completed
                elif "uploaded" in val_l or "upload" in val_l:
                    val = "completed"
                # Existing exact-match mappings (preserve behaviour)
                elif val_l in ("not needed", "existing", "duplicate"):
                    val = "cancelled"
                elif val_l == "cannot proceed":
                    val = "on hold"

            result[i] = val
        return result

    def attach_tracker_window_dialog(self, dialog) -> None:
        """Connect Tracker Window dialog buttons that are handled by StatusCollector."""
        _start = getattr(dialog, "_browse_start_dir", "")
        dialog.btn_import_tracker.clicked.connect(lambda: self.import_trackers_to_window(dialog, _start))
        dialog.btn_apply_cleanup.clicked.connect(lambda: self.apply_cleanup_to_table(dialog))
        dialog.btn_load_tsc.clicked.connect(lambda: self.load_tsc_output_to_window(dialog, _start))
        dialog.btn_export.clicked.connect(lambda: self.export_tracker_table_to_xlsx(dialog))

    def export_tracker_table_to_xlsx(self, dialog) -> None:
        """Export all visible rows of the Tracker Table to a styled .xlsx file."""
        from datetime import datetime
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        table = dialog.table
        col_count = table.columnCount()
        # Strip sort-arrow suffixes (▾ / ▾*) so the exported file has clean headers.
        headers = [
            re.sub(r'\s*▾\*?\s*$', '', table.horizontalHeaderItem(c).text()).strip()
            if table.horizontalHeaderItem(c) else f"Col{c+1}"
            for c in range(col_count)
        ]

        # Collect only visible rows
        rows: list[list[str]] = []
        for r in range(table.rowCount()):
            if table.isRowHidden(r):
                continue
            rows.append([
                table.item(r, c).text() if table.item(r, c) else ""
                for c in range(col_count)
            ])

        if not rows:
            self.show_alert("Nothing to export – the table is empty.", alert_type="info")
            return

        # Default filename – pre-fill the save directory with the Option 2
        # output location when it has been configured.
        ts = datetime.now().strftime("%Y_%m_%d_%H_%M")
        default_stem = f"tsc_data_{ts}.xlsx"
        start_path = (
            os.path.join(self.output_location, default_stem)
            if self.output_location
            else default_stem
        )

        save_path, _ = QFileDialog.getSaveFileName(
            dialog,
            "Export Tracker Table",
            start_path,
            "Excel Files (*.xlsx)",
        )
        if not save_path:
            return

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tracker Data"

        # ── Styles ──────────────────────────────────────────────────────────
        header_fill   = PatternFill(fill_type="solid", fgColor="111F35")
        header_font   = Font(color="FFFFFF", bold=True, size=11, name="Segoe UI")
        body_font     = Font(color="1B1F24", size=10, name="Segoe UI")
        alt_fill      = PatternFill(fill_type="solid", fgColor="F3F7FC")
        thin_border   = Border(
            left=Side(style="thin", color="D9E1EA"),
            right=Side(style="thin", color="D9E1EA"),
            top=Side(style="thin", color="D9E1EA"),
            bottom=Side(style="thin", color="D9E1EA"),
        )
        center_align  = Alignment(horizontal="center", vertical="center")
        left_align    = Alignment(horizontal="left",   vertical="center")

        # ── Header row ──────────────────────────────────────────────────────
        for c_idx, hdr in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=hdr)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border
        ws.row_dimensions[1].height = 24

        # ── Data rows ───────────────────────────────────────────────────────
        for r_idx, row_data in enumerate(rows, start=2):
            use_alt = (r_idx % 2 == 0)
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font      = body_font
                cell.alignment = left_align
                cell.border    = thin_border
                if use_alt:
                    cell.fill = alt_fill

        # ── Auto-filter on header row ────────────────────────────────────────
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(col_count)}{len(rows) + 1}"
        )

        # ── Freeze header and hide gridlines ────────────────────────────────
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        # ── Auto-fit column widths ───────────────────────────────────────────
        for c_idx, hdr in enumerate(headers, start=1):
            col_letter = get_column_letter(c_idx)
            max_len = len(hdr)
            for row_data in rows[:5000]:
                cell_len = len(str(row_data[c_idx - 1]))
                if cell_len > max_len:
                    max_len = cell_len
            ws.column_dimensions[col_letter].width = min(max(14, max_len + 2), 44)

        try:
            wb.save(save_path)
            self.show_alert("✅ Tracker table exported successfully.", alert_type="info")
        except Exception as exc:
            self.show_alert(f"❌ Failed to save file:\n{exc}", alert_type="error")

    def apply_cleanup_to_table(self, dialog) -> None:
        """Apply cleanup transformations to every row currently in the Tracker table.

        Rows are hidden when any of these conditions are true:
        - IDH Number is blank
        - IDH Number contains no digit
        - Build Type is blank
        - Status is blank
        All other rows receive the full _apply_row_cleanup treatment.
        """
        col_display_names = list(self.TRACKER_WINDOW_COLUMN_VARIATIONS.keys())
        try:
            idh_idx    = col_display_names.index("IDH Number")
            bt_idx     = col_display_names.index("Build Type")
            status_idx = col_display_names.index("Status")
        except ValueError:
            return

        _digits_only = re.compile(r'^\d+$')  # IDH must be purely numeric to be kept

        table = dialog.table
        table.blockSignals(True)
        try:
            for r in range(table.rowCount()):
                idh_item    = table.item(r, idh_idx)
                bt_item     = table.item(r, bt_idx)
                status_item = table.item(r, status_idx)
                idh_val     = idh_item.text().strip()    if idh_item    else ""
                bt_val      = bt_item.text().strip()     if bt_item     else ""
                status_val  = status_item.text().strip() if status_item else ""

                if (not idh_val
                        or not _digits_only.match(idh_val)
                        or not bt_val
                        or not status_val):
                    table.setRowHidden(r, True)
                    continue

                table.setRowHidden(r, False)

                row_data = [
                    (table.item(r, c).text() if table.item(r, c) else "")
                    for c in range(len(col_display_names))
                ]
                cleaned = self._apply_row_cleanup(row_data, col_display_names)
                for c, val in enumerate(cleaned):
                    item = table.item(r, c)
                    if item is None:
                        table.setItem(r, c, QTableWidgetItem(val))
                    else:
                        item.setText(val)
        finally:
            table.blockSignals(False)

        dialog._push_undo_snapshot(force=True)
        dialog._update_row_count_display()

    def import_trackers_to_window(self, dialog, start_dir: str = "") -> None:
        """Open file dialog, validate trackers, populate the Tracker Information table."""
        # Read the starting row from the dialog UI (1-based Excel row)
        starting_row = 11
        try:
            if hasattr(dialog, 'starting_row_input'):
                raw = dialog.starting_row_input.text().strip()
                if raw:
                    starting_row = max(1, int(raw))
        except (ValueError, TypeError):
            starting_row = 11

        file_dialog = QFileDialog()
        file_dialog.setWindowTitle("Select Excel Tracker File(s)")
        file_dialog.setNameFilter("Excel Files (*.xlsx *.xlsm *.xltx *.xltm)")
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)
        if start_dir:
            file_dialog.setDirectory(start_dir)

        if not file_dialog.exec():
            return

        selected_files = file_dialog.selectedFiles()
        if not selected_files:
            return

        col_variations = self.TRACKER_WINDOW_COLUMN_VARIATIONS
        col_display_names = list(col_variations.keys())

        errors: dict[str, list[str]] = {}
        all_rows: list[list[str]] = []

        dialog.import_progress.setVisible(True)
        dialog.import_progress.setRange(0, 0)  # indeterminate spinner
        QApplication.processEvents()

        try:
            for file_path in selected_files:
                file_name = os.path.basename(file_path)
                file_errors: list[str] = []

                try:
                    xl = pd.ExcelFile(file_path, engine="openpyxl")
                except Exception as exc:
                    errors[file_name] = [f"Failed to open file: {exc}"]
                    continue

                tracker_sheets = [s for s in xl.sheet_names if "tracker" in s.lower()]

                if not tracker_sheets:
                    file_errors.append("missing tracker sheet")
                    errors[file_name] = file_errors
                    continue

                for sheet in tracker_sheets:
                    # ── TSC output file detection ──────────────────────────────────────────
                    # The exported tsc_data_*.xlsx file has a 'Tracker Data' sheet whose
                    # first row already uses the display column names (TSC_HEADERS).  Skip
                    # the usual tracker-validation pipeline and import it directly.
                    if self._is_tsc_output_sheet(xl, sheet):
                        all_rows.extend(self._read_tsc_output_sheet(xl, sheet))
                        continue

                    # ── Normal tracker processing ─────────────────────────────────────────
                    try:
                        df_raw = pd.read_excel(
                            xl,
                            sheet_name=sheet,
                            nrows=self.row_threshold,
                            header=None,
                            dtype=str,
                            na_filter=False,
                            engine="openpyxl",
                        )
                    except Exception:
                        file_errors.append(f"sheet '{sheet}': could not be read")
                        continue

                    # Normalize a raw cell value to a clean lowercase string
                    def _norm(val: object) -> str:
                        return re.sub(r'\s+', ' ', str(val).strip().lower())

                    # Locate header row – require at least 2 matching known column names
                    # so that a random cell containing a common word (e.g. "status") in a
                    # title/note row does not get mistaken for the header.
                    all_known_variations = [
                        v for variations in col_variations.values() for v in variations
                    ]
                    header_row_idx = -1
                    for i in range(len(df_raw)):
                        row_vals_lower = [_norm(v) for v in df_raw.iloc[i].values]
                        match_count = sum(1 for known in all_known_variations if known in row_vals_lower)
                        if match_count >= 2:
                            header_row_idx = i
                            break

                    if header_row_idx == -1:
                        file_errors.append(f"sheet \'{sheet}\': could not find header row")
                        continue

                    col_names_in_row = [_norm(v) for v in df_raw.iloc[header_row_idx].values]

                    # Check for missing columns (variations are also normalised for safety)
                    missing_cols = [
                        display_name
                        for display_name, variations in col_variations.items()
                        if not any(re.sub(r'\s+', ' ', v.strip().lower()) in col_names_in_row
                                   for v in variations)
                    ]

                    if missing_cols:
                        file_errors.append(
                            f"sheet \'{sheet}\': missing columns: {', '.join(missing_cols)}"
                        )
                        continue

                    # Read full data from this sheet
                    try:
                        df_data = pd.read_excel(
                            xl,
                            sheet_name=sheet,
                            header=header_row_idx,
                            dtype=str,
                            na_filter=False,
                            engine="openpyxl",
                        )
                    except Exception:
                        file_errors.append(f"sheet \'{sheet}\': failed to read data")
                        continue

                    # Apply starting row: skip data rows that lie before the configured
                    # starting row in the original Excel file (1-based).  The first data
                    # row in df_data corresponds to Excel row (header_row_idx + 2) in
                    # 1-based terms, so we drop the leading rows accordingly.
                    skip_rows = max(0, starting_row - (header_row_idx + 2))
                    if skip_rows > 0:
                        df_data = df_data.iloc[skip_rows:].reset_index(drop=True)

                    # Resolve each display column to actual DataFrame column
                    normalized_cols = {
                        re.sub(r'\s+', ' ', str(c).strip().lower()): c
                        for c in df_data.columns
                    }

                    def _find_col(variations: list[str]) -> str | None:
                        for v in variations:
                            key = re.sub(r'\s+', ' ', v.strip().lower())
                            if key in normalized_cols:
                                return normalized_cols[key]
                        return None

                    col_map = {
                        name: _find_col(variations)
                        for name, variations in col_variations.items()
                    }

                    idh_col = col_map.get("IDH Number")
                    build_type_col = col_map.get("Build Type")

                    for _, row in df_data.iterrows():
                        idh_val = str(row.get(idh_col, "")).strip() if idh_col else ""
                        # Always skip admin header/meta rows
                        if idh_val.lower() == "admin":
                            continue
                        if build_type_col:
                            bt_val = str(row.get(build_type_col, "")).strip().lower()
                            if bt_val == "admin":
                                continue
                        row_data = [
                            str(row.get(col_map[name], "")).strip() if col_map[name] else ""
                            for name in col_display_names
                        ]
                        all_rows.append(row_data)

                if file_errors:
                    errors[file_name] = file_errors

        finally:
            dialog.import_progress.setRange(0, 100)
            dialog.import_progress.setValue(0)
            dialog.import_progress.setVisible(False)

        # ---- Report errors, if any ----
        if errors:
            total = len(selected_files)
            err_count = len(errors)
            msg_lines = [
                f"Trackers selected: {total}",
                f"Trackers with errors: {err_count}",
            ]
            for idx, (fname, errs) in enumerate(errors.items(), 1):
                msg_lines.append("")  # blank line between tracker entries
                msg_lines.append(f"{idx}. {fname}")
                for a_idx, err in enumerate(errs):
                    msg_lines.append(f"\t{chr(ord('a') + a_idx)}. {err}")
            error_message = "\n".join(msg_lines)
            result = self._show_tracker_import_error_dialog(error_message, parent=dialog)
            if result == 1:   # Export error details
                self.export_txt_report(error_message)
            elif result == 2:  # Ignore errors – import available columns
                self._import_with_partial_columns(selected_files, starting_row, dialog)
            return

        if not all_rows:
            self.show_alert("No valid rows found in the selected tracker(s).", alert_type="info")
            return

        # ---- Populate table ----
        dialog.table.blockSignals(True)
        try:
            dialog.table.setRowCount(len(all_rows))
            col_count = dialog.table.columnCount()
            for r, row_data in enumerate(all_rows):
                for c in range(col_count):
                    value = row_data[c] if c < len(row_data) else ""
                    item = dialog.table.item(r, c)
                    if item is None:
                        dialog.table.setItem(r, c, QTableWidgetItem(value))
                    else:
                        item.setText(value)
        finally:
            dialog.table.blockSignals(False)

        dialog.row_count_input.setText(str(dialog.table.rowCount()))
        dialog._push_undo_snapshot(force=True)
        dialog._update_row_count_display()
        dialog.table.clearSelection()

    def _show_tracker_import_error_dialog(self, message: str, parent=None) -> int:
        """Scrollable error dialog.  Returns 0=close, 1=export error details, 2=ignore errors."""
        parent = parent or self.btn_run_process

        dialog = QDialog(parent)
        dialog.setModal(True)
        dialog.setWindowTitle("Import Error")
        dialog.setStyleSheet(
            """
            QDialog {
                background-color: #111F35;
                border: 1px solid #2D3E58;
            }
            QLabel {
                color: #F3F6F8;
                font-family: "Inter";
                font-size: 12px;
            }
            QTextEdit {
                background-color: #111F35;
                color: #B2B2B2;
                border: none;
                font-family: "Inter";
                font-size: 12px;
                line-height: 1.4;
                padding: 0px;
            }
            QScrollBar:vertical {
                border: none;
                background: #0D1728;
                width: 8px;
                margin: 0px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #42546F;
                min-height: 24px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical:hover {
                background: #657A98;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
            QPushButton {
                background-color: #7A808A;
                border: 1px solid #7A808A;
                border-radius: 10px;
                color: #F3F6F8;
                font-weight: 700;
                min-width: 140px;
                max-width: 140px;
                min-height: 30px;
                max-height: 30px;
            }
            QPushButton:hover {
                background-color: #8B929D;
                border: 1px solid #8B929D;
            }
            QPushButton:pressed {
                background-color: #666C75;
                border: 1px solid #666C75;
            }
            """
        )

        app = QApplication.instance()
        if app is not None:
            screen = app.primaryScreen()
            if screen is not None:
                avail = screen.availableGeometry()
                dialog.resize(
                    min(860, int(avail.width() * 0.8)),
                    min(640, int(avail.height() * 0.75)),
                )
        dialog.setMinimumSize(520, 380)

        root = QVBoxLayout(dialog)
        root.setContentsMargins(16, 12, 16, 12)
        root.setSpacing(10)

        top_row = QHBoxLayout()
        top_row.setSpacing(12)

        icon_label = QLabel()
        icon = dialog.style().standardIcon(QStyle.StandardPixmap.SP_MessageBoxCritical)
        icon_label.setPixmap(icon.pixmap(34, 34))
        icon_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        top_row.addWidget(icon_label, 0, Qt.AlignmentFlag.AlignTop)

        # Build structured HTML from message lines
        all_lines = message.split("\n")
        html_parts: list[str] = []
        for line in all_lines:
            if not line.strip():
                # Blank spacer between tracker entries
                html_parts.append("<p style='margin:0; padding:0; line-height:0.6;'>&nbsp;</p>")
            elif line.startswith("\t") or line.startswith("    "):
                # Sub-item lines (a., b., ...)
                escaped = html.escape(line.strip())
                html_parts.append(
                    f"<p style='color:#B2B2B2; font-family:\"Inter\"; font-size:12px; "
                    f"margin:0 0 2px 32px; padding:0;'>{escaped}</p>"
                )
            elif line[:1].isdigit():
                # Numbered tracker filename lines
                escaped = html.escape(line)
                html_parts.append(
                    f"<p style='color:#F3F6F8; font-family:\"Inter\"; font-size:12px; "
                    f"font-weight:600; margin:0 0 2px 0; padding:0;'>{escaped}</p>"
                )
            else:
                # Stat lines (Trackers selected / Trackers with errors)
                escaped = html.escape(line)
                html_parts.append(
                    f"<p style='color:#FFFFFF; font-family:\"Inter\"; font-size:12px; "
                    f"font-weight:700; margin:0 0 3px 0; padding:0;'>{escaped}</p>"
                )
        full_html = "".join(html_parts)

        from PySide6.QtWidgets import QTextEdit as _QTextEdit
        text_view = _QTextEdit()
        text_view.setReadOnly(True)
        text_view.setHtml(f"<div>{full_html}</div>")
        top_row.addWidget(text_view, 1)
        root.addLayout(top_row, 1)

        button_row = QHBoxLayout()
        button_row.addStretch(1)
        export_btn = QPushButton("Export error details")
        close_btn = QPushButton("Close")
        ignore_btn = QPushButton("Ignore errors - import available columns")
        ignore_btn.setStyleSheet("min-width: 290px; max-width: 290px;")
        button_row.addWidget(export_btn)
        button_row.addSpacing(10)
        button_row.addWidget(close_btn)
        button_row.addSpacing(10)
        button_row.addWidget(ignore_btn)
        root.addLayout(button_row)

        export_btn.clicked.connect(lambda: dialog.done(1))
        close_btn.clicked.connect(lambda: dialog.done(0))
        ignore_btn.clicked.connect(lambda: dialog.done(2))

        return dialog.exec()

    def _import_with_partial_columns(self, selected_files: list, starting_row: int, dialog) -> None:
        """Re-import all selected files using only the columns that are actually present.
        Columns that are missing in a sheet are left blank for every row from that sheet.
        """
        col_variations = self.TRACKER_WINDOW_COLUMN_VARIATIONS
        col_display_names = list(col_variations.keys())

        all_rows: list[list[str]] = []

        dialog.import_progress.setVisible(True)
        dialog.import_progress.setRange(0, 0)
        QApplication.processEvents()

        try:
            for file_path in selected_files:
                file_name = os.path.basename(file_path)

                try:
                    xl = pd.ExcelFile(file_path, engine="openpyxl")
                except Exception:
                    continue  # unreadable file – skip silently

                tracker_sheets = [s for s in xl.sheet_names if "tracker" in s.lower()]
                if not tracker_sheets:
                    continue

                for sheet in tracker_sheets:
                    try:
                        df_raw = pd.read_excel(
                            xl,
                            sheet_name=sheet,
                            nrows=self.row_threshold,
                            header=None,
                            dtype=str,
                            na_filter=False,
                            engine="openpyxl",
                        )
                    except Exception:
                        continue

                    def _norm(val: object) -> str:
                        return re.sub(r'\s+', ' ', str(val).strip().lower())

                    all_known_variations = [
                        v for variations in col_variations.values() for v in variations
                    ]
                    header_row_idx = -1
                    for i in range(len(df_raw)):
                        row_vals_lower = [_norm(v) for v in df_raw.iloc[i].values]
                        match_count = sum(1 for known in all_known_variations if known in row_vals_lower)
                        if match_count >= 2:
                            header_row_idx = i
                            break

                    if header_row_idx == -1:
                        continue  # cannot locate header – skip

                    try:
                        df_data = pd.read_excel(
                            xl,
                            sheet_name=sheet,
                            header=header_row_idx,
                            dtype=str,
                            na_filter=False,
                            engine="openpyxl",
                        )
                    except Exception:
                        continue

                    # Apply starting row offset
                    skip_rows = max(0, starting_row - (header_row_idx + 2))
                    if skip_rows > 0:
                        df_data = df_data.iloc[skip_rows:].reset_index(drop=True)

                    # Resolve columns – missing ones map to None (no error raised)
                    normalized_cols = {
                        re.sub(r'\s+', ' ', str(c).strip().lower()): c
                        for c in df_data.columns
                    }

                    def _find_col(variations: list) -> str | None:
                        for v in variations:
                            key = re.sub(r'\s+', ' ', v.strip().lower())
                            if key in normalized_cols:
                                return normalized_cols[key]
                        return None

                    col_map = {
                        name: _find_col(variations)
                        for name, variations in col_variations.items()
                    }

                    idh_col       = col_map.get("IDH Number")
                    build_type_col = col_map.get("Build Type")

                    for _, row in df_data.iterrows():
                        idh_val = str(row.get(idh_col, "")).strip() if idh_col else ""
                        # Always skip admin header/meta rows
                        if idh_val.lower() == "admin":
                            continue
                        if build_type_col:
                            bt_val = str(row.get(build_type_col, "")).strip().lower()
                            if bt_val == "admin":
                                continue
                        row_data = [
                            str(row.get(col_map[name], "")).strip() if col_map[name] else ""
                            for name in col_display_names
                        ]
                        all_rows.append(row_data)

        finally:
            dialog.import_progress.setRange(0, 100)
            dialog.import_progress.setValue(0)
            dialog.import_progress.setVisible(False)

        if not all_rows:
            self.show_alert("No valid rows found in the selected tracker(s).", alert_type="info")
            return

        # Populate table
        dialog.table.blockSignals(True)
        try:
            dialog.table.setRowCount(len(all_rows))
            col_count = dialog.table.columnCount()
            for r, row_data in enumerate(all_rows):
                for c in range(col_count):
                    value = row_data[c] if c < len(row_data) else ""
                    item = dialog.table.item(r, c)
                    if item is None:
                        dialog.table.setItem(r, c, QTableWidgetItem(value))
                    else:
                        item.setText(value)
        finally:
            dialog.table.blockSignals(False)

        dialog.row_count_input.setText(str(dialog.table.rowCount()))
        dialog._push_undo_snapshot(force=True)
        dialog._update_row_count_display()
        dialog.table.clearSelection()

    def load_tsc_output_to_window(self, dialog, start_dir: str = "") -> None:
        """Open a tsc_data_*.xlsx export file and populate the Tracker table directly.

        Unlike Import Trackers, this path bypasses tracker validation and header
        searching entirely — it reads the sheet whose first row already contains
        the canonical display column names (as produced by Export).
        """
        file_dialog = QFileDialog()
        file_dialog.setWindowTitle("Select TSC Output File")
        file_dialog.setNameFilter("Excel Files (*.xlsx)")
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)
        if start_dir:
            file_dialog.setDirectory(start_dir)

        if not file_dialog.exec():
            return
        selected = file_dialog.selectedFiles()
        if not selected:
            return

        tsc_display_names = list(self.TRACKER_WINDOW_COLUMN_VARIATIONS.keys())

        all_rows: list[list[str]] = []

        dialog.import_progress.setVisible(True)
        dialog.import_progress.setRange(0, 0)
        QApplication.processEvents()

        try:
            for file_path in selected:
                try:
                    xl = pd.ExcelFile(file_path, engine="openpyxl")

                    # Prefer a sheet named "Tracker Data" (exact, case-insensitive);
                    # fall back to the first sheet in the workbook.
                    preferred = next(
                        (s for s in xl.sheet_names if s.strip().lower() == "tracker data"),
                        xl.sheet_names[0],
                    )

                    target_df = pd.read_excel(
                        file_path, sheet_name=preferred,
                        header=0, dtype=str,
                        na_filter=False, engine="openpyxl",
                    )

                    # Normalize column names: strip sort-arrow suffixes (▾, ▾*)
                    # that the Export function writes from the Qt table headers.
                    clean_cols = {
                        re.sub(r'\s*▾\*?\s*$', '', str(c)).strip(): str(c)
                        for c in target_df.columns
                    }
                    # Rename df columns to their clean versions so matching works.
                    target_df.columns = [
                        re.sub(r'\s*▾\*?\s*$', '', str(c)).strip()
                        for c in target_df.columns
                    ]

                    for _, row in target_df.iterrows():
                        row_data = [
                            str(row[name]).strip() if name in target_df.columns else ""
                            for name in tsc_display_names
                        ]
                        all_rows.append(row_data)

                except Exception as exc:
                    self.show_alert(
                        f"Could not open file:\n{os.path.basename(file_path)}\n\n{exc}",
                        alert_type="error",
                    )
        finally:
            dialog.import_progress.setRange(0, 100)
            dialog.import_progress.setValue(0)
            dialog.import_progress.setVisible(False)

        if not all_rows:
            self.show_alert("No data found in the selected TSC output file(s).", alert_type="info")
            return

        dialog.table.blockSignals(True)
        try:
            dialog.table.setRowCount(len(all_rows))
            col_count = dialog.table.columnCount()
            for r, row_data in enumerate(all_rows):
                for c in range(col_count):
                    value = row_data[c] if c < len(row_data) else ""
                    item = dialog.table.item(r, c)
                    if item is None:
                        dialog.table.setItem(r, c, QTableWidgetItem(value))
                    else:
                        item.setText(value)
        finally:
            dialog.table.blockSignals(False)

        dialog.row_count_input.setText(str(dialog.table.rowCount()))
        dialog._push_undo_snapshot(force=True)
        dialog._update_row_count_display()
        dialog.table.clearSelection()

    # -----------------------------------------------------------------------
    # TSC output-file helpers
    # -----------------------------------------------------------------------

    def _is_tsc_output_sheet(self, xl: pd.ExcelFile, sheet_name: str) -> bool:
        """Return True when *sheet_name* looks like a TSC exported output sheet.

        Detection criteria: the sheet's first row contains at least half of the
        canonical display column names (TRACKER_WINDOW_COLUMN_VARIATIONS keys).
        This distinguishes a pre-processed export from raw tracker sheets.
        """
        tsc_names = set(self.TRACKER_WINDOW_COLUMN_VARIATIONS.keys())
        try:
            df = pd.read_excel(
                xl, sheet_name=sheet_name, nrows=0,
                header=0, dtype=str, na_filter=False,
            )
            actual = {str(c).strip() for c in df.columns}
            matches = len(tsc_names & actual)
            return matches >= max(3, len(tsc_names) // 2)
        except Exception:
            return False

    def _read_tsc_output_sheet(self, xl: pd.ExcelFile, sheet_name: str) -> list[list[str]]:
        """Read a TSC output sheet whose headers already are the display column names."""
        tsc_display_names = list(self.TRACKER_WINDOW_COLUMN_VARIATIONS.keys())
        try:
            df = pd.read_excel(
                xl, sheet_name=sheet_name,
                header=0, dtype=str, na_filter=False,
            )
            rows: list[list[str]] = []
            for _, row in df.iterrows():
                row_data = [
                    str(row[name]).strip() if name in df.columns else ""
                    for name in tsc_display_names
                ]
                rows.append(row_data)
            return rows
        except Exception:
            return []

    # -----------------------------------------------------------------------

    def run_process(self)-> None:
        self.btn_select_trackers.clicked.connect(self.get_tracker_files)
        self.btn_output_location.clicked.connect(self.get_output_location)
        self.btn_menu_clear_all_fields.clicked.connect(self.clear_all_fields)

        self.btn_run_process.clicked.connect(self.process_steps)


                                                                                    
