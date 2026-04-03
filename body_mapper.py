"""
body_mapper.py
==============
All processing logic for PROJECT – SAP Data Compare (SDC).

Entry point
-----------
    run_comparison(params: CompareParams) -> CompareResult

The caller (new_ui_window.py) collects UI inputs, builds a CompareParams
dataclass and calls run_comparison().  The function returns a CompareResult
that describes the output file path and any warnings.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime
from typing import NamedTuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ── colour constants ────────────────────────────────────────────────────────
_COL_HEADER_DARK   = "111F35"   # columns 1,2,4,6,7,8,9
_COL_HEADER_RED    = "D02752"   # columns 3 (Hit Type), 5 (Master IDH Build)
_COL_IDH_MASTER    = "237227"   # clone / master → green
_COL_IDH_RESIZE    = "D97A2B"   # resizing → orange
_COL_IDH_BAD       = "C00707"   # missing / admin / upload / other → red
_COL_WHITE         = "FFFFFF"

# ── build-type classification ───────────────────────────────────────────────
_GOOD_TYPES  = frozenset({"master", "clone"})
_OK_TYPES    = frozenset({"resizing"})
# everything else (including "missing", "admin", "upload", "") → bad


def _classify_build_type(bt: str) -> str:
    """Return 'good', 'ok', or 'bad' for a Build Type string."""
    v = bt.strip().lower()
    if v in _GOOD_TYPES:
        return "good"
    if v in _OK_TYPES:
        return "ok"
    return "bad"


def _idh_color(classification: str) -> str:
    """Return hex color string (no #) for a classification."""
    if classification == "good":
        return _COL_IDH_MASTER
    if classification == "ok":
        return _COL_IDH_RESIZE
    return _COL_IDH_BAD


# ── data structs ────────────────────────────────────────────────────────────

@dataclass
class CompareParams:
    """All inputs needed to run the SDC comparison."""
    rsd_target_paths:     list[str]  # one or more .xlsx files
    rsd_master_path:      str        # single .xlsx file (unused in tsc_only mode)
    tsc_data_path:        str        # single .xlsx file
    output_dir:           str        # folder to write result into
    excel_library_path:   str = ""   # optional; empty string = library not used
    compare_mode:         str = "rsd_master"  # "rsd_master" | "tsc_only"


@dataclass
class CompareResult:
    output_path: str = ""                          # last successful output (compat)
    output_paths: list[str] = field(default_factory=list)  # one per target file
    warnings: list[str] = field(default_factory=list)


# ══════════════════════════════════════════════════════════════════════════════
# Internal helpers – reading
# ══════════════════════════════════════════════════════════════════════════════

def _norm(v: object) -> str:
    """Normalise a cell value to a stripped string."""
    if v is None or (isinstance(v, float) and v != v):   # NaN check
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_sheet_df(xl: pd.ExcelFile, sheet_name: str) -> pd.DataFrame | None:
    """Read a sheet by name (case-insensitive).  Returns None if not found."""
    lower = sheet_name.lower()
    match = next((s for s in xl.sheet_names if s.lower() == lower), None)
    if match is None:
        return None
    return pd.read_excel(xl, sheet_name=match, dtype=object)


def _find_column(df: pd.DataFrame, name: str) -> str | None:
    """Find a column in df by case-insensitive partial match."""
    lower = name.lower()
    for col in df.columns:
        if lower in str(col).lower():
            return col
    return None


def _col_values(df: pd.DataFrame, col_header: str) -> list[str]:
    """Return non-empty normalised values from a column."""
    col = _find_column(df, col_header)
    if col is None:
        return []
    return [_norm(v) for v in df[col] if _norm(v)]


# ── RSD target reading ───────────────────────────────────────────────────────

class _TargetRow(NamedTuple):
    head_bom_mat: str          # Head Bom Mat value
    basic_numbers: list[str]   # parsed from "Basic Number" cell (comma-separated)
    basic_names: list[str]     # parsed from "Basic Name" cell (pipe-separated), aligned with basic_numbers


def _parse_basic_numbers(raw: str) -> list[str]:
    """Split a 'Basic Number' cell that may contain comma/semicolon-separated values."""
    parts = re.split(r"[,;\s]+", raw.strip())
    return [p.strip() for p in parts if p.strip()]


def _parse_basic_names(raw: str) -> list[str]:
    """Split a 'Basic Name' cell whose values are pipe-separated."""
    return [p.strip() for p in raw.split("|") if p.strip()]


def _read_target_rows(target_paths: list[str]) -> list[_TargetRow]:
    """
    Read all RSD: Target files.
    For each file, reads the "reformatted_sap" sheet and look for the
    Head Bom Mat + Basic Number columns.
    Returns rows in order of appearance across all target files.
    """
    rows: list[_TargetRow] = []
    for path in target_paths:
        try:
            xl = pd.ExcelFile(path, engine="openpyxl")
        except Exception:
            continue

        df = _read_sheet_df(xl, "reformatted_sap")
        if df is None or df.empty:
            # Option 1 export uses a different sheet title
            df = _read_sheet_df(xl, "Reformatted SAP Data")
        if df is None or df.empty:
            # last resort: read whatever the first sheet is
            try:
                df = pd.read_excel(xl, sheet_name=xl.sheet_names[0], dtype=object)
            except Exception:
                continue
        if df is None or df.empty:
            continue

        hbm_col  = _find_column(df, "Head Bom Mat")
        bn_col   = _find_column(df, "Basic Number")
        name_col = _find_column(df, "Basic Name")
        if hbm_col is None or bn_col is None:
            continue

        for _, row_s in df.iterrows():
            hbm      = _norm(row_s.get(hbm_col, ""))
            bn_raw   = _norm(row_s.get(bn_col, ""))
            name_raw = _norm(row_s.get(name_col, "")) if name_col else ""
            if not hbm and not bn_raw:
                continue
            bns   = _parse_basic_numbers(bn_raw) if bn_raw else []
            names = _parse_basic_names(name_raw) if name_raw else []
            # Pad names so indices align with bns
            while len(names) < len(bns):
                names.append("")
            rows.append(_TargetRow(head_bom_mat=hbm, basic_numbers=bns, basic_names=names))
    return rows


# ── RSD Grouping Count reading ───────────────────────────────────────────────

class _GroupRow(NamedTuple):
    bc: str               # BC column value
    count: str            # Count column value (kept as string)
    basic_number: str     # Basic Number column value (single BN per row)
    basic_name: str       # Basic Name column value


def _read_grouping_rows(target_paths: list[str]) -> list[_GroupRow]:
    """Read the grouping sheet from all RSD: Target files.
    Tries 'Grouping Count' (Option 1 export) then 'grouping_data' (Option 2 export)."""
    rows: list[_GroupRow] = []
    for path in target_paths:
        try:
            xl = pd.ExcelFile(path, engine="openpyxl")
        except Exception:
            continue
        df = _read_sheet_df(xl, "Grouping Count")
        if df is None or df.empty:
            df = _read_sheet_df(xl, "grouping_data")
        if df is None or df.empty:
            continue
        bc_col   = _find_column(df, "BC")
        cnt_col  = _find_column(df, "Count")
        bn_col   = _find_column(df, "Basic Number")
        name_col = _find_column(df, "Basic Name")
        if bc_col is None:
            continue
        for _, row_s in df.iterrows():
            bc    = _norm(row_s.get(bc_col,   ""))
            cnt   = _norm(row_s.get(cnt_col,  "")) if cnt_col  else ""
            bn    = _norm(row_s.get(bn_col,   "")) if bn_col   else ""
            name  = _norm(row_s.get(name_col, "")) if name_col else ""
            if not bc:
                continue
            rows.append(_GroupRow(bc=bc, count=cnt, basic_number=bn, basic_name=name))
    return rows


def _read_master_grouping_map(master_path: str) -> dict[str, list[str]]:
    """Build basic_number -> [BC values] mapping from master Grouping Count sheet.
    Reuses the same structure as master_map but keyed the same way so the same
    SDC helpers (_determine_hit_type, _resolve_master_idh, etc.) work unchanged.
    The master Grouping Count's BC column plays the role of 'Head Bom Mat'."""
    mapping: dict[str, list[str]] = {}
    try:
        xl = pd.ExcelFile(master_path, engine="openpyxl")
    except Exception:
        return mapping
    df = _read_sheet_df(xl, "Grouping Count")
    if df is None or df.empty:
        return mapping
    bc_col = _find_column(df, "BC")
    bn_col = _find_column(df, "Basic Number")
    if bc_col is None or bn_col is None:
        return mapping
    for _, row_s in df.iterrows():
        bc = _norm(row_s.get(bc_col, ""))
        bn = _norm(row_s.get(bn_col, ""))
        if not bc or not bn:
            continue
        mapping.setdefault(bn, [])
        if bc not in mapping[bn]:
            mapping[bn].append(bc)
    return mapping

def _read_master_bn_to_hbm(master_path: str) -> dict[str, list[str]]:
    """
    Build a mapping: basic_number -> [list of Head Bom Mat values that contain it]
    from the "reformatted_sap" sheet of the master file.
    """
    mapping: dict[str, list[str]] = {}
    try:
        xl = pd.ExcelFile(master_path, engine="openpyxl")
    except Exception:
        return mapping

    df = _read_sheet_df(xl, "reformatted_sap")
    if df is None or df.empty:
        df = _read_sheet_df(xl, "Reformatted SAP Data")
    if df is None or df.empty:
        try:
            df = pd.read_excel(xl, sheet_name=xl.sheet_names[0], dtype=object)
        except Exception:
            pass
    if df is None or df.empty:
        return mapping

    hbm_col = _find_column(df, "Head Bom Mat")
    bn_col  = _find_column(df, "Basic Number")
    if hbm_col is None or bn_col is None:
        return mapping

    for _, row_s in df.iterrows():
        hbm    = _norm(row_s.get(hbm_col, ""))
        bn_raw = _norm(row_s.get(bn_col, ""))
        if not hbm or not bn_raw:
            continue
        for bn in _parse_basic_numbers(bn_raw):
            mapping.setdefault(bn, [])
            if hbm not in mapping[bn]:
                mapping[bn].append(hbm)

    return mapping


# ── TSC data reading ─────────────────────────────────────────────────────────

def _read_tsc_idh_build(tsc_path: str) -> dict[str, str]:
    """
    Build a mapping: idh_number -> build_type (lowercased).
    Reads the TSC output file.  Tries "Tracker Data" sheet first,
    then falls back to the first sheet.
    """
    mapping: dict[str, str] = {}
    try:
        xl = pd.ExcelFile(tsc_path, engine="openpyxl")
    except Exception:
        return mapping

    df = _read_sheet_df(xl, "Tracker Data")
    if df is None or df.empty:
        # fallback to first sheet
        try:
            df = pd.read_excel(xl, sheet_name=xl.sheet_names[0], dtype=object)
        except Exception:
            return mapping

    # Strip ▾ / ▾* suffixes from column names (TSC export artefact)
    df.columns = [re.sub(r"\s*▾\*?\s*$", "", str(c)).strip() for c in df.columns]

    idh_col   = _find_column(df, "IDH Number")
    build_col = _find_column(df, "Build Type")
    if idh_col is None or build_col is None:
        return mapping

    for _, row_s in df.iterrows():
        idh   = _norm(row_s.get(idh_col, ""))
        build = _norm(row_s.get(build_col, "")).lower()
        if idh:
            mapping[idh] = build

    return mapping


def _read_tsc_bn_to_idh(tsc_path: str) -> dict[str, list[str]]:
    """
    TSC-only mode: build basic_number -> [IDH Number values] from TSC Data.
    Mirrors the shape of _read_master_bn_to_hbm so the same downstream
    comparison functions work without modification.
    """
    mapping: dict[str, list[str]] = {}
    try:
        xl = pd.ExcelFile(tsc_path, engine="openpyxl")
    except Exception:
        return mapping

    df = _read_sheet_df(xl, "Tracker Data")
    if df is None or df.empty:
        try:
            df = pd.read_excel(xl, sheet_name=xl.sheet_names[0], dtype=object)
        except Exception:
            return mapping

    # Strip ▾ / ▾* suffixes from column names (TSC export artefact)
    df.columns = [re.sub(r"\s*▾\*?\s*$", "", str(c)).strip() for c in df.columns]

    idh_col = _find_column(df, "IDH Number")
    bn_col  = _find_column(df, "Basic Number")
    if idh_col is None or bn_col is None:
        return mapping

    for _, row_s in df.iterrows():
        idh    = _norm(row_s.get(idh_col, ""))
        bn_raw = _norm(row_s.get(bn_col, ""))
        if not idh or not bn_raw:
            continue
        for bn in _parse_basic_numbers(bn_raw):
            mapping.setdefault(bn, [])
            if idh not in mapping[bn]:
                mapping[bn].append(idh)

    return mapping


# ══════════════════════════════════════════════════════════════════════════════
# Core comparison logic
# ══════════════════════════════════════════════════════════════════════════════

def _filter_hbm_list(
    hbm_list: list[str],
    tsc_map: dict[str, str],
) -> list[tuple[str, str]]:
    """
    Given a flat list of Head Bom Mat values, filter down using TSC Build Type.

    Priority rules:
      1. Keep "master" and "clone" if any exist → return only those.
      2. Else keep "resizing" if any exist → return only those (excluding missing).
      3. Else return all (as "bad").

    Returns list of (hbm, classification) tuples in original order.
    """
    classified: list[tuple[str, str]] = []
    for hbm in hbm_list:
        bt = tsc_map.get(hbm, "")
        cls = _classify_build_type(bt)
        classified.append((hbm, cls))

    good_items = [(h, c) for h, c in classified if c == "good"]
    if good_items:
        return good_items

    ok_items = [(h, c) for h, c in classified if c == "ok"]
    if ok_items:
        return ok_items

    return classified


def _resolve_master_idh(
    target_bns: list[str],
    master_map: dict[str, list[str]],
    tsc_map: dict[str, str],
) -> list[tuple[str, str]]:
    """
    For a set of target basic numbers, produce the final (hbm, classification)
    list that goes into the Master IDH cell.

    Phase 1 — max-coverage primary selection:
      For every IDH that appears in at least one target BN's list, count how
      many target BNs it covers.  Keep ONLY the IDH(s) with the highest count.
      - Total Hit:   max count == len(target_bns)  → full-match IDHs only.
      - Partial Hit: max count <  len(target_bns)  → best-available IDHs.

    Phase 2 — supplemental selection for uncovered BNs (Partial Hit only):
      After phase 1, compute which target BNs are still not covered by any
      selected IDH.  For each such BN that exists in at least one other IDH,
      apply another round of max-coverage selection (restricted to IDHs not
      already chosen and to coverage of the remaining uncovered BNs).
      This ensures that a "unique" BN that doesn't appear in any best-coverage
      IDH is still represented if any IDH at all contains it.

    Build-type tier filtering (_filter_hbm_list) is applied last, to the
    full ordered list of selected IDHs.
    """
    if not target_bns:
        return []

    # Build IDH → set of target BNs it covers
    idh_to_bns: dict[str, set[str]] = {}
    for bn in target_bns:
        for hbm in master_map.get(bn, []):
            idh_to_bns.setdefault(hbm, set()).add(bn)

    if not idh_to_bns:
        return []

    # Phase 1: best-coverage IDHs
    max_count = max(len(bns) for bns in idh_to_bns.values())
    best_idh_set: set[str] = {hbm for hbm, bns in idh_to_bns.items() if len(bns) == max_count}

    covered: set[str] = set()
    for hbm in best_idh_set:
        covered |= idh_to_bns[hbm]

    # Phase 2: supplemental IDHs for BNs not covered by phase-1 winners
    supplemental_idh_set: set[str] = set()
    uncovered = set(target_bns) - covered
    if uncovered:
        # Among IDHs not already selected, count how many uncovered BNs each covers
        sup_candidates: dict[str, set[str]] = {}
        for hbm, bns in idh_to_bns.items():
            if hbm in best_idh_set:
                continue
            overlap = bns & uncovered
            if overlap:
                sup_candidates[hbm] = overlap
        if sup_candidates:
            sup_max = max(len(bns) for bns in sup_candidates.values())
            supplemental_idh_set = {hbm for hbm, bns in sup_candidates.items() if len(bns) == sup_max}

    all_selected = best_idh_set | supplemental_idh_set

    # Collect in first-seen order derived from iterating target_bns through master_map
    seen: set[str] = set()
    ordered: list[str] = []
    for bn in target_bns:
        for hbm in master_map.get(bn, []):
            if hbm in all_selected and hbm not in seen:
                seen.add(hbm)
                ordered.append(hbm)

    return _filter_hbm_list(ordered, tsc_map)


def _determine_hit_type(
    target_bns: list[str],
    master_map: dict[str, list[str]],
) -> str:
    """Return 'Total Hit', 'Partial Hit', or '0 Hit'."""
    if not target_bns:
        return "0 Hit"
    matched = [bn for bn in target_bns if bn in master_map]
    if not matched:
        return "0 Hit"
    if len(matched) == len(target_bns):
        return "Total Hit"
    return "Partial Hit"


def _build_label_for_bn(
    bn: str,
    master_map: dict[str, list[str]],
    tsc_map: dict[str, str],
) -> str:
    """Return '3D', '2D', or 'NA' for a single basic number."""
    hbm_list = master_map.get(bn, [])
    if not hbm_list:
        return "NA"
    classes = [_classify_build_type(tsc_map.get(hbm, "")) for hbm in hbm_list]
    if "good" in classes:
        return "3D"
    if "ok" in classes:
        return "2D"
    return "NA"


def _determine_master_build_label(
    target_bns: list[str],
    master_map: dict[str, list[str]],
    tsc_map: dict[str, str],
) -> tuple[str, str]:
    """Return (label, color_hex) for the Master IDH Build cell.

    Each basic number maps to one of: '3D' (clone/master), '2D' (resizing),
    'NA' (missing/bad).  Unique labels across all BNs are joined with '_'.
    Color: 3D present → green; 2D present (no 3D) → orange; NA only → red.
    """
    seen: list[str] = []
    for bn in target_bns:
        lbl = _build_label_for_bn(bn, master_map, tsc_map)
        if lbl not in seen:
            seen.append(lbl)

    if not seen:
        return "NA", _COL_IDH_BAD

    combined = "_".join(seen)
    if "3D" in seen:
        color = _COL_IDH_MASTER
    elif "2D" in seen:
        color = _COL_IDH_RESIZE
    else:
        color = _COL_IDH_BAD
    return combined, color


# ══════════════════════════════════════════════════════════════════════════════
# Excel output helpers
# ══════════════════════════════════════════════════════════════════════════════

_THIN_BORDER = Border(
    left=Side(style="thin",   color="C8C8C8"),
    right=Side(style="thin",  color="C8C8C8"),
    top=Side(style="thin",    color="C8C8C8"),
    bottom=Side(style="thin", color="C8C8C8"),
)

# Columns that receive a fixed width of 150 % of the "Head Bom Mat" auto-width
_FIXED_WIDE_COLS = frozenset({
    "Basic Number",
    "Master IDH",
    "Basic Matched",
    "Basic Matched Desc",
    "Basic Missed",
    "Basic Missed Desc",
})

_HEADERS = [
    ("Head Bom Mat",        _COL_HEADER_DARK),
    ("Basic Number",        _COL_HEADER_DARK),
    ("Hit Type",            _COL_HEADER_RED),
    ("Master IDH",          _COL_HEADER_DARK),
    ("Master IDH Build",    _COL_HEADER_RED),
    ("Basic Matched",       _COL_HEADER_DARK),
    ("Basic Matched Desc",  _COL_HEADER_DARK),
    ("Basic Missed",        _COL_HEADER_DARK),
    ("Basic Missed Desc",   _COL_HEADER_DARK),
]

_HEADERS_GROUP = [
    ("BC",                  _COL_HEADER_DARK),
    ("Count",               _COL_HEADER_DARK),
    ("Basic Number",        _COL_HEADER_DARK),
    ("Hit Type",            _COL_HEADER_RED),
    ("Master IDH",          _COL_HEADER_DARK),
    ("Master IDH Build",    _COL_HEADER_RED),
    ("Basic Matched",       _COL_HEADER_DARK),
    ("Basic Matched Desc",  _COL_HEADER_DARK),
    ("Basic Missed",        _COL_HEADER_DARK),
    ("Basic Missed Desc",   _COL_HEADER_DARK),
]

_LIB_HEADERS = [
    ("Lib Model Code", _COL_HEADER_RED),
    ("Lib IDH",        _COL_HEADER_RED),
]


def _write_header_row(ws, headers: list[tuple[str, str]]) -> None:
    for col_idx, (label, color_hex) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill   = PatternFill(fill_type="solid", fgColor=color_hex)
        cell.font   = Font(color=_COL_WHITE, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _auto_col_width(ws) -> None:
    """Set column widths.  Columns in _FIXED_WIDE_COLS receive a fixed width
    equal to 150 % of the auto-computed width of the 'Head Bom Mat' column;
    all other columns use the standard auto-width logic."""
    auto_widths: dict[str, float] = {}
    col_headers: dict[str, str]  = {}

    for col in ws.columns:
        letter = col[0].column_letter
        col_headers[letter] = str(col[0].value or "").strip()
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            longest = max((len(ln) for ln in val.split("\n")), default=0)
            if longest > max_len:
                max_len = longest
        auto_widths[letter] = min(max(max_len + 3, 14), 60)

    # Compute fixed width = 150 % of Head Bom Mat auto-width
    hbm_width = next(
        (auto_widths[ltr] for ltr, h in col_headers.items() if h == "Head Bom Mat"),
        None,
    )
    fixed_width = round(hbm_width * 1.5, 1) if hbm_width else None

    for letter, width in auto_widths.items():
        header = col_headers.get(letter, "")
        if fixed_width is not None and header in _FIXED_WIDE_COLS:
            ws.column_dimensions[letter].width = fixed_width
        else:
            ws.column_dimensions[letter].width = width


def _write_result_row(
    ws,
    row_idx: int,
    head_bom_mat: str,
    basic_number_raw: str,
    hit_type: str,
    master_idh_items: list[tuple[str, str]],   # (hbm, classification)
    master_build_label: str,
    master_build_color: str,
    bn_matched_str: str,
    desc_matched_str: str,
    bn_missed_str: str,
    desc_missed_str: str,
    lib_model_code: str | None = None,
    lib_idh: str | None = None,
) -> None:
    """Write a single data row."""

    # Hit Type cell colour
    hit_type_color = {
        "Total Hit":   "2E7D32",   # dark green
        "Partial Hit": "F57F17",   # amber
        "0 Hit":       "B71C1C",   # dark red
    }.get(hit_type, "333333")

    def _dominant_idh_color(items: list[tuple[str, str]]) -> str:
        if not items:
            return "111111"
        classes = [c for _, c in items]
        if "good" in classes:
            return _COL_IDH_MASTER
        if "ok" in classes:
            return _COL_IDH_RESIZE
        return _COL_IDH_BAD

    normal_font  = Font(color="111111", size=9)
    border       = _THIN_BORDER
    left_align   = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    # col 1 – Head Bom Mat
    c1 = ws.cell(row=row_idx, column=1, value=head_bom_mat)
    c1.font = normal_font; c1.alignment = left_align; c1.border = border

    # col 2 – Basic Number
    c2 = ws.cell(row=row_idx, column=2, value=basic_number_raw)
    c2.font = normal_font; c2.alignment = left_align; c2.border = border

    # col 3 – Hit Type
    c3 = ws.cell(row=row_idx, column=3, value=hit_type)
    c3.font       = Font(color=hit_type_color, bold=True, size=9)
    c3.alignment  = center_align
    c3.border     = border

    # col 4 – Master IDH  (comma-separated, colour reflects dominant tier)
    idh_text  = ", ".join(h for h, _ in master_idh_items)
    idh_color = _dominant_idh_color(master_idh_items)
    c4 = ws.cell(row=row_idx, column=4, value=idh_text)
    c4.font      = Font(color=idh_color, size=9)
    c4.alignment = left_align
    c4.border    = border

    # col 5 – Master IDH Build  (per-BN label: 3D / 2D / NA / combinations)
    c5 = ws.cell(row=row_idx, column=5, value=master_build_label)
    c5.font      = Font(color=master_build_color, bold=True, size=9)
    c5.alignment = center_align
    c5.border    = border

    # col 6 – Basic Matched
    c6 = ws.cell(row=row_idx, column=6, value=bn_matched_str)
    c6.font = normal_font; c6.alignment = left_align; c6.border = border

    # col 7 – Basic Matched Desc
    c7 = ws.cell(row=row_idx, column=7, value=desc_matched_str)
    c7.font = normal_font; c7.alignment = left_align; c7.border = border

    # col 8 – Basic Missed
    c8 = ws.cell(row=row_idx, column=8, value=bn_missed_str)
    c8.font = normal_font; c8.alignment = left_align; c8.border = border

    # col 9 – Basic Missed Desc
    c9 = ws.cell(row=row_idx, column=9, value=desc_missed_str)
    c9.font = normal_font; c9.alignment = left_align; c9.border = border

    # col 10 – Lib Model Code  (only present when library comparison is active)
    if lib_model_code is not None:
        c10 = ws.cell(row=row_idx, column=10, value=lib_model_code)
        c10.font = normal_font; c10.alignment = left_align; c10.border = border

    # col 11 – Lib IDH
    if lib_idh is not None:
        c11 = ws.cell(row=row_idx, column=11, value=lib_idh)
        c11.font = normal_font; c11.alignment = left_align; c11.border = border


def _write_group_row(
    ws,
    row_idx: int,
    bc: str,
    count: str,
    basic_number: str,
    hit_type: str,
    master_idh_items: list[tuple[str, str]],
    master_build_label: str,
    master_build_color: str,
    bn_matched_str: str,
    desc_matched_str: str,
    bn_missed_str: str,
    desc_missed_str: str,
    lib_model_code: str | None = None,
    lib_idh: str | None = None,
) -> None:
    """Write a single group-map data row (body_map_GROUP sheet)."""
    hit_type_color = {
        "Total Hit":   "2E7D32",
        "Partial Hit": "F57F17",
        "0 Hit":       "B71C1C",
    }.get(hit_type, "333333")

    def _dominant_idh_color(items: list[tuple[str, str]]) -> str:
        if not items:
            return "111111"
        classes = [c for _, c in items]
        if "good" in classes:
            return _COL_IDH_MASTER
        if "ok" in classes:
            return _COL_IDH_RESIZE
        return _COL_IDH_BAD

    normal_font  = Font(color="111111", size=9)
    border       = _THIN_BORDER
    left_align   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    def _w(col, val, font=None, align=None):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.font      = font  or normal_font
        c.alignment = align or left_align
        c.border    = border

    _w(1, bc)
    _w(2, count)
    _w(3, basic_number)
    _w(4, hit_type,
       font  = Font(color=hit_type_color, bold=True, size=9),
       align = center_align)

    idh_text  = ", ".join(h for h, _ in master_idh_items)
    idh_color = _dominant_idh_color(master_idh_items)
    _w(5, idh_text, font=Font(color=idh_color, size=9))
    _w(6, master_build_label,
       font  = Font(color=master_build_color, bold=True, size=9),
       align = center_align)
    _w(7,  bn_matched_str)
    _w(8,  desc_matched_str)
    _w(9,  bn_missed_str)
    _w(10, desc_missed_str)

    if lib_model_code is not None:
        _w(11, lib_model_code)
    if lib_idh is not None:
        _w(12, lib_idh)


# ══════════════════════════════════════════════════════════════════════════════
# Library helpers
# ══════════════════════════════════════════════════════════════════════════════

def _read_library_map(library_path: str) -> dict[str, list[tuple[str, str]]]:
    """Read an Excel library file and return a mapping:
        basic_number -> [(model_code, idh), ...]
    Column detection is case-insensitive and supports alternate names.
    """
    from openpyxl import load_workbook as _lw

    wb = _lw(library_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {}

    headers = [str(v or "").strip().lower() for v in rows[0]]

    # Basic Number column
    bn_col: int | None = None
    for i, h in enumerate(headers):
        if "basic number" in h or h == "basic_number":
            bn_col = i
            break

    # 3D Model Code column (many alternate names)
    _model_patterns = ["model code", "body code", "3d body", "3d code"]
    model_col: int | None = None
    for i, h in enumerate(headers):
        if any(p in h for p in _model_patterns):
            model_col = i
            break

    # IDH column – try exact first, then partial to avoid false matches
    _idh_exact = {"idh #", "idh number", "idh num", "idhs number", "idh"}
    idh_col: int | None = None
    for i, h in enumerate(headers):
        if h in _idh_exact:
            idh_col = i
            break
    if idh_col is None:
        for i, h in enumerate(headers):
            if any(p in h for p in _idh_exact):
                idh_col = i
                break

    if bn_col is None:
        return {}

    result: dict[str, list[tuple[str, str]]] = {}
    for data_row in rows[1:]:
        bn_val = _norm(data_row[bn_col]) if bn_col < len(data_row) else ""
        if not bn_val:
            continue
        model_val = _norm(data_row[model_col]) if model_col is not None and model_col < len(data_row) else ""
        idh_val   = _norm(data_row[idh_col])   if idh_col   is not None and idh_col   < len(data_row) else ""
        # Library BN cells may contain comma-separated groups (e.g. "821633, 410679").
        # Index every individual BN so per-BN lookups work correctly.
        individual_bns = [b.strip() for b in bn_val.split(",") if b.strip()]
        for bn in individual_bns:
            result.setdefault(bn, []).append((model_val, idh_val))

    return result


def _lib_lookup(
    basic_numbers: list[str],
    lib_map: dict[str, list[tuple[str, str]]],
) -> tuple[str, str]:
    """Return (lib_model_code, lib_idh) for a set of target basic numbers.

    Rules (matching the specified cases A-D):
      - ALL target basic numbers must be present in the library → show values.
      - ANY target BN missing from library → both fields return 'NA'.
      - If a BN has multiple library rows, take up to 2 per BN.
    """
    if not basic_numbers:
        return "NA", "NA"

    all_found = all(bn in lib_map for bn in basic_numbers)
    if not all_found:
        return "NA", "NA"

    model_parts: list[str] = []
    idh_parts:   list[str] = []
    for bn in basic_numbers:
        for model_val, idh_val in lib_map.get(bn, [])[:2]:
            if model_val:
                model_parts.append(model_val)
            if idh_val:
                idh_parts.append(idh_val)

    lib_model = ", ".join(model_parts) if model_parts else "NA"
    lib_idh   = ", ".join(idh_parts)   if idh_parts   else "NA"
    return lib_model, lib_idh


# ══════════════════════════════════════════════════════════════════════════════
# Public entry point
# ══════════════════════════════════════════════════════════════════════════════

def run_comparison(params: CompareParams) -> CompareResult:
    """
    Run the full SDC comparison. Each RSD: Target file is evaluated independently
    against the shared master + TSC data, producing one output workbook per file.
    """
    import re as _re
    from collections import Counter as _Counter

    result = CompareResult()

    # ── validate inputs ──────────────────────────────────────────────────────
    tsc_only = (params.compare_mode == "tsc_only")
    if not params.rsd_target_paths:
        result.warnings.append("No RSD: Target files provided.")
        return result
    if not tsc_only and not params.rsd_master_path:
        result.warnings.append("No RSD: Master file provided.")
        return result
    if not params.tsc_data_path:
        result.warnings.append("No TSC Data file provided.")
        return result
    if not params.output_dir:
        result.warnings.append("No output directory provided.")
        return result

    # ── read shared data (master + TSC) once ─────────────────────────────────
    if tsc_only:
        # TSC Data serves as both the master BN map and the build-type map.
        try:
            tsc_map = _read_tsc_idh_build(params.tsc_data_path)
        except Exception as exc:
            result.warnings.append(f"Error reading TSC Data: {exc}")
            return result
        try:
            master_map = _read_tsc_bn_to_idh(params.tsc_data_path)
        except Exception as exc:
            result.warnings.append(f"Error building BN map from TSC Data: {exc}")
            return result
    else:
        try:
            master_map = _read_master_bn_to_hbm(params.rsd_master_path)
        except Exception as exc:
            result.warnings.append(f"Error reading RSD Master: {exc}")
            return result
        try:
            tsc_map = _read_tsc_idh_build(params.tsc_data_path)
        except Exception as exc:
            result.warnings.append(f"Error reading TSC Data: {exc}")
            return result

    # ── read library (optional) ────────────────────────────────────────────
    lib_map: dict[str, list[tuple[str, str]]] | None = None
    if params.excel_library_path:
        try:
            lib_map = _read_library_map(params.excel_library_path)
        except Exception as exc:
            result.warnings.append(f"Error reading Excel Library: {exc}")
            # non-fatal – continue without library comparison

    # ── output dir + filename helpers ─────────────────────────────────────────
    out_dir = Path(params.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")

    _DATE_SUFFIX = _re.compile(
        r"[_-]\d{4}[_-]\d{2}[_-]\d{2}[_-]\d{2}[_-]\d{2}$"
        r"|[_-]\d{2}[_-]\d{2}[_-]\d{2}[_-]\d{2}$"
    )

    def _base_name(path: str) -> str:
        stem = Path(path).stem
        if stem.lower().startswith("rsd_"):
            stem = stem[4:]
        return _DATE_SUFFIX.sub("", stem)

    bases = [_base_name(p) for p in params.rsd_target_paths]
    base_counts = _Counter(bases)
    version_tracker: dict[str, int] = {}

    # ── process each target file independently ────────────────────────────────
    for target_path, base in zip(params.rsd_target_paths, bases):

        # output filename for this file
        if base_counts[base] > 1:
            version_tracker[base] = version_tracker.get(base, 0) + 1
            v = version_tracker[base]
            out_name = f"sdc_bma_{base}_v{v:02d}_{timestamp}.xlsx"
        else:
            out_name = f"sdc_bma_{base}_{timestamp}.xlsx"
        out_path = out_dir / out_name

        # read this file's target rows
        try:
            target_rows = _read_target_rows([target_path])
        except Exception as exc:
            result.warnings.append(f"{Path(target_path).name}: Error reading: {exc}")
            continue

        if not target_rows:
            result.warnings.append(f"{Path(target_path).name}: No data rows found.")
            continue

        # ── build body_map_IDH sheet ──────────────────────────────────────────
        idh_headers = list(_HEADERS)
        if lib_map is not None:
            idh_headers.append(("Lib Model Code", _COL_HEADER_RED))
            idh_headers.append(("Lib IDH",        _COL_HEADER_RED))

        wb = Workbook()
        ws = wb.active
        ws.title = "body_map_IDH"
        _write_header_row(ws, idh_headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(idh_headers))}1"
        ws.row_dimensions[1].height = 30

        for row_idx, trow in enumerate(target_rows, start=2):
            hit_type = _determine_hit_type(trow.basic_numbers, master_map)

            if hit_type == "0 Hit":
                master_idh_items: list[tuple[str, str]] = []
            else:
                master_idh_items = _resolve_master_idh(
                    trow.basic_numbers, master_map, tsc_map
                )

            build_label, build_color = _determine_master_build_label(
                trow.basic_numbers, master_map, tsc_map
            )

            bn_to_name: dict[str, str] = dict(zip(trow.basic_numbers, trow.basic_names))
            if hit_type == "0 Hit":
                bn_matched_str   = "NA"
                desc_matched_str = "NA"
                bn_missed_str    = ", ".join(trow.basic_numbers)
                desc_missed_str  = " | ".join(bn_to_name.get(bn, "") for bn in trow.basic_numbers)
            elif hit_type == "Total Hit":
                bn_matched_str   = ", ".join(trow.basic_numbers)
                desc_matched_str = " | ".join(bn_to_name.get(bn, "") for bn in trow.basic_numbers)
                bn_missed_str    = "NA"
                desc_missed_str  = "NA"
            else:  # Partial Hit
                matched_bns = [bn for bn in trow.basic_numbers if bn in master_map]
                missed_bns  = [bn for bn in trow.basic_numbers if bn not in master_map]
                bn_matched_str   = ", ".join(matched_bns)
                desc_matched_str = " | ".join(bn_to_name.get(bn, "") for bn in matched_bns)
                bn_missed_str    = ", ".join(missed_bns)
                desc_missed_str  = " | ".join(bn_to_name.get(bn, "") for bn in missed_bns)

            # library lookup
            lib_mc: str | None = None
            lib_ih: str | None = None
            if lib_map is not None:
                lib_mc, lib_ih = _lib_lookup(trow.basic_numbers, lib_map)

            _write_result_row(
                ws, row_idx,
                head_bom_mat=trow.head_bom_mat,
                basic_number_raw=", ".join(trow.basic_numbers),
                hit_type=hit_type,
                master_idh_items=master_idh_items,
                master_build_label=build_label,
                master_build_color=build_color,
                bn_matched_str=bn_matched_str,
                desc_matched_str=desc_matched_str,
                bn_missed_str=bn_missed_str,
                desc_missed_str=desc_missed_str,
                lib_model_code=lib_mc,
                lib_idh=lib_ih,
            )

        _auto_col_width(ws)

        # ── build body_map_GROUP sheet ────────────────────────────────────────
        group_rows = _read_grouping_rows([target_path])

        grp_headers = list(_HEADERS_GROUP)
        if lib_map is not None:
            grp_headers.extend(_LIB_HEADERS)

        ws_g = wb.create_sheet("body_map_GROUP")
        _write_header_row(ws_g, grp_headers)
        ws_g.freeze_panes = "A2"
        ws_g.auto_filter.ref = f"A1:{get_column_letter(len(grp_headers))}1"
        ws_g.row_dimensions[1].height = 30

        for g_idx, grow in enumerate(group_rows, start=2):
            bns = _parse_basic_numbers(grow.basic_number) if grow.basic_number else []

            g_hit_type = _determine_hit_type(bns, master_map)

            if g_hit_type == "0 Hit":
                g_master_idh_items: list[tuple[str, str]] = []
            else:
                g_master_idh_items = _resolve_master_idh(bns, master_map, tsc_map)

            g_build_label, g_build_color = _determine_master_build_label(
                bns, master_map, tsc_map
            )

            if g_hit_type == "0 Hit":
                g_bn_matched   = "NA"
                g_desc_matched = "NA"
                g_bn_missed    = grow.basic_number
                g_desc_missed  = grow.basic_name
            elif g_hit_type == "Total Hit":
                g_bn_matched   = grow.basic_number
                g_desc_matched = grow.basic_name
                g_bn_missed    = "NA"
                g_desc_missed  = "NA"
            else:  # Partial Hit
                matched_bns = [bn for bn in bns if bn in master_map]
                missed_bns  = [bn for bn in bns if bn not in master_map]
                g_bn_matched   = ", ".join(matched_bns)
                g_desc_matched = grow.basic_name if matched_bns else ""
                g_bn_missed    = ", ".join(missed_bns)
                g_desc_missed  = grow.basic_name if missed_bns else ""

            # library lookup for group row
            g_lib_mc: str | None = None
            g_lib_ih: str | None = None
            if lib_map is not None:
                g_lib_mc, g_lib_ih = _lib_lookup(bns, lib_map)

            _write_group_row(
                ws_g, g_idx,
                bc=grow.bc,
                count=grow.count,
                basic_number=grow.basic_number,
                hit_type=g_hit_type,
                master_idh_items=g_master_idh_items,
                master_build_label=g_build_label,
                master_build_color=g_build_color,
                bn_matched_str=g_bn_matched,
                desc_matched_str=g_desc_matched,
                bn_missed_str=g_bn_missed,
                desc_missed_str=g_desc_missed,
                lib_model_code=g_lib_mc,
                lib_idh=g_lib_ih,
            )

        _auto_col_width(ws_g)

        # ── save ──────────────────────────────────────────────────────────────
        try:
            wb.save(str(out_path))
            result.output_paths.append(str(out_path))
            result.output_path = str(out_path)
        except Exception as exc:
            result.warnings.append(f"Failed to save {out_name}: {exc}")

    return result
