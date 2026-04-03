from __future__ import annotations

import csv
from datetime import datetime
import re
import sys
import tempfile
from pathlib import Path
import pandas as pd
from collections import defaultdict
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
from matplotlib.figure import Figure

_BASE_DIR = Path(getattr(sys, '_MEIPASS', Path(__file__).resolve().parent))

from PySide6.QtCore import Qt, Property, QPropertyAnimation, QEasingCurve, Signal, QTimer
from PySide6.QtGui import QColor, QCursor, QKeySequence, QMovie, QPainter, QPainterPath, QPixmap, QLinearGradient, QBrush, QPen, QFont
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDialog,
    QColorDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QStackedWidget,
    QSpacerItem,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from status_collector import StatusCollector
from thumbnail_generator import ThumbnailGenerator
from packshot_naming_generator import PackshotNamingGenerator
from sap_reformat import SapTableReformatError, SapTableReformatter
from general_functions import clear_other_panel_inputs
from hat_config import HatConfig
from body_mapper import CompareParams, run_comparison
from reference_collector import RefCollectorParams, run_reference_collector
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ToggleSwitch(QWidget):
    """Animated pill toggle that mimics a native iOS/Android switch."""

    toggled = Signal(bool)
    def __init__(self, parent=None, width=46, height=26):
        super().__init__(parent)
        self._checked = False
        self._thumb_x = float(height // 2)        # centre-x of thumb
        self._w = width
        self._h = height
        self._r = height // 2                     # pill corner radius
        self._pad = 3                             # gap between thumb edge and pill edge
        self._thumb_r = self._r - self._pad       # thumb radius
        self._x_off = float(self._r)              # thumb centre when off
        self._x_on  = float(self._w - self._r)   # thumb centre when on
        self._thumb_x = self._x_off
        self.setFixedSize(self._w, self._h)
        self.setCursor(Qt.PointingHandCursor)

        self._anim = QPropertyAnimation(self, b"thumb_x", self)
        self._anim.setDuration(160)
        self._anim.setEasingCurve(QEasingCurve.Type.InOutQuad)

    # ── Qt property so QPropertyAnimation can drive it ───────────────
    def _get_thumb_x(self) -> float:
        return self._thumb_x

    def _set_thumb_x(self, val: float):
        self._thumb_x = val
        self.update()

    thumb_x = Property(float, _get_thumb_x, _set_thumb_x)

    # ── Public API ───────────────────────────────────────────────────
    def isChecked(self) -> bool:
        return self._checked

    def setChecked(self, state: bool):
        if state == self._checked:
            return
        self._checked = state
        self._animate_to(self._x_on if state else self._x_off)
        self.update()
        self.toggled.emit(state)

    def toggle(self):
        self.setChecked(not self._checked)

    # ── Interaction ──────────────────────────────────────────────────
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.toggle()

    # ── Animation ────────────────────────────────────────────────────
    def _animate_to(self, target: float):
        self._anim.stop()
        self._anim.setStartValue(self._thumb_x)
        self._anim.setEndValue(target)
        self._anim.start()

    # ── Painting ─────────────────────────────────────────────────────
    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)

        # ---- pill background ----------------------------------------
        if self._checked:
            # teal → blue-purple gradient (left to right)
            grad = QLinearGradient(0, 0, self._w, 0)
            grad.setColorAt(0.0, QColor("#5de0c8"))   # teal
            grad.setColorAt(1.0, QColor("#6b88e6"))   # blue-purple
            p.setBrush(QBrush(grad))
        else:
            p.setBrush(QBrush(QColor("#88929e")))

        p.setPen(Qt.NoPen)
        p.drawRoundedRect(0, 0, self._w, self._h, self._r, self._r)

        # ---- white thumb --------------------------------------------
        cx = int(self._thumb_x)
        cy = self._h // 2
        p.setBrush(QBrush(QColor("#ffffff")))
        # subtle drop shadow
        shadow_pen = QPen(QColor(0, 0, 0, 40))
        shadow_pen.setWidth(0)
        p.setPen(shadow_pen)
        p.drawEllipse(
            cx - self._thumb_r + 1,
            cy - self._thumb_r + 1,
            self._thumb_r * 2,
            self._thumb_r * 2,
        )
        p.setPen(Qt.NoPen)
        p.drawEllipse(
            cx - self._thumb_r,
            cy - self._thumb_r,
            self._thumb_r * 2,
            self._thumb_r * 2,
        )

        p.end()


class _PlainTextLineAdapter:
    """Compatibility wrapper so QLineEdit can be used by legacy textEdit-based logic."""

    def __init__(self, line_edit: QLineEdit) -> None:
        self._line_edit = line_edit

    def setPlainText(self, text: str) -> None:
        self._line_edit.setText(text)

    def toPlainText(self) -> str:
        return self._line_edit.text()

    def clear(self) -> None:
        self._line_edit.clear()

    def setEnabled(self, enabled: bool) -> None:
        self._line_edit.setEnabled(enabled)


class _ClipboardTableWidget(QTableWidget):
    def __init__(self, rows: int, cols: int, parent: QWidget | None = None) -> None:
        super().__init__(rows, cols, parent)
        self.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
            | QAbstractItemView.EditTrigger.AnyKeyPressed
        )
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)

    def keyPressEvent(self, event) -> None:
        if event.matches(QKeySequence.StandardKey.Copy):
            self._copy_selection()
            return
        if event.matches(QKeySequence.StandardKey.Paste):
            self._paste_clipboard()
            return
        super().keyPressEvent(event)

    def _copy_selection(self) -> None:
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return

        # Copy first selected rectangular range as tab/newline delimited text.
        selected_range = selected_ranges[0]
        lines = []
        for row in range(selected_range.topRow(), selected_range.bottomRow() + 1):
            row_values = []
            for col in range(selected_range.leftColumn(), selected_range.rightColumn() + 1):
                cell = self.item(row, col)
                row_values.append(cell.text() if cell else "")
            lines.append("\t".join(row_values))

        QApplication.clipboard().setText("\n".join(lines))

    def _paste_clipboard(self) -> None:
        if getattr(self, '_paste_disabled', False):
            return
        clip_text = QApplication.clipboard().text()
        if not clip_text:
            return

        parent = self.parent()
        begin_batch = getattr(parent, "_on_table_batch_edit_begin", None)
        end_batch = getattr(parent, "_on_table_batch_edit_end", None)

        normalized = clip_text.replace("\r\n", "\n").replace("\r", "\n")
        raw_lines = normalized.split("\n")
        if len(raw_lines) > 1 and raw_lines[-1] == "":
            raw_lines = raw_lines[:-1]

        rows = [line.split("\t") for line in raw_lines]
        if not rows:
            return

        if callable(begin_batch):
            begin_batch()

        # If multiple cells are selected, apply clipboard values across all selected cells.
        # Single value => fill all selected cells.
        # Multiple values => cycle through selected cells in row/column order.
        selected_indexes = self.selectedIndexes()
        if len(selected_indexes) > 1:
            flat_values = [value for row_values in rows for value in row_values]
            if not flat_values:
                if callable(end_batch):
                    end_batch()
                return

            ordered = sorted(selected_indexes, key=lambda idx: (idx.row(), idx.column()))
            for i, index in enumerate(ordered):
                r = index.row()
                c = index.column()
                item = self.item(r, c)
                if item is None:
                    item = QTableWidgetItem("")
                    self.setItem(r, c, item)
                item.setText(flat_values[i % len(flat_values)])
            if callable(end_batch):
                end_batch()
            return

        selected_ranges = self.selectedRanges()
        if selected_ranges:
            selected_range = sorted(selected_ranges, key=lambda r: (r.topRow(), r.leftColumn()))[0]
            start_row = selected_range.topRow()
            start_col = selected_range.leftColumn()
        elif self.currentRow() >= 0 and self.currentColumn() >= 0:
            start_row = self.currentRow()
            start_col = self.currentColumn()
        else:
            start_row = 0
            start_col = 0

        required_rows = start_row + len(rows)
        if required_rows > self.rowCount():
            self.setRowCount(required_rows)

        for r_offset, values in enumerate(rows):
            for c_offset, value in enumerate(values):
                col = start_col + c_offset
                if col >= self.columnCount():
                    break
                row = start_row + r_offset
                item = self.item(row, col)
                if item is None:
                    item = QTableWidgetItem("")
                    self.setItem(row, col, item)
                item.setText(value)

        if callable(end_batch):
            end_batch()


class _PackshotClipboardTableDialog(QDialog):
    _BLANK_FILTER = "__BLANK__"

    def __init__(self, parent: QWidget | None = None, start_dir: str = "") -> None:
        super().__init__(parent)
        self._browse_start_dir = start_dir
        self.setWindowTitle("Paste on Table")
        self.resize(1160, 620)
        self.setMinimumSize(1080, 520)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        title_row = QHBoxLayout()
        title_row.setSpacing(8)
        title = QLabel("Item Information")
        title.setObjectName("packshotTableTitle")
        title_row.addWidget(title, 0)
        title_row.addSpacing(80)
        self.label_visible_count = QLabel("Count: <b>0</b>")
        self.label_visible_count.setObjectName("packshotRowCountLabel")
        self.label_visible_count.setTextFormat(Qt.TextFormat.RichText)
        title_row.addWidget(self.label_visible_count, 0)
        title_row.addStretch(1)
        layout.addLayout(title_row)

        config_row = QHBoxLayout()
        config_row.setSpacing(8)
        row_count_label = QLabel("Row count")
        row_count_label.setObjectName("packshotRowCountLabel")
        config_row.addWidget(row_count_label, 0)

        self.row_count_input = QLineEdit("5")
        self.row_count_input.setObjectName("packshotRowCountInput")
        self.row_count_input.setFixedWidth(90)
        config_row.addWidget(self.row_count_input, 0)

        self.btn_update_rows = QPushButton("Update")
        self.btn_update_rows.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_update_rows, 0)

        self.btn_reset_table = QPushButton("Reset")
        self.btn_reset_table.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_reset_table, 0)

        self.btn_delete_row = QPushButton("Delete")
        self.btn_delete_row.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_delete_row, 0)

        self.btn_undo_table = QPushButton("Undo")
        self.btn_undo_table.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_undo_table, 0)
        config_row.addSpacing(6)

        self.btn_reset_filter = QPushButton("Reset Filter")
        self.btn_reset_filter.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_reset_filter, 0)

        starting_row_label = QLabel("Starting Row")
        starting_row_label.setObjectName("packshotRowCountLabel")
        config_row.addWidget(starting_row_label, 0)
        self.starting_row_input = QLineEdit("1")
        self.starting_row_input.setObjectName("packshotRowCountInput")
        self.starting_row_input.setFixedWidth(90)
        config_row.addWidget(self.starting_row_input, 0)

        self.btn_import_from_tracker = QPushButton("Import")
        self.btn_import_from_tracker.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_import_from_tracker, 0)

        self.btn_generate_packshot_naming = QPushButton("Generate")
        self.btn_generate_packshot_naming.setObjectName("packshotPrimaryActionBtn")
        self.btn_generate_packshot_naming.setMinimumWidth(150)
        config_row.addWidget(self.btn_generate_packshot_naming, 0)

        self.btn_export_table = QPushButton("Export")
        self.btn_export_table.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_export_table, 0)

        self.btn_read_me = QPushButton("Read Me")
        self.btn_read_me.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_read_me, 0)
        config_row.addStretch(1)
        layout.addLayout(config_row)

        self.table = _ClipboardTableWidget(5, 6, self)
        self.table.setObjectName("packshotClipboardTable")
        self._set_headers()
        self.table.verticalHeader().setVisible(True)
        self.table.verticalHeader().setDefaultSectionSize(34)
        self.table.horizontalHeader().setDefaultSectionSize(170)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(5, 300)
        self.table.horizontalHeader().setSectionsClickable(True)
        self.table.horizontalHeader().sectionClicked.connect(self._on_filter_header_clicked)

        for row in range(5):
            for col in range(6):
                self.table.setItem(row, col, QTableWidgetItem(""))

        self._undo_stack: list[list[list[str]]] = []
        self._is_restoring_undo = False
        self._is_batch_edit = False

        layout.addWidget(self.table)
        self.btn_update_rows.clicked.connect(self._on_update_rows_clicked)
        self.btn_reset_table.clicked.connect(self._on_reset_table_clicked)
        self.btn_delete_row.clicked.connect(self._on_delete_rows_clicked)
        self.btn_undo_table.clicked.connect(self._on_undo_table_clicked)
        self.btn_reset_filter.clicked.connect(self._on_reset_filter_clicked)
        self.btn_read_me.clicked.connect(self._on_read_me_clicked)
        self.table.itemChanged.connect(self._on_table_item_changed)
        self._push_undo_snapshot(force=True)
        self._apply_filters()

        self.setStyleSheet(
            """
            QDialog {
                background-color: #F4F4F4;
            }

            QTableWidget#packshotClipboardTable {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                gridline-color: #B8B8B8;
                selection-background-color: #DCE6F5;
                selection-color: #111111;
                font-family: "Segoe UI";
                font-size: 12px;
            }

            QTableWidget#packshotClipboardTable QHeaderView::section {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #7D8694;
                padding: 6px 8px;
                font-weight: 700;
            }

            QTableWidget#packshotClipboardTable QTableCornerButton::section {
                background-color: #111F35;
                border: 1px solid #7D8694;
            }

            QLabel#packshotTableTitle {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 18px;
                font-weight: 800;
            }

            QLabel#packshotRowCountLabel {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 13px;
                font-weight: 600;
            }

            QLineEdit#packshotRowCountInput {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 8px;
                font-family: "Segoe UI";
                font-size: 12px;
            }

            QProgressBar#sapImportProgress {
                background-color: #FFFFFF;
                border: 1px solid #A9A9A9;
                border-radius: 7px;
                text-align: center;
                color: #111111;
                font-family: "Segoe UI";
                font-size: 11px;
                min-height: 24px;
            }

            QProgressBar#sapImportProgress::chunk {
                background-color: #8A244B;
                border-radius: 6px;
            }

            QPushButton#packshotUpdateRowsBtn {
                background-color: #9EA3AB;
                color: #000000;
                border: 1px solid #8B9098;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }

            QPushButton#packshotUpdateRowsBtn:pressed {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #111F35;
            }

            QPushButton#packshotPrimaryActionBtn {
                background-color: #8A244B;
                color: #FFFFFF;
                border: 1px solid #8A244B;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 700;
            }

            QPushButton#packshotPrimaryActionBtn:pressed {
                background-color: #F63049;
                border: 1px solid #F63049;
            }
            """
        )

    def _set_headers(self) -> None:
        self._header_labels = [
            "Product Name",
            "IDH",
            "Packaging Type",
            "Packaging Size",
            "View",
            "Packshot Naming",
        ]
        self._column_filters: dict[int, set[str] | None] = {i: None for i in range(len(self._header_labels))}
        self._refresh_header_labels()

        last_header_item = self.table.horizontalHeaderItem(5)
        if last_header_item is not None:
            last_header_item.setBackground(Qt.GlobalColor.transparent)
            last_header_item.setBackground(QColor("#8A244B"))
            last_header_item.setForeground(QColor("#FFFFFF"))

    def _refresh_header_labels(self) -> None:
        labels = []
        for idx, base_label in enumerate(self._header_labels):
            has_active_filter = self._column_filters.get(idx) is not None
            labels.append(f"{base_label} {'▾*' if has_active_filter else '▾'}")
        self.table.setHorizontalHeaderLabels(labels)

        last_header_item = self.table.horizontalHeaderItem(5)
        if last_header_item is not None:
            last_header_item.setBackground(Qt.GlobalColor.transparent)
            last_header_item.setBackground(QColor("#8A244B"))
            last_header_item.setForeground(QColor("#FFFFFF"))

    def _on_filter_header_clicked(self, column: int) -> None:
        values: list[str] = []
        seen: set[str] = set()
        has_blank = False
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            item = self.table.item(row, column)
            cell_text = item.text().strip() if item is not None else ""
            if cell_text == "":
                has_blank = True
                continue
            if cell_text in seen:
                continue
            seen.add(cell_text)
            values.append(cell_text)

        value_pairs: list[tuple[str, str]] = [(value, value) for value in values]
        if has_blank:
            value_pairs.append((self._BLANK_FILTER, "(Blanks)"))

        popup = _MapperReformattedTableDialog._FilterPopup(value_pairs, self._column_filters.get(column), self)
        popup.move(QCursor.pos())
        if popup.exec() != QDialog.DialogCode.Accepted:
            return

        selected = popup.get_selected_values()
        if len(selected) == len(value_pairs):
            self._column_filters[column] = None
        else:
            self._column_filters[column] = selected
        self._apply_filters()

    def _apply_filters(self) -> None:
        for row in range(self.table.rowCount()):
            row_matches = True
            for col, filter_value in self._column_filters.items():
                if filter_value is None:
                    continue
                item = self.table.item(row, col)
                cell_text = item.text().strip() if item is not None else ""
                normalized = self._BLANK_FILTER if cell_text == "" else cell_text
                if normalized not in filter_value:
                    row_matches = False
                    break
            self.table.setRowHidden(row, not row_matches)

        self._refresh_header_labels()
        self._update_visible_count_label()

    def _update_visible_count_label(self) -> None:
        count = 0
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            count += 1
        self.label_visible_count.setText(f"Count: <b>{count}</b>")

    def _on_update_rows_clicked(self) -> None:
        raw_value = self.row_count_input.text().strip()
        try:
            new_count = int(raw_value)
        except ValueError:
            self.row_count_input.setText(str(self.table.rowCount()))
            return

        new_count = max(1, min(new_count, 5000))
        self.table.setRowCount(new_count)
        for row in range(new_count):
            for col in range(self.table.columnCount()):
                if self.table.item(row, col) is None:
                    self.table.setItem(row, col, QTableWidgetItem(""))
        self.row_count_input.setText(str(new_count))
        self._apply_filters()
        self._push_undo_snapshot()

    def _on_reset_table_clicked(self) -> None:
        self.row_count_input.setText("5")
        if hasattr(self, "starting_row_input"):
            self.starting_row_input.setText("1")
        self.table.clearContents()
        self.table.setRowCount(5)
        self.table.clearSelection()
        for row in range(5):
            for col in range(self.table.columnCount()):
                self.table.setItem(row, col, QTableWidgetItem(""))
        for column in self._column_filters:
            self._column_filters[column] = None
        self._apply_filters()
        self._push_undo_snapshot()

    def _on_delete_rows_clicked(self) -> None:
        selected_rows = sorted({index.row() for index in self.table.selectedIndexes()}, reverse=True)
        if not selected_rows:
            current_row = self.table.currentRow()
            if current_row >= 0:
                selected_rows = [current_row]
            else:
                return

        total_rows = self.table.rowCount()
        if total_rows <= 0:
            return

        if len(selected_rows) >= total_rows:
            self.table.setRowCount(1)
            for col in range(self.table.columnCount()):
                self.table.setItem(0, col, QTableWidgetItem(""))
        else:
            for row in selected_rows:
                if 0 <= row < self.table.rowCount():
                    self.table.removeRow(row)

        self.table.clearSelection()
        self.row_count_input.setText(str(self.table.rowCount()))
        self._apply_filters()
        self._push_undo_snapshot()

    def _on_reset_filter_clicked(self) -> None:
        for column in self._column_filters:
            self._column_filters[column] = None
        self._apply_filters()

    def _capture_table_state(self) -> list[list[str]]:
        state: list[list[str]] = []
        for row in range(self.table.rowCount()):
            row_values: list[str] = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_values.append(item.text() if item is not None else "")
            state.append(row_values)
        return state

    def _restore_table_state(self, state: list[list[str]]) -> None:
        self._is_restoring_undo = True
        self.table.blockSignals(True)
        try:
            row_count = max(1, len(state))
            col_count = self.table.columnCount()
            self.table.setRowCount(row_count)

            for row in range(row_count):
                row_values = state[row] if row < len(state) else [""] * col_count
                for col in range(col_count):
                    value = row_values[col] if col < len(row_values) else ""
                    item = self.table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem("")
                        self.table.setItem(row, col, item)
                    item.setText(value)
        finally:
            self.table.blockSignals(False)
            self._is_restoring_undo = False

        self.row_count_input.setText(str(self.table.rowCount()))
        self._apply_filters()

    def _push_undo_snapshot(self, force: bool = False) -> None:
        if self._is_restoring_undo or self._is_batch_edit:
            return

        snapshot = self._capture_table_state()
        if not force and self._undo_stack and self._undo_stack[-1] == snapshot:
            return

        self._undo_stack.append(snapshot)
        if len(self._undo_stack) > 200:
            self._undo_stack.pop(0)

    def _on_table_item_changed(self, _item: QTableWidgetItem) -> None:
        if self._is_batch_edit:
            return
        if any(value is not None for value in self._column_filters.values()):
            self._apply_filters()
        else:
            self._update_visible_count_label()
        self._push_undo_snapshot()

    def _on_table_batch_edit_begin(self) -> None:
        if self._is_restoring_undo:
            return
        self._push_undo_snapshot(force=False)
        self._is_batch_edit = True

    def _on_table_batch_edit_end(self) -> None:
        if self._is_restoring_undo:
            return
        self._is_batch_edit = False
        self._apply_filters()
        self._push_undo_snapshot(force=False)

    def _on_undo_table_clicked(self) -> None:
        if self._is_restoring_undo:
            return

        current_snapshot = self._capture_table_state()
        if not self._undo_stack:
            self._undo_stack.append(current_snapshot)
            return

        if self._undo_stack[-1] != current_snapshot:
            self._undo_stack.append(current_snapshot)

        if len(self._undo_stack) <= 1:
            return

        self._undo_stack.pop()
        previous_snapshot = self._undo_stack[-1]
        self._restore_table_state(previous_snapshot)

    def _on_read_me_clicked(self) -> None:
        msg = QMessageBox(self)
        msg.setWindowTitle("Read Me")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setStyleSheet(
            """
            QMessageBox {
                background-color: #F4F4F4;
            }
            QMessageBox QLabel {
                color: #111111;
                font-family: "Segoe UI";
                font-size: 12px;
            }
            QMessageBox QPushButton {
                background-color: #9EA3AB;
                color: #000000;
                border: 1px solid #8B9098;
                border-radius: 8px;
                min-height: 28px;
                min-width: 80px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }
            """
        )
        msg.setText(
            "Logic Info:\n"
            "1. User can copy paste data on the table similar to excel.\n"
            "2. When importing data, the following rows from the excel tracker will be ignored:\n"
            "-if IDH column is blank.\n"
            "-if Product Name and IDH column have value \"admin\"."
        )
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()


class _SingleTrackerColumnPickerDialog(QDialog):
    """Ask the user to choose one column from a list of column identifiers."""

    def __init__(self, columns: list[str], parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Select Column")
        self.setMinimumWidth(360)
        self._selected: str | None = None

        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)

        lbl = QLabel("Multiple columns detected.\nSelect the column to import:")
        lbl.setWordWrap(True)
        layout.addWidget(lbl)

        self._combo = QComboBox()
        for col in columns:
            self._combo.addItem(col)
        layout.addWidget(self._combo)

        btn_row = QHBoxLayout()
        btn_ok = QPushButton("OK")
        btn_ok.setDefault(True)
        btn_cancel = QPushButton("Cancel")
        btn_row.addStretch(1)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        btn_ok.clicked.connect(self._on_ok)
        btn_cancel.clicked.connect(self.reject)

        self.setStyleSheet(
            """
            QDialog { background-color: #F4F4F4; }
            QLabel { color: #111F35; font-family: "Segoe UI"; font-size: 13px; }
            QComboBox {
                background-color: #FFFFFF; color: #111111;
                border: 1px solid #A9A9A9; border-radius: 6px;
                min-height: 28px; padding: 0 8px;
                font-family: "Segoe UI"; font-size: 12px;
            }
            QPushButton {
                background-color: #9EA3AB; color: #000000;
                border: 1px solid #8B9098; border-radius: 8px;
                min-height: 28px; min-width: 72px; padding: 0 12px;
                font-family: "Segoe UI"; font-size: 12px; font-weight: 600;
            }
            QPushButton:pressed { background-color: #111F35; color: #FFFFFF; }
            """
        )

    def _on_ok(self) -> None:
        self._selected = self._combo.currentText()
        self.accept()

    def get_selected(self) -> str | None:
        return self._selected


class _SingleTrackerWindow(QDialog):
    """Value Duplicate Check – Single Tracker window."""

    _COL_VALUE = 0
    _COL_DUP   = 1
    _COL_PARENT = 2

    def __init__(self, parent: QWidget | None = None, start_dir: str = "") -> None:
        super().__init__(parent)
        self._browse_dir = start_dir
        self.setWindowTitle("Value Duplicate Check – Single Tracker")
        self.resize(990, 560)
        self.setMinimumSize(736, 400)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        # ── title row ────────────────────────────────────────────────────────
        title_row = QHBoxLayout()
        title_row.setSpacing(8)
        title_lbl = QLabel("Value Duplicate Check")
        title_lbl.setObjectName("stTitle")
        title_row.addWidget(title_lbl, 0)
        title_row.addSpacing(60)
        self.label_visible_count = QLabel("Count: <b>0</b>")
        self.label_visible_count.setObjectName("stCountLabel")
        self.label_visible_count.setTextFormat(Qt.TextFormat.RichText)
        title_row.addWidget(self.label_visible_count, 0)
        title_row.addStretch(1)
        layout.addLayout(title_row)

        # ── controls row ─────────────────────────────────────────────────────
        ctrl = QHBoxLayout()
        ctrl.setSpacing(8)

        row_count_label = QLabel("Row count")
        row_count_label.setObjectName("stCountLabel")
        ctrl.addWidget(row_count_label, 0)

        self.row_count_input = QLineEdit("5")
        self.row_count_input.setObjectName("stInput")
        self.row_count_input.setFixedWidth(70)
        ctrl.addWidget(self.row_count_input, 0)

        self.btn_update_rows = QPushButton("Update")
        self.btn_update_rows.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_update_rows, 0)

        self.btn_reset_table = QPushButton("Reset")
        self.btn_reset_table.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_reset_table, 0)

        self.btn_delete_row = QPushButton("Delete")
        self.btn_delete_row.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_delete_row, 0)

        self.btn_undo_table = QPushButton("Undo")
        self.btn_undo_table.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_undo_table, 0)

        ctrl.addSpacing(6)

        start_row_lbl = QLabel("Header Row")
        start_row_lbl.setObjectName("stCountLabel")
        ctrl.addWidget(start_row_lbl, 0)

        self.starting_row_input = QLineEdit("7")
        self.starting_row_input.setObjectName("stInput")
        self.starting_row_input.setFixedWidth(70)
        ctrl.addWidget(self.starting_row_input, 0)

        self.btn_import = QPushButton("Import")
        self.btn_import.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_import, 0)

        self.btn_evaluate = QPushButton("Evaluate")
        self.btn_evaluate.setObjectName("stPrimaryBtn")
        self.btn_evaluate.setMinimumWidth(120)
        ctrl.addWidget(self.btn_evaluate, 0)

        self.btn_export = QPushButton("Export")
        self.btn_export.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_export, 0)

        ctrl.addStretch(1)
        layout.addLayout(ctrl)

        # ── table ─────────────────────────────────────────────────────────────
        self.table = _ClipboardTableWidget(5, 3, self)
        self.table.setObjectName("stTable")
        self._header_labels = ["Value", "Duplicate or Not", "Parent Row"]
        self.table.setHorizontalHeaderLabels(self._header_labels)
        self.table.verticalHeader().setVisible(True)
        self.table.verticalHeader().setDefaultSectionSize(30)
        self.table.horizontalHeader().setDefaultSectionSize(200)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(0, 260)
        self.table.setColumnWidth(1, 200)

        for row in range(5):
            for col in range(3):
                self.table.setItem(row, col, QTableWidgetItem(""))
            # lock cols 1 & 2 from direct editing
            for col in (self._COL_DUP, self._COL_PARENT):
                item = self.table.item(row, col)
                if item:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

        self._undo_stack: list[list[list[str]]] = []
        self._is_restoring_undo = False
        self._is_batch_edit = False

        layout.addWidget(self.table)

        # ── connections ───────────────────────────────────────────────────────
        self.btn_update_rows.clicked.connect(self._on_update_rows)
        self.btn_reset_table.clicked.connect(self._on_reset_table)
        self.btn_delete_row.clicked.connect(self._on_delete_row)
        self.btn_undo_table.clicked.connect(self._on_undo)
        self.btn_import.clicked.connect(self._on_import)
        self.btn_evaluate.clicked.connect(self._on_evaluate)
        self.btn_export.clicked.connect(self._on_export)
        self.table.itemChanged.connect(self._on_item_changed)

        self._push_undo_snapshot(force=True)
        self._update_count_label()

        self.setStyleSheet(
            """
            QDialog { background-color: #F4F4F4; }

            QTableWidget#stTable {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                gridline-color: #B8B8B8;
                selection-background-color: #DCE6F5;
                selection-color: #111111;
                font-family: "Segoe UI";
                font-size: 12px;
            }
            QTableWidget#stTable QHeaderView::section {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #7D8694;
                padding: 6px 8px;
                font-weight: 700;
            }
            QTableWidget#stTable QTableCornerButton::section {
                background-color: #111F35;
                border: 1px solid #7D8694;
            }
            QLabel#stTitle {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 18px;
                font-weight: 800;
            }
            QLabel#stCountLabel {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 13px;
                font-weight: 600;
            }
            QLineEdit#stInput {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 8px;
                font-family: "Segoe UI";
                font-size: 12px;
            }
            QPushButton#stGrayBtn {
                background-color: #9EA3AB;
                color: #000000;
                border: 1px solid #8B9098;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }
            QPushButton#stGrayBtn:pressed {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #111F35;
            }
            QPushButton#stPrimaryBtn {
                background-color: #8A244B;
                color: #FFFFFF;
                border: 1px solid #8A244B;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 700;
            }
            QPushButton#stPrimaryBtn:pressed {
                background-color: #F63049;
                border: 1px solid #F63049;
            }
            """
        )

    # ── helpers ───────────────────────────────────────────────────────────────

    def _cell_text(self, row: int, col: int) -> str:
        item = self.table.item(row, col)
        return item.text() if item is not None else ""

    def _set_cell(self, row: int, col: int, value: str, editable: bool = False) -> None:
        item = self.table.item(row, col)
        if item is None:
            item = QTableWidgetItem("")
            self.table.setItem(row, col, item)
        flags = Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled
        if editable:
            flags |= Qt.ItemFlag.ItemIsEditable
        item.setFlags(flags)
        item.setText(value)

    def _update_count_label(self) -> None:
        count = sum(1 for r in range(self.table.rowCount()) if not self.table.isRowHidden(r))
        self.label_visible_count.setText(f"Count: <b>{count}</b>")

    def _capture_state(self) -> list[list[str]]:
        return [
            [self._cell_text(r, c) for c in range(self.table.columnCount())]
            for r in range(self.table.rowCount())
        ]

    def _restore_state(self, state: list[list[str]]) -> None:
        self._is_restoring_undo = True
        self.table.blockSignals(True)
        try:
            self.table.setRowCount(max(1, len(state)))
            for r, row_vals in enumerate(state):
                for c in range(self.table.columnCount()):
                    val = row_vals[c] if c < len(row_vals) else ""
                    editable = (c == self._COL_VALUE)
                    self._set_cell(r, c, val, editable=editable)
        finally:
            self.table.blockSignals(False)
            self._is_restoring_undo = False
        self.row_count_input.setText(str(self.table.rowCount()))
        self._update_count_label()

    def _push_undo_snapshot(self, force: bool = False) -> None:
        if self._is_restoring_undo or self._is_batch_edit:
            return
        snap = self._capture_state()
        if not force and self._undo_stack and self._undo_stack[-1] == snap:
            return
        self._undo_stack.append(snap)
        if len(self._undo_stack) > 200:
            self._undo_stack.pop(0)

    def _init_rows(self, count: int) -> None:
        self.table.setRowCount(count)
        for r in range(count):
            self._set_cell(r, self._COL_VALUE, self._cell_text(r, self._COL_VALUE) or "", editable=True)
            for c in (self._COL_DUP, self._COL_PARENT):
                self._set_cell(r, c, "", editable=False)

    # ── slot handlers ─────────────────────────────────────────────────────────

    def _on_item_changed(self, _item: QTableWidgetItem) -> None:
        if not self._is_batch_edit:
            self._update_count_label()
            self._push_undo_snapshot()

    def _on_update_rows(self) -> None:
        try:
            n = max(1, min(int(self.row_count_input.text().strip()), 5000))
        except ValueError:
            self.row_count_input.setText(str(self.table.rowCount()))
            return
        self._init_rows(n)
        self.row_count_input.setText(str(n))
        self._update_count_label()
        self._push_undo_snapshot()

    def _on_reset_table(self) -> None:
        self.row_count_input.setText("5")
        self.starting_row_input.setText("7")
        self.table.blockSignals(True)
        self.table.clearContents()
        self.table.setRowCount(5)
        for r in range(5):
            self._set_cell(r, self._COL_VALUE, "", editable=True)
            for c in (self._COL_DUP, self._COL_PARENT):
                self._set_cell(r, c, "", editable=False)
        self.table.blockSignals(False)
        self._update_count_label()
        self._push_undo_snapshot()

    def _on_delete_row(self) -> None:
        sel_rows = sorted({idx.row() for idx in self.table.selectedIndexes()}, reverse=True)
        if not sel_rows:
            cur = self.table.currentRow()
            if cur >= 0:
                sel_rows = [cur]
            else:
                return
        total = self.table.rowCount()
        if len(sel_rows) >= total:
            self.table.setRowCount(1)
            self._set_cell(0, self._COL_VALUE, "", editable=True)
            for c in (self._COL_DUP, self._COL_PARENT):
                self._set_cell(0, c, "", editable=False)
        else:
            for r in sel_rows:
                if 0 <= r < self.table.rowCount():
                    self.table.removeRow(r)
        self.table.clearSelection()
        self.row_count_input.setText(str(self.table.rowCount()))
        self._update_count_label()
        self._push_undo_snapshot()

    def _on_undo(self) -> None:
        if self._is_restoring_undo:
            return
        cur = self._capture_state()
        if not self._undo_stack:
            self._undo_stack.append(cur)
            return
        if self._undo_stack[-1] != cur:
            self._undo_stack.append(cur)
        if len(self._undo_stack) <= 1:
            return
        self._undo_stack.pop()
        self._restore_state(self._undo_stack[-1])

    # ── evaluate ──────────────────────────────────────────────────────────────

    def _on_evaluate(self) -> None:
        n = self.table.rowCount()
        # gather visible (not hidden) values; map (visual_row -> value)
        values: list[str] = [self._cell_text(r, self._COL_VALUE).strip() for r in range(n)]

        # first_seen: value -> row index (0-based display "row N+1")
        first_seen: dict[str, int] = {}
        dup_results: list[str] = [""] * n
        parent_results: list[str] = [""] * n

        for r in range(n):
            v = values[r]
            if v == "":
                continue
            if v not in first_seen:
                first_seen[v] = r
            else:
                dup_results[r] = "duplicate"
                parent_results[r] = f"row {first_seen[v] + 1}"

        self.table.blockSignals(True)
        try:
            for r in range(n):
                self._set_cell(r, self._COL_DUP, dup_results[r], editable=False)
                self._set_cell(r, self._COL_PARENT, parent_results[r], editable=False)
        finally:
            self.table.blockSignals(False)
        self._push_undo_snapshot()

    # ── export ────────────────────────────────────────────────────────────────

    def _on_export(self) -> None:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(self, "Export Error", "openpyxl is required for export.\nInstall it with: pip install openpyxl")
            return

        import datetime as _dt
        _ts = _dt.datetime.now().strftime("%Y_%m_%d_%H_%M")
        _default_name = f"duplicate_check_{_ts}.xlsx"
        import os as _os
        _default_path = _os.path.join(self._browse_dir, _default_name) if self._browse_dir else _default_name
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Table", _default_path, "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Duplicate Check"

        header_fill = PatternFill("solid", fgColor="111F35")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_side = Side(style="thin", color="7D8694")
        header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        col_labels = ["Value", "Duplicate or Not", "Parent Row"]
        col_widths  = [32, 22, 18]
        for ci, (label, width) in enumerate(zip(col_labels, col_widths), start=1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = header_align
            cell.border = header_border
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 22

        body_font   = Font(name="Segoe UI", size=10)
        body_align  = Alignment(vertical="center")
        body_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row_i in range(self.table.rowCount()):
            if self.table.isRowHidden(row_i):
                continue
            for ci in range(3):
                val = self._cell_text(row_i, ci)
                cell = ws.cell(row=ws.max_row + 1 if ci == 0 else ws.max_row, column=ci + 1, value=val)
                cell.font      = body_font
                cell.alignment = body_align
                cell.border    = body_border

        try:
            wb.save(path)
            QMessageBox.information(self, "Export", f"Exported successfully:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))

    # ── import ────────────────────────────────────────────────────────────────

    def _on_import(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Import File", self._browse_dir, "Excel / CSV Files (*.xlsx *.xls *.csv)"
        )
        if not path:
            return

        # Resolve header row (1-based from user input)
        try:
            header_row = max(1, int(self.starting_row_input.text().strip())) - 1  # 0-based index
        except ValueError:
            header_row = 0

        try:
            import pandas as _pd_imp
            if path.lower().endswith(".csv"):
                df = _pd_imp.read_csv(path, header=None, dtype=str)
            else:
                df = _pd_imp.read_excel(path, header=None, dtype=str)
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            return

        if df.empty:
            QMessageBox.information(self, "Import", "The file appears to be empty.")
            return

        # Normalise a raw cell value — treat NaN/None/empty as ""
        def _norm(v) -> str:
            s = str(v).strip() if v is not None else ""
            return "" if s.lower() in ("nan", "none", "") else s

        # Use the user-specified header row for column labels
        if header_row < df.shape[0]:
            header_cells = [_norm(v) for v in df.iloc[header_row]]
        else:
            header_cells = [f"Column {i+1}" for i in range(df.shape[1])]

        # Build col_map from the header row (only columns with a non-empty label)
        col_map: list[tuple[int, str]] = [
            (i, header_cells[i]) for i in range(df.shape[1]) if header_cells[i] != ""
        ]
        if not col_map:
            # Fallback: no recognisable headers — label by number
            col_map = [(i, f"Column {i+1}") for i in range(min(df.shape[1], 8))]

        # Data rows start immediately after the header row
        data_rows = df.iloc[header_row + 1:].reset_index(drop=True)

        col_identifiers = [label for _, label in col_map]
        col_idx_map = {label: idx for idx, label in col_map}

        # If only one meaningful column, use it directly; otherwise ask user
        if len(col_map) == 1:
            col_idx = col_map[0][0]
        else:
            picker = _SingleTrackerColumnPickerDialog(col_identifiers, self)
            if picker.exec() != QDialog.DialogCode.Accepted or picker.get_selected() is None:
                return
            chosen = picker.get_selected()
            col_idx = col_idx_map.get(chosen, col_map[0][0])

        values = [_norm(v) for v in data_rows.iloc[:, col_idx]]

        # Always paste values starting at table row 0
        start_row = 0

        required = start_row + len(values)
        if required > self.table.rowCount():
            self.table.setRowCount(required)
            for r in range(self.table.rowCount()):
                for c in range(3):
                    if self.table.item(r, c) is None:
                        editable = (c == self._COL_VALUE)
                        self._set_cell(r, c, "", editable=editable)

        self.table.blockSignals(True)
        self._is_batch_edit = True
        try:
            for i, val in enumerate(values):
                r = start_row + i
                self._set_cell(r, self._COL_VALUE, val, editable=True)
                self._set_cell(r, self._COL_DUP,   "", editable=False)
                self._set_cell(r, self._COL_PARENT, "", editable=False)
        finally:
            self.table.blockSignals(False)
            self._is_batch_edit = False

        self.row_count_input.setText(str(self.table.rowCount()))
        self._update_count_label()
        self._push_undo_snapshot()


class _MultipleTrackerWindow(QDialog):
    """Value Duplicate Check – Multiple Trackers window."""

    _COL_VALUE = 0
    _COL_TRACKERS = 1

    def __init__(self, parent: QWidget | None = None, start_dir: str = "") -> None:
        super().__init__(parent)
        self._browse_dir = start_dir
        self.setWindowTitle("Duplicate Window – Multiple Trackers")
        self.resize(990, 560)
        self.setMinimumSize(736, 400)

        # Internal state: imported per-tracker data
        self._tracker_data: dict[str, list[str]] = {}  # tracker_name -> list of values
        self._selected_column: str | None = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        # ── title row ───────────────────────────────────────────────────
        title_row = QHBoxLayout()
        title_row.setSpacing(8)
        title_lbl = QLabel("Value Duplicate Check – Multiple Trackers")
        title_lbl.setObjectName("stTitle")
        title_row.addWidget(title_lbl, 0)
        title_row.addSpacing(60)
        self.label_visible_count = QLabel("Count: <b>0</b>")
        self.label_visible_count.setObjectName("stCountLabel")
        self.label_visible_count.setTextFormat(Qt.TextFormat.RichText)
        title_row.addWidget(self.label_visible_count, 0)
        title_row.addStretch(1)
        layout.addLayout(title_row)

        # ── controls row ────────────────────────────────────────────────
        ctrl = QHBoxLayout()
        ctrl.setSpacing(8)

        row_count_label = QLabel("Row count")
        row_count_label.setObjectName("stCountLabel")
        ctrl.addWidget(row_count_label, 0)

        self.row_count_input = QLineEdit("5")
        self.row_count_input.setObjectName("stInput")
        self.row_count_input.setFixedWidth(70)
        ctrl.addWidget(self.row_count_input, 0)

        self.btn_update_rows = QPushButton("Update")
        self.btn_update_rows.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_update_rows, 0)

        self.btn_reset_table = QPushButton("Reset")
        self.btn_reset_table.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_reset_table, 0)

        self.btn_delete_row = QPushButton("Delete")
        self.btn_delete_row.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_delete_row, 0)

        self.btn_undo_table = QPushButton("Undo")
        self.btn_undo_table.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_undo_table, 0)

        ctrl.addSpacing(6)

        start_row_lbl = QLabel("Header Row")
        start_row_lbl.setObjectName("stCountLabel")
        ctrl.addWidget(start_row_lbl, 0)

        self.starting_row_input = QLineEdit("7")
        self.starting_row_input.setObjectName("stInput")
        self.starting_row_input.setFixedWidth(70)
        ctrl.addWidget(self.starting_row_input, 0)

        self.btn_import = QPushButton("Import")
        self.btn_import.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_import, 0)

        self.btn_evaluate = QPushButton("Evaluate")
        self.btn_evaluate.setObjectName("stPrimaryBtn")
        self.btn_evaluate.setMinimumWidth(120)
        ctrl.addWidget(self.btn_evaluate, 0)

        self.btn_export = QPushButton("Export")
        self.btn_export.setObjectName("stGrayBtn")
        ctrl.addWidget(self.btn_export, 0)

        ctrl.addStretch(1)
        layout.addLayout(ctrl)

        # ── table (2 columns) ──────────────────────────────────────────
        self.table = _ClipboardTableWidget(5, 2, self)
        self.table.setObjectName("stTable")
        self._header_labels = ["Value", "Tracker files sharing the value"]
        self.table.setHorizontalHeaderLabels(self._header_labels)
        self.table.verticalHeader().setVisible(True)
        self.table.verticalHeader().setDefaultSectionSize(30)
        self.table.horizontalHeader().setDefaultSectionSize(200)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(0, 260)

        for row in range(5):
            for col in range(2):
                self.table.setItem(row, col, QTableWidgetItem(""))
            item = self.table.item(row, self._COL_TRACKERS)
            if item:
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

        self._undo_stack: list[list[list[str]]] = []
        self._is_restoring_undo = False
        self._is_batch_edit = False

        layout.addWidget(self.table)

        # ── connections ─────────────────────────────────────────────────
        self.btn_update_rows.clicked.connect(self._on_update_rows)
        self.btn_reset_table.clicked.connect(self._on_reset_table)
        self.btn_delete_row.clicked.connect(self._on_delete_row)
        self.btn_undo_table.clicked.connect(self._on_undo)
        self.btn_import.clicked.connect(self._on_import)
        self.btn_evaluate.clicked.connect(self._on_evaluate)
        self.btn_export.clicked.connect(self._on_export)
        self.table.itemChanged.connect(self._on_item_changed)

        self._push_undo_snapshot(force=True)
        self._update_count_label()

        self.setStyleSheet(
            """
            QDialog { background-color: #F4F4F4; }

            QTableWidget#stTable {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                gridline-color: #B8B8B8;
                selection-background-color: #DCE6F5;
                selection-color: #111111;
                font-family: "Segoe UI";
                font-size: 12px;
            }
            QTableWidget#stTable QHeaderView::section {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #7D8694;
                padding: 6px 8px;
                font-weight: 700;
            }
            QTableWidget#stTable QTableCornerButton::section {
                background-color: #111F35;
                border: 1px solid #7D8694;
            }
            QLabel#stTitle {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 18px;
                font-weight: 800;
            }
            QLabel#stCountLabel {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 13px;
                font-weight: 600;
            }
            QLineEdit#stInput {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 8px;
                font-family: "Segoe UI";
                font-size: 12px;
            }
            QPushButton#stGrayBtn {
                background-color: #9EA3AB;
                color: #000000;
                border: 1px solid #8B9098;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }
            QPushButton#stGrayBtn:pressed {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #111F35;
            }
            QPushButton#stPrimaryBtn {
                background-color: #8A244B;
                color: #FFFFFF;
                border: 1px solid #8A244B;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 700;
            }
            QPushButton#stPrimaryBtn:pressed {
                background-color: #F63049;
                border: 1px solid #F63049;
            }
            """
        )

    # ── helpers ──────────────────────────────────────────────────────────

    def _cell_text(self, row: int, col: int) -> str:
        item = self.table.item(row, col)
        return item.text() if item is not None else ""

    def _set_cell(self, row: int, col: int, value: str, editable: bool = False) -> None:
        item = self.table.item(row, col)
        if item is None:
            item = QTableWidgetItem("")
            self.table.setItem(row, col, item)
        flags = Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled
        if editable:
            flags |= Qt.ItemFlag.ItemIsEditable
        item.setFlags(flags)
        item.setText(value)

    def _update_count_label(self) -> None:
        count = sum(1 for r in range(self.table.rowCount()) if not self.table.isRowHidden(r))
        self.label_visible_count.setText(f"Count: <b>{count}</b>")

    def _capture_state(self) -> list[list[str]]:
        return [
            [self._cell_text(r, c) for c in range(self.table.columnCount())]
            for r in range(self.table.rowCount())
        ]

    def _restore_state(self, state: list[list[str]]) -> None:
        self._is_restoring_undo = True
        self.table.blockSignals(True)
        try:
            self.table.setRowCount(max(1, len(state)))
            for r, row_vals in enumerate(state):
                for c in range(self.table.columnCount()):
                    val = row_vals[c] if c < len(row_vals) else ""
                    editable = (c == self._COL_VALUE)
                    self._set_cell(r, c, val, editable=editable)
        finally:
            self.table.blockSignals(False)
            self._is_restoring_undo = False
        self.row_count_input.setText(str(self.table.rowCount()))
        self._update_count_label()

    def _push_undo_snapshot(self, force: bool = False) -> None:
        if self._is_restoring_undo or self._is_batch_edit:
            return
        snap = self._capture_state()
        if not force and self._undo_stack and self._undo_stack[-1] == snap:
            return
        self._undo_stack.append(snap)
        if len(self._undo_stack) > 200:
            self._undo_stack.pop(0)

    def _init_rows(self, count: int) -> None:
        self.table.setRowCount(count)
        for r in range(count):
            self._set_cell(r, self._COL_VALUE, self._cell_text(r, self._COL_VALUE) or "", editable=True)
            self._set_cell(r, self._COL_TRACKERS, "", editable=False)

    def _resize_rows_to_content(self) -> None:
        """Resize each row height so multi-line cell text is fully visible."""
        fm = self.table.fontMetrics()
        line_h = fm.height() + 4  # per-line height with small padding
        pad = 10  # top+bottom cell padding
        min_h = 30
        for r in range(self.table.rowCount()):
            max_lines = 1
            for c in range(self.table.columnCount()):
                text = self._cell_text(r, c)
                lines = text.count("\n") + 1 if text else 1
                if lines > max_lines:
                    max_lines = lines
            self.table.setRowHeight(r, max(min_h, max_lines * line_h + pad))

    # ── slot handlers ───────────────────────────────────────────────────

    def _on_item_changed(self, _item: QTableWidgetItem) -> None:
        if not self._is_batch_edit:
            self._update_count_label()
            self._push_undo_snapshot()

    def _on_update_rows(self) -> None:
        try:
            n = max(1, min(int(self.row_count_input.text().strip()), 5000))
        except ValueError:
            self.row_count_input.setText(str(self.table.rowCount()))
            return
        self._init_rows(n)
        self.row_count_input.setText(str(n))
        self._update_count_label()
        self._push_undo_snapshot()

    def _on_reset_table(self) -> None:
        self.row_count_input.setText("5")
        self.starting_row_input.setText("7")
        self._tracker_data.clear()
        self._selected_column = None
        self.table.blockSignals(True)
        self.table.clearContents()
        self.table.setRowCount(5)
        for r in range(5):
            self._set_cell(r, self._COL_VALUE, "", editable=True)
            self._set_cell(r, self._COL_TRACKERS, "", editable=False)
        self.table.blockSignals(False)
        self._update_count_label()
        self._push_undo_snapshot()

    def _on_delete_row(self) -> None:
        sel_rows = sorted({idx.row() for idx in self.table.selectedIndexes()}, reverse=True)
        if not sel_rows:
            cur = self.table.currentRow()
            if cur >= 0:
                sel_rows = [cur]
            else:
                return
        total = self.table.rowCount()
        if len(sel_rows) >= total:
            self.table.setRowCount(1)
            self._set_cell(0, self._COL_VALUE, "", editable=True)
            self._set_cell(0, self._COL_TRACKERS, "", editable=False)
        else:
            for r in sel_rows:
                if 0 <= r < self.table.rowCount():
                    self.table.removeRow(r)
        self.table.clearSelection()
        self.row_count_input.setText(str(self.table.rowCount()))
        self._update_count_label()
        self._push_undo_snapshot()

    def _on_undo(self) -> None:
        if self._is_restoring_undo:
            return
        cur = self._capture_state()
        if not self._undo_stack:
            self._undo_stack.append(cur)
            return
        if self._undo_stack[-1] != cur:
            self._undo_stack.append(cur)
        if len(self._undo_stack) <= 1:
            return
        self._undo_stack.pop()
        self._restore_state(self._undo_stack[-1])

    # ── import ──────────────────────────────────────────────────────────

    def _on_import(self) -> None:
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Select Tracker Files", self._browse_dir,
            "Excel Files (*.xlsx *.xls *.csv)"
        )
        if not paths:
            return

        try:
            header_row = max(1, int(self.starting_row_input.text().strip())) - 1
        except ValueError:
            header_row = 0

        import pandas as _pd
        from pathlib import Path as _Path

        def _norm(v) -> str:
            s = str(v).strip() if v is not None else ""
            return "" if s.lower() in ("nan", "none", "") else s

        # Collect headers from all files to find common columns
        all_headers: dict[str, list[tuple[int, str]]] = {}  # path -> [(col_idx, label)]
        file_dfs: dict[str, _pd.DataFrame] = {}

        for fp in paths:
            try:
                if fp.lower().endswith(".csv"):
                    df = _pd.read_csv(fp, header=None, dtype=str)
                else:
                    df = _pd.read_excel(fp, header=None, dtype=str)
            except Exception:
                continue

            file_dfs[fp] = df

            if header_row < df.shape[0]:
                hdr_cells = [_norm(v) for v in df.iloc[header_row]]
            else:
                hdr_cells = [f"Column {i+1}" for i in range(df.shape[1])]

            col_map = [(i, hdr_cells[i]) for i in range(df.shape[1]) if hdr_cells[i] != ""]
            all_headers[fp] = col_map

        if not file_dfs:
            QMessageBox.information(self, "Import", "No valid files could be read.")
            return

        # Gather unique column labels across all files
        unique_labels: list[str] = []
        seen: set[str] = set()
        for col_map in all_headers.values():
            for _, label in col_map:
                key = label.strip().lower()
                if key not in seen:
                    seen.add(key)
                    unique_labels.append(label)

        if not unique_labels:
            QMessageBox.information(self, "Import", "No column headers detected.")
            return

        # Ask user to pick one column
        if len(unique_labels) == 1:
            chosen = unique_labels[0]
        else:
            picker = _SingleTrackerColumnPickerDialog(unique_labels, self)
            if picker.exec() != QDialog.DialogCode.Accepted or picker.get_selected() is None:
                return
            chosen = picker.get_selected()

        self._selected_column = chosen
        chosen_lower = chosen.strip().lower()

        # Read all values from the chosen column across all trackers, dedup
        tracker_data: dict[str, list[str]] = {}  # tracker_name -> values
        all_values_ordered: list[str] = []
        all_values_set: set[str] = set()

        for fp, df in file_dfs.items():
            tracker_name = _Path(fp).stem  # filename without extension
            col_map = all_headers.get(fp, [])

            # Find the column index for the chosen column in this file
            col_idx = None
            for idx, label in col_map:
                if label.strip().lower() == chosen_lower:
                    col_idx = idx
                    break

            if col_idx is None:
                tracker_data[tracker_name] = []
                continue

            data_rows = df.iloc[header_row + 1:].reset_index(drop=True)
            values = []
            for v in data_rows.iloc[:, col_idx]:
                val = _norm(v)
                if val:
                    values.append(val)
                    if val not in all_values_set:
                        all_values_set.add(val)
                        all_values_ordered.append(val)

            tracker_data[tracker_name] = values

        self._tracker_data = tracker_data

        # Populate the table with deduped values
        n = len(all_values_ordered)
        if n == 0:
            QMessageBox.information(self, "Import", "No values found in the selected column.")
            return

        self.table.blockSignals(True)
        self._is_batch_edit = True
        try:
            self.table.setRowCount(n)
            for r, val in enumerate(all_values_ordered):
                self._set_cell(r, self._COL_VALUE, val, editable=True)
                self._set_cell(r, self._COL_TRACKERS, "", editable=False)
        finally:
            self.table.blockSignals(False)
            self._is_batch_edit = False

        self.row_count_input.setText(str(n))
        self._update_count_label()
        self._resize_rows_to_content()
        self._push_undo_snapshot()

    # ── evaluate ────────────────────────────────────────────────────────

    def _on_evaluate(self) -> None:
        if not self._tracker_data:
            QMessageBox.information(self, "Evaluate", "No tracker data loaded. Import trackers first.")
            return

        n = self.table.rowCount()

        # Build a lookup: value -> set of tracker names that contain it
        value_to_trackers: dict[str, set[str]] = {}
        for tracker_name, values in self._tracker_data.items():
            for v in values:
                value_to_trackers.setdefault(v, set()).add(tracker_name)

        # Collect only rows whose value appears in 2+ trackers
        dup_rows: list[tuple[str, str]] = []
        for r in range(n):
            val = self._cell_text(r, self._COL_VALUE).strip()
            if not val:
                continue
            trackers = value_to_trackers.get(val, set())
            if len(trackers) >= 2:
                dup_rows.append((val, "\n".join(sorted(trackers))))

        # Rebuild the table with only duplicate rows
        self.table.blockSignals(True)
        self._is_batch_edit = True
        try:
            self.table.setRowCount(max(1, len(dup_rows)))
            if dup_rows:
                for r, (val, names_str) in enumerate(dup_rows):
                    self._set_cell(r, self._COL_VALUE, val, editable=True)
                    self._set_cell(r, self._COL_TRACKERS, names_str, editable=False)
            else:
                self._set_cell(0, self._COL_VALUE, "", editable=True)
                self._set_cell(0, self._COL_TRACKERS, "", editable=False)
        finally:
            self.table.blockSignals(False)
            self._is_batch_edit = False

        self.row_count_input.setText(str(self.table.rowCount()))
        self._update_count_label()
        self._resize_rows_to_content()
        self._push_undo_snapshot()

    # ── export ──────────────────────────────────────────────────────────

    def _on_export(self) -> None:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(self, "Export Error",
                                "openpyxl is required.\nInstall with: pip install openpyxl")
            return

        import datetime as _dt
        import os as _os

        _ts = _dt.datetime.now().strftime("%Y_%m_%d_%H_%M")
        _default_name = f"duplicate_check_{_ts}.xlsx"
        _default_path = _os.path.join(self._browse_dir, _default_name) if self._browse_dir else _default_name
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Table", _default_path, "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        # ── Style constants ─────────────────────────────────────────────
        header_fill   = PatternFill("solid", fgColor="111F35")
        header_font   = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_align  = Alignment(horizontal="center", vertical="center")
        thin_side     = Side(style="thin", color="7D8694")
        hdr_border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        body_font     = Font(name="Segoe UI", size=10)
        body_align    = Alignment(vertical="center", wrap_text=True)
        body_border   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        def _style_header(ws, headers, widths):
            for ci, (h, w) in enumerate(zip(headers, widths), start=1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, header_align, hdr_border
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.row_dimensions[1].height = 22

        def _write_row(ws, row_num, values):
            for ci, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.font, cell.alignment, cell.border = body_font, body_align, body_border

        wb = openpyxl.Workbook()

        # ── Sheet 1: dup_check_main (general summary) ──────────────────
        ws_main = wb.active
        ws_main.title = "dup_check_main"
        _style_header(ws_main, ["Value", "Tracker files sharing the value"], [32, 50])

        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r):
                continue
            val = self._cell_text(r, self._COL_VALUE)
            trackers_text = self._cell_text(r, self._COL_TRACKERS)
            _write_row(ws_main, ws_main.max_row + 1, [val, trackers_text])

        # ── Per-tracker sheets ──────────────────────────────────────────
        # Build value -> set(tracker_names) from _tracker_data
        value_to_trackers: dict[str, set[str]] = {}
        for tracker_name, values in self._tracker_data.items():
            for v in values:
                value_to_trackers.setdefault(v, set()).add(tracker_name)

        tracker_names = sorted(self._tracker_data.keys())
        for sheet_idx, tracker_name in enumerate(tracker_names, start=1):
            sheet_title = f"dup_check_{sheet_idx}"
            # Ensure sheet name is valid (max 31 chars)
            if len(sheet_title) > 31:
                sheet_title = sheet_title[:31]
            ws = wb.create_sheet(title=sheet_title)
            _style_header(ws, ["File Name", "Value", "Duplicate or Not",
                               "Tracker files sharing the value"],
                          [28, 32, 22, 50])

            tracker_values = self._tracker_data[tracker_name]
            seen: set[str] = set()
            for val in tracker_values:
                if val in seen or not val:
                    continue
                seen.add(val)
                trackers_with_val = value_to_trackers.get(val, set())
                # Only include values that are shared with OTHER trackers
                other_trackers = trackers_with_val - {tracker_name}
                if not other_trackers:
                    continue
                others_str = "\n".join(sorted(other_trackers))
                _write_row(ws, ws.max_row + 1, [tracker_name, val, "duplicate", others_str])

        try:
            wb.save(path)
            QMessageBox.information(self, "Export", f"Exported successfully:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))


class _SdcSelectionDialog(QDialog):
    """
    Shown when multiple SDC output files were created.
    User selects which ones to run the Reference Collector against.
    """

    def __init__(self, sdc_paths: list[str], parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Reference Collector – Select SDC File(s)")
        self.setMinimumWidth(480)

        self.selected_paths: list[str] = []
        self._checkboxes: list[tuple[QCheckBox, str]] = []

        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        lbl = QLabel("Select SDC file(s) to use for Reference Collector:")
        lbl.setWordWrap(True)
        layout.addWidget(lbl)

        for path in sdc_paths:
            cb = QCheckBox(Path(path).name)
            cb.setChecked(True)
            self._checkboxes.append((cb, path))
            layout.addWidget(cb)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_select_all = QPushButton("Select All")
        btn_select_all.clicked.connect(self._select_all)
        btn_row.addWidget(btn_select_all)
        btn_ok = QPushButton("OK")
        btn_ok.setDefault(True)
        btn_ok.clicked.connect(self._on_ok)
        btn_row.addWidget(btn_ok)
        btn_cancel = QPushButton("Cancel")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

    def _select_all(self) -> None:
        for cb, _ in self._checkboxes:
            cb.setChecked(True)

    def _on_ok(self) -> None:
        self.selected_paths = [p for cb, p in self._checkboxes if cb.isChecked()]
        self.accept()


class _TscOption1TableDialog(QDialog):
    """Tracker Information window opened from TSC Option 1."""

    _BLANK_FILTER = "__BLANK__"

    TSC_HEADERS = [
        "SBU",
        "Product Name",
        "IDH Number",
        "Build Type",
        "Status",
        "Packaging Type",
        "Packaging Size",
        "Project Name",
        "Basic Number",
        "Label Size",
        "Is Deployment",
    ]

    class _FilterPopup(QDialog):
        def __init__(
            self,
            values: list[tuple[str, str]],
            selected_values: set[str] | None,
            parent: QWidget | None = None,
        ) -> None:
            super().__init__(parent)
            self.setWindowFlags(Qt.WindowType.Popup)
            self.setMinimumSize(260, 330)
            self.resize(260, 330)

            self._is_syncing = False
            self._value_items: list[tuple[QListWidgetItem, str, str]] = []

            layout = QVBoxLayout(self)
            layout.setContentsMargins(4, 4, 4, 4)
            layout.setSpacing(4)

            self.search_input = QLineEdit(self)
            self.search_input.setPlaceholderText("Search")
            self.search_input.setClearButtonEnabled(True)
            self.search_input.setObjectName("packshotRowCountInput")
            layout.addWidget(self.search_input)

            self.list_widget = QListWidget(self)
            self.list_widget.setSelectionMode(QListWidget.SelectionMode.NoSelection)
            layout.addWidget(self.list_widget, 1)

            self.item_select_all = QListWidgetItem("(Select All)")
            self.item_select_all.setFlags(self.item_select_all.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            self.item_select_all.setCheckState(Qt.CheckState.Checked)
            self.item_select_all.setData(Qt.ItemDataRole.UserRole, "__SELECT_ALL__")
            self.list_widget.addItem(self.item_select_all)

            selected = selected_values if selected_values is not None else {value for value, _label in values}

            for value, label in values:
                item = QListWidgetItem(label)
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                item.setCheckState(Qt.CheckState.Checked if value in selected else Qt.CheckState.Unchecked)
                item.setData(Qt.ItemDataRole.UserRole, value)
                self.list_widget.addItem(item)
                self._value_items.append((item, value, label))

            button_row = QHBoxLayout()
            button_row.addStretch(1)
            self.btn_apply = QPushButton("Apply")
            self.btn_apply.setObjectName("packshotUpdateRowsBtn")
            self.btn_cancel = QPushButton("Cancel")
            self.btn_cancel.setObjectName("packshotUpdateRowsBtn")
            button_row.addWidget(self.btn_apply, 0)
            button_row.addWidget(self.btn_cancel, 0)
            layout.addLayout(button_row)

            self.search_input.textChanged.connect(self._apply_search)
            self.list_widget.itemChanged.connect(self._on_item_changed)
            self.btn_apply.clicked.connect(self.accept)
            self.btn_cancel.clicked.connect(self.reject)

            self.search_input.setFocus()
            self._sync_select_all_state()

        def _apply_search(self, text: str) -> None:
            needle = text.strip().lower()
            for item, _value, label in self._value_items:
                item.setHidden(needle not in label.lower())
            self._sync_select_all_state()

        def _on_item_changed(self, item: QListWidgetItem) -> None:
            if self._is_syncing:
                return
            role = item.data(Qt.ItemDataRole.UserRole)
            if role == "__SELECT_ALL__":
                self._is_syncing = True
                try:
                    target_state = item.checkState()
                    for value_item, _value, _label in self._value_items:
                        if value_item.isHidden():
                            continue
                        value_item.setCheckState(target_state)
                finally:
                    self._is_syncing = False
                return
            self._sync_select_all_state()

        def _sync_select_all_state(self) -> None:
            visible_items = [item for item, _value, _label in self._value_items if not item.isHidden()]
            if not visible_items:
                state = Qt.CheckState.Unchecked
            else:
                all_checked = all(i.checkState() == Qt.CheckState.Checked for i in visible_items)
                any_checked = any(i.checkState() == Qt.CheckState.Checked for i in visible_items)
                if all_checked:
                    state = Qt.CheckState.Checked
                elif any_checked:
                    state = Qt.CheckState.PartiallyChecked
                else:
                    state = Qt.CheckState.Unchecked
            self._is_syncing = True
            try:
                self.item_select_all.setCheckState(state)
            finally:
                self._is_syncing = False

        def get_selected_values(self) -> set[str]:
            selected: set[str] = set()
            for item, value, _label in self._value_items:
                if item.checkState() == Qt.CheckState.Checked:
                    selected.add(value)
            return selected

    def __init__(self, parent: QWidget | None = None, start_dir: str = "") -> None:
        super().__init__(parent)
        self._browse_start_dir = start_dir
        self.setWindowTitle("Open window")
        self.resize(1300, 620)
        self.setMinimumSize(1080, 520)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        # --- Title row with count label ---
        title_row = QHBoxLayout()
        title_row.setSpacing(8)
        title = QLabel("Tracker Information")
        title.setObjectName("packshotTableTitle")
        title_row.addWidget(title, 0)
        title_row.addSpacing(530)
        self.label_row_count_display = QLabel("Count: <b>5</b>")
        self.label_row_count_display.setObjectName("tscCountLabel")
        self.label_row_count_display.setTextFormat(Qt.TextFormat.RichText)
        title_row.addWidget(self.label_row_count_display, 0)
        title_row.addStretch(1)
        layout.addLayout(title_row)

        # --- Config row: Row count, buttons, Import, Reset Filter, Export, Generate Chart ---
        config_row = QHBoxLayout()
        config_row.setSpacing(8)

        row_count_label = QLabel("Row count")
        row_count_label.setObjectName("packshotRowCountLabel")
        config_row.addWidget(row_count_label, 0)

        self.row_count_input = QLineEdit("5")
        self.row_count_input.setObjectName("packshotRowCountInput")
        self.row_count_input.setFixedWidth(45)
        config_row.addWidget(self.row_count_input, 0)

        self.btn_update_rows = QPushButton("Update")
        self.btn_update_rows.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_update_rows, 0)

        self.btn_reset_table = QPushButton("Reset")
        self.btn_reset_table.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_reset_table, 0)

        self.btn_delete_row = QPushButton("Delete")
        self.btn_delete_row.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_delete_row, 0)

        self.btn_undo_table = QPushButton("Undo")
        self.btn_undo_table.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_undo_table, 0)

        config_row.addSpacing(8)

        starting_row_label = QLabel("Starting Row:")
        starting_row_label.setObjectName("packshotRowCountLabel")
        config_row.addWidget(starting_row_label, 0)

        self.starting_row_input = QLineEdit("11")
        self.starting_row_input.setObjectName("packshotRowCountInput")
        self.starting_row_input.setFixedWidth(45)
        config_row.addWidget(self.starting_row_input, 0)


        # Place Import Trackers and Apply Cleanup immediately after Starting Row
        self.btn_import_tracker = QPushButton("Import Trackers")
        self.btn_import_tracker.setObjectName("tscPrimaryBtn")
        config_row.addWidget(self.btn_import_tracker, 0)

        self.btn_apply_cleanup = QPushButton("Apply Cleanup")
        self.btn_apply_cleanup.setObjectName("tscPrimaryBtn")
        config_row.addWidget(self.btn_apply_cleanup, 0)

        self.btn_load_tsc = QPushButton("Load TSC")
        self.btn_load_tsc.setObjectName("tscPrimaryBtn")
        config_row.addWidget(self.btn_load_tsc, 0)

        config_row.addSpacing(8)

        self.import_progress = QProgressBar(self)
        self.import_progress.setObjectName("sapImportProgress")
        self.import_progress.setFixedWidth(140)
        self.import_progress.setRange(0, 100)
        self.import_progress.setValue(0)
        self.import_progress.setVisible(False)
        config_row.addWidget(self.import_progress, 0)

        config_row.addSpacing(20)

        self.btn_reset_filter = QPushButton("Reset Filter")
        self.btn_reset_filter.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_reset_filter, 0)

        self.btn_export = QPushButton("Export")
        # Make Export use the gray style like Reset Filter
        self.btn_export.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_export, 0)

        config_row.addSpacing(20)

        self.btn_generate_chart = QPushButton("Generate Chart")
        self.btn_generate_chart.setObjectName("tscPrimaryBtn")
        config_row.addWidget(self.btn_generate_chart, 0)

        config_row.addStretch(1)
        layout.addLayout(config_row)

        # --- Second row: Status group, Work Type group, Custom Search ---
        action_row = QHBoxLayout()
        action_row.setSpacing(12)

        # -- Status group (gray rounded frame) --
        status_frame = QFrame(self)
        status_frame.setObjectName("tscGroupFrame")
        status_layout = QHBoxLayout(status_frame)
        status_layout.setContentsMargins(10, 6, 10, 6)
        status_layout.setSpacing(8)

        status_label = QLabel("Status:")
        status_label.setObjectName("tscGroupLabel")
        status_layout.addWidget(status_label, 0)

        self.tsc_cb_to_do = QCheckBox("To Do")
        self.tsc_cb_in_progress = QCheckBox("In Progress")
        self.tsc_cb_completed = QCheckBox("Completed")
        self.tsc_cb_on_hold = QCheckBox("On Hold")
        self.tsc_cb_cancelled = QCheckBox("Cancelled")
        for cb in (self.tsc_cb_to_do, self.tsc_cb_in_progress, self.tsc_cb_completed,
                   self.tsc_cb_on_hold, self.tsc_cb_cancelled):
            cb.setObjectName("tscGroupCheck")
            status_layout.addWidget(cb, 0)

        action_row.addWidget(status_frame, 0)

        action_row.addSpacing(8)

        # -- Work Type group (gray rounded frame) --
        worktype_frame = QFrame(self)
        worktype_frame.setObjectName("tscGroupFrame")
        worktype_layout = QHBoxLayout(worktype_frame)
        worktype_layout.setContentsMargins(10, 6, 10, 6)
        worktype_layout.setSpacing(8)

        worktype_label = QLabel("Build Type:")
        worktype_label.setObjectName("tscGroupLabel")
        worktype_layout.addWidget(worktype_label, 0)

        self.tsc_cb_clone = QCheckBox("Clone")
        self.tsc_cb_master = QCheckBox("Master")
        self.tsc_cb_resizing = QCheckBox("Resizing")
        self.tsc_cb_admin = QCheckBox("Admin")
        self.tsc_cb_upload = QCheckBox("Upload")
        for cb in (self.tsc_cb_clone, self.tsc_cb_master, self.tsc_cb_resizing,
                   self.tsc_cb_admin, self.tsc_cb_upload):
            cb.setObjectName("tscGroupCheck")
            worktype_layout.addWidget(cb, 0)

        action_row.addWidget(worktype_frame, 0)

        action_row.addSpacing(12)

        # Custom Search label + input
        self.tsc_custom_value_label = QLabel("Custom Search")
        self.tsc_custom_value_label.setObjectName("tscCustomValueLabel")
        action_row.addWidget(self.tsc_custom_value_label, 0)
        self.tsc_custom_value_input = QLineEdit()
        self.tsc_custom_value_input.setObjectName("tscCustomValueInput")
        self.tsc_custom_value_input.setFixedWidth(156)
        action_row.addWidget(self.tsc_custom_value_input, 0)

        action_row.addStretch(1)
        layout.addLayout(action_row)

        # --- Table with filtering ---
        col_count = len(self.TSC_HEADERS)
        self._column_filters: dict[int, set[str] | None] = {i: None for i in range(col_count)}
        self.table = _ClipboardTableWidget(5, col_count, self)
        self.table.setObjectName("packshotClipboardTable")
        self._refresh_header_labels()
        self.table.verticalHeader().setVisible(True)
        self.table.verticalHeader().setDefaultSectionSize(34)
        self.table.horizontalHeader().setDefaultSectionSize(130)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setSectionsClickable(True)
        self.table.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        self.table.horizontalHeader().sectionDoubleClicked.connect(self._on_filter_header_clicked)

        for row in range(5):
            for col in range(col_count):
                self.table.setItem(row, col, QTableWidgetItem(""))

        self._undo_stack: list[list[list[str]]] = []
        self._is_restoring_undo = False
        self._is_batch_edit = False

        layout.addWidget(self.table)

        self.btn_update_rows.clicked.connect(self._on_update_rows_clicked)
        self.btn_reset_table.clicked.connect(self._on_reset_table_clicked)
        self.btn_delete_row.clicked.connect(self._on_delete_rows_clicked)
        self.btn_undo_table.clicked.connect(self._on_undo_table_clicked)
        self.btn_reset_filter.clicked.connect(self._on_reset_filter_clicked)
        self.table.itemChanged.connect(self._on_table_item_changed)

        # Connect Status / Build Type checkboxes and Custom Search to filtering
        for _cb in (
            self.tsc_cb_to_do, self.tsc_cb_in_progress, self.tsc_cb_completed,
            self.tsc_cb_on_hold, self.tsc_cb_cancelled,
            self.tsc_cb_clone, self.tsc_cb_master, self.tsc_cb_resizing,
            self.tsc_cb_admin, self.tsc_cb_upload,
        ):
            _cb.stateChanged.connect(self._apply_filters)
        self.tsc_custom_value_input.textChanged.connect(self._apply_filters)
        self.btn_generate_chart.clicked.connect(self._on_generate_chart_clicked)

        # Flash Import Trackers, Apply Cleanup, and Load TSC buttons to #D02752 on click
        for _flash_btn in (self.btn_import_tracker, self.btn_apply_cleanup, self.btn_load_tsc):
            _flash_btn.clicked.connect(
                lambda _checked=False, b=_flash_btn: self._flash_primary_btn(b)
            )

        self._push_undo_snapshot(force=True)
        self._update_row_count_display()

        self.setStyleSheet(
            """
            QDialog {
                background-color: #F4F4F4;
            }

            QTableWidget#packshotClipboardTable {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                gridline-color: #B8B8B8;
                selection-background-color: #DCE6F5;
                selection-color: #111111;
                font-family: "Segoe UI";
                font-size: 12px;
            }

            QTableWidget#packshotClipboardTable QHeaderView::section {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #7D8694;
                padding: 6px 8px;
                font-weight: 700;
            }

            QTableWidget#packshotClipboardTable QTableCornerButton::section {
                background-color: #111F35;
                border: 1px solid #7D8694;
            }

            QLabel#packshotTableTitle {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 18px;
                font-weight: 800;
            }

            QLabel#tscCountLabel {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 13px;
                font-weight: 600;
            }

            QLabel#packshotRowCountLabel {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 13px;
                font-weight: 600;
            }

            QLineEdit#packshotRowCountInput {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 8px;
                font-family: "Segoe UI";
                font-size: 12px;
            }

            QProgressBar#sapImportProgress {
                background-color: #FFFFFF;
                border: 1px solid #A9A9A9;
                border-radius: 7px;
                text-align: center;
                color: #111111;
                font-family: "Segoe UI";
                font-size: 11px;
                min-height: 24px;
            }

            QProgressBar#sapImportProgress::chunk {
                background-color: #8A244B;
                border-radius: 6px;
            }

            QPushButton#packshotUpdateRowsBtn {
                background-color: #9EA3AB;
                color: #000000;
                border: 1px solid #8B9098;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }

            QPushButton#packshotUpdateRowsBtn:pressed {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #111F35;
            }

            QPushButton#mapperReformatBtn {
                background-color: #8A244B;
                color: #FFFFFF;
                border: 1px solid #8A244B;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }

            QPushButton#mapperReformatBtn:pressed {
                background-color: #D02752;
                color: #FFFFFF;
                border: 1px solid #D02752;
            }

            /* Primary TSC buttons: Import, Apply Cleanup, Generate Chart */
            QPushButton#tscPrimaryBtn {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #111F35;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }

            QPushButton#tscPrimaryBtn:pressed {
                background-color: #0e1726;
                color: #FFFFFF;
                border: 1px solid #0e1726;
            }

            QFrame#tscGroupFrame {
                background-color: #9EA3AB;
                border-radius: 10px;
            }

            QLabel#tscGroupLabel {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 700;
                background: transparent;
            }

            QCheckBox#tscGroupCheck {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 500;
                spacing: 6px;
                background: transparent;
            }

            QCheckBox#tscGroupCheck::indicator {
                width: 15px;
                height: 15px;
                border: 1px solid #FFFFFF;
                border-radius: 4px;
                background: #FFFFFF;
            }

            QCheckBox#tscGroupCheck::indicator:checked {
                background: #111F35;
                border: 1px solid #111F35;
            }

            QLabel#tscCustomValueLabel {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 13px;
                font-weight: 600;
            }

            QLineEdit#tscCustomValueInput {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 8px;
                font-family: "Segoe UI";
                font-size: 12px;
            }
            """
        )

    # -- Table management helpers (same pattern as SAP window) --

    def _on_table_batch_edit_begin(self) -> None:
        self._is_batch_edit = True

    def _on_table_batch_edit_end(self) -> None:
        self._is_batch_edit = False
        self._push_undo_snapshot()

    def _push_undo_snapshot(self, force: bool = False) -> None:
        if self._is_restoring_undo or (self._is_batch_edit and not force):
            return
        snapshot: list[list[str]] = []
        for r in range(self.table.rowCount()):
            row_data: list[str] = []
            for c in range(self.table.columnCount()):
                item = self.table.item(r, c)
                row_data.append(item.text() if item else "")
            snapshot.append(row_data)
        if self._undo_stack and self._undo_stack[-1] == snapshot:
            return
        self._undo_stack.append(snapshot)
        if len(self._undo_stack) > 50:
            self._undo_stack = self._undo_stack[-50:]

    def _on_table_item_changed(self, item) -> None:
        self._push_undo_snapshot()

    def _on_update_rows_clicked(self) -> None:
        try:
            new_count = int(self.row_count_input.text())
        except ValueError:
            return
        new_count = max(1, min(new_count, 9999))
        col_count = self.table.columnCount()
        old_count = self.table.rowCount()
        self.table.setRowCount(new_count)
        if new_count > old_count:
            for r in range(old_count, new_count):
                for c in range(col_count):
                    self.table.setItem(r, c, QTableWidgetItem(""))
        self._push_undo_snapshot(force=True)
        self._update_row_count_display()

    def _on_reset_table_clicked(self) -> None:
        # Reset table to 5 blank rows
        self.table.blockSignals(True)
        self.table.setRowCount(5)
        col_count = self.table.columnCount()
        for r in range(5):
            for c in range(col_count):
                item = self.table.item(r, c)
                if item is None:
                    self.table.setItem(r, c, QTableWidgetItem(""))
                else:
                    item.setText("")
        self.table.blockSignals(False)

        # Reset row count input
        self.row_count_input.setText("5")

        # Uncheck all Status checkboxes
        for cb in (self.tsc_cb_to_do, self.tsc_cb_in_progress, self.tsc_cb_completed,
                   self.tsc_cb_on_hold, self.tsc_cb_cancelled):
            cb.setChecked(False)

        # Uncheck all Build Type checkboxes
        for cb in (self.tsc_cb_clone, self.tsc_cb_master, self.tsc_cb_resizing,
                   self.tsc_cb_admin, self.tsc_cb_upload):
            cb.setChecked(False)

        # Clear Custom Search input
        self.tsc_custom_value_input.clear()

        # Reset all column filters
        for column in self._column_filters:
            self._column_filters[column] = None
        self._apply_filters()

        self._push_undo_snapshot(force=True)
        self._update_row_count_display()

    def _on_delete_rows_clicked(self) -> None:
        selected = sorted({idx.row() for idx in self.table.selectedIndexes()}, reverse=True)
        if not selected:
            return
        self._push_undo_snapshot(force=True)
        for r in selected:
            self.table.removeRow(r)
        self.row_count_input.setText(str(self.table.rowCount()))
        self._push_undo_snapshot(force=True)
        self._update_row_count_display()

    def _on_undo_table_clicked(self) -> None:
        if len(self._undo_stack) < 2:
            return
        self._undo_stack.pop()
        snapshot = self._undo_stack[-1]
        self._is_restoring_undo = True
        col_count = self.table.columnCount()
        self.table.setRowCount(len(snapshot))
        for r, row_data in enumerate(snapshot):
            for c in range(col_count):
                value = row_data[c] if c < len(row_data) else ""
                item = self.table.item(r, c)
                if item is None:
                    self.table.setItem(r, c, QTableWidgetItem(value))
                else:
                    item.setText(value)
        self.row_count_input.setText(str(self.table.rowCount()))
        self._is_restoring_undo = False
        self._update_row_count_display()

    # -- Filter / header helpers (same pattern as Reformatted table) --

    def _refresh_header_labels(self) -> None:
        labels = []
        for idx, base_label in enumerate(self.TSC_HEADERS):
            has_active_filter = self._column_filters.get(idx) is not None
            labels.append(f"{base_label} {'▾*' if has_active_filter else '▾'}")
        self.table.setHorizontalHeaderLabels(labels)

    def _on_header_clicked(self, column: int) -> None:
        self.table.selectColumn(column)

    def _on_filter_header_clicked(self, column: int) -> None:
        values: list[str] = []
        seen: set[str] = set()
        has_blank = False
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            item = self.table.item(row, column)
            cell_text = item.text().strip() if item is not None else ""
            if cell_text == "":
                has_blank = True
                continue
            if cell_text in seen:
                continue
            seen.add(cell_text)
            values.append(cell_text)

        value_pairs: list[tuple[str, str]] = [(v, v) for v in values]
        if has_blank:
            value_pairs.append((self._BLANK_FILTER, "(Blanks)"))

        popup = self._FilterPopup(value_pairs, self._column_filters.get(column), self)
        popup.move(QCursor.pos())
        if popup.exec() != QDialog.DialogCode.Accepted:
            return

        selected = popup.get_selected_values()
        if len(selected) == len(value_pairs):
            self._column_filters[column] = None
        else:
            self._column_filters[column] = selected
        self._apply_filters()

    def _apply_filters(self) -> None:
        # Determine which Status / Build Type checkbox values are active
        status_col = self.TSC_HEADERS.index("Status")
        build_type_col = self.TSC_HEADERS.index("Build Type")

        status_cb_map: dict[str, QCheckBox] = {
            "to do":       self.tsc_cb_to_do,
            "in progress": self.tsc_cb_in_progress,
            "completed":   self.tsc_cb_completed,
            "on hold":     self.tsc_cb_on_hold,
            "cancelled":   self.tsc_cb_cancelled,
        }
        build_type_cb_map: dict[str, QCheckBox] = {
            "clone":    self.tsc_cb_clone,
            "master":   self.tsc_cb_master,
            "resizing": self.tsc_cb_resizing,
            "admin":    self.tsc_cb_admin,
            "upload":   self.tsc_cb_upload,
        }

        checked_statuses    = {v for v, cb in status_cb_map.items()    if cb.isChecked()}
        checked_build_types = {v for v, cb in build_type_cb_map.items() if cb.isChecked()}
        custom_needle = self.tsc_custom_value_input.text().strip().lower()

        for row in range(self.table.rowCount()):
            row_matches = True

            # 1. Column header popup filters
            for col, filter_value in self._column_filters.items():
                if filter_value is None:
                    continue
                item = self.table.item(row, col)
                cell_text = item.text().strip() if item is not None else ""
                normalized = self._BLANK_FILTER if cell_text == "" else cell_text
                if normalized not in filter_value:
                    row_matches = False
                    break

            # 2. Status checkboxes – only active when at least one cb is checked
            if row_matches and checked_statuses:
                item = self.table.item(row, status_col)
                cell_val = item.text().strip().lower() if item is not None else ""
                if cell_val not in checked_statuses:
                    row_matches = False

            # 3. Build Type checkboxes – only active when at least one cb is checked
            if row_matches and checked_build_types:
                item = self.table.item(row, build_type_col)
                cell_val = item.text().strip().lower() if item is not None else ""
                if cell_val not in checked_build_types:
                    row_matches = False

            # 4. Custom Search – substring match across ALL columns (case-insensitive)
            if row_matches and custom_needle:
                found = any(
                    custom_needle in (self.table.item(row, c).text().lower()
                                      if self.table.item(row, c) is not None else "")
                    for c in range(self.table.columnCount())
                )
                if not found:
                    row_matches = False

            self.table.setRowHidden(row, not row_matches)

        self._refresh_header_labels()
        self._update_row_count_display()

    def _on_reset_filter_clicked(self) -> None:
        # Clear column popup filters
        for column in self._column_filters:
            self._column_filters[column] = None
        # Uncheck all Status checkboxes
        for cb in (self.tsc_cb_to_do, self.tsc_cb_in_progress, self.tsc_cb_completed,
                   self.tsc_cb_on_hold, self.tsc_cb_cancelled):
            cb.setChecked(False)
        # Uncheck all Build Type checkboxes
        for cb in (self.tsc_cb_clone, self.tsc_cb_master, self.tsc_cb_resizing,
                   self.tsc_cb_admin, self.tsc_cb_upload):
            cb.setChecked(False)
        # Clear custom search
        self.tsc_custom_value_input.clear()
        self._apply_filters()

    def _update_row_count_display(self) -> None:
        visible = sum(1 for r in range(self.table.rowCount()) if not self.table.isRowHidden(r))
        self.label_row_count_display.setText(f"Count: <b>{visible}</b>")

    def _on_generate_chart_clicked(self) -> None:
        """Collect visible table rows and open the analytics chart dialog."""
        rows: list[list[str]] = []
        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r):
                continue
            row_data = [
                (self.table.item(r, c).text() if self.table.item(r, c) is not None else "")
                for c in range(self.table.columnCount())
            ]
            rows.append(row_data)

        if not rows:
            QMessageBox.information(self, "No Data", "The table has no visible rows to chart.")
            return

        dlg = _TscChartDialog(list(self.TSC_HEADERS), rows, parent=self)
        dlg.exec()

    def _flash_primary_btn(self, btn: QPushButton) -> None:
        """Momentarily flash a tscPrimaryBtn to #D02752 on click."""
        btn.setStyleSheet(
            "QPushButton { background-color: #D02752; color: #FFFFFF; "
            "border: 1px solid #D02752; border-radius: 8px; min-height: 30px; "
            "padding: 0 12px; font-family: 'Segoe UI'; font-size: 12px; font-weight: 600; }"
        )
        QTimer.singleShot(250, lambda: btn.setStyleSheet(""))


class _TscChartDialog(QDialog):
    """Analytics chart window for the Tracker Information table."""

    _PALETTE = [
        "#8A244B", "#1F6FAE", "#2E9E5B", "#E69C2F", "#6B4DB5",
        "#D44C37", "#3AABB5", "#A5785D", "#C557A8", "#5C7D3E",
    ]

    _PALETTE_1 = ["#32175c","#003d8f","#0062b8","#0086cd","#00a8ca","#00c9b1","#00e689","#7cff58"]
    _PALETTE_2 = ["#5c0000","#760c25","#8b2047","#9b376d","#a45193","#a66cba","#9f88de","#90a4ff"]
    _PALETTE_3 = ["#4348b9","#9745af","#ca4a9e","#eb5c8b","#ff797b","#ff9b73","#ffbd76","#ffde88"]

    _VARIABLES = [
        "SBU",
        "Build Type",
        "Status",
        "Project Name",
        "Packaging Type",
        "Packaging Size",
        "Label Size",
        "Basic Number",
        "Is Deployment",
    ]

    _COUNT_CHART_TYPES = [
        "Pie Chart",
        "Bar Chart (h)",
        "Bar Chart (v)",
        "Donut Chart",
        "Histogram",
    ]

    _COMBO_CHART_TYPES = [
        "Stacked Bar (v)",
        "Stacked Bar (h)",
        "Grouped Bar (v)",
        "Grouped Bar (h)",
    ]

    def __init__(
        self,
        headers: list[str],
        table_data: list[list[str]],
        parent=None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Analytics")
        self.resize(1180, 700)
        self.setMinimumSize(900, 540)
        self._headers = headers
        self._data = table_data
        self._setup_ui()
        self._refresh_chart()

    # ------------------------------------------------------------------ UI --

    def _setup_ui(self) -> None:
        self.setStyleSheet("""
            QDialog  { background-color: #0F1B2E; }
            QWidget#sidebar {
                background-color: #111F35;
                border-right: 1px solid #1E3050;
            }
            QLabel#sideLabel {
                color: #A0B0C4;
                font-family: "Segoe UI"; font-size: 11px; font-weight: 600;
                text-transform: uppercase; letter-spacing: 1px;
            }
            QLabel#sideTitle {
                color: #FFFFFF;
                font-family: "Segoe UI"; font-size: 15px; font-weight: 800;
                padding-bottom: 6px;
            }
            QLabel#sideFieldLabel {
                color: #C8D5E4;
                font-family: "Segoe UI"; font-size: 12px; font-weight: 600;
            }
            QRadioButton#chartModeRadio {
                color: #FFFFFF;
                font-family: "Segoe UI"; font-size: 12px; font-weight: 600;
                spacing: 8px;
            }
            QRadioButton#chartModeRadio::indicator {
                width: 15px; height: 15px;
                border-radius: 8px;
                border: 2px solid #7A8BA0;
                background: transparent;
            }
            QRadioButton#chartModeRadio::indicator:checked {
                border: 2px solid #8A244B;
                background: #8A244B;
            }
            QComboBox#sideCombo {
                background-color: #1C2E47;
                color: #FFFFFF;
                border: 1px solid #2D4260;
                border-radius: 6px;
                min-height: 30px;
                padding: 0 10px;
                font-family: "Segoe UI"; font-size: 12px;
            }
            QComboBox#sideCombo:hover { border: 1px solid #8A244B; }
            QComboBox#sideCombo::drop-down { border: none; width: 22px; }
            QComboBox#sideCombo QAbstractItemView {
                background-color: #1C2E47;
                color: #FFFFFF;
                selection-background-color: #8A244B;
                border: 1px solid #2D4260;
                outline: none;
            }
            QCheckBox#sideCheck {
                color: #C8D5E4;
                font-family: "Segoe UI"; font-size: 12px;
                spacing: 8px;
            }
            QCheckBox#sideCheck::indicator {
                width: 15px; height: 15px;
                border-radius: 4px;
                border: 1px solid #7A8BA0;
                background: transparent;
            }
            QCheckBox#sideCheck::indicator:checked {
                background: #8A244B; border: 1px solid #8A244B;
            }
            QFrame#sideDivider {
                background: #1E3050; border: none;
                max-height: 1px; min-height: 1px;
            }
            QLabel#chartTitle {
                color: #FFFFFF;
                font-family: "Segoe UI"; font-size: 16px; font-weight: 700;
            }
            QLabel#chartSubtitle {
                color: #7A8BA0;
                font-family: "Segoe UI"; font-size: 11px;
            }
            QPushButton#chartSaveBtn {
                background-color: #8A244B; color: #FFFFFF;
                border: none; border-radius: 8px;
                min-height: 32px; padding: 0 20px;
                font-family: "Segoe UI"; font-size: 12px; font-weight: 600;
            }
            QPushButton#chartSaveBtn:hover   { background-color: #A82D5C; }
            QPushButton#chartSaveBtn:pressed { background-color: #6E1A3C; }
            QPushButton#chartCloseBtn {
                background-color: #2D3E58; color: #FFFFFF;
                border: none; border-radius: 8px;
                min-height: 32px; padding: 0 20px;
                font-family: "Segoe UI"; font-size: 12px; font-weight: 600;
            }
            QPushButton#chartCloseBtn:hover { background-color: #374D6B; }
            QRadioButton#textColorRadio {
                color: #C8D5E4;
                font-family: "Segoe UI"; font-size: 12px; font-weight: 500;
                spacing: 8px;
            }
            QRadioButton#textColorRadio::indicator {
                width: 13px; height: 13px;
                border-radius: 7px;
                border: 2px solid #7A8BA0;
                background: transparent;
            }
            QRadioButton#textColorRadio::indicator:checked {
                border: 2px solid #8A244B;
                background: #8A244B;
            }
            QPushButton#colorSwatchBtn {
                border: 2px solid #FFFFFF;
                border-radius: 4px;
                min-width: 28px; max-width: 28px;
                min-height: 20px; max-height: 20px;
            }
            QRadioButton#paletteRadio {
                color: #C8D5E4;
                font-family: "Segoe UI"; font-size: 12px;
                spacing: 8px;
            }
            QRadioButton#paletteRadio::indicator {
                width: 13px; height: 13px;
                border-radius: 7px;
                border: 2px solid #7A8BA0;
                background: transparent;
            }
            QRadioButton#paletteRadio::indicator:checked {
                border: 2px solid #8A244B;
                background: #8A244B;
            }
        """)

        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Left sidebar ─────────────────────────────────────────────────
        sidebar = QWidget()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(240)
        sb = QVBoxLayout(sidebar)
        sb.setContentsMargins(16, 18, 16, 18)
        sb.setSpacing(0)

        # Title
        title_lbl = QLabel("Analytics")
        title_lbl.setObjectName("sideTitle")
        sb.addWidget(title_lbl)

        # ── Mode section label
        mode_sect = QLabel("MODE")
        mode_sect.setObjectName("sideLabel")
        mode_sect.setContentsMargins(0, 10, 0, 6)
        sb.addWidget(mode_sect)

        self._radio_count = QRadioButton("Count Chart")
        self._radio_count.setObjectName("chartModeRadio")
        self._radio_count.setChecked(True)
        self._radio_combo = QRadioButton("Combination Chart")
        self._radio_combo.setObjectName("chartModeRadio")
        mode_group = QButtonGroup(self)
        mode_group.addButton(self._radio_count)
        mode_group.addButton(self._radio_combo)
        sb.addWidget(self._radio_count)
        sb.addSpacing(4)
        sb.addWidget(self._radio_combo)

        # Divider
        div1 = QFrame()
        div1.setObjectName("sideDivider")
        div1.setContentsMargins(0, 14, 0, 14)
        sb.addSpacing(14)
        sb.addWidget(div1)
        sb.addSpacing(14)

        # ── Variable section
        var_sect = QLabel("VARIABLES")
        var_sect.setObjectName("sideLabel")
        var_sect.setContentsMargins(0, 0, 0, 6)
        sb.addWidget(var_sect)

        var_lbl = QLabel("Variable:")
        var_lbl.setObjectName("sideFieldLabel")
        sb.addWidget(var_lbl)
        self._var_combo = QComboBox()
        self._var_combo.setObjectName("sideCombo")
        self._var_combo.addItems(self._VARIABLES)
        sb.addSpacing(4)
        sb.addWidget(self._var_combo)

        # Segment-by row (Combination only)
        sb.addSpacing(10)
        self._seg_lbl = QLabel("Segment by:")
        self._seg_lbl.setObjectName("sideFieldLabel")
        sb.addWidget(self._seg_lbl)
        self._seg_combo = QComboBox()
        self._seg_combo.setObjectName("sideCombo")
        self._seg_combo.addItems(self._VARIABLES)
        self._seg_combo.setCurrentIndex(2)   # default: Status
        sb.addSpacing(4)
        sb.addWidget(self._seg_combo)

        # Divider
        div2 = QFrame()
        div2.setObjectName("sideDivider")
        sb.addSpacing(14)
        sb.addWidget(div2)
        sb.addSpacing(14)

        # ── Chart Type section
        ct_sect = QLabel("CHART TYPE")
        ct_sect.setObjectName("sideLabel")
        ct_sect.setContentsMargins(0, 0, 0, 6)
        sb.addWidget(ct_sect)

        ct_lbl = QLabel("Chart Type:")
        ct_lbl.setObjectName("sideFieldLabel")
        sb.addWidget(ct_lbl)
        self._chart_type_combo = QComboBox()
        self._chart_type_combo.setObjectName("sideCombo")
        self._chart_type_combo.addItems(self._COUNT_CHART_TYPES)
        # Default for Count Chart should be Donut Chart
        try:
            self._chart_type_combo.setCurrentIndex(self._COUNT_CHART_TYPES.index("Donut Chart"))
        except ValueError:
            pass
        sb.addSpacing(4)
        sb.addWidget(self._chart_type_combo)

        # Divider
        div3 = QFrame()
        div3.setObjectName("sideDivider")
        sb.addSpacing(14)
        sb.addWidget(div3)
        sb.addSpacing(10)

        # ── Options section
        opts_sect = QLabel("OPTIONS")
        opts_sect.setObjectName("sideLabel")
        opts_sect.setContentsMargins(0, 0, 0, 8)
        sb.addWidget(opts_sect)

        self._include_blanks_cb = QCheckBox("Include blanks")
        self._include_blanks_cb.setObjectName("sideCheck")
        sb.addWidget(self._include_blanks_cb)

        # THEME dropdown
        theme_label = QLabel("THEME")
        theme_label.setObjectName("sideLabel")
        theme_label.setContentsMargins(0, 10, 0, 6)
        sb.addWidget(theme_label)

        self._theme_combo = QComboBox()
        self._theme_combo.setObjectName("sideCombo")
        self._theme_combo.addItems(["single hue", "divergent", "palette"])
        # Default to "palette"
        self._theme_combo.setCurrentIndex(self._theme_combo.findText("palette"))
        sb.addSpacing(4)
        sb.addWidget(self._theme_combo)

        # Divider
        div_tc = QFrame()
        div_tc.setObjectName("sideDivider")
        sb.addSpacing(14)
        sb.addWidget(div_tc)
        sb.addSpacing(10)

        # TEXT COLOR
        tc_sect = QLabel("TEXT COLOR")
        tc_sect.setObjectName("sideLabel")
        tc_sect.setContentsMargins(0, 0, 0, 6)
        sb.addWidget(tc_sect)

        self._tc_radio_white = QRadioButton("White")
        self._tc_radio_white.setObjectName("textColorRadio")
        self._tc_radio_white.setChecked(True)
        self._tc_radio_black = QRadioButton("Black")
        self._tc_radio_black.setObjectName("textColorRadio")
        tc_group = QButtonGroup(self)
        tc_group.addButton(self._tc_radio_white)
        tc_group.addButton(self._tc_radio_black)
        sb.addWidget(self._tc_radio_white)
        sb.addSpacing(4)
        sb.addWidget(self._tc_radio_black)

        # stash initial single-hue / divergent colors
        self._single_hue_color = "#8A244B"
        self._divergent_color1 = "#488f31"
        self._divergent_color2 = "#de425b"
        self._palette_choice   = 1

        sb.addStretch(1)
        root.addWidget(sidebar)

        # Thin divider
        vdiv = QFrame()
        vdiv.setFrameShape(QFrame.Shape.VLine)
        vdiv.setFixedWidth(1)
        vdiv.setStyleSheet("background:#1E3050; border:none;")
        root.addWidget(vdiv)

        # ── Right chart panel ─────────────────────────────────────────────
        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(20, 18, 20, 14)
        right_layout.setSpacing(10)

        # Title row
        title_row = QHBoxLayout()
        self._title_lbl = QLabel("")
        self._title_lbl.setObjectName("chartTitle")
        self._subtitle_lbl = QLabel("")
        self._subtitle_lbl.setObjectName("chartSubtitle")
        title_row.addWidget(self._title_lbl)
        title_row.addStretch(1)
        title_row.addWidget(self._subtitle_lbl)
        right_layout.addLayout(title_row)

        # Matplotlib canvas – tight_layout disabled so figure never auto-resizes on redraws
        self._figure = Figure(facecolor="#FFFFFF")
        self._canvas = FigureCanvasQTAgg(self._figure)
        self._canvas.setMinimumHeight(400)
        self._canvas.setStyleSheet(
            "background:#FFFFFF; border:1px solid #D0D6DF; border-radius:8px;"
        )
        right_layout.addWidget(self._canvas, 1)

        # ── Bottom dynamic panel (color pickers / palette radios) ─────────
        self._bottom_panel = QWidget()
        bp = QHBoxLayout(self._bottom_panel)
        bp.setContentsMargins(0, 0, 0, 0)
        bp.setSpacing(12)

        # single-hue: 1 swatch
        self._sh_swatch = QPushButton()
        self._sh_swatch.setObjectName("colorSwatchBtn")
        self._sh_swatch.setFixedSize(34, 24)
        self._sh_swatch.setStyleSheet("background:#8A244B; border:2px solid #FFFFFF; border-radius:4px;")
        self._sh_label = QLabel("Hue color:")
        self._sh_label.setObjectName("sideFieldLabel")

        # divergent: 2 swatches
        self._dv_swatch1 = QPushButton()
        self._dv_swatch1.setObjectName("colorSwatchBtn")
        self._dv_swatch1.setFixedSize(34, 24)
        self._dv_swatch1.setStyleSheet("background:#488f31; border:2px solid #FFFFFF; border-radius:4px;")
        self._dv_label1 = QLabel("Color A:")
        self._dv_label1.setObjectName("sideFieldLabel")
        self._dv_swatch2 = QPushButton()
        self._dv_swatch2.setObjectName("colorSwatchBtn")
        self._dv_swatch2.setFixedSize(34, 24)
        self._dv_swatch2.setStyleSheet("background:#de425b; border:2px solid #FFFFFF; border-radius:4px;")
        self._dv_label2 = QLabel("Color B:")
        self._dv_label2.setObjectName("sideFieldLabel")

        # palette: 3 radio buttons
        self._pal_radio1 = QRadioButton("palette-1")
        self._pal_radio1.setObjectName("paletteRadio")
        self._pal_radio1.setChecked(True)
        self._pal_radio2 = QRadioButton("palette-2")
        self._pal_radio2.setObjectName("paletteRadio")
        self._pal_radio3 = QRadioButton("palette-3")
        self._pal_radio3.setObjectName("paletteRadio")
        pal_grp = QButtonGroup(self)
        pal_grp.addButton(self._pal_radio1)
        pal_grp.addButton(self._pal_radio2)
        pal_grp.addButton(self._pal_radio3)

        # all hidden initially; shown by _on_theme_changed
        for w in (self._sh_label, self._sh_swatch,
                  self._dv_label1, self._dv_swatch1, self._dv_label2, self._dv_swatch2,
                  self._pal_radio1, self._pal_radio2, self._pal_radio3):
            bp.addWidget(w)
            w.hide()
        bp.addStretch(1)
        # Fixed height prevents the canvas from resizing when bottom panel contents change
        self._bottom_panel.setFixedHeight(36)
        right_layout.addWidget(self._bottom_panel)

        # Button bar
        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        save_btn = QPushButton("Save Chart")
        save_btn.setObjectName("chartSaveBtn")
        close_btn = QPushButton("Close")
        close_btn.setObjectName("chartCloseBtn")
        btn_row.addWidget(save_btn)
        btn_row.addSpacing(8)
        btn_row.addWidget(close_btn)
        right_layout.addLayout(btn_row)

        root.addWidget(right, 1)

        # ── Wire up signals ───────────────────────────────────────────────
        self._radio_count.toggled.connect(self._on_mode_changed)
        self._var_combo.currentIndexChanged.connect(self._refresh_chart)
        self._seg_combo.currentIndexChanged.connect(self._refresh_chart)
        self._chart_type_combo.currentIndexChanged.connect(self._refresh_chart)
        self._include_blanks_cb.stateChanged.connect(self._refresh_chart)
        self._theme_combo.currentIndexChanged.connect(self._on_theme_changed)
        self._tc_radio_white.toggled.connect(self._refresh_chart)
        self._tc_radio_black.toggled.connect(self._refresh_chart)
        self._sh_swatch.clicked.connect(self._on_pick_single_hue)
        self._dv_swatch1.clicked.connect(lambda: self._on_pick_divergent(1))
        self._dv_swatch2.clicked.connect(lambda: self._on_pick_divergent(2))
        self._pal_radio1.toggled.connect(lambda checked: self._on_palette_radio(1) if checked else None)
        self._pal_radio2.toggled.connect(lambda checked: self._on_palette_radio(2) if checked else None)
        self._pal_radio3.toggled.connect(lambda checked: self._on_palette_radio(3) if checked else None)
        save_btn.clicked.connect(self._on_save_chart)
        close_btn.clicked.connect(self.reject)

        # Initial mode state
        self._on_mode_changed()
        self._on_theme_changed()

    # ---------------------------------------------------------------- slots --

    def _on_mode_changed(self) -> None:
        is_count = self._radio_count.isChecked()
        # Show/hide segment-by controls
        self._seg_lbl.setVisible(not is_count)
        self._seg_combo.setVisible(not is_count)
        # Swap chart-type list
        self._chart_type_combo.blockSignals(True)
        self._chart_type_combo.clear()
        self._chart_type_combo.addItems(
            self._COUNT_CHART_TYPES if is_count else self._COMBO_CHART_TYPES
        )
        self._chart_type_combo.blockSignals(False)
        # Default to Donut Chart when in Count mode
        if is_count:
            try:
                self._chart_type_combo.setCurrentIndex(self._COUNT_CHART_TYPES.index("Donut Chart"))
            except ValueError:
                pass
        self._refresh_chart()

    def _refresh_chart(self) -> None:
        self._figure.clear()
        var      = self._var_combo.currentText()
        seg      = self._seg_combo.currentText()
        ctype    = self._chart_type_combo.currentText()
        blanks   = self._include_blanks_cb.isChecked()
        is_count = self._radio_count.isChecked()

        n_rows = len(self._data)
        self._subtitle_lbl.setText(f"Based on {n_rows} visible row(s)")

        if is_count:
            self._title_lbl.setText(f"{var}  —  {ctype}")
            if ctype == "Pie Chart":
                self._draw_pie(var, blanks, donut=False)
            elif ctype == "Donut Chart":
                self._draw_pie(var, blanks, donut=True)
            elif ctype == "Bar Chart (h)":
                self._draw_bar(var, blanks, horizontal=True)
            elif ctype == "Bar Chart (v)":
                self._draw_bar(var, blanks, horizontal=False)
            elif ctype == "Histogram":
                self._draw_histogram(var, blanks)
        else:
            self._title_lbl.setText(f"{seg} by {var}  —  {ctype}")
            if "(h)" in ctype:
                self._draw_stacked(var, seg, blanks, horizontal=True,
                                   grouped="Grouped" in ctype)
            else:
                self._draw_stacked(var, seg, blanks, horizontal=False,
                                   grouped="Grouped" in ctype)

        self._canvas.draw()

    # -------------------------------------------------------- helpers --------

    def _col_values(self, col_name: str, include_blanks: bool = False) -> list[str]:
        if col_name not in self._headers:
            return []
        idx = self._headers.index(col_name)
        values = []
        for row in self._data:
            v = row[idx].strip() if idx < len(row) else ""
            if v or include_blanks:
                values.append(v if v else "(blank)")
        return values

    def _value_counts(self, col_name: str, include_blanks: bool = False) -> dict[str, int]:
        from collections import Counter
        return dict(Counter(self._col_values(col_name, include_blanks)).most_common())

    # ------------------------------------------------------ theme helpers --

    @staticmethod
    def _hex_to_rgb(h: str) -> tuple[int, int, int]:
        h = h.lstrip("#")
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)

    @staticmethod
    def _rgb_to_hex(r: int, g: int, b: int) -> str:
        return f"#{r:02x}{g:02x}{b:02x}"

    @classmethod
    def _make_single_hue_colors(cls, base_hex: str, n: int) -> list[str]:
        """Produce n shades blending from base_hex toward near-white."""
        if n <= 0:
            return []
        r0, g0, b0 = cls._hex_to_rgb(base_hex)
        # target: a very light tint
        r1, g1, b1 = 255, 220, 240
        result = []
        for i in range(n):
            t = i / max(n - 1, 1)
            r = round(r0 + (r1 - r0) * t)
            g = round(g0 + (g1 - g0) * t)
            b = round(b0 + (b1 - b0) * t)
            result.append(cls._rgb_to_hex(r, g, b))
        return result

    @classmethod
    def _make_divergent_colors(cls, hex1: str, hex2: str, n: int) -> list[str]:
        """Produce n shades from hex1 → white/light-grey → hex2."""
        if n <= 0:
            return []
        r0, g0, b0 = cls._hex_to_rgb(hex1)
        r2, g2, b2 = cls._hex_to_rgb(hex2)
        mid_r, mid_g, mid_b = 241, 241, 241  # near-white midpoint
        result = []
        for i in range(n):
            t = i / max(n - 1, 1)  # 0 → 1
            if t <= 0.5:
                s = t / 0.5
                r = round(r0 + (mid_r - r0) * s)
                g = round(g0 + (mid_g - g0) * s)
                b = round(b0 + (mid_b - b0) * s)
            else:
                s = (t - 0.5) / 0.5
                r = round(mid_r + (r2 - mid_r) * s)
                g = round(mid_g + (g2 - mid_g) * s)
                b = round(mid_b + (b2 - mid_b) * s)
            result.append(cls._rgb_to_hex(r, g, b))
        return result

    def _colors(self, n: int) -> list[str]:
        theme = self._theme_combo.currentText()
        if theme == "single hue":
            return self._make_single_hue_colors(self._single_hue_color, n)
        if theme == "divergent":
            return self._make_divergent_colors(self._divergent_color1, self._divergent_color2, n)
        # palette (also the default fallback)
        p = {1: self._PALETTE_1, 2: self._PALETTE_2, 3: self._PALETTE_3}.get(self._palette_choice, self._PALETTE_1)
        return (p * ((n // len(p)) + 1))[:n]

    def _no_data(self) -> None:
        ax = self._figure.add_subplot(111)
        ax.text(0.5, 0.5, "No data available for this chart",
                ha="center", va="center", fontsize=13,
                color="#AAAAAA", transform=ax.transAxes)
        ax.axis("off")

    # -------------------------------------------------------- chart renderers

    def _draw_pie(self, col: str, blanks: bool, donut: bool = False) -> None:
        counts = self._value_counts(col, blanks)
        if not counts:
            self._no_data()
            return
        ax = self._figure.add_subplot(111)
        # Move pie/donut leftwards to leave space for details/legend
        try:
            ax.set_position([0.05, 0.12, 0.62, 0.8])
        except Exception:
            pass
        labels = list(counts.keys())
        values = list(counts.values())
        colors = self._colors(len(labels))
        wedgeprops = {"linewidth": 1.4, "edgecolor": "#FFFFFF"}
        if donut:
            wedgeprops["width"] = 0.52
        wedges, _tx, autotexts = ax.pie(
            values, labels=None, colors=colors,
            autopct=lambda p: f"{p:.1f}%" if p >= 2 else "",
            startangle=140, pctdistance=0.78 if not donut else 0.82,
            wedgeprops=wedgeprops,
        )
        pct_color = "#FFFFFF" if self._tc_radio_white.isChecked() else "#111111"
        for at in autotexts:
            at.set_fontsize(9)
            at.set_color(pct_color)
            at.set_fontweight("bold")
        if donut:
            total = sum(values)
            ax.text(0, 0, str(total), ha="center", va="center",
                    fontsize=16, fontweight="bold", color="#333333")
        legend_labels = [f"{lbl}  ({v})" for lbl, v in zip(labels, values)]
        ax.legend(wedges, legend_labels,
                  loc="center left", bbox_to_anchor=(0.93, 0.5),
                  fontsize=9, frameon=True, framealpha=0.95, edgecolor="#CCCCCC")
        ax.set_title(f"{col} Distribution  (total {sum(values)})",
                     fontsize=13, fontweight="bold", pad=14)

    def _on_theme_changed(self) -> None:
        """Show/hide bottom-panel controls — panel has fixed height so canvas never resizes."""
        theme = self._theme_combo.currentText()
        # single-hue controls
        for w in (self._sh_label, self._sh_swatch):
            w.setVisible(theme == "single hue")
        # divergent controls
        for w in (self._dv_label1, self._dv_swatch1, self._dv_label2, self._dv_swatch2):
            w.setVisible(theme == "divergent")
        # palette controls
        for w in (self._pal_radio1, self._pal_radio2, self._pal_radio3):
            w.setVisible(theme == "palette")
        self._refresh_chart()

    def _on_pick_single_hue(self) -> None:
        color = QColorDialog.getColor(QColor(self._single_hue_color), self, "Pick single hue color")
        if color.isValid():
            self._single_hue_color = color.name()
            self._sh_swatch.setStyleSheet(
                f"background:{self._single_hue_color}; border:2px solid #FFFFFF; border-radius:4px;"
            )
            self._refresh_chart()

    def _on_pick_divergent(self, which: int) -> None:
        start = self._divergent_color1 if which == 1 else self._divergent_color2
        color = QColorDialog.getColor(QColor(start), self, f"Pick divergent color {'A' if which==1 else 'B'}")
        if color.isValid():
            hexc = color.name()
            if which == 1:
                self._divergent_color1 = hexc
                self._dv_swatch1.setStyleSheet(
                    f"background:{hexc}; border:2px solid #FFFFFF; border-radius:4px;"
                )
            else:
                self._divergent_color2 = hexc
                self._dv_swatch2.setStyleSheet(
                    f"background:{hexc}; border:2px solid #FFFFFF; border-radius:4px;"
                )
            self._refresh_chart()

    def _on_palette_radio(self, choice: int) -> None:
        self._palette_choice = choice
        self._refresh_chart()

    def _draw_bar(self, col: str, blanks: bool, horizontal: bool = True,
                  max_items: int = 25) -> None:
        counts = self._value_counts(col, blanks)
        if not counts:
            self._no_data()
            return
        items = sorted(counts.items(), key=lambda x: -x[1])[:max_items]
        labels = [k for k, _ in items]
        values = [v for _, v in items]
        colors = self._colors(len(labels))
        ax = self._figure.add_subplot(111)
        if horizontal:
            y_pos = list(range(len(labels)))
            ax.barh(y_pos, values[::-1], color=colors[::-1],
                    edgecolor="#FFFFFF", linewidth=0.5, height=0.65)
            ax.set_yticks(y_pos)
            ax.set_yticklabels(labels[::-1], fontsize=9)
            for i, val in enumerate(values[::-1]):
                ax.text(val + max(values) * 0.01, i, str(val),
                        va="center", ha="left", fontsize=9, color="#333333")
            ax.set_xlabel("Count", fontsize=10)
            ax.set_xlim(0, max(values) * 1.2)
            ax.tick_params(axis="y", length=0)
        else:
            x_pos = list(range(len(labels)))
            ax.bar(x_pos, values, color=colors,
                   edgecolor="#FFFFFF", linewidth=0.5, width=0.65)
            ax.set_xticks(x_pos)
            ax.set_xticklabels(labels, rotation=30, ha="right", fontsize=9)
            for i, val in enumerate(values):
                ax.text(i, val + max(values) * 0.01, str(val),
                        ha="center", va="bottom", fontsize=9, color="#333333")
            ax.set_ylabel("Count", fontsize=10)
            ax.set_ylim(0, max(values) * 1.15)
            ax.tick_params(axis="x", length=0)
        ax.set_title(f"{col} Count", fontsize=13, fontweight="bold", pad=14)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    def _draw_histogram(self, col: str, blanks: bool) -> None:
        """Numeric histogram – falls back to bar chart if values are non-numeric."""
        raw = self._col_values(col, blanks)
        if not raw:
            self._no_data()
            return
        numeric = []
        for v in raw:
            try:
                numeric.append(float(v))
            except ValueError:
                pass
        ax = self._figure.add_subplot(111)
        if numeric:
            ax.hist(numeric, bins="auto", color="#8A244B",
                    edgecolor="#FFFFFF", linewidth=0.6)
            ax.set_xlabel(col, fontsize=10)
            ax.set_ylabel("Frequency", fontsize=10)
        else:
            # Non-numeric – fall back to vertical bar
            from collections import Counter
            counts = dict(Counter(raw).most_common(25))
            labels = list(counts.keys())
            values = list(counts.values())
            colors = self._colors(len(labels))
            ax.bar(range(len(labels)), values, color=colors,
                   edgecolor="#FFFFFF", linewidth=0.5, width=0.65)
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, rotation=30, ha="right", fontsize=9)
            ax.set_ylabel("Count", fontsize=10)
            ax.tick_params(axis="x", length=0)
        ax.set_title(f"{col} Histogram", fontsize=13, fontweight="bold", pad=14)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    def _draw_stacked(
        self, x_col: str, seg_col: str, blanks: bool,
        horizontal: bool = False, grouped: bool = False,
        max_x: int = 18,
    ) -> None:
        xi = self._headers.index(x_col)  if x_col  in self._headers else -1
        si = self._headers.index(seg_col) if seg_col in self._headers else -1
        if xi < 0 or si < 0 or not self._data:
            self._no_data()
            return

        matrix: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for row in self._data:
            xv = row[xi].strip() if xi < len(row) else ""
            sv = row[si].strip() if si < len(row) else ""
            if not blanks and (not xv or not sv):
                continue
            xv = xv or "(blank)"
            sv = sv or "(blank)"
            matrix[xv][sv] += 1

        if not matrix:
            self._no_data()
            return

        x_labels = sorted(matrix.keys(), key=lambda k: -sum(matrix[k].values()))[:max_x]
        all_segs = sorted({s for m in matrix.values() for s in m})
        colors = self._colors(len(all_segs))
        ax = self._figure.add_subplot(111)

        n_segs = len(all_segs)
        n_x    = len(x_labels)
        bar_w  = 0.65 if not grouped else max(0.12, 0.65 / max(n_segs, 1))

        if grouped:
            offsets = [(i - (n_segs - 1) / 2) * bar_w for i in range(n_segs)]
            for (seg, color), offset in zip(zip(all_segs, colors), offsets):
                vals = [matrix[x].get(seg, 0) for x in x_labels]
                if horizontal:
                    ax.barh([p + offset for p in range(n_x)], vals,
                            color=color, edgecolor="#FFFFFF", linewidth=0.4,
                            label=seg, height=bar_w)
                else:
                    ax.bar([p + offset for p in range(n_x)], vals,
                           color=color, edgecolor="#FFFFFF", linewidth=0.4,
                           label=seg, width=bar_w)
        else:
            bottoms = [0] * n_x
            for seg, color in zip(all_segs, colors):
                vals = [matrix[x].get(seg, 0) for x in x_labels]
                if horizontal:
                    ax.barh(range(n_x), vals, left=bottoms,
                            color=color, edgecolor="#FFFFFF", linewidth=0.4,
                            label=seg, height=0.68)
                else:
                    ax.bar(range(n_x), vals, bottom=bottoms,
                           color=color, edgecolor="#FFFFFF", linewidth=0.4,
                           label=seg, width=0.68)
                bottoms = [b + v for b, v in zip(bottoms, vals)]

        if horizontal:
            ax.set_yticks(range(n_x))
            ax.set_yticklabels(x_labels, fontsize=9)
            ax.set_xlabel("Count", fontsize=10)
            ax.tick_params(axis="y", length=0)
        else:
            ax.set_xticks(range(n_x))
            ax.set_xticklabels(x_labels, rotation=30, ha="right", fontsize=9)
            ax.set_ylabel("Count", fontsize=10)
            ax.tick_params(axis="x", length=0)

        ax.legend(loc="upper right", fontsize=9,
                  framealpha=0.95, edgecolor="#CCCCCC")
        ax.set_title(f"{seg_col} by {x_col}",
                     fontsize=13, fontweight="bold", pad=14)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    def _on_save_chart(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Chart", "",
            "PNG Image (*.png);;PDF Document (*.pdf);;SVG Image (*.svg)",
        )
        if path:
            self._figure.savefig(
                path, dpi=450, bbox_inches="tight", facecolor="white"
            )


class _MapperOption1TableDialog(QDialog):
    def __init__(self, parent: QWidget | None = None, start_dir: str = "", export_dir: str = "") -> None:
        super().__init__(parent)
        self._browse_start_dir = start_dir
        self._export_dir = export_dir
        self.setWindowTitle("Open window")
        self.resize(1160, 620)
        self.setMinimumSize(1080, 520)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        title = QLabel("SAP Information")
        title.setObjectName("packshotTableTitle")
        layout.addWidget(title)

        config_row = QHBoxLayout()
        config_row.setSpacing(8)
        row_count_label = QLabel("Row count")
        row_count_label.setObjectName("packshotRowCountLabel")
        config_row.addWidget(row_count_label, 0)

        self.row_count_input = QLineEdit("5")
        self.row_count_input.setObjectName("packshotRowCountInput")
        self.row_count_input.setFixedWidth(90)
        config_row.addWidget(self.row_count_input, 0)

        self.btn_update_rows = QPushButton("Update")
        self.btn_update_rows.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_update_rows, 0)

        self.btn_reset_table = QPushButton("Reset")
        self.btn_reset_table.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_reset_table, 0)

        self.btn_delete_row = QPushButton("Delete")
        self.btn_delete_row.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_delete_row, 0)

        self.btn_undo_table = QPushButton("Undo")
        self.btn_undo_table.setObjectName("packshotUpdateRowsBtn")
        config_row.addWidget(self.btn_undo_table, 0)

        self.cleanup_radio_group = QButtonGroup(self)
        self.cleanup_radio_group.setExclusive(True)

        self.radio_cleanup_1 = QRadioButton("Cleanup 1")
        self.radio_cleanup_1.setChecked(True)
        self.radio_cleanup_1.setToolTip(
            "remove non-SMU and unneeded packaging:\n"
            "sal, flex, pall, accl, film, ship, wgl,\n"
            "sheet, shee, pl, t-secur, saco, acco_pe,\n"
            "tear, bulk"
        )
        self.cleanup_radio_group.addButton(self.radio_cleanup_1)
        config_row.addWidget(self.radio_cleanup_1, 0)

        self.radio_cleanup_2 = QRadioButton("Cleanup 2")
        self.radio_cleanup_2.setToolTip(
            "remove non-SMU, un-identifiable basic name\n"
            "and unneeded packaging:\n"
            "sal, flex, pall, accl, film, ship, wgl,\n"
            "sheet, shee, pl, t-secur, saco, bag,\n"
            "rbosac, leaflet, acco, paco, tear, bulk"
        )
        self.cleanup_radio_group.addButton(self.radio_cleanup_2)
        config_row.addWidget(self.radio_cleanup_2, 0)

        self.radio_cleanup_3 = QRadioButton("Cleanup 3")
        self.radio_cleanup_3.setToolTip("no cleanup, all info retained")
        self.cleanup_radio_group.addButton(self.radio_cleanup_3)
        config_row.addWidget(self.radio_cleanup_3, 0)

        self.btn_import_tracker = QPushButton("Import SAP Data")
        self.btn_import_tracker.setObjectName("packshotUpdateRowsBtn")

        self.import_progress = QProgressBar(self)
        self.import_progress.setObjectName("sapImportProgress")
        self.import_progress.setFixedWidth(140)
        self.import_progress.setRange(0, 100)
        self.import_progress.setValue(0)
        self.import_progress.setVisible(False)
        config_row.addWidget(self.import_progress, 0)

        config_row.addWidget(self.btn_import_tracker, 0)

        self.btn_reformat_table = QPushButton("Reformat")
        self.btn_reformat_table.setObjectName("mapperReformatBtn")
        config_row.addWidget(self.btn_reformat_table, 0)

        config_row.addStretch(1)
        layout.addLayout(config_row)

        self.table = _ClipboardTableWidget(5, 6, self)
        self.table.setObjectName("packshotClipboardTable")
        self._set_headers()
        self.table.verticalHeader().setVisible(True)
        self.table.verticalHeader().setDefaultSectionSize(34)
        self.table.horizontalHeader().setDefaultSectionSize(154)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        for row in range(5):
            for col in range(6):
                self.table.setItem(row, col, QTableWidgetItem(""))

        self.sap_table_reformatter = SapTableReformatter()
        self._last_imported_stem: str = ""
        self._undo_stack: list[list[list[str]]] = []
        self._is_restoring_undo = False
        self._is_batch_edit = False

        layout.addWidget(self.table)
        self.btn_update_rows.clicked.connect(self._on_update_rows_clicked)
        self.btn_reset_table.clicked.connect(self._on_reset_table_clicked)
        self.btn_delete_row.clicked.connect(self._on_delete_rows_clicked)
        self.btn_undo_table.clicked.connect(self._on_undo_table_clicked)
        self.btn_import_tracker.clicked.connect(self._on_import_tracker_clicked)
        self.btn_reformat_table.clicked.connect(self._on_reformat_table_clicked)
        self.table.itemChanged.connect(self._on_table_item_changed)
        self._push_undo_snapshot(force=True)
        self._update_row_count_label()

        self.setStyleSheet(
            """
            QDialog {
                background-color: #F4F4F4;
            }

            QTableWidget#packshotClipboardTable {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                gridline-color: #B8B8B8;
                selection-background-color: #DCE6F5;
                selection-color: #111111;
                font-family: "Segoe UI";
                font-size: 12px;
            }

            QTableWidget#packshotClipboardTable QHeaderView::section {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #7D8694;
                padding: 6px 8px;
                font-weight: 700;
            }

            QTableWidget#packshotClipboardTable QTableCornerButton::section {
                background-color: #111F35;
                border: 1px solid #7D8694;
            }

            QLabel#packshotTableTitle {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 18px;
                font-weight: 800;
            }

            QLabel#packshotRowCountLabel {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 13px;
                font-weight: 600;
            }

            QLineEdit#packshotRowCountInput {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 8px;
                font-family: "Segoe UI";
                font-size: 12px;
            }

            QProgressBar#sapImportProgress {
                background-color: #FFFFFF;
                border: 1px solid #A9A9A9;
                border-radius: 7px;
                text-align: center;
                color: #111111;
                font-family: "Segoe UI";
                font-size: 11px;
                min-height: 24px;
            }

            QProgressBar#sapImportProgress::chunk {
                background-color: #8A244B;
                border-radius: 6px;
            }

            QPushButton#packshotUpdateRowsBtn {
                background-color: #9EA3AB;
                color: #000000;
                border: 1px solid #8B9098;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }

            QPushButton#packshotUpdateRowsBtn:pressed {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #111F35;
            }

            QPushButton#mapperReformatBtn {
                background-color: #8A244B;
                color: #FFFFFF;
                border: 1px solid #8A244B;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }

            QPushButton#mapperReformatBtn:pressed {
                background-color: #D02752;
                color: #FFFFFF;
                border: 1px solid #D02752;
            }

            QRadioButton {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }

            QRadioButton::indicator {
                width: 14px;
                height: 14px;
            }

            QRadioButton::indicator:unchecked {
                border: 1px solid #8B9098;
                border-radius: 7px;
                background-color: #FFFFFF;
            }

            QRadioButton::indicator:checked {
                border: 1px solid #8A244B;
                border-radius: 7px;
                background-color: #8A244B;
            }
            """
        )

    def _set_headers(self) -> None:
        self.table.setHorizontalHeaderLabels(
            [
                "Head Bom Mat",
                "BOM COMPONENT",
                "Sort String",
                "Component Desc",
                "Basic Number",
                "Basic Name",
            ]
        )

    def _update_row_count_label(self) -> None:
        if hasattr(self, "label_table_count"):
            self.label_table_count.setText(f"Count: <b>{self.table.rowCount()}</b>")

    def _set_import_progress(self, value: int, visible: bool = True) -> None:
        self.import_progress.setVisible(visible)
        if visible:
            self.import_progress.setValue(max(0, min(100, value)))
        QApplication.processEvents()

    @staticmethod
    def _normalize_header_text(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip().lower()
        text = text.replace("\n", " ").replace("\r", " ")
        text = re.sub(r"[^a-z0-9]+", " ", text)
        return " ".join(text.split())

    @staticmethod
    def _cell_to_text(value: object) -> str:
        if pd.isna(value):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _read_excel_raw(self, file_path: str) -> pd.DataFrame:
        lower_path = file_path.lower()

        if lower_path.endswith(".xls"):
            # Some SAP exports are text-based files with .xls extension.
            # Try as real Excel first, then fallback to auto text->xlsx conversion.
            try:
                excel_data = pd.ExcelFile(file_path, engine="openpyxl")
                if excel_data.sheet_names:
                    return pd.read_excel(
                        excel_data,
                        sheet_name=excel_data.sheet_names[0],
                        header=None,
                        dtype=object,
                    )
            except Exception:
                pass

            try:
                excel_data = pd.ExcelFile(file_path, engine="xlrd")
            except ImportError as exc:
                excel_data = None
            except Exception as exc:
                excel_data = None

            if excel_data is not None:
                if not excel_data.sheet_names:
                    raise ValueError("No worksheet found in selected file.")
                return pd.read_excel(
                    excel_data,
                    sheet_name=excel_data.sheet_names[0],
                    header=None,
                    dtype=object,
                )

            text_df = self._read_text_style_xls(file_path)
            if text_df is None:
                raise ValueError(
                    "Unable to read .xls file. File may be corrupted or in unsupported SAP export format."
                )

            tmp_xlsx_path = self._convert_dataframe_to_temp_xlsx(text_df)
            try:
                excel_data = pd.ExcelFile(tmp_xlsx_path, engine="openpyxl")
                if not excel_data.sheet_names:
                    raise ValueError("No worksheet found in converted SAP file.")
                return pd.read_excel(
                    excel_data,
                    sheet_name=excel_data.sheet_names[0],
                    header=None,
                    dtype=object,
                )
            finally:
                try:
                    Path(tmp_xlsx_path).unlink(missing_ok=True)
                except Exception:
                    pass
        else:
            last_error: Exception | None = None
            excel_data = None
            for engine in ("openpyxl", None):
                try:
                    excel_data = pd.ExcelFile(file_path, engine=engine)
                    break
                except Exception as exc:
                    last_error = exc

            if excel_data is None:
                if last_error is not None:
                    raise last_error
                raise ValueError("Unable to read selected file.")

        if not excel_data.sheet_names:
            raise ValueError("No worksheet found in selected file.")

        return pd.read_excel(
            excel_data,
            sheet_name=excel_data.sheet_names[0],
            header=None,
            dtype=object,
        )

    def _read_text_style_xls(self, file_path: str) -> pd.DataFrame | None:
        parsed = self._try_read_sap_text_file(file_path)
        if parsed is not None:
            return parsed

        # Fallback: internally perform the manual trick (rename .xls -> .csv) and parse again.
        tmp_csv = tempfile.NamedTemporaryFile(suffix=".csv", delete=False)
        tmp_csv_path = tmp_csv.name
        tmp_csv.close()

        try:
            Path(tmp_csv_path).write_bytes(Path(file_path).read_bytes())
            return self._try_read_sap_text_file(tmp_csv_path)
        except Exception:
            return None
        finally:
            try:
                Path(tmp_csv_path).unlink(missing_ok=True)
            except Exception:
                pass

    def _try_read_sap_text_file(self, file_path: str) -> pd.DataFrame | None:
        parse_configs = [
            {"encoding": "utf-16", "sep": "\t", "skiprows": 3},
            {"encoding": "utf-16", "sep": r"\s{2,}", "skiprows": 3},
            {"encoding": "utf-16le", "sep": "\t", "skiprows": 3},
            {"encoding": "utf-16le", "sep": r"\s{2,}", "skiprows": 3},
            {"encoding": "utf-8-sig", "sep": "\t", "skiprows": 3},
            {"encoding": "cp1252", "sep": "\t", "skiprows": 3},
            {"encoding": "utf-16", "sep": "\t", "skiprows": 0},
            {"encoding": "utf-16", "sep": r"\s{2,}", "skiprows": 0},
        ]

        for config in parse_configs:
            try:
                df = pd.read_csv(
                    file_path,
                    header=None,
                    dtype=object,
                    engine="python",
                    on_bad_lines="skip",
                    quotechar='"',
                    quoting=csv.QUOTE_MINIMAL,
                    **config,
                )
            except Exception:
                continue

            if df.empty:
                continue

            df = df.dropna(how="all").reset_index(drop=True)
            if df.empty:
                continue

            # Must be reasonably tabular for SAP header detection to work.
            if df.shape[1] < 4:
                continue

            return df

        return None

    @staticmethod
    def _convert_dataframe_to_temp_xlsx(df: pd.DataFrame) -> str:
        tmp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp_path = tmp_file.name
        tmp_file.close()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "SAP Data"

        safe_df = df.fillna("")
        for row_values in safe_df.itertuples(index=False, name=None):
            worksheet.append(list(row_values))

        workbook.save(tmp_path)
        return tmp_path

    def _detect_header_row_and_columns(self, raw_df: pd.DataFrame) -> tuple[int, dict[str, int]]:
        target_columns: dict[str, list[str]] = {
            "Head Bom Mat": ["head bom mat", "head bom", "headbommat"],
            "BOM COMPONENT": ["bom component", "bom comp", "bomcomponent"],
            "Sort String": ["sort string", "sortstring"],
            "Component Desc": ["component desc", "component description", "component"],
            "Basic Number": ["basic number", "basic num", "basic no", "basic"],
            "Basic Name": ["basic name", "basicname"],
        }
        normalized_targets = {
            target: {self._normalize_header_text(alias) for alias in aliases}
            for target, aliases in target_columns.items()
        }

        max_scan_rows = min(120, raw_df.shape[0])
        for row_idx in range(max_scan_rows):
            row_values = raw_df.iloc[row_idx].tolist()
            row_normalized = [self._normalize_header_text(value) for value in row_values]

            found_columns: dict[str, int] = {}
            for target_name, aliases in normalized_targets.items():
                for col_idx, cell_value in enumerate(row_normalized):
                    if cell_value in aliases:
                        found_columns[target_name] = col_idx
                        break

            if len(found_columns) == len(target_columns):
                return row_idx, found_columns

        missing = ", ".join(target_columns.keys())
        raise ValueError(f"Unable to detect all required headers. Required headers: {missing}")

    def _extract_sap_rows(self, raw_df: pd.DataFrame, header_row_idx: int, col_map: dict[str, int]) -> list[list[str]]:
        ordered_headers = [
            "Head Bom Mat",
            "BOM COMPONENT",
            "Sort String",
            "Component Desc",
            "Basic Number",
            "Basic Name",
        ]
        out_rows: list[list[str]] = []

        for row_idx in range(header_row_idx + 1, raw_df.shape[0]):
            row_values: list[str] = []
            has_value = False
            for header in ordered_headers:
                col_idx = col_map[header]
                value = raw_df.iat[row_idx, col_idx] if col_idx < raw_df.shape[1] else ""
                text_value = self._cell_to_text(value)
                if text_value:
                    has_value = True
                row_values.append(text_value)

            if has_value:
                out_rows.append(row_values)

        return out_rows

    def _on_import_tracker_clicked(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select SAP File",
            self._browse_start_dir,
            "Excel Files (*.xls)",
        )
        if not file_path:
            return

        self._last_imported_stem = Path(file_path).stem

        self._set_import_progress(5, visible=True)
        try:
            raw_df = self._read_excel_raw(file_path)
            self._set_import_progress(35, visible=True)
            header_row_idx, col_map = self._detect_header_row_and_columns(raw_df)
            self._set_import_progress(60, visible=True)
            extracted_rows = self._extract_sap_rows(raw_df, header_row_idx, col_map)
            self._set_import_progress(80, visible=True)
        except Exception as exc:
            self._set_import_progress(0, visible=False)
            msg = QMessageBox(self)
            msg.setWindowTitle("Error")
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setText(f"Failed to import SAP data. {exc}")
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg.exec()
            return

        self.table.blockSignals(True)
        try:
            row_count = max(1, len(extracted_rows))
            self.table.setRowCount(row_count)
            for row_idx in range(row_count):
                values = extracted_rows[row_idx] if row_idx < len(extracted_rows) else [""] * self.table.columnCount()
                for col_idx in range(self.table.columnCount()):
                    text_value = values[col_idx] if col_idx < len(values) else ""
                    item = self.table.item(row_idx, col_idx)
                    if item is None:
                        item = QTableWidgetItem("")
                        self.table.setItem(row_idx, col_idx, item)
                    item.setText(text_value)
        finally:
            self.table.blockSignals(False)

        self.table.clearSelection()
        self.row_count_input.setText(str(self.table.rowCount()))
        self._update_row_count_label()
        self._push_undo_snapshot(force=True)
        self._set_import_progress(100, visible=True)
        self.import_progress.setVisible(False)

    def _on_update_rows_clicked(self) -> None:
        raw_value = self.row_count_input.text().strip()
        try:
            new_count = int(raw_value)
        except ValueError:
            self.row_count_input.setText(str(self.table.rowCount()))
            return

        new_count = max(1, min(new_count, 5000))
        self.table.setRowCount(new_count)
        for row in range(new_count):
            for col in range(self.table.columnCount()):
                if self.table.item(row, col) is None:
                    self.table.setItem(row, col, QTableWidgetItem(""))
        self.row_count_input.setText(str(new_count))
        self._update_row_count_label()
        self._push_undo_snapshot()

    def _on_reset_table_clicked(self) -> None:
        self.row_count_input.setText("5")
        self.table.clearContents()
        self.table.setRowCount(5)
        self.table.clearSelection()
        for row in range(5):
            for col in range(self.table.columnCount()):
                self.table.setItem(row, col, QTableWidgetItem(""))
        self._update_row_count_label()
        self._push_undo_snapshot()

    def _on_delete_rows_clicked(self) -> None:
        selected_rows = sorted({index.row() for index in self.table.selectedIndexes()}, reverse=True)
        if not selected_rows:
            current_row = self.table.currentRow()
            if current_row >= 0:
                selected_rows = [current_row]
            else:
                return

        total_rows = self.table.rowCount()
        if total_rows <= 0:
            return

        if len(selected_rows) >= total_rows:
            self.table.setRowCount(1)
            for col in range(self.table.columnCount()):
                self.table.setItem(0, col, QTableWidgetItem(""))
        else:
            for row in selected_rows:
                if 0 <= row < self.table.rowCount():
                    self.table.removeRow(row)

        self.table.clearSelection()
        self.row_count_input.setText(str(self.table.rowCount()))
        self._update_row_count_label()
        self._push_undo_snapshot()

    def _capture_table_state(self) -> list[list[str]]:
        state: list[list[str]] = []
        for row in range(self.table.rowCount()):
            row_values: list[str] = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_values.append(item.text() if item is not None else "")
            state.append(row_values)
        return state

    def _restore_table_state(self, state: list[list[str]]) -> None:
        self._is_restoring_undo = True
        self.table.blockSignals(True)
        try:
            row_count = max(1, len(state))
            col_count = self.table.columnCount()
            self.table.setRowCount(row_count)

            for row in range(row_count):
                row_values = state[row] if row < len(state) else [""] * col_count
                for col in range(col_count):
                    value = row_values[col] if col < len(row_values) else ""
                    item = self.table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem("")
                        self.table.setItem(row, col, item)
                    item.setText(value)
        finally:
            self.table.blockSignals(False)
            self._is_restoring_undo = False

        self.row_count_input.setText(str(self.table.rowCount()))
        self._update_row_count_label()

    def _push_undo_snapshot(self, force: bool = False) -> None:
        if self._is_restoring_undo or self._is_batch_edit:
            return

        snapshot = self._capture_table_state()
        if not force and self._undo_stack and self._undo_stack[-1] == snapshot:
            return

        self._undo_stack.append(snapshot)
        if len(self._undo_stack) > 200:
            self._undo_stack.pop(0)

    def _on_table_item_changed(self, _item: QTableWidgetItem) -> None:
        if self._is_batch_edit:
            return
        self._push_undo_snapshot()

    def _on_table_batch_edit_begin(self) -> None:
        if self._is_restoring_undo:
            return
        self._push_undo_snapshot(force=False)
        self._is_batch_edit = True

    def _on_table_batch_edit_end(self) -> None:
        if self._is_restoring_undo:
            return
        self._is_batch_edit = False
        self._push_undo_snapshot(force=False)

    def _on_undo_table_clicked(self) -> None:
        if self._is_restoring_undo:
            return

        current_snapshot = self._capture_table_state()
        if not self._undo_stack:
            self._undo_stack.append(current_snapshot)
            return

        if self._undo_stack[-1] != current_snapshot:
            self._undo_stack.append(current_snapshot)

        if len(self._undo_stack) <= 1:
            return

        self._undo_stack.pop()
        previous_snapshot = self._undo_stack[-1]
        self._restore_table_state(previous_snapshot)

    def _on_reformat_table_clicked(self) -> None:
        cleanup_mode = 1
        if self.radio_cleanup_2.isChecked():
            cleanup_mode = 2
        elif self.radio_cleanup_3.isChecked():
            cleanup_mode = 3

        try:
            reformatted_rows, basic_comb_count = self.sap_table_reformatter.reformat_from_table(self.table, cleanup_mode)
        except SapTableReformatError as exc:
            msg = QMessageBox(self)
            msg.setWindowTitle("Error")
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setText(str(exc))
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg.exec()
            return

        self._reformatted_dialog = _MapperReformattedTableDialog(
            reformatted_rows,
            basic_comb_count,
            cleanup_mode,
            self,
            source_stem=self._last_imported_stem,
            export_dir=self._export_dir,
        )
        self._reformatted_dialog.show()
        self._reformatted_dialog.raise_()
        self._reformatted_dialog.activateWindow()


class _MapperReformattedTableDialog(QDialog):
    _BLANK_FILTER = "__BLANK__"

    class _FilterPopup(QDialog):
        def __init__(
            self,
            values: list[tuple[str, str]],
            selected_values: set[str] | None,
            parent: QWidget | None = None,
        ) -> None:
            super().__init__(parent)
            self.setWindowFlags(Qt.WindowType.Popup)
            self.setMinimumSize(260, 330)
            self.resize(260, 330)

            self._is_syncing = False
            self._value_items: list[tuple[QListWidgetItem, str, str]] = []

            layout = QVBoxLayout(self)
            layout.setContentsMargins(4, 4, 4, 4)
            layout.setSpacing(4)

            self.search_input = QLineEdit(self)
            self.search_input.setPlaceholderText("Search")
            self.search_input.setClearButtonEnabled(True)
            self.search_input.setObjectName("packshotRowCountInput")
            layout.addWidget(self.search_input)

            self.list_widget = QListWidget(self)
            self.list_widget.setSelectionMode(QListWidget.SelectionMode.NoSelection)
            layout.addWidget(self.list_widget, 1)

            self.item_select_all = QListWidgetItem("(Select All)")
            self.item_select_all.setFlags(self.item_select_all.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            self.item_select_all.setCheckState(Qt.CheckState.Checked)
            self.item_select_all.setData(Qt.ItemDataRole.UserRole, "__SELECT_ALL__")
            self.list_widget.addItem(self.item_select_all)

            selected = selected_values if selected_values is not None else {value for value, _label in values}

            for value, label in values:
                item = QListWidgetItem(label)
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                item.setCheckState(Qt.CheckState.Checked if value in selected else Qt.CheckState.Unchecked)
                item.setData(Qt.ItemDataRole.UserRole, value)
                self.list_widget.addItem(item)
                self._value_items.append((item, value, label))

            button_row = QHBoxLayout()
            button_row.addStretch(1)
            self.btn_apply = QPushButton("Apply")
            self.btn_apply.setObjectName("packshotUpdateRowsBtn")
            self.btn_cancel = QPushButton("Cancel")
            self.btn_cancel.setObjectName("packshotUpdateRowsBtn")
            button_row.addWidget(self.btn_apply, 0)
            button_row.addWidget(self.btn_cancel, 0)
            layout.addLayout(button_row)

            self.search_input.textChanged.connect(self._apply_search)
            self.list_widget.itemChanged.connect(self._on_item_changed)
            self.btn_apply.clicked.connect(self.accept)
            self.btn_cancel.clicked.connect(self.reject)

            self.search_input.setFocus()
            self._sync_select_all_state()

        def _apply_search(self, text: str) -> None:
            needle = text.strip().lower()
            for item, _value, label in self._value_items:
                item.setHidden(needle not in label.lower())
            self._sync_select_all_state()

        def _on_item_changed(self, item: QListWidgetItem) -> None:
            if self._is_syncing:
                return

            role = item.data(Qt.ItemDataRole.UserRole)
            if role == "__SELECT_ALL__":
                self._is_syncing = True
                try:
                    target_state = item.checkState()
                    for value_item, _value, _label in self._value_items:
                        if value_item.isHidden():
                            continue
                        value_item.setCheckState(target_state)
                finally:
                    self._is_syncing = False
                return

            self._sync_select_all_state()

        def _sync_select_all_state(self) -> None:
            visible_items = [item for item, _value, _label in self._value_items if not item.isHidden()]
            if not visible_items:
                state = Qt.CheckState.Unchecked
            else:
                all_checked = all(i.checkState() == Qt.CheckState.Checked for i in visible_items)
                any_checked = any(i.checkState() == Qt.CheckState.Checked for i in visible_items)
                if all_checked:
                    state = Qt.CheckState.Checked
                elif any_checked:
                    state = Qt.CheckState.PartiallyChecked
                else:
                    state = Qt.CheckState.Unchecked

            self._is_syncing = True
            try:
                self.item_select_all.setCheckState(state)
            finally:
                self._is_syncing = False

        def get_selected_values(self) -> set[str]:
            selected: set[str] = set()
            for item, value, _label in self._value_items:
                if item.checkState() == Qt.CheckState.Checked:
                    selected.add(value)
            return selected

    def __init__(
        self,
        rows: list[list[str]],
        basic_comb_count: int,
        cleanup_mode: int,
        parent: QWidget | None = None,
        source_stem: str = "",
        export_dir: str = "",
    ) -> None:
        super().__init__(parent)
        self._cleanup_mode = cleanup_mode if cleanup_mode in (1, 2, 3) else 1
        self._source_stem = source_stem
        self._export_dir = export_dir
        self.setWindowTitle("Reformatted SAP Data")
        self.resize(1120, 620)
        self.setMinimumSize(1020, 520)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        header_row = QHBoxLayout()
        header_row.setSpacing(8)

        title = QLabel("Reformatted SAP Data")
        title.setObjectName("mapperReformattedTitle")
        header_row.addWidget(title, 0)
        header_row.addSpacing(16)

        self.btn_reset_filter = QPushButton("Reset Filter")
        self.btn_reset_filter.setObjectName("packshotUpdateRowsBtn")
        header_row.addWidget(self.btn_reset_filter, 0)
        header_row.addSpacing(8)

        self.btn_export_table = QPushButton("Export")
        self.btn_export_table.setObjectName("packshotUpdateRowsBtn")
        header_row.addWidget(self.btn_export_table, 0)
        header_row.addSpacing(6)

        self.export_progress = QProgressBar(self)
        self.export_progress.setObjectName("mapperExportProgress")
        self.export_progress.setFixedWidth(140)
        self.export_progress.setRange(0, 100)
        self.export_progress.setValue(0)
        self.export_progress.setVisible(False)
        header_row.addWidget(self.export_progress, 0)
        header_row.addSpacing(8)

        self.btn_grouping_count = QPushButton("Grouping Count")
        self.btn_grouping_count.setObjectName("packshotUpdateRowsBtn")
        header_row.addWidget(self.btn_grouping_count, 0)

        header_row.addSpacing(14)

        self._selected_column_for_count = 0
        self.label_selection_count = QLabel("Count: <b>0</b>")
        self.label_selection_count.setObjectName("mapperReformattedOverview")
        self.label_selection_count.setTextFormat(Qt.TextFormat.RichText)
        header_row.addWidget(self.label_selection_count, 0)
        header_row.addSpacing(20)

        self.label_hsi_count = QLabel("HSI Count: <b>0</b>")
        self.label_hsi_count.setObjectName("mapperReformattedOverview")
        self.label_hsi_count.setTextFormat(Qt.TextFormat.RichText)
        self.label_hsi_count.setVisible(self._cleanup_mode != 3)
        header_row.addWidget(self.label_hsi_count, 0)
        header_row.addSpacing(20)

        self.label_basic_comb_overview = QLabel(
            f"Total Count of Basic Combinations: <b>{basic_comb_count}</b>"
        )
        self.label_basic_comb_overview.setObjectName("mapperReformattedOverview")
        self.label_basic_comb_overview.setTextFormat(Qt.TextFormat.RichText)
        header_row.addWidget(self.label_basic_comb_overview, 0)
        header_row.addSpacing(24)

        _cu_tag = f"cu{self._cleanup_mode}"
        self.label_cleanup_tag = QLabel(_cu_tag)
        self.label_cleanup_tag.setObjectName("mapperCleanupTag")
        header_row.addWidget(self.label_cleanup_tag, 0)
        header_row.addStretch(1)
        layout.addLayout(header_row)

        self.table = _ClipboardTableWidget(max(len(rows), 1), 7, self)
        self.table.setObjectName("mapperReformattedTable")
        self._header_labels = [
            "Head Bom Mat",
            "HSI",
            "BOM COMPONENT",
            "Component Desc",
            "Basic Number",
            "BC",
            "Basic Name",
        ]
        self._column_filters: dict[int, set[str] | None] = {i: None for i in range(len(self._header_labels))}
        self._refresh_header_labels()
        self.table.verticalHeader().setVisible(True)
        self.table.verticalHeader().setDefaultSectionSize(34)
        self.table.horizontalHeader().setDefaultSectionSize(200)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(0, 100)  # Head Bom Mat
        self.table.setColumnWidth(1, 100)  # HSI
        self.table.setColumnWidth(5, 100)  # BC
        self.table.horizontalHeader().setSectionsClickable(True)
        self.table.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        self.table.horizontalHeader().sectionDoubleClicked.connect(self._on_filter_header_clicked)

        for row_idx, values in enumerate(rows):
            for col_idx, value in enumerate(values):
                self.table.setItem(row_idx, col_idx, QTableWidgetItem(value))

        if not rows:
            for col_idx in range(7):
                self.table.setItem(0, col_idx, QTableWidgetItem(""))

        layout.addWidget(self.table)
        self.table.selectColumn(0)
        self._update_selection_count_label()
        self._update_hsi_count_label()

        self.btn_export_table.clicked.connect(self._on_export_table_clicked)
        self.btn_reset_filter.clicked.connect(self._on_reset_filter_clicked)
        self.btn_grouping_count.clicked.connect(self._on_grouping_count_clicked)

        self.setStyleSheet(
            """
            QDialog {
                background-color: #F4F4F4;
            }

            QLabel#mapperReformattedTitle {
                color: #8A244B;
                font-family: "Segoe UI";
                font-size: 18px;
                font-weight: 800;
            }

            QLabel#mapperReformattedOverview {
                color: #111F35;
                font-family: "Segoe UI";
                font-size: 13px;
                font-weight: 600;
            }

            QLabel#mapperCleanupTag {
                color: #D02752;
                font-family: "Segoe UI";
                font-size: 13px;
                font-weight: 700;
            }

            QTableWidget#mapperReformattedTable {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                gridline-color: #B8B8B8;
                selection-background-color: #DCE6F5;
                selection-color: #111111;
                font-family: "Segoe UI";
                font-size: 12px;
            }

            QTableWidget#mapperReformattedTable QHeaderView::section {
                background-color: #8A244B;
                color: #FFFFFF;
                border: 1px solid #7D8694;
                padding: 6px 8px;
                font-weight: 700;
            }

            QTableWidget#mapperReformattedTable QTableCornerButton::section {
                background-color: #8A244B;
                border: 1px solid #7D8694;
            }

            QPushButton#packshotUpdateRowsBtn {
                background-color: #9EA3AB;
                color: #000000;
                border: 1px solid #8B9098;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }

            QPushButton#packshotUpdateRowsBtn:pressed {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #111F35;
            }

            QLineEdit#packshotRowCountInput {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 8px;
                font-family: "Segoe UI";
                font-size: 12px;
            }

            QProgressBar#mapperExportProgress {
                background-color: #FFFFFF;
                border: 1px solid #A9A9A9;
                border-radius: 7px;
                text-align: center;
                color: #111111;
                font-family: "Segoe UI";
                font-size: 11px;
                min-height: 24px;
            }

            QProgressBar#mapperExportProgress::chunk {
                background-color: #8A244B;
                border-radius: 6px;
            }
            """
        )

    def _refresh_header_labels(self) -> None:
        labels = []
        for idx, base_label in enumerate(self._header_labels):
            has_active_filter = self._column_filters.get(idx) is not None
            labels.append(f"{base_label} {'▾*' if has_active_filter else '▾'}")
        self.table.setHorizontalHeaderLabels(labels)

    def _on_header_clicked(self, column: int) -> None:
        self._selected_column_for_count = column
        self.table.selectColumn(column)
        self._update_selection_count_label()

    def _on_filter_header_clicked(self, column: int) -> None:
        values: list[str] = []
        seen: set[str] = set()
        has_blank = False
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            item = self.table.item(row, column)
            cell_text = item.text().strip() if item is not None else ""
            if cell_text == "":
                has_blank = True
                continue
            if cell_text in seen:
                continue
            seen.add(cell_text)
            values.append(cell_text)

        value_pairs: list[tuple[str, str]] = [(value, value) for value in values]
        if has_blank:
            value_pairs.append((self._BLANK_FILTER, "(Blanks)"))

        popup = self._FilterPopup(value_pairs, self._column_filters.get(column), self)
        popup.move(QCursor.pos())
        if popup.exec() != QDialog.DialogCode.Accepted:
            return

        selected = popup.get_selected_values()
        if len(selected) == len(value_pairs):
            self._column_filters[column] = None
        else:
            self._column_filters[column] = selected
        self._apply_filters()

    def _update_selection_count_label(self) -> None:
        count = 0
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            count += 1
        self.label_selection_count.setText(f"Count: <b>{count}</b>")

    def _update_hsi_count_label(self) -> None:
        yes_count = 0
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            item = self.table.item(row, 1)
            hsi_value = item.text().strip().upper() if item is not None else ""
            if hsi_value == "YES":
                yes_count += 1
        self.label_hsi_count.setText(f"HSI Count: <b>{yes_count}</b>")

    def _apply_filters(self) -> None:
        for row in range(self.table.rowCount()):
            row_matches = True
            for col, filter_value in self._column_filters.items():
                if filter_value is None:
                    continue
                item = self.table.item(row, col)
                cell_text = item.text().strip() if item is not None else ""
                normalized = self._BLANK_FILTER if cell_text == "" else cell_text
                if normalized not in filter_value:
                    row_matches = False
                    break
            self.table.setRowHidden(row, not row_matches)

        self._refresh_header_labels()
        self._update_selection_count_label()
        self._update_hsi_count_label()

    def _on_reset_filter_clicked(self) -> None:
        for column in self._column_filters:
            self._column_filters[column] = None
        self._apply_filters()

    def _on_export_table_clicked(self) -> None:
        import os as _os
        timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
        stem = self._source_stem if self._source_stem else "manual_entry"
        cu_tag = f"cu{self._cleanup_mode}"
        _filename = f"rsd_{stem}_{cu_tag}_{timestamp}.xlsx"
        _start_path = _os.path.join(self._export_dir, _filename) if self._export_dir else _filename
        output_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Reformatted SAP Data",
            _start_path,
            "Excel Files (*.xlsx)",
        )
        if not output_path:
            return

        if not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"

        header = [
            "Head Bom Mat",
            "HSI",
            "BOM COMPONENT",
            "Component Desc",
            "Basic Number",
            "BC",
            "Basic Name",
        ]

        self.export_progress.setVisible(True)
        self.export_progress.setValue(10)
        QApplication.processEvents()

        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Reformatted SAP Data"

            output_rows: list[list[str]] = []
            for row in range(self.table.rowCount()):
                if self.table.isRowHidden(row):
                    continue
                values: list[str] = []
                row_has_value = False
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    text = item.text().strip() if item is not None else ""
                    if text:
                        row_has_value = True
                    values.append(text)
                if row_has_value:
                    output_rows.append(values)

            self.export_progress.setValue(45)
            QApplication.processEvents()

            worksheet.append(header)
            for values in output_rows:
                worksheet.append(values)

            accent_fill = PatternFill(fill_type="solid", fgColor="8A244B")
            white_font = Font(color="FFFFFF", bold=True)
            normal_font = Font(color="111111")
            border = Border(
                left=Side(style="thin", color="B8B8B8"),
                right=Side(style="thin", color="B8B8B8"),
                top=Side(style="thin", color="B8B8B8"),
                bottom=Side(style="thin", color="B8B8B8"),
            )

            max_row = worksheet.max_row
            max_col = worksheet.max_column

            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = accent_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    data_cell = worksheet.cell(row=row, column=col)
                    data_cell.font = normal_font
                    data_cell.alignment = Alignment(horizontal="left", vertical="center")
                    data_cell.border = border

            self.export_progress.setValue(80)
            QApplication.processEvents()

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=max_col).coordinate}"

            for col_cells in worksheet.columns:
                col_letter = col_cells[0].column_letter
                max_len = 0
                for cell in col_cells:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_len:
                        max_len = len(value)
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 56)

            # ── Grouping Count sheet (second sheet) ──────────────────────
            gc_counts: dict[str, int] = {}
            gc_bn_map: dict[str, dict[str, int]] = {}
            gc_bname_map: dict[str, dict[str, int]] = {}
            for row_vals in output_rows:
                comb = row_vals[5].strip() if len(row_vals) > 5 else ""
                if not comb:
                    continue
                gc_counts[comb] = gc_counts.get(comb, 0) + 1
                bn = row_vals[4].strip() if len(row_vals) > 4 else ""
                gc_bn_map.setdefault(comb, {})
                gc_bn_map[comb][bn] = gc_bn_map[comb].get(bn, 0) + 1
                bname = row_vals[6].strip() if len(row_vals) > 6 else ""
                gc_bname_map.setdefault(comb, {})
                gc_bname_map[comb][bname] = gc_bname_map[comb].get(bname, 0) + 1

            gc_rows: list[tuple[str, int, str, str]] = []
            for comb, cnt in gc_counts.items():
                best_bn = max(gc_bn_map[comb].items(), key=lambda p: (p[1], p[0]))[0] if gc_bn_map.get(comb) else ""
                best_name = max(gc_bname_map[comb].items(), key=lambda p: (p[1], p[0]))[0] if gc_bname_map.get(comb) else ""
                gc_rows.append((comb, cnt, best_bn, best_name))

            def _gc_sort(r):
                m = re.search(r"(\d+)", r[0])
                return (-r[1], int(m.group(1)) if m else 10**9, r[0])

            gc_rows.sort(key=_gc_sort)

            grouping_ws = workbook.create_sheet("Grouping Count")
            grouping_ws.append(["BC", "Count", "Basic Number", "Basic Name"])
            for comb_name, count, bn, bname in gc_rows:
                grouping_ws.append([comb_name, count, bn, bname])

            gc_max_row = grouping_ws.max_row
            gc_max_col = grouping_ws.max_column
            for col in range(1, gc_max_col + 1):
                cell = grouping_ws.cell(row=1, column=col)
                cell.fill = accent_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            for row in range(2, gc_max_row + 1):
                for col in range(1, gc_max_col + 1):
                    c = grouping_ws.cell(row=row, column=col)
                    c.font = normal_font
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = border
            grouping_ws.freeze_panes = "A2"
            grouping_ws.auto_filter.ref = f"A1:{grouping_ws.cell(row=1, column=gc_max_col).coordinate}"
            for col_cells in grouping_ws.columns:
                letter = col_cells[0].column_letter
                mx = max((len(str(cell.value)) for cell in col_cells if cell.value is not None), default=0)
                grouping_ws.column_dimensions[letter].width = min(max(mx + 2, 12), 56)

            workbook.save(output_path)
            self.export_progress.setValue(100)
            QApplication.processEvents()
        except Exception as exc:
            self.export_progress.setVisible(False)
            msg = QMessageBox(self)
            msg.setWindowTitle("Error")
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setText(f"Failed to export file. {exc}")
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg.exec()
            return

        self.export_progress.setVisible(False)

        msg = QMessageBox(self)
        msg.setWindowTitle("Done")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setText("Reformatted SAP data exported successfully.")
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def _on_grouping_count_clicked(self) -> None:
        grouping_counts: dict[str, int] = {}
        grouping_basic_numbers: dict[str, dict[str, int]] = {}
        grouping_basic_names: dict[str, dict[str, int]] = {}
        combination_column = 5
        basic_number_column = 4
        basic_name_column = 6

        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            item = self.table.item(row, combination_column)
            combination = item.text().strip() if item is not None else ""
            if not combination:
                continue
            grouping_counts[combination] = grouping_counts.get(combination, 0) + 1

            basic_number_item = self.table.item(row, basic_number_column)
            basic_number = basic_number_item.text().strip() if basic_number_item is not None else ""
            if combination not in grouping_basic_numbers:
                grouping_basic_numbers[combination] = {}
            grouping_basic_numbers[combination][basic_number] = grouping_basic_numbers[combination].get(basic_number, 0) + 1

            basic_name_item = self.table.item(row, basic_name_column)
            basic_name = basic_name_item.text().strip() if basic_name_item is not None else ""
            if combination not in grouping_basic_names:
                grouping_basic_names[combination] = {}
            grouping_basic_names[combination][basic_name] = grouping_basic_names[combination].get(basic_name, 0) + 1

        grouped_rows: list[tuple[str, int, str, str]] = []
        for combination, count in grouping_counts.items():
            best_basic_number = ""
            if combination in grouping_basic_numbers and grouping_basic_numbers[combination]:
                best_basic_number = max(
                    grouping_basic_numbers[combination].items(),
                    key=lambda pair: (pair[1], pair[0]),
                )[0]

            best_basic_name = ""
            if combination in grouping_basic_names and grouping_basic_names[combination]:
                best_basic_name = max(
                    grouping_basic_names[combination].items(),
                    key=lambda pair: (pair[1], pair[0]),
                )[0]
            grouped_rows.append((combination, count, best_basic_number, best_basic_name))

        sorted_counts = sorted(
            grouped_rows,
            key=lambda row: (-row[1], self._combination_sort_key(row[0])),
        )

        self._grouping_count_dialog = _MapperGroupingCountDialog(sorted_counts, self)
        self._grouping_count_dialog.show()
        self._grouping_count_dialog.raise_()
        self._grouping_count_dialog.activateWindow()

    @staticmethod
    def _combination_sort_key(value: str) -> tuple[int, str]:
        # "comb 0" sorts first (numeric), then BC codes sort alphabetically.
        match = re.search(r"(\d+)", value)
        if match:
            return int(match.group(1)), value
        return 10**9, value


class _MapperGroupingCountDialog(QDialog):
    def __init__(self, grouping_counts: list[tuple[str, int, str, str]], parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Grouping Count")
        self.resize(420, 560)
        self.setMinimumSize(360, 460)

        self._grouping_counts = grouping_counts

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        header_row = QHBoxLayout()
        header_row.setSpacing(8)

        title = QLabel("Grouping Count")
        title.setObjectName("mapperReformattedTitle")
        header_row.addWidget(title, 0)

        self.btn_export = QPushButton("Export")
        self.btn_export.setObjectName("packshotUpdateRowsBtn")
        header_row.addWidget(self.btn_export, 0)

        header_row.addStretch(1)

        layout.addLayout(header_row)

        self.table = _ClipboardTableWidget(max(len(grouping_counts), 1), 4, self)
        self.table._paste_disabled = True
        self.table.setObjectName("mapperReformattedTable")
        self.table.setHorizontalHeaderLabels(["BC", "Count", "Basic Number", "Basic Name"])
        self.table.verticalHeader().setVisible(True)
        self.table.verticalHeader().setDefaultSectionSize(30)
        self.table.horizontalHeader().setDefaultSectionSize(140)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        if grouping_counts:
            for row_idx, (name, count, basic_number, basic_name) in enumerate(grouping_counts):
                self.table.setItem(row_idx, 0, QTableWidgetItem(name))
                self.table.setItem(row_idx, 1, QTableWidgetItem(str(count)))
                self.table.setItem(row_idx, 2, QTableWidgetItem(basic_number))
                self.table.setItem(row_idx, 3, QTableWidgetItem(basic_name))
        else:
            self.table.setItem(0, 0, QTableWidgetItem(""))
            self.table.setItem(0, 1, QTableWidgetItem("0"))
            self.table.setItem(0, 2, QTableWidgetItem(""))
            self.table.setItem(0, 3, QTableWidgetItem(""))

        layout.addWidget(self.table)

        self.btn_export.clicked.connect(self._on_export_clicked)

        self.setStyleSheet(
            """
            QDialog {
                background-color: #F4F4F4;
            }

            QLabel#mapperReformattedTitle {
                color: #8A244B;
                font-family: "Segoe UI";
                font-size: 16px;
                font-weight: 800;
            }

            QTableWidget#mapperReformattedTable {
                background-color: #FFFFFF;
                color: #111111;
                border: 1px solid #A9A9A9;
                gridline-color: #B8B8B8;
                selection-background-color: #DCE6F5;
                selection-color: #111111;
                font-family: "Segoe UI";
                font-size: 11px;
            }

            QTableWidget#mapperReformattedTable QHeaderView::section {
                background-color: #8A244B;
                color: #FFFFFF;
                border: 1px solid #7D8694;
                padding: 6px 8px;
                font-weight: 700;
            }

            QTableWidget#mapperReformattedTable QTableCornerButton::section {
                background-color: #8A244B;
                border: 1px solid #7D8694;
            }

            QPushButton#packshotUpdateRowsBtn {
                background-color: #9EA3AB;
                color: #000000;
                border: 1px solid #8B9098;
                border-radius: 8px;
                min-height: 30px;
                padding: 0 12px;
                font-family: "Segoe UI";
                font-size: 12px;
                font-weight: 600;
            }

            QPushButton#packshotUpdateRowsBtn:pressed {
                background-color: #111F35;
                color: #FFFFFF;
                border: 1px solid #111F35;
            }
            """
        )

    def _on_export_clicked(self) -> None:
        timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
        output_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Save Grouping Count",
            f"grouping_count_{timestamp}.xlsx",
            "Excel Files (*.xlsx);;Text Files (*.txt)",
        )
        if not output_path:
            return

        is_txt = selected_filter.startswith("Text Files") or output_path.lower().endswith(".txt")
        is_xlsx = selected_filter.startswith("Excel Files") or output_path.lower().endswith(".xlsx")

        if not is_txt and not is_xlsx:
            output_path += ".xlsx"
            is_xlsx = True
        elif is_txt and not output_path.lower().endswith(".txt"):
            output_path += ".txt"
        elif is_xlsx and not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"

        try:
            if is_txt:
                with open(output_path, "w", encoding="utf-8") as handle:
                    handle.write("BC\tCount\tBasic Number\tBasic Name\n")
                    for name, count, basic_number, basic_name in self._grouping_counts:
                        handle.write(f"{name}\t{count}\t{basic_number}\t{basic_name}\n")
            else:
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Grouping Count"

                worksheet.append(["BC", "Count", "Basic Number", "Basic Name"])
                for name, count, basic_number, basic_name in self._grouping_counts:
                    worksheet.append([name, str(count), basic_number, basic_name])

                accent_fill = PatternFill(fill_type="solid", fgColor="8A244B")
                white_font = Font(color="FFFFFF", bold=True)
                normal_font = Font(color="111111")
                border = Border(
                    left=Side(style="thin", color="B8B8B8"),
                    right=Side(style="thin", color="B8B8B8"),
                    top=Side(style="thin", color="B8B8B8"),
                    bottom=Side(style="thin", color="B8B8B8"),
                )

                for col in range(1, 5):
                    header_cell = worksheet.cell(row=1, column=col)
                    header_cell.fill = accent_fill
                    header_cell.font = white_font
                    header_cell.alignment = Alignment(horizontal="center", vertical="center")
                    header_cell.border = border

                for row in range(2, worksheet.max_row + 1):
                    for col in range(1, 5):
                        body_cell = worksheet.cell(row=row, column=col)
                        body_cell.font = normal_font
                        if col == 2:
                            body_cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            body_cell.alignment = Alignment(horizontal="left", vertical="center")
                        body_cell.border = border

                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = "A1:D1"
                worksheet.column_dimensions["A"].width = 30
                worksheet.column_dimensions["B"].width = 12
                worksheet.column_dimensions["C"].width = 30
                worksheet.column_dimensions["D"].width = 48
                workbook.save(output_path)
        except Exception as exc:
            msg = QMessageBox(self)
            msg.setWindowTitle("Error")
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setText(f"Failed to export file. {exc}")
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg.exec()
            return

        msg = QMessageBox(self)
        msg.setWindowTitle("Done")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setText("Grouping count exported successfully.")
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()


class _SettingsDialog(QDialog):
    """Settings window — Configure Root Folders and Projects."""

    _ROOT_HINT = (
        "HAT Dashboard will create a 'HAT DASHBOARD ROOT' folder at the chosen location "
        "containing a MASTER sub-folder with: Tracker Status Collector, "
        "Thumbnail Generator, and SAP Data Reformat. "
        "Settings are saved to AppData and persist across sessions."
    )
    _PROJECT_HINT = (
        "HAT Dashboard will create the named project folder at the chosen location "
        "with sub-folders: Packshot Naming Generator, SAP Data Reformat, "
        "SAP Data Compare, and Project Viewer. "
        "The project folder path is saved to AppData."
    )

    def __init__(self, config: "HatConfig", parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("HAT Dashboard — Settings")
        self.resize(650, 340)
        self.setMinimumSize(520, 280)
        self._config = config

        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(28, 24, 28, 24)
        root_layout.setSpacing(14)

        # ── section title ──────────────────────────────────────────────
        section_label = QLabel("Configure root folders")
        section_label.setObjectName("settingsSectionTitle")
        root_layout.addWidget(section_label)

        # ── mode radio buttons ─────────────────────────────────────────
        radio_row = QHBoxLayout()
        radio_row.setSpacing(24)
        self._radio_root = QRadioButton("Setup root")
        self._radio_root.setObjectName("settingsRadio")
        self._radio_root.setChecked(True)
        self._radio_project = QRadioButton("Setup a project")
        self._radio_project.setObjectName("settingsRadio")
        radio_row.addWidget(self._radio_root)
        radio_row.addWidget(self._radio_project)
        radio_row.addStretch(1)
        root_layout.addLayout(radio_row)

        # ── location row ───────────────────────────────────────────────
        loc_row = QHBoxLayout()
        loc_row.setSpacing(10)
        self._lbl_location = QLabel("Root folder location:")
        self._lbl_location.setObjectName("settingsLabel")
        loc_row.addWidget(self._lbl_location, 0)

        self._btn_browse = QPushButton("Browse…")
        self._btn_browse.setObjectName("settingsBrowseBtn")
        self._btn_browse.setFixedHeight(34)
        self._btn_browse.setMinimumWidth(90)
        loc_row.addWidget(self._btn_browse, 0)

        self._input_location = QLineEdit()
        self._input_location.setObjectName("settingsLineEdit")
        self._input_location.setPlaceholderText("Select a folder…")
        self._input_location.setText(config.root_folder())
        loc_row.addWidget(self._input_location, 1)
        root_layout.addLayout(loc_row)

        # ── project name row (project mode only) ──────────────────────
        self._proj_name_row = QWidget()
        proj_name_layout = QHBoxLayout(self._proj_name_row)
        proj_name_layout.setContentsMargins(0, 0, 0, 0)
        proj_name_layout.setSpacing(10)
        lbl_proj = QLabel("Project name:")
        lbl_proj.setObjectName("settingsLabel")
        proj_name_layout.addWidget(lbl_proj, 0)
        self._input_project_name = QLineEdit()
        self._input_project_name.setObjectName("settingsLineEdit")
        self._input_project_name.setPlaceholderText("Enter project folder name…")
        proj_name_layout.addWidget(self._input_project_name, 1)
        self._proj_name_row.setVisible(False)
        root_layout.addWidget(self._proj_name_row)

        # ── hint ───────────────────────────────────────────────────────
        self._hint = QLabel(self._ROOT_HINT)
        self._hint.setObjectName("settingsHint")
        self._hint.setWordWrap(True)
        root_layout.addWidget(self._hint)

        # ── config path label ──────────────────────────────────────────
        self._path_label = QLabel(f"Config file: {config.config_path()}")
        self._path_label.setObjectName("settingsPathLabel")
        self._path_label.setWordWrap(True)
        root_layout.addWidget(self._path_label)

        root_layout.addStretch(1)

        # ── bottom buttons ─────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.addStretch(1)

        self._btn_open_folder = QPushButton("Open config folder")
        self._btn_open_folder.setObjectName("settingsSecondaryBtn")
        self._btn_open_folder.setFixedHeight(34)
        btn_row.addWidget(self._btn_open_folder)

        self._btn_save = QPushButton("Save and Apply")
        self._btn_save.setObjectName("settingsSaveBtn")
        self._btn_save.setFixedHeight(34)
        self._btn_save.setMinimumWidth(120)
        btn_row.addWidget(self._btn_save)

        root_layout.addLayout(btn_row)

        # ── connections ────────────────────────────────────────────────
        self._radio_root.toggled.connect(self._sync_mode)
        self._btn_browse.clicked.connect(self._browse_folder)
        self._btn_save.clicked.connect(self._save)
        self._btn_open_folder.clicked.connect(self._open_config_folder)

        self.setStyleSheet("""
            QDialog { background-color: #F4F4F4; }
            #settingsSectionTitle {
                font-family: 'Segoe UI'; font-size: 17px; font-weight: 700;
                color: #111F35;
            }
            #settingsLabel {
                font-family: 'Segoe UI'; font-size: 13px; color: #333333;
            }
            #settingsHint {
                font-family: 'Segoe UI'; font-size: 11px; color: #666666;
            }
            #settingsPathLabel {
                font-family: 'Segoe UI'; font-size: 10px; color: #999999;
            }
            #settingsRadio {
                font-family: 'Segoe UI'; font-size: 13px; color: #333333;
            }
            #settingsLineEdit {
                background-color: #FFFFFF; color: #111111;
                border: 1px solid #BBBBBB; border-radius: 6px;
                padding: 0 8px; min-height: 34px;
                font-family: 'Segoe UI'; font-size: 12px;
            }
            #settingsBrowseBtn, #settingsSecondaryBtn {
                background-color: #9EA3AB; color: #000000;
                border: 1px solid #8B9098; border-radius: 6px;
                padding: 0 12px;
                font-family: 'Segoe UI'; font-size: 12px; font-weight: 600;
            }
            #settingsBrowseBtn:hover, #settingsSecondaryBtn:hover {
                background-color: #ACB1B8;
            }
            #settingsBrowseBtn:pressed, #settingsSecondaryBtn:pressed {
                background-color: #111F35; color: #ffffff;
            }
            #settingsSaveBtn {
                background-color: #111F35; color: #ffffff;
                border: none; border-radius: 6px; padding: 0 16px;
                font-family: 'Segoe UI'; font-size: 13px; font-weight: 700;
            }
            #settingsSaveBtn:pressed { background-color: #D02752; }
        """)

    # ── slots ──────────────────────────────────────────────────────────────

    def _sync_mode(self, root_checked: bool) -> None:
        if root_checked:
            self._lbl_location.setVisible(True)
            self._btn_browse.setVisible(True)
            self._input_location.setVisible(True)
            self._lbl_location.setText("Root folder location:")
            self._input_location.setPlaceholderText("Select a folder…")
            self._input_location.setText(self._config.root_folder())
            self._proj_name_row.setVisible(False)
            self._hint.setText(self._ROOT_HINT)
        else:
            self._lbl_location.setVisible(False)
            self._btn_browse.setVisible(False)
            self._input_location.setVisible(False)
            self._proj_name_row.setVisible(True)
            self._hint.setText(self._PROJECT_HINT)
        self.adjustSize()

    def _browse_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self, "Select folder", self._input_location.text() or ""
        )
        if folder:
            self._input_location.setText(folder)

    def _save(self) -> None:
        from pathlib import Path

        if self._radio_root.isChecked():
            location = self._input_location.text().strip()
            if not location:
                QMessageBox.warning(self, "No folder selected",
                                    "Please select a folder before saving.")
                return
            if not Path(location).exists():
                QMessageBox.warning(self, "Folder not found",
                                    f"The folder does not exist:\n{location}")
                return
            created = self._config.create_folder_hierarchy(location)
            self._config.set_root_folder(location)
            self._config.save()
            msg = QMessageBox(self)
            msg.setWindowTitle("Settings saved")
            msg.setIcon(QMessageBox.Icon.Information)
            if created:
                msg.setText(
                    f"Root folder set to:\n{location}\n\n"
                    f"Created {len(created)} new sub-folder(s)."
                )
            else:
                msg.setText(f"Root folder set to:\n{location}\n\nAll folders already exist.")
            msg.exec()
        else:
            project_name = self._input_project_name.text().strip()
            if not project_name:
                QMessageBox.warning(self, "No project name",
                                    "Please enter a project name.")
                return
            root = self._config.root_folder()
            if not root or not Path(root).exists():
                QMessageBox.warning(self, "Root folder not found",
                                    "Root folder not found, configure in Setup root.")
                return
            # Place project at <root>/HAT DASHBOARD ROOT/ — same level as MASTER
            location = str(Path(root) / "HAT DASHBOARD ROOT")
            created, project_path = self._config.create_project_folder_hierarchy(
                location, project_name
            )
            self._config.save()
            msg = QMessageBox(self)
            msg.setWindowTitle("Project created")
            msg.setIcon(QMessageBox.Icon.Information)
            if created:
                msg.setText(
                    f"Project folder created:\n{project_path}\n\n"
                    f"Created {len(created)} sub-folder(s)."
                )
            else:
                msg.setText(
                    f"Project folder:\n{project_path}\n\nAll folders already exist."
                )
            msg.exec()

        self._path_label.setText(f"Config file: {self._config.config_path()}")

    def _open_config_folder(self) -> None:
        import subprocess
        folder = str(self._config.config_path().parent)
        subprocess.Popen(f'explorer "{folder}"')


class NewUIWindow(QMainWindow):
    """UI-only dashboard shell. No business logic is connected yet."""

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("HAT Dashboard")
        self.resize(910, 792)
        self.setMinimumSize(910, 792)

        root = QWidget()
        root.setObjectName("root")
        self.setCentralWidget(root)

        outer = QVBoxLayout(root)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        shell = QFrame()
        shell.setObjectName("shell")
        outer.addWidget(shell)

        shell_layout = QVBoxLayout(shell)
        shell_layout.setContentsMargins(12, 12, 12, 12)
        shell_layout.setSpacing(14)

        self.btn_home = self._make_top_button("Home")
        self.btn_master = self._make_top_button("Master")
        self.btn_project = self._make_top_button("Project")
        self.btn_search = self._make_top_button("Search")

        main_row = QHBoxLayout()
        main_row.setSpacing(14)
        shell_layout.addLayout(main_row, 1)

        icon_rail = QVBoxLayout()
        icon_rail.setSpacing(10)
        icon_rail.setAlignment(Qt.AlignTop)
        main_row.addLayout(icon_rail)

        self.btn_settings = self._make_icon_button("\u2699")  # ⚙ cog
        self.btn_settings.setToolTip("Settings")
        icon_rail.addWidget(self.btn_settings)

        self.btn_rail_toggle = ToggleSwitch(width=46, height=26)
        self.btn_rail_toggle.setToolTip("Use configured root folders")
        icon_rail.addWidget(self.btn_rail_toggle)
        icon_rail.addStretch(1)

        content = QFrame()
        content.setObjectName("contentCard")
        main_row.addWidget(content, 1)

        self._content_layout = QVBoxLayout(content)
        self._content_layout.setContentsMargins(20, 18, 20, 18)
        self._content_layout.setSpacing(0)
        content_layout = self._content_layout

        header_band = QFrame()
        header_band.setObjectName("headerBand")
        header_band.setFixedHeight(110)
        content_layout.addWidget(header_band)

        # ── Project selector row (visible only when PROJECT is active) ──────────
        self.project_selector_row = QWidget()
        self.project_selector_row.setVisible(False)
        proj_sel_outer = QHBoxLayout(self.project_selector_row)
        proj_sel_outer.setContentsMargins(0, 0, 0, 0)
        proj_sel_outer.setSpacing(0)

        self._proj_sel_band = QFrame()
        self._proj_sel_band.setObjectName("projectSelectorBand")
        self._proj_sel_band.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        proj_band_layout = QHBoxLayout(self._proj_sel_band)
        proj_band_layout.setContentsMargins(14, 6, 14, 6)
        proj_band_layout.setSpacing(10)

        lbl_proj_name = QLabel("Project Name:")
        lbl_proj_name.setObjectName("projectSelectorLabel")
        self.combo_project = QComboBox()
        self.combo_project.setObjectName("projectSelectorCombo")
        self.combo_project.setMinimumWidth(400)
        self.combo_project.setFixedHeight(28)
        proj_band_layout.addWidget(lbl_proj_name)
        proj_band_layout.addWidget(self.combo_project)
        self.combo_project.currentIndexChanged.connect(self._on_project_combo_changed)
        proj_band_layout.addStretch(1)

        proj_sel_outer.addWidget(self._proj_sel_band, 1)
        # (added to active_tools_page VBox below)

        header_row = QHBoxLayout(header_band)
        header_row.setContentsMargins(16, 14, 16, 14)
        header_row.setSpacing(0)

        left_slot = QWidget()
        left_slot.setFixedWidth(250)
        left_slot_layout = QHBoxLayout(left_slot)
        left_slot_layout.setContentsMargins(16, 0, 0, 0)
        left_slot_layout.setSpacing(0)

        hat_badge = QLabel()
        hat_badge.setObjectName("hatBadge")
        hat_badge.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        hat_badge.setMinimumWidth(231)
        _logo_path = str(_BASE_DIR / "logo" / "hat_logo.png")
        _logo_pix = QPixmap(_logo_path)
        if not _logo_pix.isNull():
            hat_badge.setPixmap(_logo_pix.scaled(
                231, 84, Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation))
        left_slot_layout.addWidget(hat_badge)
        left_slot_layout.addStretch(1)
        header_row.addWidget(left_slot)

        buttons_slot = QWidget()
        buttons_row = QHBoxLayout(buttons_slot)
        buttons_row.setContentsMargins(0, 0, 0, 0)
        buttons_row.setSpacing(8)
        buttons_row.addWidget(self.btn_home)
        buttons_row.addWidget(self.btn_master)
        buttons_row.addWidget(self.btn_project)
        buttons_row.addWidget(self.btn_search)
        buttons_row.addItem(QSpacerItem(20, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        header_row.addWidget(buttons_slot, 1)

        right_slot = QWidget()
        right_slot.setFixedWidth(40)
        header_row.addWidget(right_slot)

        self.section_title = QLabel("HOME")
        self.section_title.setObjectName("sectionTitle")

        # ── Active tools page (shared by MASTER and PROJECT) ─────────────────────
        self.active_tools_page = QWidget()
        self._active_page_layout = QVBoxLayout(self.active_tools_page)
        self._active_page_layout.setContentsMargins(0, 16, 0, 0)
        self._active_page_layout.setSpacing(8)
        active_page_layout = self._active_page_layout

        # Body row: left nav + right column (selector + right panel)
        active_body_widget = QWidget()
        self._active_body_layout = QHBoxLayout(active_body_widget)
        self._active_body_layout.setContentsMargins(0, 0, 0, 0)
        self._active_body_layout.setSpacing(18)
        active_page_layout.addWidget(active_body_widget, 1)
        active_body_layout = self._active_body_layout

        # Left nav stack – switches between master / project button sets
        self.left_nav_stack = QStackedWidget()
        self.left_nav_stack.setMaximumWidth(178)
        self.left_nav_blank = QWidget()  # HOME / SEARCH
        self.left_nav_stack.addWidget(self.left_nav_blank)  # index 0

        # ── MASTER left buttons: TSC, TG, SDR ────────────────────────────────
        self.master_left = QWidget()
        master_actions_layout = QVBoxLayout(self.master_left)
        master_actions_layout.setContentsMargins(0, 0, 0, 0)
        master_actions_layout.setSpacing(14)

        self.btn_action_status_collector = QPushButton("Trackers Status Collector")
        self.btn_action_thumbnail_generator = QPushButton("Thumbnail Generator")
        self.btn_action_sdr_master = QPushButton("SAP Data Reformat")

        master_action_buttons = (
            self.btn_action_status_collector,
            self.btn_action_thumbnail_generator,
            self.btn_action_sdr_master,
        )
        self.master_action_group = QButtonGroup(self)
        self.master_action_group.setExclusive(True)
        for btn in master_action_buttons:
            btn.setObjectName("basicToolActionBtn")
            btn.setCheckable(True)
            btn.setText(self._wrap_button_text(btn.text(), words_per_line=2))
            btn.setFixedSize(160, 132)
            btn.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
            self.master_action_group.addButton(btn)
        for btn in master_action_buttons:
            master_actions_layout.addWidget(btn, 0, Qt.AlignmentFlag.AlignLeft)
        master_actions_layout.addStretch(1)
        self.left_nav_stack.addWidget(self.master_left)  # index 1

        # ── PROJECT left buttons: PNG, SDR, SDC, Project Viewer ───────────────────
        self.project_left = QWidget()
        project_actions_layout = QVBoxLayout(self.project_left)
        project_actions_layout.setContentsMargins(0, 10, 0, 0)
        project_actions_layout.setSpacing(14)

        self.btn_action_packshot_naming = QPushButton("Packshot Naming Generator")
        self.btn_action_sap_data_reformat = QPushButton("SAP Data Reformat")
        self.btn_action_sap_data_compare = QPushButton("SAP Data Compare")
        self.btn_action_project_viewer = QPushButton("Review Project")
        self.btn_action_other_tools = QPushButton("Other Tools")

        project_action_buttons = (
            self.btn_action_packshot_naming,
            self.btn_action_sap_data_reformat,
            self.btn_action_sap_data_compare,
            self.btn_action_project_viewer,
            self.btn_action_other_tools,
        )
        self.project_action_group = QButtonGroup(self)
        self.project_action_group.setExclusive(True)
        for btn in project_action_buttons:
            btn.setObjectName("basicToolActionBtn")
            btn.setCheckable(True)
            btn.setText(self._wrap_button_text(btn.text(), words_per_line=2))
            btn.setFixedSize(160, 83)
            btn.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
            self.project_action_group.addButton(btn)
        for btn in project_action_buttons:
            project_actions_layout.addWidget(btn, 0, Qt.AlignmentFlag.AlignLeft)
        project_actions_layout.addStretch(1)
        self.left_nav_stack.addWidget(self.project_left)  # index 2

        active_body_layout.addWidget(self.left_nav_stack, 0)

        # ── Right column: project selector (top, hidden by default) + right panel ─
        right_col_widget = QWidget()
        right_col_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._right_col_layout = QVBoxLayout(right_col_widget)
        self._right_col_layout.setContentsMargins(0, 0, 0, 0)
        self._right_col_layout.setSpacing(6)
        self._right_col_layout.addWidget(self.project_selector_row)

        # ── Combined right panel (all tool panels in one stack) ──────────────────
        self.combined_right_panel = QFrame()
        self.combined_right_panel.setObjectName("basicToolsRightPanel")
        self.combined_right_panel.setVisible(False)
        self.combined_right_panel.setMinimumSize(382, 437)
        self.combined_right_panel.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._right_col_layout.addWidget(self.combined_right_panel, 1)
        active_body_layout.addWidget(right_col_widget, 1)
        self._build_combined_right_panel()

        # MASTER button connections
        self.btn_action_status_collector.clicked.connect(self._on_status_panel_clicked)
        self.btn_action_thumbnail_generator.clicked.connect(self._on_thumbnail_panel_clicked)
        self.btn_action_sdr_master.clicked.connect(self._on_mapper_reformat_clicked)
        # PROJECT button connections
        self.btn_action_packshot_naming.clicked.connect(self._on_packshot_panel_clicked)
        self.btn_action_sap_data_reformat.clicked.connect(self._on_mapper_reformat_clicked)
        self.btn_action_sap_data_compare.clicked.connect(self._on_mapper_compare_clicked)
        self.btn_action_project_viewer.clicked.connect(self._on_project_viewer_clicked)
        self.btn_action_other_tools.clicked.connect(self._on_other_tools_clicked)

        self.empty_page = QWidget()
        _home_outer = QVBoxLayout(self.empty_page)
        _home_outer.setContentsMargins(0, 0, 0, 0)
        _home_outer.setSpacing(0)

        # White card panel matching other pages
        _home_card = QFrame()
        _home_card.setObjectName("homeCard")
        _home_card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        _home_card_layout = QHBoxLayout(_home_card)
        _home_card_layout.setContentsMargins(40, 0, 40, 0)
        _home_card_layout.setSpacing(24)
        _home_card_layout.addStretch(1)

        # Text graphic: home_page_text.png
        _home_img_label = QLabel()
        _home_img_label.setAlignment(Qt.AlignTop | Qt.AlignRight)
        _home_img_label.setScaledContents(False)
        _home_img_label.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Preferred)
        _home_img_path = str(_BASE_DIR / "logo" / "home_page_text.png")
        _home_pixmap = QPixmap(_home_img_path)
        if not _home_pixmap.isNull():
            _home_img_label.setPixmap(_home_pixmap)
        # Wrapper so we can shift the text PNG via top margin
        _text_wrapper = QWidget()
        _text_wrapper.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
        _text_wrapper_layout = QVBoxLayout(_text_wrapper)
        _text_wrapper_layout.setContentsMargins(0, 0, 0, 0)
        _text_wrapper_layout.setSpacing(0)
        _text_wrapper_layout.addWidget(_home_img_label, 0, Qt.AlignTop | Qt.AlignRight)
        _text_wrapper_layout.addStretch(1)
        _home_card_layout.addWidget(_text_wrapper, 0, Qt.AlignRight)

        # Animated GIF: 360_Bottle 1.gif
        _gif_label = QLabel()
        _gif_label.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        _gif_label.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Preferred)
        _gif_path = str(_BASE_DIR / "logo" / "360_Bottle 1.gif")
        _gif_movie = QMovie(_gif_path)
        if _gif_movie.isValid():
            _gif_label.setMovie(_gif_movie)
            _gif_movie.start()
        # Wrapper so we can shift the GIF upward via top margin
        _gif_wrapper = QWidget()
        _gif_wrapper.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
        _gif_wrapper_layout = QVBoxLayout(_gif_wrapper)
        _gif_wrapper_layout.setContentsMargins(0, 0, 0, 0)
        _gif_wrapper_layout.setSpacing(0)
        _gif_wrapper_layout.addWidget(_gif_label, 0, Qt.AlignTop | Qt.AlignLeft)
        _gif_wrapper_layout.addStretch(1)
        _home_card_layout.addWidget(_gif_wrapper, 0, Qt.AlignLeft)

        _home_card_layout.addStretch(1)
        _home_outer.addWidget(_home_card, 1)

        self._home_img_label = _home_img_label
        self._home_pixmap = _home_pixmap
        self._text_wrapper_layout = _text_wrapper_layout
        self._gif_movie = _gif_movie  # keep reference alive
        self._gif_label = _gif_label
        self._gif_wrapper_layout = _gif_wrapper_layout
        self._home_card = _home_card
        self.content_stack = QStackedWidget()
        self.content_stack.addWidget(self.empty_page)
        self.content_stack.addWidget(self.active_tools_page)
        content_layout.addWidget(self.content_stack, 1)
        content_layout.addWidget(self.section_title, 0, Qt.AlignLeft | Qt.AlignBottom)

        self.top_group = QButtonGroup(self)
        self.top_group.setExclusive(True)
        self._register_top_button(self.btn_home, "HOME")
        self._register_top_button(self.btn_master, "MASTER")
        self._register_top_button(self.btn_project, "PROJECT")
        self._register_top_button(self.btn_search, "SEARCH")
        self.btn_home.setChecked(True)
        self._set_section_title("HOME")

        self._config = HatConfig()
        self._use_root_folders: bool = False
        self.btn_settings.clicked.connect(self._open_settings_dialog)
        self.btn_rail_toggle.toggled.connect(self._on_root_folders_toggle)

        self._init_modules()
        self.setStyleSheet(self._stylesheet())
        QTimer.singleShot(0, self._update_home_image)

    def _make_top_button(self, text: str) -> QPushButton:
        btn = QPushButton(text.upper())
        btn.setObjectName("topHeaderBtn")
        btn.setCheckable(True)
        btn.setCursor(Qt.PointingHandCursor)
        btn.setMinimumHeight(50)
        width_map = {
            "HOME": 90,
            "MASTER": 120,
            "PROJECT": 120,
            "SEARCH": 110,
        }
        btn.setMinimumWidth(width_map.get(text.upper(), 120))
        return btn

    def _make_icon_button(self, symbol: str) -> QPushButton:
        btn = QPushButton(symbol)
        btn.setObjectName("iconBtn")
        btn.setCursor(Qt.PointingHandCursor)
        btn.setFixedSize(38, 38)
        return btn

    def _register_top_button(self, button: QPushButton, title: str) -> None:
        self.top_group.addButton(button)
        button.clicked.connect(lambda _checked=False, t=title: self._set_section_title(t))

    def _set_section_title(self, title: str) -> None:
        self.section_title.setText(title)
        if title == "MASTER":
            self.project_selector_row.setVisible(False)
            self._content_layout.setSpacing(0)
            self._active_page_layout.setContentsMargins(0, 16, 0, 0)
            self.content_stack.setCurrentWidget(self.active_tools_page)
            self.left_nav_stack.setCurrentWidget(self.master_left)
            self.btn_action_status_collector.setChecked(True)
            self._show_tracker_status_panel()
        elif title == "PROJECT":
            self._populate_project_combo()
            self.project_selector_row.setVisible(True)
            self._content_layout.setSpacing(0)
            self._active_page_layout.setContentsMargins(0, 4, 0, 0)
            self.content_stack.setCurrentWidget(self.active_tools_page)
            self.left_nav_stack.setCurrentWidget(self.project_left)
            self.btn_action_packshot_naming.setChecked(True)
            self._show_packshot_panel()
        else:  # HOME, SEARCH
            self.project_selector_row.setVisible(False)
            self._content_layout.setSpacing(16)
            self.content_stack.setCurrentWidget(self.empty_page)
            self.combined_right_panel.setVisible(False)

    def _update_home_image(self) -> None:
        """Scale home_page_text.png and the GIF to fit the card height."""
        from PySide6.QtCore import QSize
        card_h = self._home_card.height()
        if card_h <= 0:
            return

        # Text PNG: ~19.97% of card height (+20%), shifted up from centre
        text_target_h = max(20, int(card_h * 0.1997))
        text_top_margin = max(0, (card_h - text_target_h) // 2 - 20)
        self._text_wrapper_layout.setContentsMargins(0, text_top_margin, 0, 0)
        if not self._home_pixmap.isNull():
            scaled_pix = self._home_pixmap.scaledToHeight(
                text_target_h, Qt.TransformationMode.SmoothTransformation
            )
            self._home_img_label.setPixmap(scaled_pix)

        # GIF: ~46.59% of card height (+20%), fixed position
        gif_target_h = max(60, int(card_h * 0.4659))
        if self._gif_movie.isValid():
            nat = self._gif_movie.currentImage().size()
            if nat.height() > 0:
                gif_w = int(nat.width() * gif_target_h / nat.height())
                self._gif_movie.setScaledSize(QSize(gif_w, gif_target_h))
        gif_top_margin = max(0, (card_h - gif_target_h) // 2 - 67)
        self._gif_wrapper_layout.setContentsMargins(0, gif_top_margin, 0, 0)

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self._update_home_image()

    def _populate_project_combo(self) -> None:
        """Fill combo_project with folder names inside HAT DASHBOARD ROOT that are not MASTER."""
        from pathlib import Path
        self.combo_project.blockSignals(True)
        self.combo_project.clear()
        root = self._config.root_folder()
        if not root:
            self.combo_project.addItem("(no root configured)")
            self.combo_project.blockSignals(False)
            return
        hat_root = Path(root) / "HAT DASHBOARD ROOT"
        if not hat_root.is_dir():
            self.combo_project.addItem("(root not found)")
            self.combo_project.blockSignals(False)
            return
        projects = sorted(
            d.name for d in hat_root.iterdir()
            if d.is_dir() and d.name.upper() != "MASTER"
        )
        if projects:
            for name in projects:
                self.combo_project.addItem(name)
        else:
            self.combo_project.addItem("(no projects found)")
        self.combo_project.blockSignals(False)
        if self._use_root_folders:
            self._autofill_project_page()

    def _on_project_combo_changed(self) -> None:
        """Re-autofill PROJECT fields when the selected project changes."""
        if self._use_root_folders:
            self._autofill_project_page()

    def _show_basic_tools_right_panel(self) -> None:
        self.combined_panel_stack.setCurrentWidget(self.right_panel_blank)
        self.combined_right_panel.setVisible(True)

    def _show_tracker_status_panel(self) -> None:
        self.combined_panel_stack.setCurrentWidget(self.right_panel_tracker)
        self.combined_right_panel.setVisible(True)

    def _show_thumbnail_panel(self) -> None:
        self.combined_panel_stack.setCurrentWidget(self.right_panel_thumbnail)
        self.combined_right_panel.setVisible(True)

    def _show_packshot_panel(self) -> None:
        self.combined_panel_stack.setCurrentWidget(self.right_panel_packshot)
        self.combined_right_panel.setVisible(True)

    def _on_status_panel_clicked(self) -> None:
        clear_other_panel_inputs(self, active_panel="status")
        if self._use_root_folders:
            self._apply_root_folder_inputs()
        self._show_tracker_status_panel()

    def _on_thumbnail_panel_clicked(self) -> None:
        clear_other_panel_inputs(self, active_panel="thumbnail")
        if self._use_root_folders:
            self._apply_root_folder_inputs()
        self._show_thumbnail_panel()

    def _on_packshot_panel_clicked(self) -> None:
        clear_other_panel_inputs(self, active_panel="packshot")
        if self._use_root_folders:
            self._apply_root_folder_inputs()
        self._show_packshot_panel()

    def _on_mapper_reformat_clicked(self) -> None:
        if self._use_root_folders:
            self._apply_root_folder_inputs()
        self._show_mapper_reformat_panel()

    def _on_mapper_compare_clicked(self) -> None:
        self._reset_mapper_reformat_inputs()
        if self._use_root_folders:
            self._apply_root_folder_inputs()
        self._show_mapper_compare_panel()

    def _on_project_viewer_clicked(self) -> None:
        self.combined_panel_stack.setCurrentWidget(self.right_panel_project_viewer)
        self.combined_right_panel.setVisible(True)

    def _on_other_tools_clicked(self) -> None:
        self.combined_panel_stack.setCurrentWidget(self.right_panel_other_tools)
        self.combined_right_panel.setVisible(True)

    def _reset_mapper_reformat_inputs(self) -> None:
        """Return the SAP Data Reformat panel to its default state."""
        self.radio_mapper_option_1.setChecked(True)  # also triggers _sync_mapper_reformat_mode_ui
        self.input_mapper_sap_data_files.clear()
        self.input_mapper_output_location.clear()
        self._mapper_sap_file_paths = []
        self._mapper_output_location = ""
        self.radio_mapper_cleanup_1.setChecked(True)
        self.checkbox_mapper_include_grouping_report.setChecked(True)
        self.mapper_reformat_progress_bar.setValue(0)
        self.mapper_reformat_progress_bar.setVisible(False)
        self.btn_run_process_mapper_reformat.setEnabled(False)

    def _build_combined_right_panel(self) -> None:
        panel_layout = QVBoxLayout(self.combined_right_panel)
        panel_layout.setContentsMargins(16, 16, 16, 16)
        panel_layout.setSpacing(0)

        self.combined_panel_stack = QStackedWidget()
        self.right_panel_blank = QWidget()
        self.right_panel_tracker = self._create_tracker_status_collector_page()
        self.right_panel_thumbnail = self._create_thumbnail_generator_page()
        self.right_panel_packshot = self._create_packshot_naming_page()
        self.mapper_right_panel_reformat = self._create_mapper_reformat_page()
        self.mapper_right_panel_compare = self._create_mapper_compare_page()
        self.right_panel_project_viewer = QWidget()
        self.right_panel_other_tools = self._create_other_tools_page()
        self.combined_panel_stack.addWidget(self.right_panel_blank)
        self.combined_panel_stack.addWidget(self.right_panel_tracker)
        self.combined_panel_stack.addWidget(self.right_panel_thumbnail)
        self.combined_panel_stack.addWidget(self.right_panel_packshot)
        self.combined_panel_stack.addWidget(self.mapper_right_panel_reformat)
        self.combined_panel_stack.addWidget(self.mapper_right_panel_compare)
        self.combined_panel_stack.addWidget(self.right_panel_project_viewer)
        self.combined_panel_stack.addWidget(self.right_panel_other_tools)
        panel_layout.addWidget(self.combined_panel_stack)

    def _create_tracker_status_collector_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(14)

        title = QLabel("Trackers Status Collector")
        title.setObjectName("collectorTitle")
        title_row = QHBoxLayout()
        title_row.setContentsMargins(0, 0, 0, 0)
        title_row.addWidget(title, 1)
        layout.addLayout(title_row)

        # --- Option 1 / Option 2 radios ---
        self.radio_tsc_option_1 = QRadioButton("Option 1")
        self.radio_tsc_option_1.setObjectName("collectorModeRadio")
        self.radio_tsc_option_2 = QRadioButton("Option 2")
        self.radio_tsc_option_2.setObjectName("collectorModeRadio")
        self.radio_tsc_option_1.setChecked(True)

        self.tsc_option_group = QButtonGroup(page)
        self.tsc_option_group.setExclusive(True)
        self.tsc_option_group.addButton(self.radio_tsc_option_1)
        self.tsc_option_group.addButton(self.radio_tsc_option_2)

        layout.addWidget(self.radio_tsc_option_1)

        # Option 1 – Open Window button
        opt1_row = QHBoxLayout()
        opt1_row.setSpacing(12)
        self.btn_tsc_open_window = QPushButton("Open window")
        self.btn_tsc_open_window.setObjectName("collectorGrayBtn")
        opt1_row.addWidget(self.btn_tsc_open_window, 0)
        opt1_row.addStretch(1)
        layout.addLayout(opt1_row)

        layout.addSpacing(8)

        tsc_opt2_header = QHBoxLayout()
        tsc_opt2_header.setSpacing(0)
        tsc_opt2_header.addWidget(self.radio_tsc_option_2)
        tsc_opt2_header.addSpacing(80)
        self.lbl_tsc_input_count = QLabel("")
        self.lbl_tsc_input_count.setObjectName("inputCountLabel")
        tsc_opt2_header.addWidget(self.lbl_tsc_input_count, 0)
        tsc_opt2_header.addStretch(1)
        layout.addLayout(tsc_opt2_header)

        # Option 2 – existing elements
        trackers_row = QHBoxLayout()
        trackers_row.setSpacing(12)
        self.btn_sc_select_trackers = QPushButton("Select Trackers")
        self.btn_sc_select_trackers.setObjectName("collectorGrayBtn")
        trackers_row.addWidget(self.btn_sc_select_trackers, 0)
        self.input_trackers = QLineEdit()
        self.input_trackers.setObjectName("collectorLineEdit")
        trackers_row.addWidget(self.input_trackers, 1)
        layout.addLayout(trackers_row)

        output_row = QHBoxLayout()
        output_row.setSpacing(12)
        self.btn_sc_output_location = QPushButton("Output Report Location")
        self.btn_sc_output_location.setObjectName("collectorGrayBtn")
        output_row.addWidget(self.btn_sc_output_location, 0)
        self.input_output = QLineEdit()
        self.input_output.setObjectName("collectorLineEdit")
        output_row.addWidget(self.input_output, 1)
        layout.addLayout(output_row)

        # Select Status / Input Status sub-radios (styled like Cleanup radios)
        self.radioButton_status_select = QRadioButton("Select Status")
        self.radioButton_status_select.setObjectName("collectorCleanupRadio")
        self.radioButton_status_input = QRadioButton("Input Status")
        self.radioButton_status_input.setObjectName("collectorCleanupRadio")
        self.radioButton_status_select.setChecked(True)
        self.collector_mode_group = QButtonGroup(page)
        self.collector_mode_group.setExclusive(True)
        self.collector_mode_group.addButton(self.radioButton_status_select)
        self.collector_mode_group.addButton(self.radioButton_status_input)
        layout.addWidget(self.radioButton_status_select)

        statuses_box = QVBoxLayout()
        statuses_box.setSpacing(6)
        # Row 1 – the 5 status checkboxes + Apply Cleanup
        statuses_row1 = QHBoxLayout()
        statuses_row1.setSpacing(8)
        self.checkbox_status_collector_to_do = QCheckBox("To Do")
        self.checkbox_status_collector_cancelled = QCheckBox("Cancelled")
        self.checkbox_status_collector_completed = QCheckBox("Completed")
        self.checkbox_status_collector_completed.setChecked(True)
        self.checkbox_status_collector_in_progress = QCheckBox("In Progress")
        self.checkbox_status_collector_on_hold = QCheckBox("On Hold")
        status_checkboxes = [
            self.checkbox_status_collector_to_do,
            self.checkbox_status_collector_in_progress,
            self.checkbox_status_collector_completed,
            self.checkbox_status_collector_on_hold,
            self.checkbox_status_collector_cancelled,
        ]
        for cb in status_checkboxes:
            cb.setObjectName("collectorCheckSmall")
            statuses_row1.addWidget(cb)
        # Apply Cleanup on the same row, after Cancelled
        self.checkbox_sc_apply_cleanup = QCheckBox("Apply Cleanup")
        self.checkbox_sc_apply_cleanup.setObjectName("collectorCheckSmall")
        self.checkbox_sc_apply_cleanup.setChecked(True)
        statuses_row1.addWidget(self.checkbox_sc_apply_cleanup)
        statuses_row1.addStretch(1)
        statuses_box.addLayout(statuses_row1)
        statuses_indent_row = QHBoxLayout()
        statuses_indent_row.setContentsMargins(24, 0, 0, 0)
        statuses_indent_row.addLayout(statuses_box)
        layout.addLayout(statuses_indent_row)
        layout.addSpacing(4)
        layout.addWidget(self.radioButton_status_input)

        self._input_status_line = QLineEdit()
        self._input_status_line.setObjectName("collectorLineEdit")
        self._input_status_line.setPlaceholderText("completed")
        self._input_status_line.setText("completed")
        layout.addWidget(self._input_status_line)

        layout.addStretch(1)
        run_row = QHBoxLayout()
        run_row.addStretch(1)
        self.btn_run_process_all_trackers_status_collector = QPushButton("Run Process")
        self.btn_run_process_all_trackers_status_collector.setObjectName("collectorRunBtn")
        self.btn_run_process_all_trackers_status_collector.setFixedHeight(50)
        self.btn_run_process_all_trackers_status_collector.setMinimumWidth(180)
        run_row.addWidget(self.btn_run_process_all_trackers_status_collector)
        run_row.addStretch(1)
        layout.addLayout(run_row)
        layout.addSpacing(8)

        progress_row = QHBoxLayout()
        progress_row.addStretch(1)
        self.collector_progress_bar = QProgressBar()
        self.collector_progress_bar.setObjectName("collectorProgressBar")
        self.collector_progress_bar.setMinimumWidth(260)
        self.collector_progress_bar.setMaximumWidth(360)
        self.collector_progress_bar.setFixedHeight(14)
        self.collector_progress_bar.setTextVisible(False)
        self.collector_progress_bar.setRange(0, 100)
        self.collector_progress_bar.setValue(0)
        self.collector_progress_bar.setVisible(False)
        progress_row.addWidget(self.collector_progress_bar)
        progress_row.addStretch(1)
        layout.addLayout(progress_row)

        self.radio_tsc_option_1.toggled.connect(self._sync_tsc_mode_ui)
        self.radio_tsc_option_2.toggled.connect(self._sync_tsc_mode_ui)
        self.btn_tsc_open_window.clicked.connect(self._open_tsc_option1_table)
        self.input_trackers.textChanged.connect(lambda: self._update_tsc_input_count())
        self._sync_tsc_mode_ui()

        return page

    def _sync_tsc_mode_ui(self) -> None:
        is_opt1 = self.radio_tsc_option_1.isChecked()

        self.btn_tsc_open_window.setEnabled(is_opt1)

        opt2_widgets = [
            self.btn_sc_select_trackers,
            self.input_trackers,
            self.btn_sc_output_location,
            self.input_output,
            self.radioButton_status_select,
            self.radioButton_status_input,
            self.checkbox_status_collector_to_do,
            self.checkbox_status_collector_cancelled,
            self.checkbox_status_collector_completed,
            self.checkbox_status_collector_in_progress,
            self.checkbox_status_collector_on_hold,
            self._input_status_line,
            self.checkbox_sc_apply_cleanup,
            self.btn_run_process_all_trackers_status_collector,
        ]
        for w in opt2_widgets:
            w.setEnabled(not is_opt1)

        if is_opt1:
            self.input_trackers.clear()
            self.input_output.clear()
            self.radioButton_status_select.setChecked(True)
            self.checkbox_status_collector_completed.setChecked(True)
            self.checkbox_status_collector_to_do.setChecked(False)
            self.checkbox_status_collector_cancelled.setChecked(False)
            self.checkbox_status_collector_in_progress.setChecked(False)
            self.checkbox_status_collector_on_hold.setChecked(False)
            self._input_status_line.setText("completed")
        else:
            self.checkbox_sc_apply_cleanup.setChecked(True)

    def _create_thumbnail_generator_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(16)

        title = QLabel("Thumbnail Generator")
        title.setObjectName("thumbnailTitle")
        title_row = QHBoxLayout()
        title_row.setContentsMargins(0, 0, 0, 0)
        title_row.addWidget(title, 1)
        layout.addLayout(title_row)

        description = QLabel(
            "NOTE: Thumbnail sized packshots cropped edge to edge used for library\n"
            "and reference collector."
        )
        description.setObjectName("thumbnailDescription")
        layout.addWidget(description)
        layout.addSpacing(12)

        images_row = QHBoxLayout()
        images_row.setSpacing(12)
        self.btn_pg7_images_folder = QPushButton("Images Folder")
        self.btn_pg7_images_folder.setObjectName("collectorGrayBtn")
        images_row.addWidget(self.btn_pg7_images_folder, 0)
        self.input_pg7_images_folder = QLineEdit()
        self.input_pg7_images_folder.setObjectName("collectorLineEdit")
        images_row.addWidget(self.input_pg7_images_folder, 1)
        layout.addLayout(images_row)

        output_row = QHBoxLayout()
        output_row.setSpacing(12)
        self.btn_pg7_output = QPushButton("Output Location")
        self.btn_pg7_output.setObjectName("collectorGrayBtn")
        output_row.addWidget(self.btn_pg7_output, 0)
        self.input_pg7_output = QLineEdit()
        self.input_pg7_output.setObjectName("collectorLineEdit")
        output_row.addWidget(self.input_pg7_output, 1)
        layout.addLayout(output_row)

        layout.addStretch(1)
        run_row = QHBoxLayout()
        run_row.addStretch(1)
        self.btn_run_process_generate_thumbnails = QPushButton("Run Process")
        self.btn_run_process_generate_thumbnails.setObjectName("collectorRunBtn")
        self.btn_run_process_generate_thumbnails.setFixedHeight(50)
        self.btn_run_process_generate_thumbnails.setMinimumWidth(180)
        run_row.addWidget(self.btn_run_process_generate_thumbnails)
        run_row.addStretch(1)
        layout.addLayout(run_row)
        layout.addSpacing(8)

        progress_row = QHBoxLayout()
        progress_row.addStretch(1)
        self.thumbnail_progress_bar = QProgressBar()
        self.thumbnail_progress_bar.setObjectName("collectorProgressBar")
        self.thumbnail_progress_bar.setMinimumWidth(260)
        self.thumbnail_progress_bar.setMaximumWidth(360)
        self.thumbnail_progress_bar.setFixedHeight(14)
        self.thumbnail_progress_bar.setTextVisible(False)
        self.thumbnail_progress_bar.setRange(0, 100)
        self.thumbnail_progress_bar.setValue(0)
        self.thumbnail_progress_bar.setVisible(False)
        progress_row.addWidget(self.thumbnail_progress_bar)
        progress_row.addStretch(1)
        layout.addLayout(progress_row)

        return page

    def _show_mapper_reformat_panel(self) -> None:
        self.combined_panel_stack.setCurrentWidget(self.mapper_right_panel_reformat)
        self.combined_right_panel.setVisible(True)

    def _show_mapper_compare_panel(self) -> None:
        self.combined_panel_stack.setCurrentWidget(self.mapper_right_panel_compare)
        self.combined_right_panel.setVisible(True)

    def _create_mapper_reformat_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(14)

        title = QLabel("SAP Data Reformat")
        title.setObjectName("mapperTitle")
        title_row = QHBoxLayout()
        title_row.setContentsMargins(0, 0, 0, 0)
        title_row.addWidget(title, 1)
        layout.addLayout(title_row)

        note = QLabel("NOTE: Do not alter SAP data file. This tool is heavily based on the default state.")
        note.setObjectName("mapperDescription")
        note.setWordWrap(True)
        layout.addWidget(note)
        layout.addSpacing(8)

        self.radio_mapper_option_1 = QRadioButton("Option 1")
        self.radio_mapper_option_1.setObjectName("collectorModeRadio")
        self.radio_mapper_option_2 = QRadioButton("Option 2")
        self.radio_mapper_option_2.setObjectName("collectorModeRadio")
        self.radio_mapper_option_1.setChecked(True)

        self.mapper_option_group = QButtonGroup(page)
        self.mapper_option_group.setExclusive(True)
        self.mapper_option_group.addButton(self.radio_mapper_option_1)
        self.mapper_option_group.addButton(self.radio_mapper_option_2)

        layout.addWidget(self.radio_mapper_option_1)

        option1_row = QHBoxLayout()
        option1_row.setSpacing(12)
        self.btn_mapper_open_window = QPushButton("Open window")
        self.btn_mapper_open_window.setObjectName("collectorGrayBtn")
        option1_row.addWidget(self.btn_mapper_open_window, 0)
        option1_row.addStretch(1)
        layout.addLayout(option1_row)

        layout.addSpacing(8)

        sap_opt2_header = QHBoxLayout()
        sap_opt2_header.setSpacing(0)
        sap_opt2_header.addWidget(self.radio_mapper_option_2)
        sap_opt2_header.addSpacing(100)
        self.lbl_sap_input_count = QLabel("")
        self.lbl_sap_input_count.setObjectName("inputCountLabel")
        sap_opt2_header.addWidget(self.lbl_sap_input_count, 0)
        sap_opt2_header.addStretch(1)
        layout.addLayout(sap_opt2_header)

        sap_data_row = QHBoxLayout()
        sap_data_row.setSpacing(12)
        self.btn_mapper_sap_data_files = QPushButton("SAP Data Files")
        self.btn_mapper_sap_data_files.setObjectName("collectorGrayBtn")
        sap_data_row.addWidget(self.btn_mapper_sap_data_files, 0)
        self.input_mapper_sap_data_files = QLineEdit()
        self.input_mapper_sap_data_files.setObjectName("collectorLineEdit")
        self.input_mapper_sap_data_files.setReadOnly(True)
        sap_data_row.addWidget(self.input_mapper_sap_data_files, 1)
        layout.addLayout(sap_data_row)

        output_row = QHBoxLayout()
        output_row.setSpacing(12)
        self.btn_mapper_output_location = QPushButton("Output Location")
        self.btn_mapper_output_location.setObjectName("collectorGrayBtn")
        output_row.addWidget(self.btn_mapper_output_location, 0)
        self.input_mapper_output_location = QLineEdit()
        self.input_mapper_output_location.setObjectName("collectorLineEdit")
        self.input_mapper_output_location.setReadOnly(True)
        output_row.addWidget(self.input_mapper_output_location, 1)
        layout.addLayout(output_row)

        cleanup_radio_row = QHBoxLayout()
        cleanup_radio_row.setSpacing(16)

        self.mapper_cleanup_radio_group = QButtonGroup(page)
        self.mapper_cleanup_radio_group.setExclusive(True)

        self.radio_mapper_cleanup_1 = QRadioButton("Cleanup 1")
        self.radio_mapper_cleanup_1.setChecked(True)
        self.radio_mapper_cleanup_1.setObjectName("collectorCleanupRadio")
        self.radio_mapper_cleanup_1.setToolTip(
            "remove non-SMU and unneeded packaging:\n"
            "sal, flex, pall, accl, film, ship, wgl,\n"
            "sheet, shee, pl, t-secur, saco, acco_pe,\n"
            "tear, bulk"
        )
        self.mapper_cleanup_radio_group.addButton(self.radio_mapper_cleanup_1)
        cleanup_radio_row.addWidget(self.radio_mapper_cleanup_1, 0)

        self.radio_mapper_cleanup_2 = QRadioButton("Cleanup 2")
        self.radio_mapper_cleanup_2.setObjectName("collectorCleanupRadio")
        self.radio_mapper_cleanup_2.setToolTip(
            "remove non-SMU, un-identifiable basic name\n"
            "and unneeded packaging:\n"
            "sal, flex, pall, accl, film, ship, wgl,\n"
            "sheet, shee, pl, t-secur, saco, bag,\n"
            "rbosac, leaflet, acco, paco, tear, bulk"
        )
        self.mapper_cleanup_radio_group.addButton(self.radio_mapper_cleanup_2)
        cleanup_radio_row.addWidget(self.radio_mapper_cleanup_2, 0)

        self.radio_mapper_cleanup_3 = QRadioButton("Cleanup 3")
        self.radio_mapper_cleanup_3.setObjectName("collectorCleanupRadio")
        self.radio_mapper_cleanup_3.setToolTip("no cleanup, all info retained")
        self.mapper_cleanup_radio_group.addButton(self.radio_mapper_cleanup_3)
        cleanup_radio_row.addWidget(self.radio_mapper_cleanup_3, 0)

        cleanup_radio_row.addStretch(1)
        layout.addLayout(cleanup_radio_row)

        self.checkbox_mapper_include_grouping_report = QCheckBox("Include Grouping Report")
        self.checkbox_mapper_include_grouping_report.setObjectName("collectorCheck")
        self.checkbox_mapper_include_grouping_report.setChecked(True)
        layout.addWidget(self.checkbox_mapper_include_grouping_report)

        layout.addStretch(1)
        run_row = QHBoxLayout()
        run_row.addStretch(1)
        self.btn_run_process_mapper_reformat = QPushButton("Run Process")
        self.btn_run_process_mapper_reformat.setObjectName("collectorRunBtn")
        self.btn_run_process_mapper_reformat.setFixedHeight(50)
        self.btn_run_process_mapper_reformat.setMinimumWidth(180)
        run_row.addWidget(self.btn_run_process_mapper_reformat)
        run_row.addStretch(1)
        layout.addLayout(run_row)
        layout.addSpacing(8)

        progress_row = QHBoxLayout()
        progress_row.addStretch(1)
        self.mapper_reformat_progress_bar = QProgressBar()
        self.mapper_reformat_progress_bar.setObjectName("collectorProgressBar")
        self.mapper_reformat_progress_bar.setMinimumWidth(260)
        self.mapper_reformat_progress_bar.setMaximumWidth(360)
        self.mapper_reformat_progress_bar.setFixedHeight(14)
        self.mapper_reformat_progress_bar.setTextVisible(False)
        self.mapper_reformat_progress_bar.setRange(0, 100)
        self.mapper_reformat_progress_bar.setValue(0)
        self.mapper_reformat_progress_bar.setVisible(False)
        progress_row.addWidget(self.mapper_reformat_progress_bar)
        progress_row.addStretch(1)
        layout.addLayout(progress_row)

        self.radio_mapper_option_1.toggled.connect(self._sync_mapper_reformat_mode_ui)
        self.radio_mapper_option_2.toggled.connect(self._sync_mapper_reformat_mode_ui)
        self.btn_mapper_open_window.clicked.connect(self._open_mapper_option1_table)
        self.btn_mapper_sap_data_files.clicked.connect(self._on_mapper_select_sap_files)
        self.btn_mapper_output_location.clicked.connect(self._on_mapper_select_output_location)
        self.btn_run_process_mapper_reformat.clicked.connect(self._on_mapper_run_process)
        self.input_mapper_sap_data_files.textChanged.connect(lambda: self._update_sap_input_count())
        self._sync_mapper_reformat_mode_ui()

        return page

    # ------------------------------------------------------------------
    # SAP Data Compare page
    # ------------------------------------------------------------------

    def _create_mapper_compare_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(7)

        title = QLabel("SAP Data Compare")
        title.setObjectName("mapperTitle")
        title_row = QHBoxLayout()
        title_row.setContentsMargins(0, 0, 0, 0)
        title_row.addWidget(title, 0)
        title_row.addSpacing(140)
        self.lbl_sdc_input_count = QLabel("")
        self.lbl_sdc_input_count.setObjectName("inputCountLabel")
        title_row.addWidget(self.lbl_sdc_input_count, 0)
        title_row.addStretch(1)
        layout.addLayout(title_row)

        def _make_pair(btn_label: str):
            row = QHBoxLayout()
            row.setSpacing(8)
            btn = QPushButton(btn_label)
            btn.setObjectName("compareGrayBtn")
            btn.setFixedWidth(160)
            row.addWidget(btn, 0)
            lbl = QLineEdit()
            lbl.setObjectName("comparePathLabel")
            lbl.setFixedHeight(34)
            row.addWidget(lbl, 1)
            layout.addLayout(row)
            return btn, lbl

        self.btn_compare_rsd_target, self.lbl_compare_rsd_target = _make_pair("RSD: Target")

        layout.addSpacing(12)

        master_row = QHBoxLayout()
        master_row.setSpacing(0)
        self.checkbox_compare_with_master = QCheckBox("Compare with Master Data")
        self.checkbox_compare_with_master.setObjectName("collectorCheckSmall")
        self.checkbox_compare_with_master.setChecked(True)
        master_row.addWidget(self.checkbox_compare_with_master, 0)
        master_row.addSpacing(40)
        self.radio_compare_with_rsd_master = QRadioButton("Compare with RSD: Master")
        self.radio_compare_with_rsd_master.setObjectName("compareModeRadioSmall")
        self.radio_compare_with_rsd_master.setChecked(True)
        master_row.addWidget(self.radio_compare_with_rsd_master, 0)
        master_row.addSpacing(16)
        self.radio_compare_only_tsc = QRadioButton("Compare only with TSC")
        self.radio_compare_only_tsc.setObjectName("compareModeRadioSmall")
        master_row.addWidget(self.radio_compare_only_tsc, 0)
        master_row.addStretch(1)
        self.compare_master_mode_group = QButtonGroup(page)
        self.compare_master_mode_group.setExclusive(True)
        self.compare_master_mode_group.addButton(self.radio_compare_with_rsd_master)
        self.compare_master_mode_group.addButton(self.radio_compare_only_tsc)
        layout.addLayout(master_row)

        self.btn_compare_rsd_master, self.lbl_compare_rsd_master = _make_pair("RSD: Master")
        self.btn_compare_tsc_data, self.lbl_compare_tsc_data = _make_pair("TSC Data")

        # ── Compare with Library ──────────────────────────────────────────
        layout.addSpacing(10)
        self.checkbox_compare_with_library = QCheckBox("Compare with Library")
        self.checkbox_compare_with_library.setObjectName("collectorCheckSmall")
        self.checkbox_compare_with_library.setChecked(True)
        layout.addWidget(self.checkbox_compare_with_library)
        self.btn_compare_library, self.lbl_compare_library = _make_pair("Excel Library")

        # ── Run reference collector ───────────────────────────────────────
        layout.addSpacing(10)
        self.checkbox_compare_run_ref_collector = QCheckBox("Run reference collector")
        self.checkbox_compare_run_ref_collector.setObjectName("collectorCheckSmall")
        self.checkbox_compare_run_ref_collector.setChecked(True)
        layout.addWidget(self.checkbox_compare_run_ref_collector)
        self.btn_compare_packshot_location, self.lbl_compare_packshot_location = _make_pair("Packshot Image Library")

        max_packshot_row = QHBoxLayout()
        max_packshot_row.setSpacing(8)
        self.lbl_max_packshot = QLabel("Maximum count:")
        self.lbl_max_packshot.setObjectName("compareSmallLabel")
        max_packshot_row.addWidget(self.lbl_max_packshot, 0)
        self.input_compare_max_packshot = QLineEdit("5")
        self.input_compare_max_packshot.setObjectName("compareSmallInput")
        self.input_compare_max_packshot.setFixedWidth(50)
        self.input_compare_max_packshot.setFixedHeight(28)
        max_packshot_row.addWidget(self.input_compare_max_packshot, 0)
        max_packshot_row.addStretch(1)
        layout.addLayout(max_packshot_row)

        # ── Output Location ───────────────────────────────────────────────
        self.btn_compare_output_location, self.lbl_compare_output_location = _make_pair("Output Location")

        layout.addStretch(1)

        run_row = QHBoxLayout()
        run_row.addStretch(1)
        self.btn_run_process_mapper_compare = QPushButton("Run Process")
        self.btn_run_process_mapper_compare.setObjectName("compareRunBtn")
        self.btn_run_process_mapper_compare.setFixedHeight(35)
        self.btn_run_process_mapper_compare.setMinimumWidth(126)
        run_row.addWidget(self.btn_run_process_mapper_compare)
        run_row.addStretch(1)
        layout.addLayout(run_row)
        layout.addSpacing(4)

        progress_row = QHBoxLayout()
        progress_row.addStretch(1)
        self.mapper_compare_progress_bar = QProgressBar()
        self.mapper_compare_progress_bar.setObjectName("compareProgressBar")
        self.mapper_compare_progress_bar.setMinimumWidth(180)
        self.mapper_compare_progress_bar.setMaximumWidth(260)
        self.mapper_compare_progress_bar.setFixedHeight(10)
        self.mapper_compare_progress_bar.setTextVisible(False)
        self.mapper_compare_progress_bar.setRange(0, 100)
        self.mapper_compare_progress_bar.setValue(0)
        self.mapper_compare_progress_bar.setVisible(False)
        progress_row.addWidget(self.mapper_compare_progress_bar)
        progress_row.addStretch(1)
        layout.addLayout(progress_row)

        self.checkbox_compare_with_master.stateChanged.connect(self._sync_mapper_compare_mode_ui)
        self.radio_compare_with_rsd_master.toggled.connect(self._sync_mapper_compare_mode_ui)
        self.radio_compare_only_tsc.toggled.connect(self._sync_mapper_compare_mode_ui)
        self.checkbox_compare_with_library.stateChanged.connect(self._sync_mapper_compare_mode_ui)
        self.checkbox_compare_run_ref_collector.stateChanged.connect(self._sync_mapper_compare_mode_ui)
        self._sync_mapper_compare_mode_ui()

        self.lbl_compare_rsd_target.textChanged.connect(lambda: self._update_sdc_input_count())
        self.btn_compare_rsd_target.clicked.connect(self._on_compare_browse_rsd_target)
        self.btn_compare_rsd_master.clicked.connect(self._on_compare_browse_rsd_master)
        self.btn_compare_tsc_data.clicked.connect(self._on_compare_browse_tsc_data)
        self.btn_compare_library.clicked.connect(self._on_compare_browse_library)
        self.btn_compare_packshot_location.clicked.connect(self._on_compare_browse_packshot_location)
        self.btn_compare_output_location.clicked.connect(self._on_compare_browse_output_location)
        self.btn_run_process_mapper_compare.clicked.connect(self._on_mapper_compare_run)

        return page

    def _on_mapper_compare_run(self) -> None:
        """Collect UI inputs and run the SDC comparison via body_mapper."""
        from pathlib import Path as _Path

        # ── gather RSD: Target paths ─────────────────────────────────────
        raw_target = self.lbl_compare_rsd_target.text().strip()
        rsd_target_paths = [p.strip() for p in raw_target.split(",") if p.strip()] if raw_target else []

        rsd_master_path = self.lbl_compare_rsd_master.text().strip()
        tsc_data_path   = self.lbl_compare_tsc_data.text().strip()
        output_dir      = self.lbl_compare_output_location.text().strip()

        # ── library validation ───────────────────────────────────────────
        _LIB_ERROR_STYLE = (
            "background-color: rgba(208, 39, 82, 128); color: #000000; "
            "border: 1px solid #6F6F6F; border-radius: 11px; "
            "padding: 0 8px; font-family: 'Segoe UI'; font-size: 10px;"
        )
        excel_library_path = ""
        if self.checkbox_compare_with_library.isChecked():
            lib_text = self.lbl_compare_library.text().strip()
            lib_paths = [p.strip() for p in lib_text.split(",") if p.strip()]
            lib_is_single_file = (
                len(lib_paths) == 1
                and lib_paths[0].lower().endswith(".xlsx")
                and _Path(lib_paths[0]).is_file()
            )
            if len(lib_paths) > 1 or (lib_text and not lib_is_single_file and "multiple" in lib_text.lower()):
                self.lbl_compare_library.setStyleSheet(_LIB_ERROR_STYLE)
                self.lbl_compare_library.setText("multiple excel files in library folder")
                return
            elif not lib_is_single_file:
                # no valid library file – warn but continue without library
                pass
            else:
                excel_library_path = lib_paths[0]

        # ── basic validation ─────────────────────────────────────────────
        missing: list[str] = []
        compare_mode = (
            "tsc_only"
            if self.radio_compare_only_tsc.isChecked()
            else "rsd_master"
        )
        if not rsd_target_paths:
            missing.append("RSD: Target")
        if compare_mode == "rsd_master" and not rsd_master_path:
            missing.append("RSD: Master")
        if not tsc_data_path:
            missing.append("TSC Data")
        if not output_dir:
            missing.append("Output Location")
        if missing:
            QMessageBox.warning(
                self, "Missing Input",
                "Please fill in the following fields:\n" + "\n".join(f"  • {m}" for m in missing)
            )
            return

        # ── run ──────────────────────────────────────────────────────────
        self.mapper_compare_progress_bar.setValue(0)
        self.mapper_compare_progress_bar.setVisible(True)
        self.btn_run_process_mapper_compare.setEnabled(False)
        QApplication.processEvents()

        params = CompareParams(
            rsd_target_paths=rsd_target_paths,
            rsd_master_path=rsd_master_path,
            tsc_data_path=tsc_data_path,
            output_dir=output_dir,
            excel_library_path=excel_library_path,
            compare_mode=compare_mode,
        )

        self.mapper_compare_progress_bar.setValue(30)
        QApplication.processEvents()

        compare_result = run_comparison(params)

        # ── reference collector (optional) ──────────────────────────────────
        rc_result = None
        if (
            self.checkbox_compare_run_ref_collector.isChecked()
            and compare_result.output_paths
        ):
            packshot_folder = self.lbl_compare_packshot_location.text().strip()
            if not packshot_folder or not Path(packshot_folder).is_dir():
                compare_result.warnings.append(
                    "Reference collector skipped: Packshot Image Library folder "
                    "is missing or invalid."
                )
            else:
                try:
                    max_imgs = int(self.input_compare_max_packshot.text().strip())
                except ValueError:
                    max_imgs = 5

                # Determine which SDC files to process
                sdc_paths_for_rc: list[str] = []
                if len(compare_result.output_paths) == 1:
                    sdc_paths_for_rc = compare_result.output_paths
                else:
                    dlg = _SdcSelectionDialog(compare_result.output_paths, parent=self)
                    if dlg.exec() == QDialog.Accepted and dlg.selected_paths:
                        sdc_paths_for_rc = dlg.selected_paths
                    # If cancelled or nothing checked → skip RC silently

                if sdc_paths_for_rc:
                    rc_params = RefCollectorParams(
                        sdc_output_paths   = sdc_paths_for_rc,
                        image_library_path = packshot_folder,
                        output_dir         = output_dir,
                        tsc_data_path      = tsc_data_path,
                        max_images         = max(1, max_imgs),
                    )
                    rc_result = run_reference_collector(rc_params)
                    compare_result.warnings.extend(rc_result.warnings)

        self.mapper_compare_progress_bar.setValue(100)
        QApplication.processEvents()
        self.mapper_compare_progress_bar.setVisible(False)
        self.btn_run_process_mapper_compare.setEnabled(True)

        if compare_result.warnings:
            warn_text = "\n".join(compare_result.warnings)
            if not compare_result.output_paths:
                QMessageBox.critical(self, "SDC Error", warn_text)
                return
            QMessageBox.warning(self, "SDC completed with warnings", warn_text)

        if compare_result.output_paths:
            paths_text = "\n".join(compare_result.output_paths)
            rc_info = ""
            if rc_result and rc_result.pdf_paths:
                folders_text = "\n".join(rc_result.output_folders)
                rc_info = (
                    f"\n\nReference Collector PDFs ({len(rc_result.pdf_paths)}) "
                    f"in {len(rc_result.output_folders)} folder(s):"
                    f"\n{folders_text}"
                )
            elif rc_result and not rc_result.pdf_paths:
                rc_info = "\n\nReference Collector: no PDFs generated."
            QMessageBox.information(
                self, "SDC Complete",
                f"Output saved to:\n{paths_text}{rc_info}"
            )

    def _on_compare_browse_rsd_target(self) -> None:
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Select RSD: Target File(s)",
            self._get_browse_dir("sap_compare"),
            "Excel Files (*.xlsx)")
        if paths:
            self.lbl_compare_rsd_target.setText(", ".join(paths))

    def _on_compare_browse_rsd_master(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Select RSD: Master File",
            self._get_browse_dir("sap_compare"),
            "Excel Files (*.xlsx *.xls)")
        if path:
            self.lbl_compare_rsd_master.setStyleSheet("")
            self.lbl_compare_rsd_master.setText(path)

    def _on_compare_browse_tsc_data(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Select TSC Data File",
            self._get_browse_dir("sap_compare"),
            "Excel Files (*.xlsx *.xls)")
        if path:
            self.lbl_compare_tsc_data.setText(path)

    def _on_compare_browse_library(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel Library File",
            self._get_browse_dir("excel_library"),
            "Excel Files (*.xlsx)")
        if path:
            self.lbl_compare_library.setText(path)

    def _on_compare_browse_packshot_location(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self, "Select Packshot Location",
            self._get_browse_dir("sap_compare"))
        if folder:
            self.lbl_compare_packshot_location.setText(folder)

    def _on_compare_browse_output_location(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self, "Select Output Location",
            self._get_browse_dir("sap_compare"))
        if folder:
            self.lbl_compare_output_location.setText(folder)

    def _on_root_folders_toggle(self, state: bool) -> None:
        """Validate config, update shared flag, then auto-fill or clear inputs."""
        if state and not self._config.is_root_configured():
            self._use_root_folders = False
            if self.btn_rail_toggle.isChecked():
                self.btn_rail_toggle.setChecked(False)
            QMessageBox.warning(
                self,
                "Root folders not configured",
                "No root folders configured.\n\nGo to Settings (\u2699) and set up root folders first."
            )
            return

        self._use_root_folders = state

        if not state:
            self._clear_root_folder_inputs()
            return

        self._apply_root_folder_inputs()

    def _clear_root_folder_inputs(self) -> None:
        """Clear all fields that were auto-filled by the root folder toggle."""
        for field in (
            self.input_trackers,
            self.input_output,
            self.input_pg7_images_folder,
            self.input_pg7_output,
            self.input_pg6_excel_tracker,
            self.input_pg6_output,
            self.input_mapper_sap_data_files,
            self.input_mapper_output_location,
        ):
            field.clear()
        self._update_tsc_input_count()
        self._update_png_input_count()
        self._update_sap_input_count()
        for lbl in (
            self.lbl_compare_rsd_target,
            self.lbl_compare_rsd_master,
            self.lbl_compare_tsc_data,
            self.lbl_compare_library,
            self.lbl_compare_packshot_location,
            self.lbl_compare_output_location,
        ):
            lbl.setText("")
            lbl.setStyleSheet("")

    def _apply_root_folder_inputs(self) -> None:
        """Route autofill to MASTER or PROJECT depending on the active nav page."""
        is_project = (self.left_nav_stack.currentWidget() is self.project_left)
        if is_project:
            self._autofill_project_page()
        else:
            self._autofill_master_page()

    def _autofill_master_page(self) -> None:
        """Fill MASTER tool input fields from configured root folder paths."""
        from pathlib import Path

        # ── TSC ───────────────────────────────────────────────────────────
        tsc_folder = self._config.tsc_input()
        if tsc_folder:
            p = Path(tsc_folder)
            if p.is_dir():
                files = sorted(
                    f for f in p.glob("*.xlsx")
                    if "tsc" not in f.stem.lower()
                )
                if files:
                    file_strs = [str(f) for f in files]
                    self.input_trackers.setText(", ".join(file_strs))
                    if hasattr(self, "status_collector"):
                        sc = self.status_collector
                        sc.excel_trackers = file_strs
                        sc.excel_tracker_names = [f.name for f in files]
                        sc.editText_selected_trackers.setPlainText(
                            ",\n".join(f.name for f in files)
                        )
            self.input_output.setText(tsc_folder)
            if hasattr(self, "status_collector"):
                self.status_collector.output_location = tsc_folder
                self.status_collector.editText_output_location.setPlainText(tsc_folder)
        self._update_tsc_input_count()

        # ── Thumbnail Generator ───────────────────────────────────────────
        tg_folder = self._config.thumbnail_input()
        if tg_folder:
            self.input_pg7_images_folder.setText(tg_folder)
            self.input_pg7_output.setText(tg_folder)

        # ── SAP Data Reformat (MASTER) ────────────────────────────────────
        sdr_folder = self._config.sap_reformat_input()
        if sdr_folder:
            p = Path(sdr_folder)
            if p.is_dir():
                files = sorted(
                    f for f in p.iterdir()
                    if f.suffix.lower() == ".xls"
                    and "rsd" not in f.stem.lower()
                )
                if files:
                    file_strs = [str(f) for f in files]
                    self.input_mapper_sap_data_files.setText(", ".join(file_strs))
                    self._mapper_sap_file_paths = file_strs
            self.input_mapper_output_location.setText(sdr_folder)
            self._mapper_output_location = sdr_folder
        self._update_sap_input_count()

    def _autofill_project_page(self) -> None:
        """Fill PROJECT tool input fields based on the selected project folder."""
        from pathlib import Path
        project_folder = self._get_active_project_folder()
        if not project_folder:
            return

        # ── Packshot Naming Generator ──────────────────────────────────────
        png_folder = str(Path(project_folder) / "Packshot Naming Generator")
        p_png = Path(png_folder)
        if p_png.is_dir():
            files = sorted(
                f for f in p_png.glob("*.xlsx")
                if "packshot_naming" not in f.stem.lower()
            )
            if files:
                first = files[0]
                self.input_pg6_excel_tracker.setText(str(first))
                if hasattr(self, "packshot_naming_generator"):
                    png = self.packshot_naming_generator
                    png.excel_values.out_excel_file = str(first)
                    png.excel_values.out_excel_file_name = first.name
                    png.excel_values.check_tracker_loaded = True
            self.input_pg6_output.setText(png_folder)
        self._update_png_input_count()

        # ── SAP Data Reformat (PROJECT) ───────────────────────────────────
        sdr_folder = str(Path(project_folder) / "SAP Data Reformat")
        p_sdr = Path(sdr_folder)
        if p_sdr.is_dir():
            files = sorted(
                f for f in p_sdr.iterdir()
                if f.suffix.lower() == ".xls"
                and "rsd" not in f.stem.lower()
            )
            if files:
                file_strs = [str(f) for f in files]
                self.input_mapper_sap_data_files.setText(", ".join(file_strs))
                self._mapper_sap_file_paths = file_strs
            self.input_mapper_output_location.setText(sdr_folder)
            self._mapper_output_location = sdr_folder
        self._update_sap_input_count()

        # ── SAP Data Compare (PROJECT) ────────────────────────────────────
        # RSD: Target — project SAP Data Reformat folder, all .xlsx files with "rsd" in name
        p_rsd_target = Path(sdr_folder)
        if p_rsd_target.is_dir():
            rsd_target_files = sorted(
                f for f in p_rsd_target.iterdir()
                if f.suffix.lower() == ".xlsx"
                and "rsd" in f.stem.lower()
            )
            if rsd_target_files:
                self.lbl_compare_rsd_target.setText(", ".join(str(f) for f in rsd_target_files))

        # RSD: Master — [sap_compare] rsd_master_folder → sap_reformat_input() → Desktop default
        master_rsd_folder = self._config.rsd_master_folder()
        if not master_rsd_folder:
            master_rsd_folder = str(Path.home() / "Desktop" / "HAT DASHBOARD ROOT" / "MASTER" / "SAP Data Reformat")
        if master_rsd_folder:
            p_master_rsd = Path(master_rsd_folder)
            if p_master_rsd.is_dir():
                master_rsd_files = sorted(
                    f for f in p_master_rsd.iterdir()
                    if f.suffix.lower() == ".xlsx"
                    and "rsd" in f.stem.lower()
                )
                if len(master_rsd_files) == 1:
                    self.lbl_compare_rsd_master.setStyleSheet("")
                    self.lbl_compare_rsd_master.setText(str(master_rsd_files[0]))
                elif len(master_rsd_files) > 1:
                    self.lbl_compare_rsd_master.setStyleSheet(
                        "background-color: rgba(208, 39, 82, 128); color: #111111; "
                        "border: 1px solid #6F6F6F; border-radius: 11px; "
                        "padding: 0 8px; font-family: 'Segoe UI'; font-size: 10px;"
                    )
                    self.lbl_compare_rsd_master.setText("Multiple rsd files found in master")

        # TSC Data — MASTER Tracker Status Collector folder, single .xlsx file with "tsc" in name
        tsc_folder = self._config.tsc_input()
        if tsc_folder:
            p_tsc = Path(tsc_folder)
            if p_tsc.is_dir():
                tsc_files = sorted(
                    f for f in p_tsc.iterdir()
                    if f.suffix.lower() == ".xlsx"
                    and "tsc" in f.stem.lower()
                )
                if len(tsc_files) == 1:
                    self.lbl_compare_tsc_data.setStyleSheet("")
                    self.lbl_compare_tsc_data.setText(str(tsc_files[0]))
                elif len(tsc_files) > 1:
                    self.lbl_compare_tsc_data.setStyleSheet("color: rgba(208, 39, 82, 128);")
                    self.lbl_compare_tsc_data.setText("Multiple tsc files found in Master")

        # Excel Library — INI override or MASTER/Excel Library folder
        excel_lib_folder = self._config.excel_library_folder()
        if excel_lib_folder:
            p_lib = Path(excel_lib_folder)
            if p_lib.is_dir():
                lib_files = sorted(f for f in p_lib.glob("*.xlsx"))
                if len(lib_files) == 1:
                    self.lbl_compare_library.setStyleSheet("")
                    self.lbl_compare_library.setText(str(lib_files[0]))
                elif len(lib_files) > 1:
                    self.lbl_compare_library.setStyleSheet(
                        "background-color: rgba(208, 39, 82, 128); color: #000000; "
                        "border: 1px solid #6F6F6F; border-radius: 11px; "
                        "padding: 0 8px; font-family: 'Segoe UI'; font-size: 10px;"
                    )
                    self.lbl_compare_library.setText("multiple excel files in library folder")

        # Packshot Image Library — MASTER Thumbnail Generator folder
        tg_folder = self._config.thumbnail_input()
        if tg_folder and Path(tg_folder).is_dir():
            self.lbl_compare_packshot_location.setText(tg_folder)

        # Output Location — project SAP Data Compare folder
        sdc_out = str(Path(project_folder) / "SAP Data Compare")
        if Path(sdc_out).is_dir():
            self.lbl_compare_output_location.setText(sdc_out)

    def _get_active_project_folder(self) -> str:
        """Return full path to the currently selected project folder, or '' if unavailable."""
        from pathlib import Path
        root = self._config.root_folder()
        if not root:
            return ""
        project_name = self.combo_project.currentText().strip()
        if not project_name or project_name.startswith("("):
            return ""
        p = Path(root) / "HAT DASHBOARD ROOT" / project_name
        return str(p) if p.is_dir() else ""

    def _update_sdc_input_count(self) -> None:
        text = self.lbl_compare_rsd_target.text().strip()
        if text:
            count = len([p for p in text.split(",") if p.strip()])
        else:
            count = 0
        lbl = getattr(self, "lbl_sdc_input_count", None)
        if lbl is not None:
            lbl.setText(f"Input files count: {count}" if count > 0 else "")

    def _update_tsc_input_count(self) -> None:
        text = self.input_trackers.text().strip()
        if text:
            count = len([p for p in text.split(",") if p.strip()])
        else:
            count = 0
        lbl = getattr(self, "lbl_tsc_input_count", None)
        if lbl is not None:
            lbl.setText(f"Input files count: {count}" if count > 0 else "")

    def _update_png_input_count(self) -> None:
        text = self.input_pg6_excel_tracker.text().strip()
        count = 1 if text else 0
        lbl = getattr(self, "lbl_png_input_count", None)
        if lbl is not None:
            lbl.setText(f"Input files count: {count}" if count > 0 else "")

    def _update_sap_input_count(self) -> None:
        text = self.input_mapper_sap_data_files.text().strip()
        if text:
            count = len([p for p in text.split(",") if p.strip()])
        else:
            count = 0
        lbl = getattr(self, "lbl_sap_input_count", None)
        if lbl is not None:
            lbl.setText(f"Input files count: {count}" if count > 0 else "")

    def _sync_mapper_compare_mode_ui(self) -> None:
        on = self.checkbox_compare_with_master.isChecked()
        rsd_mode = self.radio_compare_with_rsd_master.isChecked()
        self.radio_compare_with_rsd_master.setEnabled(on)
        self.radio_compare_only_tsc.setEnabled(on)
        self.btn_compare_rsd_master.setEnabled(on and rsd_mode)
        self.lbl_compare_rsd_master.setEnabled(on and rsd_mode)
        self.btn_compare_tsc_data.setEnabled(on)
        self.lbl_compare_tsc_data.setEnabled(on)

        lib_on = self.checkbox_compare_with_library.isChecked()
        self.btn_compare_library.setEnabled(lib_on)
        self.lbl_compare_library.setEnabled(lib_on)

        ref_on = self.checkbox_compare_run_ref_collector.isChecked()
        self.btn_compare_packshot_location.setEnabled(ref_on)
        self.lbl_compare_packshot_location.setEnabled(ref_on)
        self.lbl_max_packshot.setEnabled(ref_on)
        self.input_compare_max_packshot.setEnabled(ref_on)

    def _sync_mapper_reformat_mode_ui(self) -> None:
        is_option_1 = self.radio_mapper_option_1.isChecked()

        self.btn_mapper_open_window.setEnabled(is_option_1)

        self.btn_mapper_sap_data_files.setEnabled(not is_option_1)
        self.input_mapper_sap_data_files.setEnabled(not is_option_1)
        self.btn_mapper_output_location.setEnabled(not is_option_1)
        self.input_mapper_output_location.setEnabled(not is_option_1)

        self.radio_mapper_cleanup_1.setEnabled(not is_option_1)
        self.radio_mapper_cleanup_2.setEnabled(not is_option_1)
        self.radio_mapper_cleanup_3.setEnabled(not is_option_1)
        self.checkbox_mapper_include_grouping_report.setEnabled(not is_option_1)

        self.btn_run_process_mapper_reformat.setEnabled(not is_option_1)

        if is_option_1:
            self.input_mapper_sap_data_files.clear()
            self.input_mapper_output_location.clear()
            self.radio_mapper_cleanup_1.setChecked(True)
            self.checkbox_mapper_include_grouping_report.setChecked(True)

    # ------------------------------------------------------------------
    # Option 2 handlers
    # ------------------------------------------------------------------

    def _on_mapper_select_sap_files(self) -> None:
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Select SAP Data Files",
            self._get_browse_dir("sap_reformat"),
            "Excel Files (*.xls)",
        )
        if not file_paths:
            return
        self._mapper_sap_file_paths = file_paths
        names = [Path(p).name for p in file_paths]
        self.input_mapper_sap_data_files.setText(", ".join(names))

    def _on_mapper_select_output_location(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self, "Select Output Folder", self._get_browse_dir("sap_reformat"))
        if not folder:
            return
        self._mapper_output_location = folder
        self.input_mapper_output_location.setText(folder)

    # ---- helpers reused from the SAP-table dialog (static-compatible) ----

    @staticmethod
    def _opt2_normalize_header(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip().lower()
        text = text.replace("\n", " ").replace("\r", " ")
        text = re.sub(r"[^a-z0-9]+", " ", text)
        return " ".join(text.split())

    @staticmethod
    def _opt2_cell_to_text(value: object) -> str:
        if pd.isna(value):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _opt2_read_excel_raw(self, file_path: str) -> pd.DataFrame:
        """Read a raw SAP .xls / .xlsx file into a headerless DataFrame."""
        lower = file_path.lower()

        if lower.endswith(".xls"):
            # Try openpyxl first (in case the .xls is actually xlsx)
            try:
                ef = pd.ExcelFile(file_path, engine="openpyxl")
                if ef.sheet_names:
                    return pd.read_excel(ef, sheet_name=ef.sheet_names[0], header=None, dtype=object)
            except Exception:
                pass

            # Try xlrd (genuine BIFF .xls)
            try:
                ef = pd.ExcelFile(file_path, engine="xlrd")
                if ef.sheet_names:
                    return pd.read_excel(ef, sheet_name=ef.sheet_names[0], header=None, dtype=object)
            except Exception:
                pass

            # Fallback – text-style SAP export
            text_df = self._opt2_read_text_style_xls(file_path)
            if text_df is not None:
                return text_df

            raise ValueError(f"Unable to read .xls file: {Path(file_path).name}")
        else:
            for engine in ("openpyxl", None):
                try:
                    ef = pd.ExcelFile(file_path, engine=engine)
                    if ef.sheet_names:
                        return pd.read_excel(ef, sheet_name=ef.sheet_names[0], header=None, dtype=object)
                except Exception:
                    pass
            raise ValueError(f"Unable to read file: {Path(file_path).name}")

    def _opt2_read_text_style_xls(self, file_path: str) -> pd.DataFrame | None:
        import csv as _csv
        parse_configs = [
            {"encoding": "utf-16", "sep": "\t", "skiprows": 3},
            {"encoding": "utf-16", "sep": r"\s{2,}", "skiprows": 3},
            {"encoding": "utf-16le", "sep": "\t", "skiprows": 3},
            {"encoding": "utf-8-sig", "sep": "\t", "skiprows": 3},
            {"encoding": "cp1252", "sep": "\t", "skiprows": 3},
            {"encoding": "utf-16", "sep": "\t", "skiprows": 0},
        ]
        for cfg in parse_configs:
            try:
                df = pd.read_csv(
                    file_path, header=None, dtype=object, engine="python",
                    on_bad_lines="skip", quotechar='"', quoting=_csv.QUOTE_MINIMAL, **cfg,
                )
            except Exception:
                continue
            if df.empty:
                continue
            df = df.dropna(how="all").reset_index(drop=True)
            if df.empty or df.shape[1] < 4:
                continue
            return df
        return None

    def _opt2_detect_headers(self, raw_df: pd.DataFrame) -> tuple[int, dict[str, int]]:
        target_columns: dict[str, list[str]] = {
            "Head Bom Mat": ["head bom mat", "head bom", "headbommat"],
            "BOM COMPONENT": ["bom component", "bom comp", "bomcomponent"],
            "Sort String": ["sort string", "sortstring"],
            "Component Desc": ["component desc", "component description", "component"],
            "Basic Number": ["basic number", "basic num", "basic no", "basic"],
            "Basic Name": ["basic name", "basicname"],
        }
        normalized_targets = {
            t: {self._opt2_normalize_header(a) for a in aliases}
            for t, aliases in target_columns.items()
        }

        max_scan = min(120, raw_df.shape[0])
        for row_idx in range(max_scan):
            row_norm = [self._opt2_normalize_header(v) for v in raw_df.iloc[row_idx].tolist()]
            found: dict[str, int] = {}
            for tname, aliases in normalized_targets.items():
                for ci, cv in enumerate(row_norm):
                    if cv in aliases:
                        found[tname] = ci
                        break
            if len(found) == len(target_columns):
                return row_idx, found

        raise ValueError("Unable to detect required SAP headers in file.")

    def _opt2_extract_rows(self, raw_df: pd.DataFrame, header_row: int, col_map: dict[str, int]) -> list[list[str]]:
        ordered = ["Head Bom Mat", "BOM COMPONENT", "Sort String", "Component Desc", "Basic Number", "Basic Name"]
        out: list[list[str]] = []
        for ri in range(header_row + 1, raw_df.shape[0]):
            vals: list[str] = []
            has = False
            for h in ordered:
                ci = col_map[h]
                v = raw_df.iat[ri, ci] if ci < raw_df.shape[1] else ""
                t = self._opt2_cell_to_text(v)
                if t:
                    has = True
                vals.append(t)
            if has:
                out.append(vals)
        return out

    @staticmethod
    def _opt2_build_grouping_data(reformatted_rows: list[list[str]]) -> list[tuple[str, int, str, str]]:
        """Build grouping counts from reformatted rows (same columns as the dialog table)."""
        # Reformatted row layout: [Head Bom Mat, HSI, BOM COMPONENT, Component Desc,
        #                          Basic Number, BC, Basic Name]
        combination_col = 5  # BC
        basic_number_col = 4
        basic_name_col = 6

        counts: dict[str, int] = {}
        bn_map: dict[str, dict[str, int]] = {}
        bname_map: dict[str, dict[str, int]] = {}

        for row in reformatted_rows:
            comb = row[combination_col].strip() if len(row) > combination_col else ""
            if not comb:
                continue
            counts[comb] = counts.get(comb, 0) + 1

            bn = row[basic_number_col].strip() if len(row) > basic_number_col else ""
            bn_map.setdefault(comb, {})
            bn_map[comb][bn] = bn_map[comb].get(bn, 0) + 1

            bname = row[basic_name_col].strip() if len(row) > basic_name_col else ""
            bname_map.setdefault(comb, {})
            bname_map[comb][bname] = bname_map[comb].get(bname, 0) + 1

        grouped: list[tuple[str, int, str, str]] = []
        for comb, cnt in counts.items():
            best_bn = ""
            if comb in bn_map and bn_map[comb]:
                best_bn = max(bn_map[comb].items(), key=lambda p: (p[1], p[0]))[0]
            best_name = ""
            if comb in bname_map and bname_map[comb]:
                best_name = max(bname_map[comb].items(), key=lambda p: (p[1], p[0]))[0]
            grouped.append((comb, cnt, best_bn, best_name))

        match = re.search  # local ref
        def sort_key(r):
            m = re.search(r"(\d+)", r[0])
            return (-r[1], int(m.group(1)) if m else 10**9, r[0])

        return sorted(grouped, key=sort_key)

    @staticmethod
    def _opt2_style_worksheet(ws, accent_color: str = "8A244B") -> None:
        """Apply the standard modern styling to a worksheet."""
        accent_fill = PatternFill(fill_type="solid", fgColor=accent_color)
        white_font = Font(color="FFFFFF", bold=True)
        normal_font = Font(color="111111")
        border = Border(
            left=Side(style="thin", color="B8B8B8"),
            right=Side(style="thin", color="B8B8B8"),
            top=Side(style="thin", color="B8B8B8"),
            bottom=Side(style="thin", color="B8B8B8"),
        )
        max_row = ws.max_row
        max_col = ws.max_column

        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = accent_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                c = ws.cell(row=row, column=col)
                c.font = normal_font
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = border

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=max_col).coordinate}"

        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            mx = 0
            for cell in col_cells:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > mx:
                    mx = len(v)
            ws.column_dimensions[letter].width = min(max(mx + 2, 12), 56)

    def _on_mapper_run_process(self) -> None:
        # Validate inputs
        file_paths = getattr(self, "_mapper_sap_file_paths", None) or []
        if not file_paths:
            QMessageBox.warning(self, "Missing Input", "Please select SAP data files first.")
            return
        output_dir = getattr(self, "_mapper_output_location", None) or ""
        if not output_dir:
            QMessageBox.warning(self, "Missing Input", "Please select an output location first.")
            return

        # Determine cleanup mode
        cleanup_mode = 1
        if self.radio_mapper_cleanup_2.isChecked():
            cleanup_mode = 2
        elif self.radio_mapper_cleanup_3.isChecked():
            cleanup_mode = 3

        include_grouping = self.checkbox_mapper_include_grouping_report.isChecked()
        reformatter = SapTableReformatter()
        timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        errors: list[str] = []
        success_count = 0
        total_files = len(file_paths)

        # Show and reset progress bar
        self.mapper_reformat_progress_bar.setValue(0)
        self.mapper_reformat_progress_bar.setVisible(True)
        self.btn_run_process_mapper_reformat.setEnabled(False)
        QApplication.processEvents()

        for file_idx, file_path in enumerate(file_paths):
            fname = Path(file_path).stem
            try:
                # 1. Read raw file
                self.mapper_reformat_progress_bar.setValue(int((file_idx / total_files) * 30))
                QApplication.processEvents()
                raw_df = self._opt2_read_excel_raw(file_path)

                # 2. Detect headers & extract rows
                self.mapper_reformat_progress_bar.setValue(int((file_idx / total_files) * 100 + 30 / total_files))
                QApplication.processEvents()
                header_row, col_map = self._opt2_detect_headers(raw_df)
                extracted_rows = self._opt2_extract_rows(raw_df, header_row, col_map)
                if not extracted_rows:
                    errors.append(f"{Path(file_path).name}: No data rows found.")
                    self.mapper_reformat_progress_bar.setValue(int((file_idx + 1) / total_files * 100))
                    QApplication.processEvents()
                    continue

                # 3. Reformat
                self.mapper_reformat_progress_bar.setValue(int((file_idx / total_files) * 100 + 50 / total_files))
                QApplication.processEvents()
                reformatted_rows, basic_comb_count = reformatter.reformat_from_rows(extracted_rows, cleanup_mode)

                # 4. Build workbook
                wb = Workbook()
                ws_reformat = wb.active
                ws_reformat.title = "reformatted_sap"

                reformat_header = [
                    "Head Bom Mat", "HSI", "BOM COMPONENT", "Component Desc",
                    "Basic Number", "BC", "Basic Name",
                ]
                ws_reformat.append(reformat_header)
                for row_vals in reformatted_rows:
                    ws_reformat.append(row_vals)

                self._opt2_style_worksheet(ws_reformat)

                # 5. Grouping report sheet (optional)
                if include_grouping:
                    grouping_data = self._opt2_build_grouping_data(reformatted_rows)
                    ws_grouping = wb.create_sheet("grouping_data")
                    ws_grouping.append(["BC", "Count", "Basic Number", "Basic Name"])
                    for comb_name, count, bn, bname in grouping_data:
                        ws_grouping.append([comb_name, count, bn, bname])
                    self._opt2_style_worksheet(ws_grouping)

                # 6. Save
                self.mapper_reformat_progress_bar.setValue(int((file_idx / total_files) * 100 + 90 / total_files))
                QApplication.processEvents()
                cu_tag = f"cu{cleanup_mode}"
                out_name = f"rsd_{fname}_{cu_tag}_{timestamp}.xlsx"
                out_path = out_dir / out_name
                wb.save(str(out_path))
                success_count += 1

            except Exception as exc:
                errors.append(f"{Path(file_path).name}: {exc}")

            self.mapper_reformat_progress_bar.setValue(int((file_idx + 1) / total_files * 100))
            QApplication.processEvents()

        self.mapper_reformat_progress_bar.setVisible(False)
        self.btn_run_process_mapper_reformat.setEnabled(True)

        # Summary message
        parts: list[str] = []
        if success_count:
            parts.append(f"Successfully processed {success_count} file(s).")
        if errors:
            parts.append("Errors:\n" + "\n".join(errors))

        msg = QMessageBox(self)
        msg.setWindowTitle("Process Complete" if success_count else "Error")
        msg.setIcon(QMessageBox.Icon.Information if success_count else QMessageBox.Icon.Warning)
        msg.setText("\n\n".join(parts))
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()

    def _create_other_tools_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(14)

        title = QLabel("Other Tools")
        title.setObjectName("collectorTitle")
        title_row = QHBoxLayout()
        title_row.setContentsMargins(0, 0, 0, 0)
        title_row.addWidget(title, 1)
        layout.addLayout(title_row)

        # ── Duplicate Check ───────────────────────────────────────────────────
        section_label = QLabel("Duplicate Check")
        section_label.setObjectName("collectorSectionLabel")
        layout.addWidget(section_label)
        layout.addSpacing(-8)

        self.radio_idh_dup_single = QRadioButton("Single Tracker")
        self.radio_idh_dup_single.setObjectName("idhDupRadio")
        self.radio_idh_dup_single.setChecked(True)
        self.radio_idh_dup_multiple = QRadioButton("Multiple Trackers")
        self.radio_idh_dup_multiple.setObjectName("idhDupRadio")

        self.idh_dup_mode_group = QButtonGroup(page)
        self.idh_dup_mode_group.setExclusive(True)
        self.idh_dup_mode_group.addButton(self.radio_idh_dup_single)
        self.idh_dup_mode_group.addButton(self.radio_idh_dup_multiple)

        self.btn_idh_dup_open_window = QPushButton("Open Window")
        self.btn_idh_dup_open_window.setObjectName("collectorGrayBtn")

        controls_row = QHBoxLayout()
        controls_row.setSpacing(12)
        controls_row.addWidget(self.radio_idh_dup_single)
        controls_row.addWidget(self.radio_idh_dup_multiple)
        controls_row.addWidget(self.btn_idh_dup_open_window)
        controls_row.addStretch(1)
        layout.addLayout(controls_row)

        self.btn_idh_dup_open_window.clicked.connect(self._open_idh_dup_window)

        layout.addStretch(1)
        return page

    def _create_packshot_naming_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(16)

        title = QLabel("Packshot Naming Generator")
        title.setObjectName("packshotTitle")
        title_row = QHBoxLayout()
        title_row.setContentsMargins(0, 0, 0, 0)
        title_row.addWidget(title, 1)
        layout.addLayout(title_row)

        description = QLabel("NOTE: All fields must have input.")
        description.setObjectName("packshotDescription")
        layout.addWidget(description)
        layout.addSpacing(12)

        self.radioButton_pnm_from_clipboard = QRadioButton("From clipboard")
        self.radioButton_pnm_from_clipboard.setObjectName("collectorModeRadio")
        self.radioButton_pnm_generate_from_tracker = QRadioButton("From tracker")
        self.radioButton_pnm_generate_from_tracker.setObjectName("collectorModeRadio")
        self.radioButton_pnm_from_clipboard.setChecked(True)

        self.packshot_mode_group = QButtonGroup(page)
        self.packshot_mode_group.setExclusive(True)
        self.packshot_mode_group.addButton(self.radioButton_pnm_from_clipboard)
        self.packshot_mode_group.addButton(self.radioButton_pnm_generate_from_tracker)

        layout.addWidget(self.radioButton_pnm_from_clipboard)

        clipboard_row = QHBoxLayout()
        clipboard_row.setSpacing(12)
        self.btn_pnm_paste_on_table = QPushButton("Paste on Table")
        self.btn_pnm_paste_on_table.setObjectName("collectorGrayBtn")
        clipboard_row.addWidget(self.btn_pnm_paste_on_table, 0)
        clipboard_row.addStretch(1)
        layout.addLayout(clipboard_row)

        layout.addSpacing(8)

        pnm_tracker_header = QHBoxLayout()
        pnm_tracker_header.setSpacing(0)
        pnm_tracker_header.addWidget(self.radioButton_pnm_generate_from_tracker)
        pnm_tracker_header.addSpacing(80)
        self.lbl_png_input_count = QLabel("")
        self.lbl_png_input_count.setObjectName("inputCountLabel")
        pnm_tracker_header.addWidget(self.lbl_png_input_count, 0)
        pnm_tracker_header.addStretch(1)
        layout.addLayout(pnm_tracker_header)

        self.packshot_tracker_row = QHBoxLayout()
        self.packshot_tracker_row.setSpacing(12)
        self.btn_pg6_excel_tracker = QPushButton("Excel Tracker")
        self.btn_pg6_excel_tracker.setObjectName("collectorGrayBtn")
        self.packshot_tracker_row.addWidget(self.btn_pg6_excel_tracker, 0)
        self.input_pg6_excel_tracker = QLineEdit()
        self.input_pg6_excel_tracker.setObjectName("collectorLineEdit")
        self.packshot_tracker_row.addWidget(self.input_pg6_excel_tracker, 1)
        layout.addLayout(self.packshot_tracker_row)

        output_row = QHBoxLayout()
        output_row.setSpacing(12)
        self.btn_pg6_output = QPushButton("Output Location")
        self.btn_pg6_output.setObjectName("collectorGrayBtn")
        output_row.addWidget(self.btn_pg6_output, 0)
        self.input_pg6_output = QLineEdit()
        self.input_pg6_output.setObjectName("collectorLineEdit")
        output_row.addWidget(self.input_pg6_output, 1)
        layout.addLayout(output_row)

        layout.addStretch(1)
        run_row = QHBoxLayout()
        run_row.addStretch(1)
        self.btn_run_process_generate_packshot_naming = QPushButton("Run Process")
        self.btn_run_process_generate_packshot_naming.setObjectName("collectorRunBtn")
        self.btn_run_process_generate_packshot_naming.setFixedHeight(50)
        self.btn_run_process_generate_packshot_naming.setMinimumWidth(180)
        run_row.addWidget(self.btn_run_process_generate_packshot_naming)
        run_row.addStretch(1)
        layout.addLayout(run_row)

        self.radioButton_pnm_from_clipboard.toggled.connect(self._sync_packshot_mode_ui)
        self.radioButton_pnm_generate_from_tracker.toggled.connect(self._sync_packshot_mode_ui)
        self.btn_pnm_paste_on_table.clicked.connect(self._open_packshot_clipboard_table)
        self.input_pg6_excel_tracker.textChanged.connect(lambda: self._update_png_input_count())
        self._sync_packshot_mode_ui()

        return page

    def _open_packshot_clipboard_table(self) -> None:
        self._packshot_clipboard_dialog = _PackshotClipboardTableDialog(self, start_dir=self._get_browse_dir("packshot"))
        if hasattr(self, "packshot_naming_generator"):
            self.packshot_naming_generator.attach_table_dialog(self._packshot_clipboard_dialog)
        self._packshot_clipboard_dialog.show()
        self._packshot_clipboard_dialog.raise_()
        self._packshot_clipboard_dialog.activateWindow()

    def _open_idh_dup_window(self) -> None:
        start_dir = self._get_browse_dir("packshot")
        if getattr(self, "radio_idh_dup_single", None) and self.radio_idh_dup_single.isChecked():
            self._idh_dup_window = _SingleTrackerWindow(self, start_dir=start_dir)
        else:
            self._idh_dup_window = _MultipleTrackerWindow(self, start_dir=start_dir)
        self._idh_dup_window.show()
        self._idh_dup_window.raise_()
        self._idh_dup_window.activateWindow()

    def _open_tsc_option1_table(self) -> None:
        self._tsc_option1_dialog = _TscOption1TableDialog(self, start_dir=self._get_browse_dir("tsc"))
        # Connect Import Trackers button to status_collector logic
        if hasattr(self, "status_collector"):
            self.status_collector.attach_tracker_window_dialog(self._tsc_option1_dialog)
        self._tsc_option1_dialog.show()
        self._tsc_option1_dialog.raise_()
        self._tsc_option1_dialog.activateWindow()

    def _open_mapper_option1_table(self) -> None:
        from pathlib import Path as _Path
        _export_dir = ""
        _is_project = (self.left_nav_stack.currentWidget() is self.project_left)
        if _is_project:
            _proj = self._get_active_project_folder()
            if _proj:
                _export_dir = str(_Path(_proj) / "SAP Data Compare")
        else:
            _export_dir = self._config.sap_reformat_input() or ""
        self._mapper_option1_dialog = _MapperOption1TableDialog(
            self, start_dir=self._get_browse_dir("sap_reformat"),
            export_dir=_export_dir)
        self._mapper_option1_dialog.show()
        self._mapper_option1_dialog.raise_()
        self._mapper_option1_dialog.activateWindow()

    def _sync_packshot_mode_ui(self) -> None:
        from_clipboard = getattr(self, "radioButton_pnm_from_clipboard", None)
        if from_clipboard is None:
            return
        is_clipboard = from_clipboard.isChecked()

        if hasattr(self, "btn_pnm_paste_on_table"):
            self.btn_pnm_paste_on_table.setEnabled(is_clipboard)
        if hasattr(self, "btn_pg6_excel_tracker"):
            self.btn_pg6_excel_tracker.setEnabled(not is_clipboard)
        if hasattr(self, "input_pg6_excel_tracker"):
            self.input_pg6_excel_tracker.setEnabled(not is_clipboard)
            if is_clipboard:
                self.input_pg6_excel_tracker.clear()
        if hasattr(self, "btn_pg6_output"):
            self.btn_pg6_output.setEnabled(not is_clipboard)
        if hasattr(self, "input_pg6_output"):
            self.input_pg6_output.setEnabled(not is_clipboard)
            if is_clipboard:
                self.input_pg6_output.clear()
        if hasattr(self, "btn_run_process_generate_packshot_naming"):
            self.btn_run_process_generate_packshot_naming.setEnabled(not is_clipboard)

    def _wrap_button_text(self, text: str, words_per_line: int = 2) -> str:
        words = text.split()
        lines = []
        for i in range(0, len(words), words_per_line):
            lines.append(" ".join(words[i : i + words_per_line]))
        return "\n".join(lines)

    def _get_browse_dir(self, tool: str) -> str:
        """Return the configured start directory for a file dialog.
        When the configure TSB is ON returns the tool's configured folder
        (falls back to Desktop if the path doesn't exist). When OFF returns
        an empty string so the OS uses its default last-used location."""
        if not self._use_root_folders:
            return ""
        from pathlib import Path
        desktop = str(Path.home() / "Desktop")

        if tool == "tsc":
            path = self._config.tsc_input()
            return path if path and Path(path).is_dir() else desktop

        if tool == "thumbnail":
            path = self._config.thumbnail_input()
            return path if path and Path(path).is_dir() else desktop

        is_project = (self.left_nav_stack.currentWidget() is self.project_left)

        if tool == "excel_library":
            path = self._config.excel_library_folder()
            return path if path and Path(path).is_dir() else desktop

        if tool == "sap_compare":
            if is_project:
                proj = self._get_active_project_folder()
                if proj:
                    p = Path(proj) / "SAP Data Reformat"
                    return str(p) if p.is_dir() else desktop
            else:
                path = self._config.sap_reformat_input()
                return path if path and Path(path).is_dir() else desktop

        if tool == "sap_reformat":
            if is_project:
                proj = self._get_active_project_folder()
                if proj:
                    p = Path(proj) / "SAP Data Reformat"
                    return str(p) if p.is_dir() else desktop
            else:
                path = self._config.sap_reformat_input()
                return path if path and Path(path).is_dir() else desktop

        if tool == "packshot":
            proj = self._get_active_project_folder()
            if proj:
                p = Path(proj) / "Packshot Naming Generator"
                return str(p) if p.is_dir() else desktop

        return desktop

    def _open_settings_dialog(self) -> None:
        btn = self.btn_settings
        btn.setStyleSheet(
            "QPushButton#iconBtn { background-color: #8A244B; color: #ffffff; "
            "border: 1px solid #8A244B; border-radius: 19px; font-size: 16px; font-weight: 700; padding: 0; }"
        )
        QTimer.singleShot(300, lambda: btn.setStyleSheet(""))
        dlg = _SettingsDialog(self._config, self)
        dlg.exec()

    def _init_modules(self) -> None:
        self._init_status_collector_module()
        self._init_thumbnail_generator_module()
        self._init_packshot_naming_module()

    def _init_status_collector_module(self) -> None:
        # Bridge line edits to legacy API expected by status_collector.py.
        self.textEdit_sc_selected_trackers = _PlainTextLineAdapter(self.input_trackers)
        self.textEdit_sc_output_location = _PlainTextLineAdapter(self.input_output)
        self.textEdit_status_collector_status_input = _PlainTextLineAdapter(self._input_status_line)

        # Optional clear-all control expected by the module; hidden in this UI.
        self.btn_menu_clear_all_fields = QPushButton("Clear All")
        self.btn_menu_clear_all_fields.setVisible(False)

        self.status_collector = StatusCollector(self)
        self.status_collector.bind_input_type_to_radioboxes()
        self.status_collector.run_process()

        # Override browse buttons to respect the configure TSB start directory.
        self.status_collector.btn_select_trackers.clicked.disconnect()
        self.status_collector.btn_select_trackers.clicked.connect(
            lambda: self.status_collector.get_tracker_files(
                self._get_browse_dir("tsc")))
        self.status_collector.btn_output_location.clicked.disconnect()
        self.status_collector.btn_output_location.clicked.connect(
            lambda: self.status_collector.get_output_location(
                self._get_browse_dir("tsc")))

    def _init_thumbnail_generator_module(self) -> None:
        self.thumbnail_generator = ThumbnailGenerator(self)
        self.thumbnail_generator.run_process()

        # Override browse button to respect the configure TSB start directory.
        self.thumbnail_generator.btn_images_folder.clicked.disconnect()
        self.thumbnail_generator.btn_images_folder.clicked.connect(
            lambda: self.thumbnail_generator.browse_images_folder(
                self._get_browse_dir("thumbnail")))

    def _init_packshot_naming_module(self) -> None:
        self.packshot_naming_generator = PackshotNamingGenerator(self)
        self.packshot_naming_generator.run_process()

        # Override browse buttons to respect the configure TSB start directory.
        if self.packshot_naming_generator.btn_excel_tracker is not None:
            self.packshot_naming_generator.btn_excel_tracker.clicked.disconnect()
            self.packshot_naming_generator.btn_excel_tracker.clicked.connect(
                lambda: self.packshot_naming_generator.name_gen_get_tracker_file(
                    self._get_browse_dir("packshot")))
        if self.packshot_naming_generator.btn_output_location is not None:
            self.packshot_naming_generator.btn_output_location.clicked.disconnect()
            self.packshot_naming_generator.btn_output_location.clicked.connect(
                lambda: self.packshot_naming_generator.name_gen_get_output_location(
                    self._get_browse_dir("packshot")))

    def _stylesheet(self) -> str:
        return """
        #root {
            background-color: #E5E5E5;
        }

        #shell {
            background-color: #E5E5E5;
            border-radius: 0px;
            border: none;
        }

        #topHeaderBtn {
            background-color: transparent;
            color: #ffffff;
            border: none;
            border-radius: 22px;
            padding: 0 18px;
            font-family: "Bahnschrift SemiCondensed", "Arial Narrow", "Segoe UI";
            font-size: 14px;
            font-weight: 800;
        }

        #topHeaderBtn:hover {
            background-color: transparent;
            color: #ffffff;
            border: none;
        }

        #topHeaderBtn:checked {
            background-color: #F63049;
            border: none;
            color: #ffffff;
        }

        #topHeaderBtn:checked:hover {
            background-color: #F63049;
            border: none;
            color: #ffffff;
        }

        #iconBtn {
            background-color: #171b21;
            color: #ffffff;
            border: 1px solid #4b5664;
            border-radius: 19px;
            font-size: 16px;
            font-weight: 700;
            padding: 0;
        }

        #iconBtn:hover {
            border: 1px solid #707c8b;
        }

        #contentCard {
            background-color: #E5E5E5;
            border: 1px solid #C0C0C0;
            border-radius: 18px;
        }

        #homeCard {
            background-color: #FFFFFF;
            border: 1px solid #C0C0C0;
            border-radius: 18px;
        }

        #headerBand {
            background-color: #000000;
            border: 1px solid #1F1F1F;
            border-radius: 12px;
        }

        #sectionTitle {
            color: #111F35;
            font-family: "Segoe UI";
            font-size: 21px;
            font-weight: 900;
            letter-spacing: 1px;
        }

        #projectSelectorBand {
            background-color: #9EA3AB;
            border: 1px solid #8B9098;
            border-radius: 8px;
        }

        #projectSelectorLabel {
            color: #000000;
            font-family: "Segoe UI";
            font-size: 13px;
            font-weight: 600;
            background: transparent;
        }

        #projectSelectorCombo {
            background-color: transparent;
            color: #000000;
            border: none;
            padding: 0 4px;
            font-family: "Segoe UI";
            font-size: 13px;
            font-weight: 700;
        }

        #projectSelectorCombo::drop-down {
            border: none;
            width: 22px;
        }

        #projectSelectorCombo QAbstractItemView {
            background-color: #9EA3AB;
            color: #000000;
            selection-background-color: #6B7280;
            selection-color: #FFFFFF;
            border: 1px solid #8B9098;
        }

        #basicToolActionBtn {
            background-color: #8A244B;
            color: #ffffff;
            border: none;
            border-radius: 16px;
            padding: 10px;
            text-align: center;
            font-family: "Roboto", "Segoe UI", Arial, sans-serif;
            font-size: 14px;
            font-weight: 700;
            letter-spacing: 0.2px;
        }

        #basicToolActionBtn:hover {
            background-color: #8A244B;
        }

        #basicToolActionBtn:checked {
            background-color: #D02752;
        }

        #basicToolActionBtn:pressed {
            background-color: #D02752;
        }

        #basicToolsRightPanel {
            background-color: #EDEDED;
            border: 1px solid #CDCDCD;
            border-radius: 18px;
        }

        #collectorTitle {
            color: #8A244B;
            font-family: "Segoe UI";
            font-size: 30px;
            font-weight: 800;
        }

        #collectorSectionLabel {
            color: #111F35;
            font-family: "Segoe UI";
            font-size: 15px;
            font-weight: 700;
        }

        #collectorDescription {
            color: #000000;
            font-family: "Segoe UI";
            font-size: 14px;
            font-weight: 500;
        }

        #thumbnailTitle {
            color: #8A244B;
            font-family: "Segoe UI";
            font-size: 30px;
            font-weight: 800;
        }

        #thumbnailDescription {
            color: #000000;
            font-family: "Segoe UI";
            font-size: 14px;
            font-weight: 500;
        }

        #packshotTitle {
            color: #8A244B;
            font-family: "Segoe UI";
            font-size: 30px;
            font-weight: 800;
        }

        #packshotDescription {
            color: #000000;
            font-family: "Segoe UI";
            font-size: 14px;
            font-weight: 500;
        }

        #mapperTitle {
            color: #8A244B;
            font-family: "Segoe UI";
            font-size: 30px;
            font-weight: 800;
        }

        #mapperDescription {
            color: #000000;
            font-family: "Segoe UI";
            font-size: 14px;
            font-weight: 500;
        }

        #collectorGrayBtn {
            background-color: #9EA3AB;
            color: #000000;
            border: 1px solid #8B9098;
            border-radius: 12px;
            padding: 0 14px;
            min-height: 42px;
            font-family: "Segoe UI";
            font-size: 14px;
            font-weight: 600;
        }

        #collectorGrayBtn:hover {
            background-color: #9EA3AB;
            color: #000000;
            border: 1px solid #8B9098;
        }

        #collectorGrayBtn:pressed {
            background-color: #111F35;
            color: #ffffff;
            border: 1px solid #111F35;
        }

        #collectorLineEdit {
            background-color: #D3D3D3;
            color: #000000;
            border: 1px solid #6F6F6F;
            border-radius: 16px;
            min-height: 40px;
            padding: 0 12px;
            font-family: "Segoe UI";
            font-size: 14px;
        }

        QLineEdit#comparePathLabel {
            background-color: #D3D3D3;
            color: #000000;
            border: 1px solid #6F6F6F;
            border-radius: 11px;
            padding: 0 8px;
            font-family: "Segoe UI";
            font-size: 10px;
        }

        QLineEdit#comparePathLabel:disabled {
            background-color: #E8E8E8;
            color: #aaaaaa;
            border: 1px solid #c0c0c0;
        }

        #compareSmallLabel {
            color: #000000;
            font-family: "Segoe UI";
            font-size: 11px;
        }

        QLineEdit#compareSmallInput {
            background-color: #D3D3D3;
            color: #000000;
            border: 1px solid #6F6F6F;
            border-radius: 6px;
            padding: 0 6px;
            font-family: "Segoe UI";
            font-size: 11px;
        }

        QLineEdit#compareSmallInput:disabled {
            background-color: #E8E8E8;
            color: #aaaaaa;
            border: 1px solid #c0c0c0;
        }

        #compareGrayBtn {
            background-color: #9EA3AB;
            color: #000000;
            border: 1px solid #8B9098;
            border-radius: 8px;
            padding: 0 10px;
            min-height: 34px;
            font-family: "Segoe UI";
            font-size: 11px;
            font-weight: 600;
        }

        #compareGrayBtn:hover { background-color: #ACB1B8; }

        #compareGrayBtn:pressed {
            background-color: #111F35;
            color: #ffffff;
            border: 1px solid #111F35;
        }

        #compareGrayBtn:disabled {
            background-color: #C8CBD0;
            color: #888888;
            border: 1px solid #B0B3B8;
        }

        #compareRunBtn {
            background-color: #111F35;
            color: #ffffff;
            border: none;
            border-radius: 8px;
            padding: 0 14px;
            font-family: "Segoe UI";
            font-size: 13px;
            font-weight: 700;
        }

        #compareRunBtn:pressed { background-color: #D02752; }

        #compareProgressBar {
            background-color: #C4C7CC;
            border: 1px solid #8D939D;
            border-radius: 5px;
        }

        #compareProgressBar::chunk {
            background-color: #111F35;
            border-radius: 5px;
            margin: 0px;
        }

        #collectorModeRadio {
            color: #8A244B;
            font-family: "Segoe UI";
            font-size: 16px;
            font-weight: 700;
            spacing: 8px;
        }

        #inputCountLabel {
            color: #8A244B;
            font-family: "Segoe UI";
            font-size: 13px;
            font-weight: 600;
        }

        #collectorModeRadio::indicator {
            width: 18px;
            height: 18px;
            border: 1px solid #8A244B;
            border-radius: 9px;
            background: transparent;
        }

        #collectorModeRadio::indicator:checked {
            background: #8A244B;
            border: 1px solid #8A244B;
        }

        #idhDupRadio {
            color: #8A244B;
            font-family: "Segoe UI";
            font-size: 13px;
            font-weight: 400;
            spacing: 8px;
        }

        #idhDupRadio::indicator {
            width: 14px;
            height: 14px;
            border: 1px solid #8A244B;
            border-radius: 7px;
            background: transparent;
        }

        #idhDupRadio::indicator:checked {
            background: #8A244B;
            border: 1px solid #8A244B;
        }

        #compareModeRadioSmall {
            color: #8A244B;
            font-family: "Segoe UI";
            font-size: 13px;
            font-weight: 400;
            spacing: 6px;
        }

        #compareModeRadioSmall::indicator {
            width: 14px;
            height: 14px;
            border: 1px solid #8A244B;
            border-radius: 7px;
            background: transparent;
        }

        #compareModeRadioSmall::indicator:checked {
            background: #8A244B;
            border: 1px solid #8A244B;
        }

        #compareModeRadioSmall:disabled {
            color: #aaaaaa;
        }

        #compareModeRadioSmall::indicator:disabled {
            border: 1px solid #aaaaaa;
        }

        #collectorCleanupRadio {
            color: #000000;
            font-family: "Segoe UI";
            font-size: 14px;
            font-weight: 500;
            spacing: 8px;
        }

        #collectorCleanupRadio::indicator {
            width: 18px;
            height: 18px;
            border: 1px solid #6F6F6F;
            border-radius: 9px;
            background: #D3D3D3;
        }

        #collectorCleanupRadio::indicator:checked {
            background: #111F35;
            border: 1px solid #111F35;
        }

        #collectorCheck {
            color: #000000;
            font-family: "Segoe UI";
            font-size: 14px;
            font-weight: 500;
            spacing: 8px;
        }

        #collectorCheck::indicator {
            width: 18px;
            height: 18px;
            border: 1px solid #6F6F6F;
            border-radius: 6px;
            background: #D3D3D3;
        }

        #collectorCheck::indicator:checked {
            background: #111F35;
            border: 1px solid #111F35;
        }

        #collectorCheckSmall {
            color: #000000;
            font-family: "Segoe UI";
            font-size: 11px;
            font-weight: 500;
            spacing: 5px;
        }

        #collectorCheckSmall::indicator {
            width: 14px;
            height: 14px;
            border: 1px solid #6F6F6F;
            border-radius: 4px;
            background: #D3D3D3;
        }

        #collectorCheckSmall::indicator:checked {
            background: #111F35;
            border: 1px solid #111F35;
        }

        #collectorRunBtn {
            background-color: #111F35;
            color: #ffffff;
            border: none;
            border-radius: 12px;
            padding: 0 18px;
            font-family: "Segoe UI";
            font-size: 16px;
            font-weight: 700;
        }

        #collectorRunBtn:pressed {
            background-color: #D02752;
            color: #ffffff;
        }

        #collectorProgressBar {
            background-color: #C4C7CC;
            border: 1px solid #8D939D;
            border-radius: 7px;
        }

        #collectorProgressBar::chunk {
            background-color: #111F35;
            border-radius: 7px;
            margin: 0px;
        }

        #hatBadge {
            background: transparent;
            color: #ff1e1e;
            border: none;
            font-family: "Segoe UI";
            font-size: 40px;
            font-weight: 900;
            letter-spacing: 1px;
            padding: 0;
        }
        """
